import json
import os.path
import openpyxl as xl

from openpyxl.styles import Border, Side

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service

from webdriver_manager.chrome import ChromeDriverManager

config = json.load(open('config.json', encoding='UTF8'))

if True:
# if not os.path.isfile('./data/'+config['dailyTestFileName']):
    print('Making DailyTest Result Form...')

    iniWb = xl.Workbook()
    iniWs = iniWb.active
    iniWs.title = 'DailyTestForm'
    iniWs['A1'] = '요일'
    iniWs['B1'] = '시간'
    iniWs['C1'] = '반'
    iniWs['D1'] = '이름'
    iniWs['E1'] = '담당T'
    iniWs['F1'] = '시험명'
    iniWs['G1'] = '점수'
    iniWs['H1'] = '평균'
    iniWs['I1'] = '기타 시험명'
    iniWs['J1'] = '기타 시험 점수'
    iniWs['K1'] = '기타 시험 평균'
    iniWs.auto_filter.ref = 'A:B'

    classWb = xl.load_workbook("class.xlsx")
    classWs = classWb.active

    options = webdriver.ChromeOptions()
    options.add_argument('headless')
    driver = webdriver.Chrome(service = Service(ChromeDriverManager().install()), options = options)

    # 아이소식 접속
    driver.get(config['url'])
    tableNames = driver.find_elements(By.CLASS_NAME, 'style1')

    #반 루프
    for i in range(3, len(tableNames)):
        trs = driver.find_element(By.ID, 'table_' + str(i)).find_elements(By.CLASS_NAME, 'style12')
        writeLocation = start = iniWs.max_row + 1

        className = tableNames[i].text.split('(')[0].rstrip()

        iniWs.cell(writeLocation, 3).value = className
        for j in range(2, classWs.max_row + 1):
            if classWs.cell(j, 1).value == className:
                teacher = classWs.cell(j, 2).value
                date = classWs.cell(j, 3).value
                time = classWs.cell(j, 4).value
        iniWs.cell(writeLocation, 5).value = teacher
        #학생 루프
        for tr in trs:
            iniWs.cell(writeLocation, 1).value = date
            iniWs.cell(writeLocation, 2).value = time
            iniWs.cell(writeLocation, 4).value = tr.find_element(By.CLASS_NAME, 'style9').text
            writeLocation = iniWs.max_row + 1
        
        end = writeLocation - 1

        # 시험 평균
        iniWs.cell(start, 8).value = '=ROUND(AVERAGE(G' + str(start) + ':G' + str(end) + '), 0)'
        # 기타 시험 평균
        iniWs.cell(start, 11).value = '=ROUND(AVERAGE(J' + str(start) + ':J' + str(end) + '), 0)'
        
        iniWs.merge_cells('C' + str(start) + ':C' + str(end))
        iniWs.merge_cells('E' + str(start) + ':E' + str(end))
        iniWs.merge_cells('F' + str(start) + ':F' + str(end))
        iniWs.merge_cells('H' + str(start) + ':H' + str(end))
        iniWs.merge_cells('K' + str(start) + ':K' + str(end))
        

    iniWb.save('./dailyTestForm.xlsx')
    print('done')

else:
    print('이미 파일이 존재합니다')