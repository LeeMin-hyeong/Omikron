import os
import json
import win32com.client
import openpyxl as xl
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Border, Color, PatternFill, Side

config = json.load(open('config.json', encoding='UTF8'))

def applyColor():
    if not os.path.isfile('./data/' + config['dataFileName'] + '.xlsx'):
        print('[오류]'+ config['dataFileName'] +'.xlsx 파일이 존재하지 않습니다.')
        return
    
    excel = win32com.client.Dispatch('Excel.Application')
    excel.Visible = False
    wb = excel.Workbooks.Open(os.getcwd() + '\\data\\' + config['dataFileName'] + '.xlsx')
    wb.Save()
    wb.Close()

    dataFileWb = xl.load_workbook('./data/' + config['dataFileName'] + '.xlsx')
    dataFileColorWb = xl.load_workbook('./data/' + config['dataFileName'] + '.xlsx', data_only=True)
    for sheetName in dataFileWb.sheetnames:
        dataFileWs = dataFileWb[sheetName]
        dataFileColorWs = dataFileColorWb[sheetName]
        for i in range(2, dataFileColorWs.max_row+1):
            if i > 6:
                dataFileWs.column_dimensions[get_column_letter(i)].width = 14
            if dataFileColorWs.cell(i, 5).value is None:
                break
            for j in range(7, dataFileColorWs.max_column+1):
                if dataFileWs.cell(i, 5).value == '시험 평균' and dataFileWs(i, j).value is not None:
                    dataFileWs.cell(i, j).border = Border(bottom=Side(border_style='medium', color='000000'))
                if dataFileWs.cell(i, 5).value == '날짜' and dataFileWs(i, j).value is not None:
                    dataFileWs.cell(i, j).border = Border(top=Side(border_style='medium', color='000000'))
                # 입력 데이터 조건부 서식
                if type(dataFileColorWs.cell(i, j).value) == int:
                    if dataFileColorWs.cell(i, j).value < 60:
                        dataFileWs.cell(i, j).fill = PatternFill(fill_type='solid', fgColor=Color('EC7E31'))
                    elif dataFileColorWs.cell(i, j).value < 70:
                        dataFileWs.cell(i, j).fill = PatternFill(fill_type='solid', fgColor=Color('F5AF85'))
                    elif dataFileColorWs.cell(i, j).value < 80:
                        dataFileWs.cell(i, j).fill = PatternFill(fill_type='solid', fgColor=Color('FCE4D6'))
                    elif dataFileColorWs.cell(i, 5).value == '시험 평균':
                        dataFileWs.cell(i, j).fill = PatternFill(fill_type='solid', fgColor=Color('DDEBF7'))
                    else:
                        dataFileWs.cell(i, j).fill = PatternFill(fill_type=None, fgColor=Color('00FFFFFF'))


                # 학생별 평균 조건부 서식
                if type(dataFileColorWs.cell(i, 6).value) == int:
                    if dataFileColorWs.cell(i, 6).value < 60:
                        dataFileWs.cell(i, 6).fill = PatternFill(fill_type='solid', fgColor=Color('EC7E31'))
                    elif dataFileColorWs.cell(i, 6).value < 70:
                        dataFileWs.cell(i, 6).fill = PatternFill(fill_type='solid', fgColor=Color('F5AF85'))
                    elif dataFileColorWs.cell(i, 6).value < 80:
                        dataFileWs.cell(i, 6).fill = PatternFill(fill_type='solid', fgColor=Color('FCE4D6'))
                    elif dataFileColorWs.cell(i, 5).value == '시험 평균':
                        dataFileWs.cell(i, 6).fill = PatternFill(fill_type='solid', fgColor=Color('DDEBF7'))
                    else:
                        dataFileWs.cell(i, 6).fill = PatternFill(fill_type='solid', fgColor=Color('E2EFDA'))

    dataFileWb.save('./data/' + config['dataFileName'] + '.xlsx')
    excel.Visible = True
    wb = excel.Workbooks.Open(os.getcwd() + '\\data\\' + config['dataFileName'] + '.xlsx')

applyColor()