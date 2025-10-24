import os
import openpyxl as xl
import pythoncom  
import win32com.client

from datetime import datetime
from openpyxl.utils.cell import get_column_letter as gcl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Border, Side

import omikron.chrome
import omikron.config

from omikron.defs import ClassInfo
from omikron.exception import NoMatchingSheetException, FileOpenException


# 파일 기본 작업
def make_file():
    wb = xl.Workbook()
    ws = wb.worksheets[0]
    ws.title = ClassInfo.DEFAULT_NAME
    ws[gcl(ClassInfo.CLASS_NAME_COLUMN)+"1"]    = "반명"
    ws[gcl(ClassInfo.TEACHER_NAME_COLUMN)+"1"]  = "선생님명"
    ws[gcl(ClassInfo.CLASS_WEEKDAY_COLUMN)+"1"] = "요일"
    ws[gcl(ClassInfo.TEST_TIME_COLUMN)+"1"]     = "시간"

    ws.freeze_panes = "A2"

    # 반 루프
    for class_name in omikron.chrome.get_class_names():
        WRITE_LOCATION = ws.max_row + 1
        ws.cell(WRITE_LOCATION, 1).value = class_name

    # 정렬 및 테두리
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row, col).border    = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    save(wb)


def open(data_only:bool=True) -> xl.Workbook:
    return xl.load_workbook(f"{omikron.config.DATA_DIR}/{ClassInfo.DEFAULT_NAME}.xlsx", data_only=data_only)

def open_temp(data_only:bool=True) -> xl.Workbook:
    return xl.load_workbook(f"{omikron.config.DATA_DIR}/{ClassInfo.TEMP_FILE_NAME}.xlsx", data_only=data_only)

def open_worksheet(wb:xl.Workbook):
    try:
        return wb[ClassInfo.DEFAULT_NAME]
    except:
        raise NoMatchingSheetException(f"'{ClassInfo.DEFAULT_NAME}.xlsx'의 시트명을 '{ClassInfo.DEFAULT_NAME}'로 변경해 주세요.")

def save(wb:xl.Workbook):
    try:
        wb.save(f"{omikron.config.DATA_DIR}/{ClassInfo.DEFAULT_NAME}.xlsx")
    except:
        raise FileOpenException(f"{ClassInfo.DEFAULT_NAME} 파일을 닫은 뒤 다시 시도해주세요")

def save_to_temp(wb:xl.Workbook):
    wb.save(f"{omikron.config.DATA_DIR}/{ClassInfo.TEMP_FILE_NAME}.xlsx")

def delete_temp():
    os.remove(f"{omikron.config.DATA_DIR}/{ClassInfo.TEMP_FILE_NAME}.xlsx")

def isopen() -> bool:
    return os.path.isfile(f"{omikron.config.DATA_DIR}/~${ClassInfo.DEFAULT_NAME}.xlsx")

# 파일 유틸리티
def make_backup_file():
    wb = open()
    wb.save(f"{omikron.config.DATA_DIR}/data/backup/{ClassInfo.DEFAULT_NAME}({datetime.today().strftime('%Y%m%d')}).xlsx")

def get_class_info(class_name:str, ws:Worksheet = None):
    """
    반 정보 파일로부터 특정 반의 정보 추출

    return `반 정보 존재 여부`, `담당 선생님`, `수업 요일`, `테스트 응시 시간`
    """
    if ws is None:
        wb = open()
        ws = open_worksheet(wb)

    for row in range(2, ws.max_row + 1):
        if ws.cell(row, ClassInfo.CLASS_NAME_COLUMN).value == class_name:
            teacher_name  = ws.cell(row, ClassInfo.TEACHER_NAME_COLUMN).value
            class_weekday = ws.cell(row, ClassInfo.CLASS_WEEKDAY_COLUMN).value
            test_time     = ws.cell(row, ClassInfo.TEST_TIME_COLUMN).value
            break
    else:
        return False, None, None, None
    
    return True, teacher_name, class_weekday, test_time

def get_class_names(ws:Worksheet = None) -> list[str]:
    """
    반 정보 기준 반 이름 리스트 추출
    """
    if ws is None:
        wb = open()
        ws = open_worksheet(wb)

    class_names = []
    for row in range(2, ws.max_row + 1):
        class_name = ws.cell(row, ClassInfo.CLASS_NAME_COLUMN).value
        if class_name is not None:
            class_names.append(class_name)

    return sorted(class_names)

def get_new_class_names():
    """
    임시 반 정보 파일에서 새 반 리스트를 리턴
    """
    temp_wb = open_temp()
    temp_ws = open_worksheet(temp_wb)

    return get_class_names(temp_ws)

# 파일 작업
def make_temp_file_for_update(new_class_list:list[str]):
    """
    반 업데이트 작업에 필요한 임시 반 정보 파일 생성

    등록되지 않은 반을 반 리스트 최하단에 작성
    """
    make_backup_file()

    wb = open()
    ws = open_worksheet(wb)

    class_names = get_class_names(ws)

    unregistered_class_names = sorted(list(set(new_class_list).difference(class_names)))

    for row in range(2, ws.max_row+1):
        while ws.cell(row, ClassInfo.CLASS_NAME_COLUMN).value is not None and ws.cell(row, ClassInfo.CLASS_NAME_COLUMN).value not in new_class_list:
            ws.delete_rows(row)

    temp_path = os.path.abspath(f'{omikron.config.DATA_DIR}/{ClassInfo.TEMP_FILE_NAME}.xlsx')

    if len(unregistered_class_names) == 0:
        save_to_temp(wb)
        return temp_path

    for row in range(ws.max_row+1, 1, -1):
        if ws.cell(row-1, ClassInfo.CLASS_NAME_COLUMN).value is not None:
            WRITE_RANGE = WRITE_ROW = row
            break

    for row, class_name in enumerate(unregistered_class_names, start=WRITE_ROW):
        ws.cell(row, ClassInfo.CLASS_NAME_COLUMN).value = class_name

    for row in range(WRITE_RANGE, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row, col).border    = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    save_to_temp(wb)

    return temp_path

def change_class_info(target_class_name:str, target_teacher_name:str):
    """
    특정 반의 담당 선생님 변경
    """
    make_backup_file()

    wb = open()
    ws = open_worksheet(wb)

    for row in range(2, ws.max_row + 1):
        if ws.cell(row, ClassInfo.CLASS_NAME_COLUMN).value == target_class_name:
            ws.cell(row, ClassInfo.TEACHER_NAME_COLUMN).value = target_teacher_name
            break
    else:
        raise Exception(f"'{target_class_name}' 반이 존재하지 않습니다.")

    save(wb)

def update_class():
    save(open_temp())
