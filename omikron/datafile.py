import os
import openpyxl as xl
import pythoncom       # only works in Windows
import win32com.client # only works in Windows

from datetime import datetime
from openpyxl.utils.cell import get_column_letter as gcl
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet

import omikron.chrome
import omikron.classinfo
import omikron.config
import omikron.dataform
import omikron.studentinfo

from omikron.defs import DataFile, DataForm
from omikron.exception import NoMatchingSheetException, FileOpenException
from omikron.util import copy_cell, class_average_color, student_average_color, test_score_color
from omikron.progress import Progress
from omikron.style import *

class NoReservedColumnError(Exception):
    """
    예약된 열이 없을 경우
    """
    pass

# 파일 기본 작업
def make_file():
    wb = xl.Workbook()
    ws = wb.worksheets[0]
    ws.title = DataFile.DEFAULT_SHEET_NAME
    # ws[gcl(DataFile.TEST_TIME_COLUMN)+"1"]     = "시간"
    # ws[gcl(DataFile.CLASS_WEEKDAY_COLUMN)+"1"] = "요일"
    ws[gcl(DataFile.CLASS_NAME_COLUMN)+"1"]    = "반"
    ws[gcl(DataFile.TEACHER_NAME_COLUMN)+"1"]  = "담당"
    ws[gcl(DataFile.STUDENT_NAME_COLUMN)+"1"]  = "이름"
    ws[gcl(DataFile.AVERAGE_SCORE_COLUMN)+"1"] = "학생 평균"
    ws.freeze_panes    = f"{gcl(DataFile.DATA_COLUMN)}2"
    ws.auto_filter.ref = f"A:{gcl(DataFile.MAX)}"

    for col in range(1, DataFile.DATA_COLUMN):
        ws.cell(1, col).border = BORDER_BOTTOM_MEDIUM_000

    class_wb = omikron.classinfo.open(True)
    class_ws = omikron.classinfo.open_worksheet(class_wb)

    # 반 루프
    for class_name, student_list in omikron.chrome.get_class_student_dict().items():
        if len(student_list) == 0:
            continue

        exist, teacher_name, _, _, mock_test_check = omikron.classinfo.get_class_info(class_name, ws=class_ws)
        if not exist: continue

        for i in range(2):
            if i == 1 and not mock_test_check:
                continue

            if i == 1:
                class_name = class_name + " (모의고사)"

            WRITE_LOCATION = ws.max_row + 1

            # 시험명
            # ws.cell(WRITE_LOCATION, DataFile.TEST_TIME_COLUMN).value     = test_time
            # ws.cell(WRITE_LOCATION, DataFile.CLASS_WEEKDAY_COLUMN).value = class_weekday
            ws.cell(WRITE_LOCATION, DataFile.CLASS_NAME_COLUMN).value    = class_name
            ws.cell(WRITE_LOCATION, DataFile.TEACHER_NAME_COLUMN).value  = teacher_name
            ws.cell(WRITE_LOCATION, DataFile.STUDENT_NAME_COLUMN).value  = "날짜"
            
            WRITE_LOCATION = ws.max_row + 1
            # ws.cell(WRITE_LOCATION, DataFile.TEST_TIME_COLUMN).value     = test_time
            # ws.cell(WRITE_LOCATION, DataFile.CLASS_WEEKDAY_COLUMN).value = class_weekday
            ws.cell(WRITE_LOCATION, DataFile.CLASS_NAME_COLUMN).value    = class_name
            ws.cell(WRITE_LOCATION, DataFile.TEACHER_NAME_COLUMN).value  = teacher_name
            ws.cell(WRITE_LOCATION, DataFile.STUDENT_NAME_COLUMN).value  = "시험명"

            for col in range(1, DataFile.DATA_COLUMN):
                ws.cell(WRITE_LOCATION, col).border = BORDER_BOTTOM_THIN_9090

            class_start = WRITE_LOCATION + 1

            # 학생 루프
            for student_name in student_list:
                WRITE_LOCATION = ws.max_row + 1
                # ws.cell(WRITE_LOCATION, DataFile.TEST_TIME_COLUMN).value     = test_time
                # ws.cell(WRITE_LOCATION, DataFile.CLASS_WEEKDAY_COLUMN).value = class_weekday
                ws.cell(WRITE_LOCATION, DataFile.CLASS_NAME_COLUMN).value    = class_name
                ws.cell(WRITE_LOCATION, DataFile.TEACHER_NAME_COLUMN).value  = teacher_name
                ws.cell(WRITE_LOCATION, DataFile.STUDENT_NAME_COLUMN).value  = student_name
                ws.cell(WRITE_LOCATION, DataFile.AVERAGE_SCORE_COLUMN).value = f"=ROUND(AVERAGE({gcl(DataFile.DATA_COLUMN)}{WRITE_LOCATION}:XFD{WRITE_LOCATION}), 0)"
                ws.cell(WRITE_LOCATION, DataFile.AVERAGE_SCORE_COLUMN).font  = FONT_BOLD
            
            # 시험별 평균
            class_end = WRITE_LOCATION
            WRITE_LOCATION = ws.max_row + 1
            # ws.cell(WRITE_LOCATION, DataFile.TEST_TIME_COLUMN).value     = test_time
            # ws.cell(WRITE_LOCATION, DataFile.CLASS_WEEKDAY_COLUMN).value = class_weekday
            ws.cell(WRITE_LOCATION, DataFile.CLASS_NAME_COLUMN).value    = class_name
            ws.cell(WRITE_LOCATION, DataFile.TEACHER_NAME_COLUMN).value  = teacher_name
            ws.cell(WRITE_LOCATION, DataFile.STUDENT_NAME_COLUMN).value  = "시험 평균"
            ws[f"{gcl(DataFile.AVERAGE_SCORE_COLUMN)}{WRITE_LOCATION}"] = ArrayFormula(f"{gcl(DataFile.AVERAGE_SCORE_COLUMN)}{WRITE_LOCATION}", f"=ROUND(AVERAGE(IFERROR({gcl(DataFile.AVERAGE_SCORE_COLUMN)}{class_start}:{gcl(DataFile.AVERAGE_SCORE_COLUMN)}{class_end}, \"\")), 0)")
            ws.cell(WRITE_LOCATION, DataFile.AVERAGE_SCORE_COLUMN).font = FONT_BOLD

            for col in range(1, DataFile.DATA_COLUMN):
                ws.cell(WRITE_LOCATION, col).border = BORDER_TOP_THIN_9090_BOTTOM_MEDIUM_000

    # 정렬
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).alignment = ALIGN_CENTER
    
    # 모의고사 sheet 생성
    # copy_ws                 = wb.copy_worksheet(wb[DataFile.DEFAULT_SHEET_NAME])
    # copy_ws.title           = DataFile.SECOND_SHEET_NAME
    # copy_ws.freeze_panes    = f"{gcl(DataFile.DATA_COLUMN)}2"
    # copy_ws.auto_filter.ref = f"A:{gcl(DataFile.MAX)}"

    save(wb)

def open(data_only:bool=False, read_only:bool=False) -> xl.Workbook:
    return xl.load_workbook(f"{omikron.config.DATA_DIR}/data/{omikron.config.DATA_FILE_NAME}.xlsx", data_only=data_only, read_only=read_only)

def open_temp(data_only:bool=False, read_only:bool=False) -> xl.Workbook:
    return xl.load_workbook(f"{omikron.config.DATA_DIR}/data/{DataFile.TEMP_FILE_NAME}.xlsx", data_only=data_only, read_only=read_only)

def save(wb:xl.Workbook):
    try:
        wb.save(f"{omikron.config.DATA_DIR}/data/{omikron.config.DATA_FILE_NAME}.xlsx")
    except:
        raise FileOpenException(f"{omikron.config.DATA_FILE_NAME} 파일을 닫은 뒤 다시 시도해주세요")

def save_to_temp(wb:xl.Workbook):
    wb.save(f"{omikron.config.DATA_DIR}/data/{DataFile.TEMP_FILE_NAME}.xlsx")
    os.system(f"attrib +h {omikron.config.DATA_DIR}/data/{DataFile.TEMP_FILE_NAME}.xlsx")

def delete_temp():
    os.remove(f"{omikron.config.DATA_DIR}/data/{DataFile.TEMP_FILE_NAME}.xlsx")

def isopen() -> bool:
    return os.path.isfile(f"{omikron.config.DATA_DIR}/data/~${omikron.config.DATA_FILE_NAME}.xlsx")

def file_validation():
    wb = open(read_only=True)

    if DataFile.DEFAULT_SHEET_NAME not in wb.sheetnames:
        raise NoMatchingSheetException(f"데이터 파일: {DataFile.DEFAULT_SHEET_NAME} 시트가 존재하지 않습니다.")

    wb.close()

# 파일 유틸리티
def make_backup_file():
    wb = open()
    wb.save(f"{omikron.config.DATA_DIR}/data/backup/{omikron.config.DATA_FILE_NAME}({datetime.today().strftime('%Y%m%d%H%M%S')}).xlsx")

def get_data_sorted_dict(mocktest = False):
    """
    데이터 파일의 대략적 정보를 `dict` 형태로 추출

    return `dict[반:학생]`, `dict[반:시험명]`
    """
    wb = open()

    ws = wb[DataFile.DEFAULT_SHEET_NAME]

    CLASS_NAME_COLUMN, _, STUDENT_NAME_COLUMN, AVERAGE_SCORE_COLUMN = find_dynamic_columns(ws)

    class_wb = omikron.classinfo.open()
    class_ws = omikron.classinfo.open_worksheet(class_wb)

    class_student_dict = {}
    class_test_dict    = {}

    for class_name in omikron.classinfo.get_class_names(class_ws, mocktest=mocktest):
        student_index_dict = {}
        test_index_dict    = {}
        for row in range(2, ws.max_row+1):
            if ws.cell(row, CLASS_NAME_COLUMN).value != class_name:
                continue
            if ws.cell(row, STUDENT_NAME_COLUMN).value == "날짜":
                for col in range(AVERAGE_SCORE_COLUMN+1, ws.max_row+1):
                    test_date = ws.cell(row, col).value
                    test_name = ws.cell(row+1, col).value
                    if test_date is None and test_name is None:
                        break
                    if type(test_date) == datetime:
                        test_date = test_date.strftime("%y.%m.%d")
                    else:
                        test_date = str(test_date).split()[0][2:10].replace("-", ".").replace(",", ".").replace("/", ".")
                    test_index_dict[f"[{test_date}] {test_name}"] = col
                continue
            if ws.cell(row, STUDENT_NAME_COLUMN).value in ("시험명", "시험 평균"):
                continue
            if ws.cell(row, STUDENT_NAME_COLUMN).font.strike:
                continue
            if ws.cell(row, STUDENT_NAME_COLUMN).font.color is not None and ws.cell(row, STUDENT_NAME_COLUMN).font.color.rgb == "FFFF0000":
                continue
            student_index_dict[ws.cell(row, STUDENT_NAME_COLUMN).value] = row

        test_index_dict = dict(sorted(test_index_dict.items(), reverse=True))

        class_student_dict[class_name] = student_index_dict
        class_test_dict[class_name]    = test_index_dict

    class_student_dict = dict(sorted(class_student_dict.items()))

    return class_student_dict, class_test_dict

def find_dynamic_columns(ws:Worksheet):
    """
    파일 열(column) 정보 동적 탐색

    '반' 열, '담당' 열, '이름' 열, '학생 평균' 열

    return `CLASS_NAME_COLUMN`, `TEACHER_NAME_COLUMN`, `STUDENT_NAME_COLUMN`, `AVERAGE_SCORE_COLUMN`
    """

    # for col in range(1, ws.max_column+1):
    #     if ws.cell(1, col).value == "시간":
    #         TEST_TIME_COLUMN = col
    #         break
    # else:
    #     raise NoReservedColumnError(f"{ws.title} 시트에 '시간' 열이 없습니다.")

    # for col in range(1, ws.max_column+1):
    #     if ws.cell(1, col).value == "요일":
    #         CLASS_WEEKDAY_COLUMN = col
    #         break
    # else:
    #     raise NoReservedColumnError(f"{ws.title} 시트에 '요일' 열이 없습니다.")

    for col in range(1, ws.max_column+1):
        if ws.cell(1, col).value == "반":
            CLASS_NAME_COLUMN = col
            break
    else:
        raise NoReservedColumnError(f"{ws.title} 시트에 '반' 열이 없습니다.")

    for col in range(1, ws.max_column+1):
        if ws.cell(1, col).value == "담당":
            TEACHER_NAME_COLUMN = col
            break
    else:
        raise NoReservedColumnError(f"{ws.title} 시트에 '담당' 열이 없습니다.")

    for col in range(1, ws.max_column+1):
        if ws.cell(1, col).value == "이름":
            STUDENT_NAME_COLUMN = col
            break
    else:
        raise NoReservedColumnError(f"{ws.title} 시트에 '이름' 열이 없습니다.")

    for col in range(1, ws.max_column+1):
        if ws.cell(1, col).value == "학생 평균":
            AVERAGE_SCORE_COLUMN = col
            break
    else:
        raise NoReservedColumnError(f"{ws.title} 시트에 '학생 평균' 열이 없습니다.")
    
    return CLASS_NAME_COLUMN, TEACHER_NAME_COLUMN, STUDENT_NAME_COLUMN, AVERAGE_SCORE_COLUMN

def is_cell_empty(row:int, col:int) -> bool:
    """
    데이터 파일이 열려있지 않을 때 특정 셀의 값이 비어있는 지 확인

    데일리테스트 시트 한정 기능
    """
    wb = open(data_only=True, read_only=True)
    ws = wb[DataFile.DEFAULT_SHEET_NAME]

    if ws.cell(row, col).value is None:
        return True, None

    value = ws.cell(row, col).value

    return False, value

def get_class_names(ws:Worksheet):
    class_names = []

    CLASS_NAME_COLUMN, _, _, _ = find_dynamic_columns(ws)

    for row in range(2, ws.max_row+1):
        if ws.cell(row, CLASS_NAME_COLUMN).value  not in class_names:
            class_names.append(ws.cell(row, CLASS_NAME_COLUMN).value)

    return class_names

# 파일 작업
def save_test_data(filepath:str, prog: Progress):
    """
    데이터 양식에 작성된 데이터를 데이터 파일에 저장
    """
    # 임시 파일 삭제
    if os.path.isfile(f"{omikron.config.DATA_DIR}/data/{DataFile.TEMP_FILE_NAME}.xlsx"):
        delete_temp()

    form_wb = omikron.dataform.open(filepath)
    form_ws = omikron.dataform.open_worksheet(form_wb)

    # 학생 정보 열기
    student_wb = omikron.studentinfo.open(True)
    student_ws = omikron.studentinfo.open_worksheet(student_wb)

    file_validation()

    # 백업 생성
    make_backup_file()
    prog.step("백업 생성 완료")

    wb = open()
    ws = wb[DataFile.DEFAULT_SHEET_NAME]

    CLASS_NAME_COLUMN, _, STUDENT_NAME_COLUMN, AVERAGE_SCORE_COLUMN = find_dynamic_columns(ws)

    for t in range(2):
        if t == 0:
            TEST_NAME_COLUMN    = DataForm.DAILYTEST_NAME_COLUMN
            TEST_SCORE_COLUMN   = DataForm.DAILYTEST_SCORE_COLUMN
            TEST_AVERAGE_COLUMN = DataForm.DAILYTEST_AVERAGE_COLUMN
        else:
            TEST_NAME_COLUMN    = DataForm.MOCKTEST_NAME_COLUMN
            TEST_SCORE_COLUMN   = DataForm.MOCKTEST_SCORE_COLUMN
            TEST_AVERAGE_COLUMN = DataForm.MOCKTEST_AVERAGE_COLUMN

        for i in range(2, form_ws.max_row+1): # 데일리데이터 기록 양식 루프
            # 반 필터링
            if (form_ws.cell(i, DataForm.CLASS_NAME_COLUMN).value is not None) and (form_ws.cell(i, TEST_NAME_COLUMN).value is not None):
                class_name   = form_ws.cell(i, DataForm.CLASS_NAME_COLUMN).value
                if t == 1: class_name += " (모의고사)"
                test_name    = form_ws.cell(i, TEST_NAME_COLUMN).value
                test_average = form_ws.cell(i, TEST_AVERAGE_COLUMN).value

                no_class = False

                #반 시작 찾기
                for row in range(2, ws.max_row+1):
                    if ws.cell(row, CLASS_NAME_COLUMN).value == class_name:
                        CLASS_START = row
                        break
                else:
                    prog.warning(f"{class_name} 반이 존재하지 않습니다.")
                    no_class = True
                    continue

                # 반 끝 찾기
                for row in range(CLASS_START, ws.max_row+1):
                    if ws.cell(row, STUDENT_NAME_COLUMN).value == "시험 평균":
                        CLASS_END = row
                        break

                # 데이터 작성 열 찾기
                for col in range(AVERAGE_SCORE_COLUMN+1, ws.max_column+2):
                    test_date = ws.cell(CLASS_START, col).value
                    if type(test_date) == datetime and test_date.strftime("%y%m%d") == datetime.today().strftime("%y%m%d"):
                        WRITE_COLUMN = col
                        break
                    elif test_date is None:
                        WRITE_COLUMN = col
                        break

                # 입력 틀 작성
                AVERAGE_FORMULA = f"=ROUND(AVERAGE({gcl(WRITE_COLUMN)+str(CLASS_START + 2)}:{gcl(WRITE_COLUMN)+str(CLASS_END - 1)}), 0)"
                ws.column_dimensions[gcl(WRITE_COLUMN)].width    = 14
                ws.cell(CLASS_START, WRITE_COLUMN).value         = datetime.today().date()
                ws.cell(CLASS_START, WRITE_COLUMN).number_format = "yyyy.mm.dd(aaa)"
                ws.cell(CLASS_START, WRITE_COLUMN).alignment     = ALIGN_CENTER
                ws.cell(CLASS_START, WRITE_COLUMN).border        = BORDER_TOP_MEDIUM_000

                ws.cell(CLASS_START + 1, WRITE_COLUMN).value     = test_name
                ws.cell(CLASS_START + 1, WRITE_COLUMN).alignment = ALIGN_CENTER_WRAP
                ws.cell(CLASS_START + 1, WRITE_COLUMN).border    = BORDER_BOTTOM_THIN_9090

                ws.cell(CLASS_END, WRITE_COLUMN).value           = AVERAGE_FORMULA
                ws.cell(CLASS_END, WRITE_COLUMN).font            = FONT_BOLD
                ws.cell(CLASS_END, WRITE_COLUMN).alignment       = ALIGN_CENTER
                ws.cell(CLASS_END, WRITE_COLUMN).border          = BORDER_TOP_THIN_9090_BOTTOM_MEDIUM_000
                
                if type(test_average) in (int, float):
                    ws.cell(CLASS_END, WRITE_COLUMN).fill = class_average_color(test_average)

            test_score   = form_ws.cell(i, TEST_SCORE_COLUMN).value
            student_name = form_ws.cell(i, DataForm.STUDENT_NAME_COLUMN).value

            if test_score is None:
                continue
            if no_class:
                continue

            # 학생 찾기
            for row in range(CLASS_START + 2, CLASS_END):
                if ws.cell(row, STUDENT_NAME_COLUMN).value == student_name:
                    ws.cell(row, WRITE_COLUMN).value = test_score
                    if type(test_score) in (int, float):
                        ws.cell(row, WRITE_COLUMN).fill = test_score_color(test_score)

                    ws.cell(row, WRITE_COLUMN).alignment = ALIGN_CENTER
                    break
            else:
                prog.warning(f"{class_name} 반에 {student_name} 학생이 존재하지 않습니다.")

    ws = wb[DataFile.DEFAULT_SHEET_NAME]
    save_to_temp(wb)
    prog.step("데이터 저장 완료")

    # 조건부 서식 수식 로딩
    pythoncom.CoInitialize()
    excel = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        abs_path = os.path.abspath(f"{omikron.config.DATA_DIR}/data/{DataFile.TEMP_FILE_NAME}.xlsx")
        wb_com = excel.Workbooks.Open(abs_path)
        wb_com.Save()
        wb_com.Close()
    finally:
        try:
            if excel is not None:
                excel.Quit()
        except:
            pass
        pythoncom.CoUninitialize()

    wb           = open_temp()
    data_only_wb = open_temp(data_only=True)

    # for sheet_name in wb.sheetnames:
    #     if sheet_name not in (DataFile.DEFAULT_SHEET_NAME, DataFile.SECOND_SHEET_NAME):
    #         continue

    ws           = wb[DataFile.DEFAULT_SHEET_NAME]
    data_only_ws = data_only_wb[DataFile.DEFAULT_SHEET_NAME]

    _, _, STUDENT_NAME_COLUMN, AVERAGE_SCORE_COLUMN = find_dynamic_columns(ws)

    for row in range(2, data_only_ws.max_row+1):
        if data_only_ws.cell(row, STUDENT_NAME_COLUMN).value is None:
            break

        # 학생 별 평균 점수에 대한 조건부 서식
        student_average = data_only_ws.cell(row, AVERAGE_SCORE_COLUMN).value
        if type(student_average) in (int, float):
            if ws.cell(row, STUDENT_NAME_COLUMN).value == "시험 평균":
                ws.cell(row, AVERAGE_SCORE_COLUMN).fill = class_average_color(student_average)
            else:
                ws.cell(row, AVERAGE_SCORE_COLUMN).fill = student_average_color(student_average)

        # 신규생 하이라이트
        if ws.cell(row, STUDENT_NAME_COLUMN).value in ("날짜", "시험명", "시험 평균"):
            continue
        if ws.cell(row, STUDENT_NAME_COLUMN).font.strike:
            continue
        if ws.cell(row, STUDENT_NAME_COLUMN).font.color is not None and ws.cell(row, STUDENT_NAME_COLUMN).font.color.rgb == "FFFF0000":
            continue

        exist, _, _, new_student = omikron.studentinfo.get_student_info(student_ws, ws.cell(row, STUDENT_NAME_COLUMN).value)
        if exist:
            if new_student:
                ws.cell(row, STUDENT_NAME_COLUMN).fill = FILL_NEW_STUDENT
            else:
                ws.cell(row, STUDENT_NAME_COLUMN).fill = FILL_NONE
        else:
            ws.cell(row, STUDENT_NAME_COLUMN).fill = FILL_NONE
            prog.warning(f"{ws.cell(row, STUDENT_NAME_COLUMN).value} 학생 정보가 존재하지 않습니다.")

    ws = wb[DataFile.DEFAULT_SHEET_NAME]
    prog.step("조건부 서식 로딩 완료")

    return wb

def save_individual_test_data(target_row:int, target_col:int, test_score:int|float):
    """정규 시험에 미응시한 학생의 결과를 입력하고 해당 반의 평균을 반환"""
    # 임시 파일 삭제
    if os.path.isfile(f"{omikron.config.DATA_DIR}/data/{DataFile.TEMP_FILE_NAME}.xlsx"):
        delete_temp()

    file_validation()

    # 백업 생성
    make_backup_file()

    wb = open()
    ws = wb[DataFile.DEFAULT_SHEET_NAME]

    # 시험 점수 기록
    ws.cell(target_row, target_col).value     = test_score
    ws.cell(target_row, target_col).fill      = test_score_color(test_score)
    ws.cell(target_row, target_col).alignment = ALIGN_CENTER

    save_to_temp(wb)

    pythoncom.CoInitialize()
    excel = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        abs_path = os.path.abspath(f"{omikron.config.DATA_DIR}/data/{DataFile.TEMP_FILE_NAME}.xlsx")
        wb_com = excel.Workbooks.Open(abs_path)
        wb_com.Save()
        wb_com.Close()
    finally:
        try:
            if excel is not None:
                excel.Quit()
        except:
            pass
        pythoncom.CoUninitialize()

    wb           = open_temp()
    data_only_wb = open_temp(True)

    ws           = wb[DataFile.DEFAULT_SHEET_NAME]
    data_only_ws = data_only_wb[DataFile.DEFAULT_SHEET_NAME]

    _, _, STUDENT_NAME_COLUMN, AVERAGE_SCORE_COLUMN = find_dynamic_columns(ws)

    # 학생 평균 조건부 서식 반영
    student_average = data_only_ws.cell(target_row, AVERAGE_SCORE_COLUMN).value
    if type(student_average)  in (int, float):
        ws.cell(target_row, AVERAGE_SCORE_COLUMN).fill = student_average_color(student_average)

    # 시험 평균 조건부 서식 반영
    test_average_row = target_row
    while data_only_ws.cell(test_average_row, STUDENT_NAME_COLUMN).value != "시험 평균":
        test_average_row += 1

    test_average = data_only_ws.cell(test_average_row, target_col).value
    if type(test_average) in (int, float):
        ws.cell(test_average_row, target_col).fill = class_average_color(test_average)

    # 반 평균 조건부 서식 반영
    class_average = data_only_ws.cell(test_average_row, AVERAGE_SCORE_COLUMN).value
    if type(class_average) in (int, float):
        ws.cell(test_average_row, AVERAGE_SCORE_COLUMN).fill = class_average_color(test_average)

    save(wb)
    delete_temp()

    return test_average

def conditional_formatting():
    file_validation()

    pythoncom.CoInitialize()
    excel = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        abs_path = os.path.abspath(f"{omikron.config.DATA_DIR}/data/{DataFile.TEMP_FILE_NAME}.xlsx")
        wb_com = excel.Workbooks.Open(abs_path)
        wb_com.Save()
        wb_com.Close()
    finally:
        try:
            if excel is not None:
                excel.Quit()
        except:
            pass
        pythoncom.CoUninitialize()

    warnings = []

    wb           = open()
    data_only_wb = open(data_only=True, read_only=True)
    student_wb   = omikron.studentinfo.open()
    student_ws   = omikron.studentinfo.open_worksheet(student_wb)

    # for sheet_name in wb.sheetnames:
    #     if sheet_name not in (DataFile.DEFAULT_SHEET_NAME, DataFile.SECOND_SHEET_NAME):
    #         continue

    ws           = wb[DataFile.DEFAULT_SHEET_NAME]
    data_only_ws = data_only_wb[DataFile.DEFAULT_SHEET_NAME]

    _, _, STUDENT_NAME_COLUMN, AVERAGE_SCORE_COLUMN = find_dynamic_columns(ws)

    for row in range(2, ws.max_row+1):
        if ws.cell(row, STUDENT_NAME_COLUMN).value is None:
            break
        if ws.cell(row, STUDENT_NAME_COLUMN).value == "날짜":
            DATE_ROW = row
        if ws.cell(row, STUDENT_NAME_COLUMN).value != "시험명":
            ws.row_dimensions[row].height = 18

        # 데이터 조건부 서식
        for col in range(1, data_only_ws.max_column+1):
            try:
                if col > AVERAGE_SCORE_COLUMN and ws.cell(DATE_ROW, col).value is None:
                    break
            except:
                if col > AVERAGE_SCORE_COLUMN and ws.cell(row, col).value is None:
                    break

            ws.column_dimensions[gcl(col)].width = 14
            if ws.cell(row, STUDENT_NAME_COLUMN).value == "날짜":
                ws.cell(row, col).border = BORDER_BOTTOM_MEDIUM_000
            elif ws.cell(row, STUDENT_NAME_COLUMN).value == "시험명":
                ws.cell(row, col).border = BORDER_BOTTOM_THIN_9090
            elif ws.cell(row, STUDENT_NAME_COLUMN).value == "시험 평균":
                ws.cell(row, col).border = BORDER_TOP_THIN_9090_BOTTOM_MEDIUM_000
            else:
                ws.cell(row, col).border = None

            # 학생 평균 점수 열 기준 분기   
            if col <= AVERAGE_SCORE_COLUMN:
                continue

            if ws.cell(row, STUDENT_NAME_COLUMN).value == "시험명":
                ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            elif data_only_ws.cell(row, STUDENT_NAME_COLUMN).value == "시험 평균":
                ws.cell(row, col).font = FONT_BOLD
                if type(data_only_ws.cell(row, col).value) in (int, float):
                    ws.cell(row, col).fill = class_average_color(data_only_ws.cell(row, col).value)
            elif type(data_only_ws.cell(row, col).value) in (int, float):
                ws.cell(row, col).fill = test_score_color(data_only_ws.cell(row, col).value)
            else:
                ws.cell(row, col).fill = FILL_NONE

        # 학생별 평균 조건부 서식
        if type(data_only_ws.cell(row, AVERAGE_SCORE_COLUMN).value) in (int, float):
            if ws.cell(row, STUDENT_NAME_COLUMN).value == "시험 평균":
                ws.cell(row, AVERAGE_SCORE_COLUMN).fill = class_average_color(data_only_ws.cell(row, AVERAGE_SCORE_COLUMN).value)
            else:
                ws.cell(row, AVERAGE_SCORE_COLUMN).fill = student_average_color(data_only_ws.cell(row, AVERAGE_SCORE_COLUMN).value)
        else:
            ws.cell(row, col).fill = FILL_NONE

        # 학생별 평균 폰트 설정
        if ws.cell(row, STUDENT_NAME_COLUMN).value in ("날짜", "시험명", "시험 평균"):
            ws.cell(row, AVERAGE_SCORE_COLUMN).font = FONT_BOLD
            continue
        if ws.cell(row, STUDENT_NAME_COLUMN).font.strike:
            ws.cell(row, AVERAGE_SCORE_COLUMN).font = FONT_BOLD_STRIKE
            continue
        if ws.cell(row, STUDENT_NAME_COLUMN).font.color is not None and ws.cell(row, STUDENT_NAME_COLUMN).font.color.rgb == "FFFF0000":
            ws.cell(row, AVERAGE_SCORE_COLUMN).font = FONT_BOLD_RED
            continue

        # 신규생 하이라이트
        exist, _, _, new_student = omikron.studentinfo.get_student_info(student_ws, ws.cell(row, STUDENT_NAME_COLUMN).value)
        if exist:
            if new_student:
                ws.cell(row, STUDENT_NAME_COLUMN).fill = FILL_NEW_STUDENT
            else:
                ws.cell(row, STUDENT_NAME_COLUMN).fill = FILL_NONE
        else:
            ws.cell(row, STUDENT_NAME_COLUMN).fill = FILL_NONE
            warnings.append(f"{ws.cell(row, STUDENT_NAME_COLUMN).value} 학생 정보가 존재하지 않습니다.")

    save(wb)

    return warnings

def update_class():
    """
    수정된 반 정보 파일을 바탕으로 데이터 파일 업데이트
    """
    file_validation()

    make_backup_file()

    new_class_names = set(omikron.classinfo.get_new_class_names())

    # 조건부 서식 수식 로딩
    pythoncom.CoInitialize()
    excel = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        abs_path = os.path.abspath(f"{omikron.config.DATA_DIR}/data/{omikron.config.DATA_FILE_NAME}.xlsx")
        wb_com = excel.Workbooks.Open(abs_path)
        wb_com.Save()
        wb_com.Close()
    finally:
        try:
            if excel is not None:
                excel.Quit()
        except:
            pass
        pythoncom.CoUninitialize()

    # 지난 데이터 파일이 없으면 새로 생성
    if not os.path.isfile(f"{omikron.config.DATA_DIR}/data/{DataFile.PRE_DATA_FILE_NAME}.xlsx"):
        pre_data_wb = xl.Workbook()
        pre_data_ws = pre_data_wb.worksheets[0]
        pre_data_ws.title = DataFile.DEFAULT_SHEET_NAME

        pre_data_ws[gcl(DataFile.CLASS_NAME_COLUMN)+"1"]    = "반"
        pre_data_ws[gcl(DataFile.TEACHER_NAME_COLUMN)+"1"]  = "담당"
        pre_data_ws[gcl(DataFile.STUDENT_NAME_COLUMN)+"1"]  = "이름"
        pre_data_ws[gcl(DataFile.AVERAGE_SCORE_COLUMN)+"1"] = "학생 평균"
        pre_data_ws.freeze_panes    = f"{gcl(DataFile.DATA_COLUMN)}2"
        pre_data_ws.auto_filter.ref = f"A:{gcl(DataFile.MAX)}"

        for col in range(1, DataFile.DATA_COLUMN):
            pre_data_ws.cell(1, col).alignment = ALIGN_CENTER
            pre_data_ws.cell(1, col).border    = BORDER_BOTTOM_MEDIUM_000

        pre_data_wb.save(f"{omikron.config.DATA_DIR}/data/{DataFile.PRE_DATA_FILE_NAME}.xlsx")
    else:
        pre_data_wb = xl.load_workbook(f"{omikron.config.DATA_DIR}/data/{DataFile.PRE_DATA_FILE_NAME}.xlsx")

    # 지난 데이터 이동
    data_only_wb = open(data_only=True, read_only=True) # 데이터가 더이상 수정되지 않으므로 읽기 전용으로 불러옴

    data_only_ws = data_only_wb[DataFile.DEFAULT_SHEET_NAME]
    pre_data_ws  = pre_data_wb[DataFile.DEFAULT_SHEET_NAME]

    CLASS_NAME_COLUMN, TEACHER_NAME_COLUMN, STUDENT_NAME_COLUMN, AVERAGE_SCORE_COLUMN = find_dynamic_columns(data_only_ws)

    for row in range(2, data_only_ws.max_row+1):
        if data_only_ws.cell(row, CLASS_NAME_COLUMN).value not in new_class_names:
            PRE_DATA_WRITE_ROW = pre_data_ws.max_row+1
            copy_cell(pre_data_ws.cell(PRE_DATA_WRITE_ROW, DataFile.CLASS_NAME_COLUMN),    data_only_ws.cell(row, CLASS_NAME_COLUMN))
            copy_cell(pre_data_ws.cell(PRE_DATA_WRITE_ROW, DataFile.TEACHER_NAME_COLUMN),  data_only_ws.cell(row, TEACHER_NAME_COLUMN))
            copy_cell(pre_data_ws.cell(PRE_DATA_WRITE_ROW, DataFile.STUDENT_NAME_COLUMN),  data_only_ws.cell(row, STUDENT_NAME_COLUMN))
            copy_cell(pre_data_ws.cell(PRE_DATA_WRITE_ROW, DataFile.AVERAGE_SCORE_COLUMN), data_only_ws.cell(row, AVERAGE_SCORE_COLUMN))
            PRE_DATA_WRITE_COLUMN = DataFile.MAX+1
            for col in range(AVERAGE_SCORE_COLUMN+1, data_only_ws.max_column+1):
                copy_cell(pre_data_ws.cell(PRE_DATA_WRITE_ROW, PRE_DATA_WRITE_COLUMN), data_only_ws.cell(row, col))
                PRE_DATA_WRITE_COLUMN += 1

    for col in range(DataFile.MAX + 1, pre_data_ws.max_column + 1):
        pre_data_ws.column_dimensions[gcl(col)].width = 14

    data_only_wb.close()
    data_only_wb = None
    pre_data_wb.save(f"{omikron.config.DATA_DIR}/data/{DataFile.PRE_DATA_FILE_NAME}.xlsx")
    pre_data_wb = None

    # 데이터 파일 지난 데이터 삭제 및 신규 반 추가
    wb = open()
    ws = wb[DataFile.DEFAULT_SHEET_NAME]

    to_delete = []
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row, CLASS_NAME_COLUMN).value
        if v is not None and v not in new_class_names:
            to_delete.append(row)

    for row in reversed(to_delete):
        ws.delete_rows(row)
    ws.auto_filter.ref = f"A:{gcl(AVERAGE_SCORE_COLUMN)}"

    old_class_names = set(get_class_names(ws))
    unregistered_class_names = list(new_class_names.difference(old_class_names))

    if len(unregistered_class_names) > 0:
        class_wb = omikron.classinfo.open_temp()
        class_ws = omikron.classinfo.open_worksheet(class_wb)

        class_student_dict = omikron.chrome.get_class_student_dict()

        ws = wb[DataFile.DEFAULT_SHEET_NAME]

        CLASS_NAME_COLUMN, TEACHER_NAME_COLUMN, STUDENT_NAME_COLUMN, AVERAGE_SCORE_COLUMN = find_dynamic_columns(ws)

        for row in range(ws.max_row+1, 1, -1):
            if ws.cell(row-1, STUDENT_NAME_COLUMN).value is not None:
                WRITE_RANGE = WRITE_LOCATION = row
                break

        for class_name in unregistered_class_names:
            temp_name = class_name
            if " (모의고사)" in class_name:
                temp_name = class_name[:-7]
            if len(class_student_dict[temp_name]) == 0 :
                continue
            exist, teacher_name, _, _, _ = omikron.classinfo.get_class_info(temp_name, ws=class_ws)
            if not exist: continue

            # 시험명
            ws.cell(WRITE_LOCATION, CLASS_NAME_COLUMN).value    = class_name
            ws.cell(WRITE_LOCATION, TEACHER_NAME_COLUMN).value  = teacher_name
            ws.cell(WRITE_LOCATION, STUDENT_NAME_COLUMN).value  = "날짜"
            WRITE_LOCATION += 1
            
            ws.cell(WRITE_LOCATION, CLASS_NAME_COLUMN).value    = class_name
            ws.cell(WRITE_LOCATION, TEACHER_NAME_COLUMN).value  = teacher_name
            ws.cell(WRITE_LOCATION, STUDENT_NAME_COLUMN).value  = "시험명"

            for col in range(1, AVERAGE_SCORE_COLUMN + 1):
                ws.cell(WRITE_LOCATION, col).border = BORDER_BOTTOM_THIN_9090

            WRITE_LOCATION += 1

            # 학생 루프
            for studnet_name in class_student_dict[temp_name]:
                ws.cell(WRITE_LOCATION, CLASS_NAME_COLUMN).value    = class_name
                ws.cell(WRITE_LOCATION, TEACHER_NAME_COLUMN).value  = teacher_name
                ws.cell(WRITE_LOCATION, STUDENT_NAME_COLUMN).value  = studnet_name
                WRITE_LOCATION += 1
            
            # 시험별 평균
            ws.cell(WRITE_LOCATION, CLASS_NAME_COLUMN).value    = class_name
            ws.cell(WRITE_LOCATION, TEACHER_NAME_COLUMN).value  = teacher_name
            ws.cell(WRITE_LOCATION, STUDENT_NAME_COLUMN).value  = "시험 평균"

            for col in range(1, AVERAGE_SCORE_COLUMN+1):
                ws.cell(WRITE_LOCATION, col).border = BORDER_TOP_THIN_9090_BOTTOM_MEDIUM_000

            WRITE_LOCATION += 1

        # 정렬
        for row in range(WRITE_RANGE, ws.max_row + 1):
            for col in range(1, AVERAGE_SCORE_COLUMN + 1):
                ws.cell(row, col).alignment = ALIGN_CENTER

        # 필터 범위 재지정
        ws.auto_filter.ref = f"A:{gcl(AVERAGE_SCORE_COLUMN)}"

    return rescoping_formula(wb)

def add_student(student_name:str, target_class_name:str, wb:xl.Workbook=None):
    """
    학생 추가
    
    `move_student` 작업 시 `wb`로 작업중인 파일 정보 전달
    """
    file_validation()

    if wb is None:
        wb = open()

    ws = wb[DataFile.DEFAULT_SHEET_NAME]

    warnings = []

    # for ws in wb.worksheets:
    #     if ws.title not in (DataFile.DEFAULT_SHEET_NAME, DataFile.SECOND_SHEET_NAME):
    #         continue

    #     exist = False

    for i in range(2):
        if i == 1: target_class_name += " (모의고사)"

        CLASS_NAME_COLUMN, TEACHER_NAME_COLUMN, STUDENT_NAME_COLUMN, AVERAGE_SCORE_COLUMN = find_dynamic_columns(ws)

        # 목표 반에 학생 추가
        for row in range(2, ws.max_row+1):
            if ws.cell(row, CLASS_NAME_COLUMN).value == target_class_name:
                class_index = row+2
                break
        else:
            # warnings.append(f"{ws.title} 시트에 {target_class_name} 반이 존재하지 않습니다.")
            continue

        while ws.cell(class_index, STUDENT_NAME_COLUMN).value != "시험 평균":
            if ws.cell(class_index, STUDENT_NAME_COLUMN).value > student_name:
                break
            elif ws.cell(class_index, STUDENT_NAME_COLUMN).font.strike:
                class_index += 1
            elif ws.cell(class_index, STUDENT_NAME_COLUMN).font.color is not None and ws.cell(class_index, STUDENT_NAME_COLUMN).font.color.rgb == "FFFF0000":
                class_index += 1
            elif ws.cell(class_index, STUDENT_NAME_COLUMN).value == student_name:
                warnings.append(f"{student_name} 학생이 이미 존재합니다.")
                break
            else:
                class_index += 1

        ws.insert_rows(class_index)
        # ws.cell(class_index, TEST_TIME_COLUMN).value         = ws.cell(class_index-1, TEST_TIME_COLUMN).value
        # ws.cell(class_index, CLASS_WEEKDAY_COLUMN).value     = ws.cell(class_index-1, CLASS_WEEKDAY_COLUMN).value
        ws.cell(class_index, CLASS_NAME_COLUMN).value        = ws.cell(class_index-1, CLASS_NAME_COLUMN).value
        ws.cell(class_index, TEACHER_NAME_COLUMN).value      = ws.cell(class_index-1, TEACHER_NAME_COLUMN).value
        ws.cell(class_index, STUDENT_NAME_COLUMN).value      = student_name

        # ws.cell(class_index, TEST_TIME_COLUMN).alignment     = ALIGN_CENTER
        # ws.cell(class_index, CLASS_WEEKDAY_COLUMN).alignment = ALIGN_CENTER
        ws.cell(class_index, CLASS_NAME_COLUMN).alignment    = ALIGN_CENTER
        ws.cell(class_index, TEACHER_NAME_COLUMN).alignment  = ALIGN_CENTER
        ws.cell(class_index, STUDENT_NAME_COLUMN).alignment  = ALIGN_CENTER

        ws.cell(class_index, AVERAGE_SCORE_COLUMN).alignment = ALIGN_CENTER
        ws.cell(class_index, AVERAGE_SCORE_COLUMN).font      = FONT_BOLD

    rescoping_formula(wb)

    return warnings

def delete_student(student_name:str):
    """
    학생 퇴원 처리
    
    퇴원 처리된 학생은 모든 데이터에 취소선 적용
    """
    file_validation()

    wb = open()
    ws = wb[DataFile.DEFAULT_SHEET_NAME]

    _, _, STUDENT_NAME_COLUMN, AVERAGE_SCORE_COLUMN = find_dynamic_columns(ws)

    for row in range(2, ws.max_row+1):
        if ws.cell(row, STUDENT_NAME_COLUMN).value == student_name:
            for col in range(1, ws.max_column+1):
                if ws.cell(row, col).font.bold:
                    ws.cell(row, col).font = FONT_BOLD_STRIKE
                else:
                    ws.cell(row, col).font = FONT_STRIKE
            
            # 퇴원한 학생이 반 평균에 영향을 주지 않도록 수정
            ws.cell(row, AVERAGE_SCORE_COLUMN).value = ""

    save(wb)

def move_student(student_name:str, target_class_name:str, current_class_name:str):
    """
    학생 반 이동

    학생의 기존 반 데이터 글꼴 색을 빨간색으로 변경 후 목표 반에 학생 추가
    """
    file_validation()

    wb = open()
    ws = wb[DataFile.DEFAULT_SHEET_NAME]

    # for ws in wb.worksheets:
    #     if ws.title not in (DataFile.DEFAULT_SHEET_NAME, DataFile.SECOND_SHEET_NAME):
    #         continue

    CLASS_NAME_COLUMN, _, STUDENT_NAME_COLUMN, _ = find_dynamic_columns(ws)

    # 기존 반 데이터 빨간색 처리
    for row in range(2, ws.max_row+1):
        if ws.cell(row, STUDENT_NAME_COLUMN).value == student_name and ws.cell(row, CLASS_NAME_COLUMN).value in (current_class_name, current_class_name+" (모의고사)"):
            for col in range(1, ws.max_column+1):
                if ws.cell(row, col).font.bold:
                    ws.cell(row, col).font = FONT_BOLD_RED
                else:
                    ws.cell(row, col).font = FONT_RED
            # break

    return add_student(student_name, target_class_name, wb)

def rescoping_formula(wb:xl.Workbook=None):
    """
    데이터 파일 내 평균 산출 수식의 범위 재조정
    """
    file_validation()

    if wb is None:
        wb = open()

    ws = wb[DataFile.DEFAULT_SHEET_NAME]

    _, _, STUDENT_NAME_COLUMN, AVERAGE_SCORE_COLUMN = find_dynamic_columns(ws)

    # 평균 범위 재지정
    for row in range(2, ws.max_row+1):
        if ws.cell(row, STUDENT_NAME_COLUMN).value is None:
            break
        striked = False
        colored = False
        if ws.cell(row, STUDENT_NAME_COLUMN).font.strike:
            striked = True
        if ws.cell(row, STUDENT_NAME_COLUMN).font.color is not None:
            if ws.cell(row, STUDENT_NAME_COLUMN).font.color.rgb == "FFFF0000":
                colored = True

        if ws.cell(row, STUDENT_NAME_COLUMN).value == "날짜":
            DATE_ROW = row
            CLASS_START = row+2
        elif ws.cell(row, STUDENT_NAME_COLUMN).value == "시험 평균":
            CLASS_END = row-1
            ws[f"{gcl(AVERAGE_SCORE_COLUMN)}{row}"] = ArrayFormula(f"{gcl(AVERAGE_SCORE_COLUMN)}{row}", f"=ROUND(AVERAGE(IFERROR({gcl(AVERAGE_SCORE_COLUMN)}{CLASS_START}:{gcl(AVERAGE_SCORE_COLUMN)}{CLASS_END}, \"\")), 0)")
            if CLASS_START >= CLASS_END:
                continue
            for col in range(AVERAGE_SCORE_COLUMN+1, ws.max_column+1):
                if ws.cell(DATE_ROW, col).value is None:
                    break
                ws.cell(row, col).value = f"=ROUND(AVERAGE({gcl(col)}{CLASS_START}:{gcl(col)}{CLASS_END}), 0)"
                ws.cell(row, col).font  = FONT_BOLD
        elif ws.cell(row, STUDENT_NAME_COLUMN).value == "시험명":
            continue
        else:
            ws.cell(row, AVERAGE_SCORE_COLUMN).value = f"=ROUND(AVERAGE({gcl(AVERAGE_SCORE_COLUMN+1)}{row}:XFD{row}), 0)"

        if striked:
            ws.cell(row, AVERAGE_SCORE_COLUMN).font = FONT_BOLD_STRIKE
        elif colored:
            ws.cell(row, AVERAGE_SCORE_COLUMN).font = FONT_BOLD_RED
        else:
            ws.cell(row, AVERAGE_SCORE_COLUMN).font = FONT_BOLD

    save(wb)

def change_class_info(target_class_name:str, target_teacher_name:str):
    """
    특정 반의 담당 선생님 변경
    """
    wb = open()
    ws = wb[DataFile.DEFAULT_SHEET_NAME]

    CLASS_NAME_COLUMN, TEACHER_NAME_COLUMN, _, _ = find_dynamic_columns(ws)

    for row in range(2, ws.max_row+1):
        if ws.cell(row, CLASS_NAME_COLUMN).value in (target_class_name, target_class_name+" (모의고사)"):
            ws.cell(row, TEACHER_NAME_COLUMN).value = target_teacher_name

    save(wb)
