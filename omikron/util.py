from copy import copy
from datetime import datetime
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
from omikron.style import FILL_BELOW_60, FILL_BELOW_70, FILL_BELOW_80, FILL_CLASS_AVG, FILL_STUDENT_AVG, FILL_NONE

def calculate_makeup_test_schedule(makeup_test_weekday:str, makeup_test_date:dict[str:datetime]):
    """
    학생의 재시험 응시 희망 요일에 따라 가장 가까운 재시험 일정을 계산

    return `계산 성공 여부`, `계산된 날짜`, `계산된 시간`
    """
    try:
        weekday_list = makeup_test_weekday.split("/")
        calculated_date = makeup_test_date[weekday_list[0].replace(" ", "")]
        time_index = 0
        for tmp_idx in range(len(weekday_list)):
            if calculated_date > makeup_test_date[weekday_list[tmp_idx].replace(" ", "")]:
                calculated_date = makeup_test_date[weekday_list[tmp_idx].replace(" ", "")]
                time_index = tmp_idx
    except KeyError:
        return False, None, 0

    return True, makeup_test_date[weekday_list[time_index]], time_index

def date_to_kor_date(date:datetime) -> str:
    """
    `datetime.strftime` 한글 인코딩 오류 우회

    return `mm월 dd일`
    """
    month, day = date.strftime("%m %d").split()

    return f"{month}월 {day}일"

def copy_cell(dst:Cell, src:Cell):
    dst.value         = src.value
    dst.font          = copy(src.font)
    dst.fill          = copy(src.fill)
    dst.border        = copy(src.border)
    dst.alignment     = copy(src.alignment)
    dst.number_format = copy(src.number_format)

def class_average_color(score:int|float) -> PatternFill:
    """
    반 전체 평균에 대한 점수 기반 색 채우기 (`시험 평균` 행)
    """
    if score < 60:
        return FILL_BELOW_60
    elif score < 70:
        return FILL_BELOW_70
    elif score < 80:
        return FILL_BELOW_80
    else:
        return FILL_CLASS_AVG

def student_average_color(score:int|float) -> PatternFill:
    """
    학생 평균에 대한 점수 기반 색 채우기 (`학생 평균` 열 중 `시험 평균` 행 제외)
    """
    if score < 60:
        return FILL_BELOW_60
    elif score < 70:
        return FILL_BELOW_70
    elif score < 80:
        return FILL_BELOW_80
    else:
        return FILL_STUDENT_AVG

def test_score_color(score:int|float) -> PatternFill:
    """
    각 시험 결과에 대한 점수 기반 색 채우기
    """
    if score < 60:
        return FILL_BELOW_60
    elif score < 70:
        return FILL_BELOW_70
    elif score < 80:
        return FILL_BELOW_80
    else:
        return FILL_NONE
