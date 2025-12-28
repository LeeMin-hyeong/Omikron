import os
import openpyxl as xl

from datetime import datetime
from openpyxl.utils.cell import get_column_letter as gcl

import omikron.classinfo
import omikron.dataform
import omikron.studentinfo
import omikron.config

from omikron.defs import MakeupTestList, DataForm
from omikron.exception import NoMatchingSheetException, FileOpenException
from omikron.util import calculate_makeup_test_schedule
from omikron.progress import Progress
from omikron.style import ALIGN_CENTER, ALIGN_CENTER_WRAP, FILL_NEW_STUDENT, BORDER_ALL


# 파일 기본 작업
def make_file():
    wb = xl.Workbook()
    ws = wb.worksheets[0]
    ws.title = MakeupTestList.DEFAULT_NAME
    ws[gcl(MakeupTestList.TEST_DATE_COLUMN)+"1"]          = "응시일"
    ws[gcl(MakeupTestList.CLASS_NAME_COLUMN)+"1"]         = "반"
    ws[gcl(MakeupTestList.TEACHER_NAME_COLUMN)+"1"]       = "담당T"
    ws[gcl(MakeupTestList.STUDENT_NAME_COLUMN)+"1"]       = "이름"
    ws[gcl(MakeupTestList.TEST_NAME_COLUMN)+"1"]          = "시험명"
    ws[gcl(MakeupTestList.MAKEUPTEST_DATE_COLUMN)+"1"]    = "재시 날짜"
    ws[gcl(MakeupTestList.MAKEUPTEST_SCORE_COLUMN)+"1"]   = "재시 점수"
    ws[gcl(MakeupTestList.ETC_COLUMN)+"1"]                = "비고"

    ws.column_dimensions[gcl(MakeupTestList.TEST_DATE_COLUMN)].width = 14
    ws.auto_filter.ref = "A:"+gcl(MakeupTestList.MAX)
    ws.freeze_panes    = "A2"

    for col in range(1, DataForm.MAX+1):
        ws.cell(1, col).alignment = ALIGN_CENTER_WRAP
        ws.cell(1, col).border    = BORDER_ALL

    wb.save(f"{omikron.config.DATA_DIR}/data/{MakeupTestList.DEFAULT_NAME}.xlsx")

def open(data_only:bool=False) -> xl.Workbook:
    return xl.load_workbook(f"{omikron.config.DATA_DIR}/data/{MakeupTestList.DEFAULT_NAME}.xlsx", data_only=data_only)

def open_worksheet(wb:xl.Workbook):
    try:
        return wb[MakeupTestList.DEFAULT_NAME]
    except:
        raise NoMatchingSheetException(f"'{MakeupTestList.DEFAULT_NAME}.xlsx'의 시트명을 '{MakeupTestList.DEFAULT_NAME}'으로 변경해 주세요.")

def save(wb:xl.Workbook):
    try:
        wb.save(f"{omikron.config.DATA_DIR}/data/{MakeupTestList.DEFAULT_NAME}.xlsx")
    except:
        raise FileOpenException(f"{MakeupTestList.DEFAULT_NAME} 파일을 닫은 뒤 다시 시도해주세요")

def isopen():
    return os.path.isfile(f"{omikron.config.DATA_DIR}/data/~${MakeupTestList.DEFAULT_NAME}.xlsx")

# 파일 유틸리티
def get_studnet_test_index_dict():
    """
    1st key: 학생 이름

    2nd key: 시험명

    value: 행 인덱스
    """
    wb = open(True)
    ws = open_worksheet(wb)

    student_test_index_dict:dict[str, dict[str, int]] = {}
    for row in range(2, ws.max_row+1):
        if ws.cell(row, MakeupTestList.MAKEUPTEST_SCORE_COLUMN).value is None:
            student_name     = ws.cell(row, MakeupTestList.STUDENT_NAME_COLUMN).value
            makeup_test_name = ws.cell(row, MakeupTestList.TEST_NAME_COLUMN).value
            try:
                student_test_index_dict[student_name]
            except:
                student_test_index_dict[student_name] = {}
            
            student_test_index_dict[student_name][makeup_test_name] = row

    return student_test_index_dict

# 파일 작업
def save_makeup_test_list(filepath: str, makeup_test_date: dict, prog: Progress):
    form_wb = None
    student_wb = None
    wb = None

    try:
        form_wb = omikron.dataform.open(filepath)
        form_ws = omikron.dataform.open_worksheet(form_wb)

        # 재시험 정보 파일 없으면 생성
        if not os.path.isfile(f"{omikron.config.DATA_DIR}/data/{MakeupTestList.DEFAULT_NAME}.xlsx"):
            make_file()

        wb = open()
        ws = open_worksheet(wb)

        # 학생 정보
        student_wb = omikron.studentinfo.open(True)
        student_ws = omikron.studentinfo.open_worksheet(student_wb)

        # ✅ 오늘 날짜 캐시 (루프 밖)
        today = datetime.today().date()
        today_key = today.strftime("%y%m%d")

        # 재시험 데이터 작성 시작 위치 탐색
        for row in range(ws.max_row + 1, 1, -1):
            if ws.cell(row - 1, MakeupTestList.TEST_DATE_COLUMN).value is not None:
                MAKEUP_TEST_RANGE = MAKEUP_TEST_WRITE_ROW = row
                break
        else:
            # 시트가 비어있는 특이 케이스 방어
            MAKEUP_TEST_RANGE = MAKEUP_TEST_WRITE_ROW = 2

        # ✅ (핵심) 중복 검사 캐시: "오늘 날짜인 행"만 스캔해서 set 구축
        #     기존 로직은 '오늘자 영역에서 같은 학생+반이면 duplicated'였음
        today_existing = set()  # (student_name, class_name)

        # 뒤에서 앞으로 훑되, 날짜가 오늘보다 과거로 내려가면 break
        check = ws.max_row
        while check > 1:
            test_date = ws.cell(check, MakeupTestList.TEST_DATE_COLUMN).value

            if test_date is None or type(test_date) != datetime:
                check -= 1
                continue

            dkey = test_date.strftime("%y%m%d")
            if dkey == today_key:
                sname = ws.cell(check, MakeupTestList.STUDENT_NAME_COLUMN).value
                cname = ws.cell(check, MakeupTestList.CLASS_NAME_COLUMN).value
                if sname is not None and cname is not None:
                    today_existing.add((sname, cname))
                check -= 1
                continue

            if dkey < today_key:
                break

            check -= 1

        for test_type in range(2):
            if test_type == 0:
                TEST_NAME_COLUMN = DataForm.DAILYTEST_NAME_COLUMN
                TEST_SCORE_COLUMN = DataForm.DAILYTEST_SCORE_COLUMN
            else:
                TEST_NAME_COLUMN = DataForm.MOCKTEST_NAME_COLUMN
                TEST_SCORE_COLUMN = DataForm.MOCKTEST_SCORE_COLUMN

            # 데일리데이터 기록 양식 루프
            class_name = test_name = teacher_name = None

            for i in range(2, form_ws.max_row + 1):
                # 반/시험명 갱신
                c = form_ws.cell(i, DataForm.CLASS_NAME_COLUMN).value
                tn = form_ws.cell(i, TEST_NAME_COLUMN).value
                if c is not None and tn is not None:
                    class_name = c
                    test_name = tn
                    teacher_name = form_ws.cell(i, DataForm.TEACHER_NAME_COLUMN).value

                test_score = form_ws.cell(i, TEST_SCORE_COLUMN).value
                if test_score is None or type(test_score) not in (int, float) or test_score >= 80:
                    continue

                makeup_test_check = form_ws.cell(i, DataForm.MAKEUP_TEST_CHECK_COLUMN).value
                if makeup_test_check in ("x", "X"):
                    continue

                student_name = form_ws.cell(i, DataForm.STUDENT_NAME_COLUMN).value
                if not student_name or not class_name:
                    continue

                # ✅ O(1) 중복 검사 (기존 while check 루프 제거)
                key = (student_name, class_name)
                if key in today_existing:
                    continue

                # 학생 재시험 정보 검색
                complete, makeup_test_weekday, _, new_student = omikron.studentinfo.get_student_info(student_ws, student_name)
                if not complete:
                    prog.warning(f"{student_name}의 학생 정보가 존재하지 않습니다.")

                ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.TEST_DATE_COLUMN).value = today
                ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.CLASS_NAME_COLUMN).value = class_name
                ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.TEACHER_NAME_COLUMN).value = teacher_name
                ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.STUDENT_NAME_COLUMN).value = student_name
                ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.TEST_NAME_COLUMN).value = test_name

                if new_student:
                    ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.STUDENT_NAME_COLUMN).fill = FILL_NEW_STUDENT

                if makeup_test_weekday is not None:
                    ok, calculated_schedule, _ = calculate_makeup_test_schedule(makeup_test_weekday, makeup_test_date)
                    if not ok:
                        prog.warning(f"{student_name}의 재시험 일정이 올바른 양식이 아닙니다.")

                    ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.MAKEUPTEST_DATE_COLUMN).value = calculated_schedule
                    ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.MAKEUPTEST_DATE_COLUMN).number_format = "mm월 dd일(aaa)"

                # ✅ 오늘자 중복 캐시에 즉시 반영(같은 실행에서 중복 추가 방지)
                today_existing.add(key)

                MAKEUP_TEST_WRITE_ROW += 1

        # ✅ 정렬 및 테두리: "추가된 행 범위만" 적용
        for row in range(MAKEUP_TEST_RANGE, MAKEUP_TEST_WRITE_ROW):
            for col in range(1, MakeupTestList.MAX + 1):
                cell = ws.cell(row, col)
                cell.alignment = ALIGN_CENTER
                cell.border = BORDER_ALL

        return wb
    finally:
        # ✅ close (잠김/메모리 누수 방지)
        try:
            if form_wb is not None:
                form_wb.close()
        except Exception:
            pass
        try:
            if student_wb is not None:
                student_wb.close()
        except Exception:
            pass

def save_makeup_test_result(target_row:int, makeup_test_score:str) -> bool:
    wb = open()
    ws = open_worksheet(wb)

    ws.cell(target_row, MakeupTestList.MAKEUPTEST_SCORE_COLUMN).value = makeup_test_score

    save(wb)

    return True

def save_individual_makeup_test(student_name:str, class_name:str, test_name:str, test_score:int|float, makeup_test_date:dict, prog:Progress):
    wb = open()
    ws = open_worksheet(wb)

    student_wb = omikron.studentinfo.open(True)
    student_ws = omikron.studentinfo.open_worksheet(student_wb)

    class_wb = omikron.classinfo.open(True)
    class_ws = omikron.classinfo.open_worksheet(class_wb)

    for row in range(ws.max_row+1, 1, -1):
        if ws.cell(row-1, MakeupTestList.TEST_DATE_COLUMN).value is not None:
            MAKEUP_TEST_WRITE_ROW = row
            break

    exist, teacher_name, _, _, _ = omikron.classinfo.get_class_info(class_name, class_ws)
    if not exist:
        prog.warning(f"{class_name}의 반 정보가 존재하지 않습니다.")

    exist, makeup_test_weekday, _, new_student = omikron.studentinfo.get_student_info(student_ws, student_name)
    if not exist:
        prog.warning(f"{student_name}의 학생 정보가 존재하지 않습니다.")

    ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.TEST_DATE_COLUMN).value    = datetime.today().date()
    ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.CLASS_NAME_COLUMN).value   = class_name
    ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.TEACHER_NAME_COLUMN).value = teacher_name
    ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.STUDENT_NAME_COLUMN).value = student_name
    ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.TEST_NAME_COLUMN).value    = test_name
    # ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.TEST_SCORE_COLUMN).value   = test_score

    if new_student:
        ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.STUDENT_NAME_COLUMN).fill = FILL_NEW_STUDENT

    if makeup_test_weekday is not None:
        # ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.MAKEUPTEST_WEEKDAY_COLUMN).value = makeup_test_weekday

        complete, calculated_schedule, _ = calculate_makeup_test_schedule(makeup_test_weekday, makeup_test_date)
        if not complete:
            prog.warning(f"{student_name}의 재시험 일정이 올바른 양식이 아닙니다.")

        ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.MAKEUPTEST_DATE_COLUMN).value         = calculated_schedule
        ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.MAKEUPTEST_DATE_COLUMN).number_format = "mm월 dd일(aaa)"

        # if makeup_test_time is not None:
            # ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.MAKEUPTEST_TIME_COLUMN).value = makeup_test_time

    for col in range(1, MakeupTestList.MAX + 1):
        ws.cell(MAKEUP_TEST_WRITE_ROW, col).alignment = ALIGN_CENTER
        ws.cell(MAKEUP_TEST_WRITE_ROW, col).border    = BORDER_ALL

    save(wb)
