import os
import openpyxl as xl

from datetime import datetime
from openpyxl.styles import Alignment, Border, Side, Protection
from openpyxl.utils.cell import get_column_letter as gcl
from openpyxl.worksheet.datavalidation import DataValidation

import omikron.chrome
import omikron.classinfo
import omikron.config

from omikron.defs import DataForm
from omikron.exception import NoMatchingSheetException

class DataValidationException(Exception):
    pass

# 파일 기본 작업
def make_file() -> bool:
    wb = xl.Workbook()
    ws = wb.worksheets[0]
    ws.title = DataForm.DEFAULT_NAME
    ws[gcl(DataForm.CLASS_WEEKDAY_COLUMN)+"1"]     = "요일"
    ws[gcl(DataForm.TEST_TIME_COLUMN)+"1"]         = "시간"
    ws[gcl(DataForm.CLASS_NAME_COLUMN)+"1"]        = "반"
    ws[gcl(DataForm.STUDENT_NAME_COLUMN)+"1"]      = "이름"
    ws[gcl(DataForm.TEACHER_NAME_COLUMN)+"1"]      = "담당T"
    ws[gcl(DataForm.DAILYTEST_NAME_COLUMN)+"1"]    = "시험명"
    ws[gcl(DataForm.DAILYTEST_SCORE_COLUMN)+"1"]   = "점수"
    ws[gcl(DataForm.DAILYTEST_AVERAGE_COLUMN)+"1"] = "평균"
    ws[gcl(DataForm.MOCKTEST_NAME_COLUMN)+"1"]     = "모의고사 시험명"
    ws[gcl(DataForm.MOCKTEST_SCORE_COLUMN)+"1"]    = "모의고사 점수"
    ws[gcl(DataForm.MOCKTEST_AVERAGE_COLUMN)+"1"]  = "모의고사 평균"
    ws[gcl(DataForm.MAKEUP_TEST_CHECK_COLUMN)+"1"] = "재시험 응시 여부"
    ws["Y1"] = "X"
    ws["Z1"] = "x"
    ws.column_dimensions.group("Y", "Z", hidden=True)
    ws.auto_filter.ref = "A:"+gcl(DataForm.TEST_TIME_COLUMN)
    ws.freeze_panes    = "A2"
    
    for col in range(1, DataForm.MAX+1):
        ws.cell(1, col).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        ws.cell(1, col).border    = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    class_wb = omikron.classinfo.open(True)
    class_ws = omikron.classinfo.open_worksheet(class_wb)

    for class_name, student_names in omikron.chrome.get_class_student_dict().items():
        if len(student_names) == 0:
            continue

        exist, teacher_name, class_weekday, test_time, _ = omikron.classinfo.get_class_info(class_name, ws=class_ws)
        if not exist: continue

        WRITE_LOCATION = start = ws.max_row + 1

        ws.cell(WRITE_LOCATION, DataForm.CLASS_NAME_COLUMN).value   = class_name
        ws.cell(WRITE_LOCATION, DataForm.TEACHER_NAME_COLUMN).value = teacher_name

        #학생 루프
        for student_name in student_names:
            ws.cell(WRITE_LOCATION, DataForm.CLASS_WEEKDAY_COLUMN).value = class_weekday
            ws.cell(WRITE_LOCATION, DataForm.TEST_TIME_COLUMN).value     = test_time
            ws.cell(WRITE_LOCATION, DataForm.STUDENT_NAME_COLUMN).value  = student_name
            dv = DataValidation(type="list", formula1="=Y1:Z1", showDropDown=True, allow_blank=True, showErrorMessage=True)
            dv.error = "이 셀의 값은 'x' 또는 'X'이어야 합니다."
            ws.add_data_validation(dv)
            dv.add(ws.cell(WRITE_LOCATION,DataForm.MAKEUP_TEST_CHECK_COLUMN))
            WRITE_LOCATION = ws.max_row + 1
        
        end = WRITE_LOCATION - 1

        # 시험 평균
        ws.cell(start, DataForm.DAILYTEST_AVERAGE_COLUMN).value = f"=ROUND(AVERAGE({gcl(DataForm.DAILYTEST_SCORE_COLUMN)}{start}:{gcl(DataForm.DAILYTEST_SCORE_COLUMN)}{end}), 0)"
        # 모의고사 평균
        ws.cell(start, DataForm.MOCKTEST_AVERAGE_COLUMN).value  = f"=ROUND(AVERAGE({gcl(DataForm.MOCKTEST_SCORE_COLUMN)}{start}:{gcl(DataForm.MOCKTEST_SCORE_COLUMN)}{end}), 0)"
        
        # 정렬 및 테두리
        for row in range(start, end + 1):
            for col in range(1, DataForm.MAX+1):
                ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row, col).border    = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        
        # 셀 병합
        if start < end:
            ws.merge_cells(f"{gcl(DataForm.CLASS_NAME_COLUMN)}{start}:{gcl(DataForm.CLASS_NAME_COLUMN)}{end}")
            ws.merge_cells(f"{gcl(DataForm.TEACHER_NAME_COLUMN)}{start}:{gcl(DataForm.TEACHER_NAME_COLUMN)}{end}")
            ws.merge_cells(f"{gcl(DataForm.DAILYTEST_NAME_COLUMN)}{start}:{gcl(DataForm.DAILYTEST_NAME_COLUMN)}{end}")
            ws.merge_cells(f"{gcl(DataForm.DAILYTEST_AVERAGE_COLUMN)}{start}:{gcl(DataForm.DAILYTEST_AVERAGE_COLUMN)}{end}")
            ws.merge_cells(f"{gcl(DataForm.MOCKTEST_NAME_COLUMN)}{start}:{gcl(DataForm.MOCKTEST_NAME_COLUMN)}{end}")
            ws.merge_cells(f"{gcl(DataForm.MOCKTEST_AVERAGE_COLUMN)}{start}:{gcl(DataForm.MOCKTEST_AVERAGE_COLUMN)}{end}")
        
    ws.protection.sheet         = True
    ws.protection.autoFilter    = False
    ws.protection.formatColumns = False
    for row in range(2, ws.max_row + 1):
        ws.cell(row, DataForm.CLASS_NAME_COLUMN).alignment         = Alignment(horizontal="center", vertical="center", wrapText=True)
        ws.cell(row, DataForm.DAILYTEST_NAME_COLUMN).alignment     = Alignment(horizontal="center", vertical="center", wrapText=True)
        ws.cell(row, DataForm.MOCKTEST_NAME_COLUMN).alignment      = Alignment(horizontal="center", vertical="center", wrapText=True)
        ws.cell(row, DataForm.DAILYTEST_NAME_COLUMN).protection    = Protection(locked=False)
        ws.cell(row, DataForm.DAILYTEST_SCORE_COLUMN).protection   = Protection(locked=False)
        ws.cell(row, DataForm.MOCKTEST_NAME_COLUMN).protection     = Protection(locked=False)
        ws.cell(row, DataForm.MOCKTEST_SCORE_COLUMN).protection    = Protection(locked=False)
        ws.cell(row, DataForm.MAKEUP_TEST_CHECK_COLUMN).protection = Protection(locked=False)

    if os.path.isfile(f"{omikron.config.DATA_DIR}/데일리테스트 기록 양식({datetime.today().strftime('%m.%d')}).xlsx"):
        i = 1
        while True:
            if not os.path.isfile(f"{omikron.config.DATA_DIR}/데일리테스트 기록 양식({datetime.today().strftime('%m.%d')}) ({i}).xlsx"):
                wb.save(f"{omikron.config.DATA_DIR}/데일리테스트 기록 양식({datetime.today().strftime('%m.%d')}) ({i}).xlsx")
                break
            i += 1
    else:
        wb.save(f"{omikron.config.DATA_DIR}/데일리테스트 기록 양식({datetime.today().strftime('%m.%d')}).xlsx")

    return True

def open(filepath, data_only=True) -> xl.Workbook:
    return xl.load_workbook(filepath, data_only=data_only)

def open_worksheet(wb:xl.Workbook):
    try:
        return wb[DataForm.DEFAULT_NAME]
    except:
        raise NoMatchingSheetException(f"'{DataForm.DEFAULT_NAME}.xlsx'의 시트명을 '{DataForm.DEFAULT_NAME}'로 변경해 주세요.")

# 파일 유틸리티
def data_validation(filepath:str) -> bool:
    """
    데이터 입력 양식의 데이터가 올바르게 입력되었는지 확인
    """
    errors = []
    wb = open(filepath)
    ws = open_worksheet(wb)
    
    form_checked      = True
    dailytest_checked = False
    mocktest_checked  = False
    for i in range(1, ws.max_row+1):
        if ws.cell(i, DataForm.CLASS_NAME_COLUMN).value is not None:
            class_name        = ws.cell(i, DataForm.CLASS_NAME_COLUMN).value
            dailytest_checked = False
            mocktest_checked  = False
            dailytest_name    = ws.cell(i, DataForm.DAILYTEST_NAME_COLUMN).value
            mocktest_name     = ws.cell(i, DataForm.MOCKTEST_NAME_COLUMN).value
        
        if dailytest_checked and mocktest_checked: continue
        
        if not dailytest_checked and ws.cell(i, DataForm.DAILYTEST_SCORE_COLUMN).value is not None and dailytest_name is None:
            errors.append(f"{class_name}의 시험명이 작성되지 않았습니다.")
            dailytest_checked = True
            form_checked      = False
        if not mocktest_checked and ws.cell(i, DataForm.MOCKTEST_SCORE_COLUMN).value is not None and mocktest_name is None:
            errors.append(f"{class_name}의 모의고사명이 작성되지 않았습니다.")
            mocktest_checked = True
            form_checked     = False

    if errors:
        raise DataValidationException("\n".join(errors))

    return form_checked
