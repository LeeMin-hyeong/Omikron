from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side

_SIDE_THIN_9090 = Side(border_style="thin", color="909090")
_SIDE_MEDIUM_000 = Side(border_style="medium", color="000000")

BORDER_BOTTOM_MEDIUM_000 = Border(bottom=_SIDE_MEDIUM_000)
BORDER_BOTTOM_THIN_9090  = Border(bottom=_SIDE_THIN_9090)
BORDER_TOP_THIN_9090_BOTTOM_MEDIUM_000 = Border(top=_SIDE_THIN_9090, bottom=_SIDE_MEDIUM_000)
BORDER_TOP_MEDIUM_000 = Border(top=_SIDE_MEDIUM_000)
BORDER_ALL = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrapText=True)

FONT_BOLD = Font(bold=True)
FONT_STRIKE = Font(strike=True)
FONT_BOLD_STRIKE = Font(bold=True, strike=True)
FONT_RED = Font(color="FFFF0000")
FONT_BOLD_RED = Font(bold=True, color="FFFF0000")

FILL_NONE = PatternFill(fill_type=None)
FILL_NEW_STUDENT = PatternFill(fill_type="solid", fgColor=Color("FFFF00"))
FILL_BELOW_60 = PatternFill(fill_type="solid", fgColor=Color("EC7E31"))
FILL_BELOW_70 = PatternFill(fill_type="solid", fgColor=Color("F5AF85"))
FILL_BELOW_80 = PatternFill(fill_type="solid", fgColor=Color("FCE4D6"))
FILL_CLASS_AVG = PatternFill(fill_type="solid", fgColor=Color("DDEBF7"))
FILL_STUDENT_AVG = PatternFill(fill_type="solid", fgColor=Color("E2EFDA"))