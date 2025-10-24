import os.path
import openpyxl as xl

from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils.cell import get_column_letter as gcl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation

import omikron.chrome

from omikron.defs import StudentInfo
from omikron.exception import NoMatchingSheetException, FileOpenException

# 파일 기본 작업
def make_file() -> bool:
    wb = xl.Workbook()
    ws = wb.worksheets[0]
    ws.title = StudentInfo.DEFAULT_NAME

    ws[gcl(StudentInfo.STUDENT_NAME_COLUMN)+"1"]       = "이름"
    ws[gcl(StudentInfo.MAKEUPTEST_WEEKDAY_COLUMN)+"1"] = "재시험 응시 요일"
    ws[gcl(StudentInfo.MAKEUPTEST_TIME_COLUMN)+"1"]    = "재시험 응시 시간"
    ws[gcl(StudentInfo.NEW_STUDENT_CHECK_COLUMN)+"1"]  = "기수 신규생"

    ws["Z1"] = "N"
    ws.auto_filter.ref = "A:"+gcl(StudentInfo.MAX)
    ws.freeze_panes    = "A2"
    ws.column_dimensions.group("Z", hidden=True)

    # 첫 행 정렬 및 자동 줄 바꿈
    for col in range(1, StudentInfo.MAX+1):
        ws.cell(1, col).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        ws.cell(1, col).border    = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    return update_student(wb)

def open(data_only:bool=False) -> xl.Workbook:
    return xl.load_workbook(f"{omikron.config.DATA_DIR}/{StudentInfo.DEFAULT_NAME}.xlsx", data_only=data_only)

def open_worksheet(wb:xl.Workbook):
    try:
        return wb[StudentInfo.DEFAULT_NAME]
    except:
        raise NoMatchingSheetException(f"'{StudentInfo.DEFAULT_NAME}.xlsx'의 시트명을 '{StudentInfo.DEFAULT_NAME}'으로 변경해 주세요.")

def save(wb:xl.Workbook):
    try:
        wb.save(f"{omikron.config.DATA_DIR}/{StudentInfo.DEFAULT_NAME}.xlsx")
    except:
        raise FileOpenException()

def isopen() -> bool:
    return os.path.isfile(f"{omikron.config.DATA_DIR}/~${StudentInfo.DEFAULT_NAME}.xlsx")

# 파일 유틸리티
def get_student_info(ws:Worksheet, student_name:str):
    """
    학생 정보 파일로부터 학생 정보 추출

    return 파일 내 학생 존재 여부, 재시험 요일, 재시험 시간, 신규생 여부
    """
    for row in range(2, ws.max_row+1):
        if ws.cell(row, StudentInfo.STUDENT_NAME_COLUMN).value == student_name:
            makeup_test_weekday = ws.cell(row, StudentInfo.MAKEUPTEST_WEEKDAY_COLUMN).value
            makeup_test_time    = ws.cell(row, StudentInfo.MAKEUPTEST_TIME_COLUMN).value
            new_studnet         = ws.cell(row, StudentInfo.NEW_STUDENT_CHECK_COLUMN).value
            break
    else:
        return False, None, None, False
    
    return True, makeup_test_weekday, makeup_test_time, new_studnet == 'N'

# 파일 작업
def add_student(target_student_name:str):
    """
    학생 정보 파일 내 신규생 추가
    """
    wb = open()
    ws = open_worksheet(wb)

    for row in range(ws.max_row+1, 1, -1):
        if ws.cell(row-1, StudentInfo.STUDENT_NAME_COLUMN).value is not None:
            ws.cell(row, StudentInfo.STUDENT_NAME_COLUMN).value      = target_student_name
            # ws.cell(row, StudentInfo.CLASS_NAME_COLUMN).value      = target_class_name
            ws.cell(row, StudentInfo.NEW_STUDENT_CHECK_COLUMN).value = "N"
            for col in range(1, StudentInfo.MAX+1):
                ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row, col).border    = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
            break

    save(wb)

def delete_student(target_student_name:str):
    """
    학생 정보 파일에서 학생 정보 삭제
    """
    wb = open()
    ws = open_worksheet(wb)

    for row in range(2, ws.max_row+1):
        if ws.cell(row, StudentInfo.STUDENT_NAME_COLUMN).value == target_student_name:
            ws.delete_rows(row)

    save(wb)

def update_student(wb:xl.Workbook=None):
    latest_student_names = omikron.chrome.get_student_names()

    if wb is None:
        wb = open()

    ws = open_worksheet(wb)

    student_names = [ws.cell(row, StudentInfo.STUDENT_NAME_COLUMN).value for row in range(2, ws.max_row+1)]
    
    deleted_student_names      = list(set(student_names).difference(latest_student_names))
    unregistered_student_names = list(set(latest_student_names).difference(student_names))
    
    for row in range(ws.max_row+1, 1, -1):
        if ws.cell(row-1, StudentInfo.STUDENT_NAME_COLUMN).value is not None:
            WRITE_ROW = row
            break
    
    for student_name in sorted(unregistered_student_names):
        ws.cell(WRITE_ROW, StudentInfo.STUDENT_NAME_COLUMN).value = student_name

        dv = DataValidation(type="list", formula1="=Z1", allow_blank=True, errorStyle="stop", showErrorMessage=True)
        dv.error = "이 셀의 값은 'N'이어야 합니다."
        ws.add_data_validation(dv)
        dv.add(ws.cell(WRITE_ROW, StudentInfo.NEW_STUDENT_CHECK_COLUMN))

        for col in range(1, StudentInfo.MAX+1):
            ws.cell(WRITE_ROW, col).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(WRITE_ROW, col).border    = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        WRITE_ROW += 1

    for row in range(2, ws.max_row+1):
        while ws.cell(row, StudentInfo.STUDENT_NAME_COLUMN).value in deleted_student_names:
            ws.delete_rows(row)

    save(wb)
