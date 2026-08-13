"""Creation, opening, saving, and backup of the main data workbook."""

import os

import openpyxl as xl
from openpyxl.utils.cell import get_column_letter as gcl
from openpyxl.worksheet.formula import ArrayFormula

import tdm.aisosik.reader
import tdm.excel.class_info
from tdm.domain.errors import FileOpenException, TDMError
from tdm.domain.models import DataFile
from tdm.excel.atomic import atomic_save_workbook
from tdm.excel.backup import create_backup
from tdm.excel.paths import WorkbookPaths
from tdm.excel.styles import ALIGN_CENTER, BORDER_BOTTOM_MEDIUM_000, BORDER_BOTTOM_THIN_9090, BORDER_TOP_THIN_9090_BOTTOM_MEDIUM_000, FONT_BOLD
from tdm.excel.validation import validate_workbook_structure
from tdm.excel.workbook_io import load_workbook

# 파일 기본 작업

def make_file():
    wb = xl.Workbook()
    ws = wb.worksheets[0]
    ws.title = DataFile.DEFAULT_SHEET_NAME
    ws[gcl(DataFile.CLASS_NAME_COLUMN)+"1"]    = "반"
    ws[gcl(DataFile.TEACHER_NAME_COLUMN)+"1"]  = "담당"
    ws[gcl(DataFile.STUDENT_NAME_COLUMN)+"1"]  = "이름"
    ws[gcl(DataFile.AVERAGE_SCORE_COLUMN)+"1"] = "학생 평균"
    ws.freeze_panes    = f"{gcl(DataFile.DATA_COLUMN)}2"
    ws.auto_filter.ref = f"A:{gcl(DataFile.MAX)}"

    for col in range(1, DataFile.DATA_COLUMN):
        ws.cell(1, col).border = BORDER_BOTTOM_MEDIUM_000

    class_wb = tdm.excel.class_info.open(True)
    class_ws = tdm.excel.class_info.open_worksheet(class_wb)

    # 반 루프
    for class_name, student_list in tdm.aisosik.reader.get_class_student_dict().items():
        if len(student_list) == 0:
            continue

        exist, teacher_name, _, _, mock_test_check = tdm.excel.class_info.get_class_info(class_name, ws=class_ws)
        if not exist:
            continue

        for i in range(2):
            if i == 1 and not mock_test_check:
                continue

            if i == 1:
                class_name = class_name + " (모의고사)"

            WRITE_LOCATION = ws.max_row + 1

            # 시험명
            ws.cell(WRITE_LOCATION, DataFile.CLASS_NAME_COLUMN).value    = class_name
            ws.cell(WRITE_LOCATION, DataFile.TEACHER_NAME_COLUMN).value  = teacher_name
            ws.cell(WRITE_LOCATION, DataFile.STUDENT_NAME_COLUMN).value  = "날짜"
            
            WRITE_LOCATION = ws.max_row + 1
            ws.cell(WRITE_LOCATION, DataFile.CLASS_NAME_COLUMN).value    = class_name
            ws.cell(WRITE_LOCATION, DataFile.TEACHER_NAME_COLUMN).value  = teacher_name
            ws.cell(WRITE_LOCATION, DataFile.STUDENT_NAME_COLUMN).value  = "시험명"

            for col in range(1, DataFile.DATA_COLUMN):
                ws.cell(WRITE_LOCATION, col).border = BORDER_BOTTOM_THIN_9090

            class_start = WRITE_LOCATION + 1

            # 학생 루프
            for student_name in student_list:
                WRITE_LOCATION = ws.max_row + 1
                ws.cell(WRITE_LOCATION, DataFile.CLASS_NAME_COLUMN).value    = class_name
                ws.cell(WRITE_LOCATION, DataFile.TEACHER_NAME_COLUMN).value  = teacher_name
                ws.cell(WRITE_LOCATION, DataFile.STUDENT_NAME_COLUMN).value  = student_name
                ws.cell(WRITE_LOCATION, DataFile.AVERAGE_SCORE_COLUMN).value = f"=ROUND(AVERAGE({gcl(DataFile.DATA_COLUMN)}{WRITE_LOCATION}:XFD{WRITE_LOCATION}), 0)"
                ws.cell(WRITE_LOCATION, DataFile.AVERAGE_SCORE_COLUMN).font  = FONT_BOLD
            
            # 시험별 평균
            class_end = WRITE_LOCATION
            WRITE_LOCATION = ws.max_row + 1
            ws.cell(WRITE_LOCATION, DataFile.CLASS_NAME_COLUMN).value    = class_name
            ws.cell(WRITE_LOCATION, DataFile.TEACHER_NAME_COLUMN).value  = teacher_name
            ws.cell(WRITE_LOCATION, DataFile.STUDENT_NAME_COLUMN).value  = "시험 평균"
            ws[f"{gcl(DataFile.AVERAGE_SCORE_COLUMN)}{WRITE_LOCATION}"] = ArrayFormula(
                f"{gcl(DataFile.AVERAGE_SCORE_COLUMN)}{WRITE_LOCATION}",
                f"=ROUND(SUM(IFERROR({gcl(DataFile.AVERAGE_SCORE_COLUMN)}{class_start}:{gcl(DataFile.AVERAGE_SCORE_COLUMN)}{class_end},0))/COUNT({gcl(DataFile.AVERAGE_SCORE_COLUMN)}{class_start}:{gcl(DataFile.AVERAGE_SCORE_COLUMN)}{class_end}),0)",
            )
            ws.cell(WRITE_LOCATION, DataFile.AVERAGE_SCORE_COLUMN).font = FONT_BOLD

            for col in range(1, DataFile.DATA_COLUMN):
                ws.cell(WRITE_LOCATION, col).border = BORDER_TOP_THIN_9090_BOTTOM_MEDIUM_000

    # 정렬
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).alignment = ALIGN_CENTER

    class_wb.close()
    save(wb)

def open(data_only:bool=False, read_only:bool=False) -> xl.Workbook:
    paths = WorkbookPaths.current()
    return load_workbook(
        paths.data_file,
        display_name=paths.data_file_name,
        data_only=data_only,
        read_only=read_only,
    )

def open_temp(data_only:bool=False, read_only:bool=False) -> xl.Workbook:
    return load_workbook(
        WorkbookPaths.current().data_temp,
        display_name=DataFile.TEMP_FILE_NAME,
        data_only=data_only,
        read_only=read_only,
    )

def save(wb:xl.Workbook):
    paths = WorkbookPaths.current()
    try:
        paths.data_dir.mkdir(parents=True, exist_ok=True)
        atomic_save_workbook(wb, paths.data_file)
    except TDMError:
        raise
    except Exception as exc:
        raise FileOpenException(
            f"{paths.data_file_name} 파일을 닫은 뒤 다시 시도해주세요"
        ) from exc
    finally:
        wb.close()

def save_to_temp(wb:xl.Workbook):
    paths = WorkbookPaths.current()
    paths.data_dir.mkdir(parents=True, exist_ok=True)
    try:
        atomic_save_workbook(wb, paths.data_temp, update_source=False)
        os.system(f'attrib +h "{paths.data_temp}"')
    finally:
        wb.close()

def delete_temp():
    try:
        WorkbookPaths.current().data_temp.unlink(missing_ok=True)
    except OSError:
        pass

def file_validation():
    validate_workbook_structure(
        WorkbookPaths.current().data_file,
        sheet_name=DataFile.DEFAULT_SHEET_NAME,
        required_headers=("반", "담당", "이름", "학생 평균"),
    )

# 파일 유틸리티

def make_backup_file():
    paths = WorkbookPaths.current()
    return create_backup(
        paths.data_file,
        stem=paths.data_file_name,
        backup_dir=paths.backup_dir,
    )
