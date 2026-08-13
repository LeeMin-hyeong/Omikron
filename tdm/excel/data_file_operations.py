"""Exam, formatting, class, and student mutations for the data workbook."""

from datetime import datetime

import openpyxl as xl
from openpyxl.utils.cell import get_column_letter as gcl
from openpyxl.worksheet.formula import ArrayFormula

import tdm.aisosik.reader
import tdm.excel.class_info
import tdm.excel.data_form
import tdm.excel.student_info
from tdm.domain.errors import InvalidOperationError
from tdm.domain.models import DataFile, DataForm
from tdm.domain.progress import Progress
from tdm.excel.data_file_queries import find_dynamic_columns, get_class_names
from tdm.excel.atomic import (
    atomic_save_workbook,
    track_workbook_source,
    workbook_source_revision,
)
from tdm.excel.com_adapter import recalculate_workbook
from tdm.excel.data_file_storage import (
    delete_temp,
    file_validation,
    make_backup_file,
    open,
    open_temp,
    save,
    save_to_temp,
)
from tdm.excel.paths import WorkbookPaths
from tdm.excel.styles import (
    ALIGN_CENTER,
    ALIGN_CENTER_WRAP,
    BORDER_BOTTOM_MEDIUM_000,
    BORDER_BOTTOM_THIN_9090,
    BORDER_TOP_MEDIUM_000,
    BORDER_TOP_THIN_9090_BOTTOM_MEDIUM_000,
    FILL_NEW_STUDENT,
    FILL_NONE,
    FONT_BOLD,
    FONT_BOLD_RED,
    FONT_BOLD_STRIKE,
    FONT_RED,
    FONT_STRIKE,
)
from tdm.excel.utils import class_average_color, copy_cell, student_average_color, test_score_color
from tdm.excel.workbook_io import close_workbooks, load_workbook

def save_test_data(filepath:str, prog: Progress):
    """
    데이터 양식에 작성된 데이터를 데이터 파일에 저장
    """
    # 임시 파일 삭제
    if WorkbookPaths.current().data_temp.is_file():
        delete_temp()

    form_wb = tdm.excel.data_form.open(filepath)
    form_ws = tdm.excel.data_form.open_worksheet(form_wb)

    # 학생 정보 열기
    student_wb = tdm.excel.student_info.open(True)
    student_ws = tdm.excel.student_info.open_worksheet(student_wb)

    file_validation()

    # 백업 생성
    make_backup_file()
    prog.step("백업 생성 완료")

    wb = open()
    ws = wb[DataFile.DEFAULT_SHEET_NAME]
    source_revision = workbook_source_revision(wb, WorkbookPaths.current().data_file)

    CLASS_NAME_COLUMN, _, STUDENT_NAME_COLUMN, AVERAGE_SCORE_COLUMN = find_dynamic_columns(ws)

    for t in range(2):
        if t == 0:
            TEST_NAME_COLUMN    = DataForm.DAILYTEST_NAME_COLUMN
            TEST_SCORE_COLUMN   = DataForm.DAILYTEST_SCORE_COLUMN
            TEST_AVERAGE_COLUMN = DataForm.DAILYTEST_AVERAGE_COLUMN
        else:
            TEST_NAME_COLUMN    = DataForm.MOCKTEST_NAME_COLUMN
            TEST_SCORE_COLUMN   = DataForm.MOCKTEST_SCORE_COLUMN
            TEST_AVERAGE_COLUMN = DataForm.MOCKTEST_AVERAGE_COLUMN

        for i in range(2, form_ws.max_row+1): # 데일리데이터 기록 양식 루프
            # 반 필터링
            if (form_ws.cell(i, DataForm.CLASS_NAME_COLUMN).value is not None) and (form_ws.cell(i, TEST_NAME_COLUMN).value is not None):
                class_name   = form_ws.cell(i, DataForm.CLASS_NAME_COLUMN).value
                if t == 1:
                    class_name += " (모의고사)"
                test_name    = form_ws.cell(i, TEST_NAME_COLUMN).value
                test_average = form_ws.cell(i, TEST_AVERAGE_COLUMN).value

                no_class = False

                if test_name is None:
                    continue

                #반 시작 찾기
                for row in range(2, ws.max_row+1):
                    if ws.cell(row, CLASS_NAME_COLUMN).value == class_name:
                        CLASS_START = row
                        break
                else:
                    prog.warning(f"{class_name} 반이 존재하지 않습니다.")
                    no_class = True
                    continue

                # 반 끝 찾기
                for row in range(CLASS_START, ws.max_row+1):
                    if ws.cell(row, STUDENT_NAME_COLUMN).value == "시험 평균":
                        CLASS_END = row
                        break

                # 데이터 작성 열 찾기
                for col in range(AVERAGE_SCORE_COLUMN+1, ws.max_column+2):
                    test_date = ws.cell(CLASS_START, col).value
                    if isinstance(test_date, datetime) and test_date.strftime(
                        "%y%m%d"
                    ) == datetime.today().strftime("%y%m%d"):
                        WRITE_COLUMN = col
                        break
                    elif test_date is None:
                        WRITE_COLUMN = col
                        break

                # 입력 틀 작성
                AVERAGE_FORMULA = f"=ROUND(AVERAGE({gcl(WRITE_COLUMN)+str(CLASS_START + 2)}:{gcl(WRITE_COLUMN)+str(CLASS_END - 1)}), 0)"
                ws.column_dimensions[gcl(WRITE_COLUMN)].width    = 14
                ws.cell(CLASS_START, WRITE_COLUMN).value         = datetime.today().date()
                ws.cell(CLASS_START, WRITE_COLUMN).number_format = "yyyy.mm.dd(aaa)"
                ws.cell(CLASS_START, WRITE_COLUMN).alignment     = ALIGN_CENTER
                ws.cell(CLASS_START, WRITE_COLUMN).border        = BORDER_TOP_MEDIUM_000

                ws.cell(CLASS_START + 1, WRITE_COLUMN).value     = test_name
                ws.cell(CLASS_START + 1, WRITE_COLUMN).alignment = ALIGN_CENTER_WRAP
                ws.cell(CLASS_START + 1, WRITE_COLUMN).border    = BORDER_BOTTOM_THIN_9090

                ws.cell(CLASS_END, WRITE_COLUMN).value           = AVERAGE_FORMULA
                ws.cell(CLASS_END, WRITE_COLUMN).font            = FONT_BOLD
                ws.cell(CLASS_END, WRITE_COLUMN).alignment       = ALIGN_CENTER
                ws.cell(CLASS_END, WRITE_COLUMN).border          = BORDER_TOP_THIN_9090_BOTTOM_MEDIUM_000
                
                if type(test_average) in (int, float):
                    ws.cell(CLASS_END, WRITE_COLUMN).fill = class_average_color(test_average)

            test_score   = form_ws.cell(i, TEST_SCORE_COLUMN).value
            student_name = form_ws.cell(i, DataForm.STUDENT_NAME_COLUMN).value

            if test_score is None:
                continue
            if no_class:
                continue
            if test_name is None:
                continue

            # 학생 찾기
            for row in range(CLASS_START + 2, CLASS_END):
                if ws.cell(row, STUDENT_NAME_COLUMN).value == student_name:
                    ws.cell(row, WRITE_COLUMN).value = test_score
                    if type(test_score) in (int, float):
                        ws.cell(row, WRITE_COLUMN).fill = test_score_color(test_score)

                    ws.cell(row, WRITE_COLUMN).alignment = ALIGN_CENTER
                    break
            else:
                prog.warning(f"{class_name} 반에 {student_name} 학생이 존재하지 않습니다.")

    ws = wb[DataFile.DEFAULT_SHEET_NAME]
    save_to_temp(wb)
    prog.step("데이터 저장 완료")

    # 조건부 서식 수식 로딩
    recalculate_workbook(WorkbookPaths.current().data_temp)

    wb           = open_temp()
    track_workbook_source(wb, WorkbookPaths.current().data_file, source_revision)
    data_only_wb = open_temp(data_only=True)

    ws           = wb[DataFile.DEFAULT_SHEET_NAME]
    data_only_ws = data_only_wb[DataFile.DEFAULT_SHEET_NAME]

    _, _, STUDENT_NAME_COLUMN, AVERAGE_SCORE_COLUMN = find_dynamic_columns(ws)

    for row in range(2, data_only_ws.max_row+1):
        if data_only_ws.cell(row, STUDENT_NAME_COLUMN).value is None:
            break

        # 학생 별 평균 점수에 대한 조건부 서식
        student_average = data_only_ws.cell(row, AVERAGE_SCORE_COLUMN).value
        if type(student_average) in (int, float):
            if ws.cell(row, STUDENT_NAME_COLUMN).value == "시험 평균":
                ws.cell(row, AVERAGE_SCORE_COLUMN).fill = class_average_color(student_average)
            else:
                ws.cell(row, AVERAGE_SCORE_COLUMN).fill = student_average_color(student_average)

        # 신규생 하이라이트
        if ws.cell(row, STUDENT_NAME_COLUMN).value in ("날짜", "시험명", "시험 평균"):
            continue
        if ws.cell(row, STUDENT_NAME_COLUMN).font.strike:
            continue
        if ws.cell(row, STUDENT_NAME_COLUMN).font.color is not None and ws.cell(row, STUDENT_NAME_COLUMN).font.color.rgb == "FFFF0000":
            continue

        exist, _, _, new_student = tdm.excel.student_info.get_student_info(student_ws, ws.cell(row, STUDENT_NAME_COLUMN).value)
        if exist:
            if new_student:
                ws.cell(row, STUDENT_NAME_COLUMN).fill = FILL_NEW_STUDENT
            else:
                ws.cell(row, STUDENT_NAME_COLUMN).fill = FILL_NONE
        else:
            ws.cell(row, STUDENT_NAME_COLUMN).fill = FILL_NONE
            prog.warning(f"{ws.cell(row, STUDENT_NAME_COLUMN).value} 학생 정보가 존재하지 않습니다.")

    ws = wb[DataFile.DEFAULT_SHEET_NAME]
    prog.step("조건부 서식 로딩 완료")

    close_workbooks(form_wb, student_wb, data_only_wb)
    return wb

def prepare_individual_test_data(
    target_row: int,
    target_col: int,
    test_score: int | float,
) -> tuple[xl.Workbook, int | float | None]:
    """Prepare an individual result without committing the main workbook."""
    # 임시 파일 삭제
    if WorkbookPaths.current().data_temp.is_file():
        delete_temp()

    file_validation()

    # 백업 생성
    make_backup_file()

    wb = open()
    ws = wb[DataFile.DEFAULT_SHEET_NAME]
    source_revision = workbook_source_revision(wb, WorkbookPaths.current().data_file)

    # 시험 점수 기록
    ws.cell(target_row, target_col).value     = test_score
    ws.cell(target_row, target_col).fill      = test_score_color(test_score)
    ws.cell(target_row, target_col).alignment = ALIGN_CENTER

    save_to_temp(wb)

    recalculate_workbook(WorkbookPaths.current().data_temp)

    wb           = open_temp()
    track_workbook_source(wb, WorkbookPaths.current().data_file, source_revision)
    data_only_wb = open_temp(True)

    ws           = wb[DataFile.DEFAULT_SHEET_NAME]
    data_only_ws = data_only_wb[DataFile.DEFAULT_SHEET_NAME]

    _, _, STUDENT_NAME_COLUMN, AVERAGE_SCORE_COLUMN = find_dynamic_columns(ws)

    # 학생 평균 조건부 서식 반영
    student_average = data_only_ws.cell(target_row, AVERAGE_SCORE_COLUMN).value
    if type(student_average)  in (int, float):
        ws.cell(target_row, AVERAGE_SCORE_COLUMN).fill = student_average_color(student_average)

    # 시험 평균 조건부 서식 반영
    test_average_row = target_row
    while data_only_ws.cell(test_average_row, STUDENT_NAME_COLUMN).value != "시험 평균":
        test_average_row += 1

    test_average = data_only_ws.cell(test_average_row, target_col).value
    if type(test_average) in (int, float):
        ws.cell(test_average_row, target_col).fill = class_average_color(test_average)

    # 반 평균 조건부 서식 반영
    class_average = data_only_ws.cell(test_average_row, AVERAGE_SCORE_COLUMN).value
    if type(class_average) in (int, float):
        ws.cell(test_average_row, AVERAGE_SCORE_COLUMN).fill = class_average_color(class_average)

    data_only_wb.close()
    delete_temp()

    return wb, test_average


def save_individual_test_data(target_row:int, target_col:int, test_score:int|float):
    """정규 시험에 미응시한 학생의 결과를 입력하고 해당 반의 평균을 반환"""
    wb, test_average = prepare_individual_test_data(
        target_row,
        target_col,
        test_score,
    )
    try:
        save(wb)
    finally:
        wb.close()

    return test_average

def conditional_formatting():
    file_validation()

    recalculate_workbook(WorkbookPaths.current().data_file)

    warnings = []

    wb           = open()
    data_only_wb = open(data_only=True, read_only=True)
    student_wb   = tdm.excel.student_info.open()
    student_ws   = tdm.excel.student_info.open_worksheet(student_wb)

    ws           = wb[DataFile.DEFAULT_SHEET_NAME]
    data_only_ws = data_only_wb[DataFile.DEFAULT_SHEET_NAME]

    _, _, STUDENT_NAME_COLUMN, AVERAGE_SCORE_COLUMN = find_dynamic_columns(ws)
    date_row: int | None = None

    for row in range(2, ws.max_row+1):
        if ws.cell(row, STUDENT_NAME_COLUMN).value is None:
            break
        if ws.cell(row, STUDENT_NAME_COLUMN).value == "날짜":
            date_row = row
        if ws.cell(row, STUDENT_NAME_COLUMN).value != "시험명":
            ws.row_dimensions[row].height = 18

        # 데이터 조건부 서식
        for col in range(1, data_only_ws.max_column+1):
            reference_row = date_row if date_row is not None else row
            if (
                col > AVERAGE_SCORE_COLUMN
                and ws.cell(reference_row, col).value is None
            ):
                break

            ws.column_dimensions[gcl(col)].width = 14
            if ws.cell(row, STUDENT_NAME_COLUMN).value == "날짜":
                ws.cell(row, col).border = BORDER_BOTTOM_MEDIUM_000
            elif ws.cell(row, STUDENT_NAME_COLUMN).value == "시험명":
                ws.cell(row, col).border = BORDER_BOTTOM_THIN_9090
            elif ws.cell(row, STUDENT_NAME_COLUMN).value == "시험 평균":
                ws.cell(row, col).border = BORDER_TOP_THIN_9090_BOTTOM_MEDIUM_000
            else:
                ws.cell(row, col).border = None

            # 학생 평균 점수 열 기준 분기   
            if col <= AVERAGE_SCORE_COLUMN:
                continue

            if ws.cell(row, STUDENT_NAME_COLUMN).value == "시험명":
                ws.cell(row, col).alignment = ALIGN_CENTER_WRAP
            elif data_only_ws.cell(row, STUDENT_NAME_COLUMN).value == "시험 평균":
                ws.cell(row, col).font = FONT_BOLD
                if type(data_only_ws.cell(row, col).value) in (int, float):
                    ws.cell(row, col).fill = class_average_color(data_only_ws.cell(row, col).value)
            elif type(data_only_ws.cell(row, col).value) in (int, float):
                ws.cell(row, col).fill = test_score_color(data_only_ws.cell(row, col).value)
            else:
                ws.cell(row, col).fill = FILL_NONE

        # 학생별 평균 조건부 서식
        if type(data_only_ws.cell(row, AVERAGE_SCORE_COLUMN).value) in (int, float):
            if ws.cell(row, STUDENT_NAME_COLUMN).value == "시험 평균":
                ws.cell(row, AVERAGE_SCORE_COLUMN).fill = class_average_color(data_only_ws.cell(row, AVERAGE_SCORE_COLUMN).value)
            else:
                ws.cell(row, AVERAGE_SCORE_COLUMN).fill = student_average_color(data_only_ws.cell(row, AVERAGE_SCORE_COLUMN).value)
        else:
            ws.cell(row, col).fill = FILL_NONE

        # 학생별 평균 폰트 설정
        if ws.cell(row, STUDENT_NAME_COLUMN).value in ("날짜", "시험명", "시험 평균"):
            ws.cell(row, AVERAGE_SCORE_COLUMN).font = FONT_BOLD
            continue
        if ws.cell(row, STUDENT_NAME_COLUMN).font.strike:
            ws.cell(row, AVERAGE_SCORE_COLUMN).font = FONT_BOLD_STRIKE
            continue
        if ws.cell(row, STUDENT_NAME_COLUMN).font.color is not None and ws.cell(row, STUDENT_NAME_COLUMN).font.color.rgb == "FFFF0000":
            ws.cell(row, AVERAGE_SCORE_COLUMN).font = FONT_BOLD_RED
            continue

        # 신규생 하이라이트
        exist, _, _, new_student = tdm.excel.student_info.get_student_info(student_ws, ws.cell(row, STUDENT_NAME_COLUMN).value)
        if exist:
            if new_student:
                ws.cell(row, STUDENT_NAME_COLUMN).fill = FILL_NEW_STUDENT
            else:
                ws.cell(row, STUDENT_NAME_COLUMN).fill = FILL_NONE
        else:
            ws.cell(row, STUDENT_NAME_COLUMN).fill = FILL_NONE
            warnings.append(f"{ws.cell(row, STUDENT_NAME_COLUMN).value} 학생 정보가 존재하지 않습니다.")

    # A read-only openpyxl workbook keeps the source ZIP handle open.  On
    # Windows that prevents the atomic replacement of the same file.
    close_workbooks(data_only_wb, student_wb)
    save(wb)

    return warnings

def update_class(prog: Progress | None = None):
    """
    수정된 반 정보 파일을 바탕으로 데이터 파일 업데이트
    """
    file_validation()

    if prog:
        prog.step("백업 생성 중...")

    make_backup_file()

    new_class_names = set(tdm.excel.class_info.get_new_class_names())

    # 조건부 서식 수식 로딩
    paths = WorkbookPaths.current()
    recalculate_workbook(paths.data_file)

    # 지난 데이터 파일이 없으면 새로 생성
    if not paths.previous_data.is_file():
        pre_data_wb = xl.Workbook()
        pre_data_ws = pre_data_wb.worksheets[0]
        pre_data_ws.title = DataFile.DEFAULT_SHEET_NAME

        pre_data_ws[gcl(DataFile.CLASS_NAME_COLUMN)+"1"]    = "반"
        pre_data_ws[gcl(DataFile.TEACHER_NAME_COLUMN)+"1"]  = "담당"
        pre_data_ws[gcl(DataFile.STUDENT_NAME_COLUMN)+"1"]  = "이름"
        pre_data_ws[gcl(DataFile.AVERAGE_SCORE_COLUMN)+"1"] = "학생 평균"
        pre_data_ws.freeze_panes    = f"{gcl(DataFile.DATA_COLUMN)}2"
        pre_data_ws.auto_filter.ref = f"A:{gcl(DataFile.MAX)}"

        for col in range(1, DataFile.DATA_COLUMN):
            pre_data_ws.cell(1, col).alignment = ALIGN_CENTER
            pre_data_ws.cell(1, col).border    = BORDER_BOTTOM_MEDIUM_000

        atomic_save_workbook(
            pre_data_wb,
            paths.previous_data,
        )
    else:
        pre_data_wb = load_workbook(
            paths.previous_data,
            display_name=DataFile.PRE_DATA_FILE_NAME,
        )

    # 지난 데이터 이동
    data_only_wb = open(data_only=True) # 데이터가 더이상 수정되지 않으므로 읽기 전용으로 불러옴

    data_only_ws = data_only_wb[DataFile.DEFAULT_SHEET_NAME]
    pre_data_ws  = pre_data_wb[DataFile.DEFAULT_SHEET_NAME]

    CLASS_NAME_COLUMN, TEACHER_NAME_COLUMN, STUDENT_NAME_COLUMN, AVERAGE_SCORE_COLUMN = find_dynamic_columns(data_only_ws)

    if prog:
        prog.step("지난 데이터 파일로 데이터 이동 중...")

    to_delete = []
    for row in range(2, max(2, data_only_ws.max_row + 1)):
        v = data_only_ws.cell(row, CLASS_NAME_COLUMN).value
        if v is not None and v not in new_class_names:
            to_delete.append(row)

    for idx, row in enumerate(to_delete, start=1):
        prog.phase(idx, len(to_delete), f"지난 데이터 이동 중... ({idx}/{len(to_delete)})")
        PRE_DATA_WRITE_ROW = pre_data_ws.max_row+1
        copy_cell(pre_data_ws.cell(PRE_DATA_WRITE_ROW, DataFile.CLASS_NAME_COLUMN),    data_only_ws.cell(row, CLASS_NAME_COLUMN))
        copy_cell(pre_data_ws.cell(PRE_DATA_WRITE_ROW, DataFile.TEACHER_NAME_COLUMN),  data_only_ws.cell(row, TEACHER_NAME_COLUMN))
        copy_cell(pre_data_ws.cell(PRE_DATA_WRITE_ROW, DataFile.STUDENT_NAME_COLUMN),  data_only_ws.cell(row, STUDENT_NAME_COLUMN))
        copy_cell(pre_data_ws.cell(PRE_DATA_WRITE_ROW, DataFile.AVERAGE_SCORE_COLUMN), data_only_ws.cell(row, AVERAGE_SCORE_COLUMN))
        PRE_DATA_WRITE_COLUMN = DataFile.MAX+1
        for col in range(AVERAGE_SCORE_COLUMN+1, data_only_ws.max_column+1):
            if data_only_ws.cell(row, col).value is not None:
                copy_cell(pre_data_ws.cell(PRE_DATA_WRITE_ROW, PRE_DATA_WRITE_COLUMN), data_only_ws.cell(row, col))
            PRE_DATA_WRITE_COLUMN += 1

    for col in range(DataFile.MAX + 1, pre_data_ws.max_column + 1):
        pre_data_ws.column_dimensions[gcl(col)].width = 14

    data_only_wb.close()
    del data_only_wb
    atomic_save_workbook(
        pre_data_wb,
        paths.previous_data,
    )
    del pre_data_wb

    # 데이터 파일 지난 데이터 삭제 및 신규 반 추가
    if prog:
        prog.step("지난 데이터 삭제 중...")

    delete_ranges: list[tuple[int, int]] = []
    if to_delete:
        to_delete.sort()
        start = prev = to_delete[0]
        for row in to_delete[1:]:
            if row == prev + 1:
                prev = row
                continue
            delete_ranges.append((start, prev))
            start = prev = row
        delete_ranges.append((start, prev))

    wb = open()
    ws = wb[DataFile.DEFAULT_SHEET_NAME]

    for idx, (start, end) in enumerate(reversed(delete_ranges), start=1):
        if prog:
            prog.phase(idx, len(delete_ranges), f"지난 데이터 삭제 중... ({idx}/{len(delete_ranges)})")
        ws.delete_rows(start, end - start + 1)

    ws.auto_filter.ref = f"A:{gcl(AVERAGE_SCORE_COLUMN)}"

    if prog:
        prog.step("신규 반 추가 중...")

    old_class_names = set(get_class_names(ws))
    unregistered_class_names = sorted(list(new_class_names.difference(old_class_names)))

    if len(unregistered_class_names) > 0:
        class_wb = tdm.excel.class_info.open_temp()
        class_ws = tdm.excel.class_info.open_worksheet(class_wb)

        class_student_dict = tdm.aisosik.reader.get_class_student_dict()

        ws = wb[DataFile.DEFAULT_SHEET_NAME]

        CLASS_NAME_COLUMN, TEACHER_NAME_COLUMN, STUDENT_NAME_COLUMN, AVERAGE_SCORE_COLUMN = find_dynamic_columns(ws)

        for row in range(ws.max_row+1, 1, -1):
            if ws.cell(row-1, STUDENT_NAME_COLUMN).value is not None:
                WRITE_RANGE = WRITE_LOCATION = row
                break

        for idx, class_name in enumerate(unregistered_class_names, start=1):
            if prog:
                prog.phase(idx, len(unregistered_class_names), f"신규 반 추가 중... ({idx}/{len(unregistered_class_names)})")
            temp_name = class_name
            if " (모의고사)" in class_name:
                temp_name = class_name[:-7]
            if len(class_student_dict[temp_name]) == 0 :
                continue
            exist, teacher_name, _, _, _ = tdm.excel.class_info.get_class_info(temp_name, ws=class_ws)
            if not exist:
                continue

            # 시험명
            ws.cell(WRITE_LOCATION, CLASS_NAME_COLUMN).value    = class_name
            ws.cell(WRITE_LOCATION, TEACHER_NAME_COLUMN).value  = teacher_name
            ws.cell(WRITE_LOCATION, STUDENT_NAME_COLUMN).value  = "날짜"
            WRITE_LOCATION += 1
            
            ws.cell(WRITE_LOCATION, CLASS_NAME_COLUMN).value    = class_name
            ws.cell(WRITE_LOCATION, TEACHER_NAME_COLUMN).value  = teacher_name
            ws.cell(WRITE_LOCATION, STUDENT_NAME_COLUMN).value  = "시험명"

            for col in range(1, AVERAGE_SCORE_COLUMN + 1):
                ws.cell(WRITE_LOCATION, col).border = BORDER_BOTTOM_THIN_9090

            WRITE_LOCATION += 1

            # 학생 루프
            for student_name in class_student_dict[temp_name]:
                ws.cell(WRITE_LOCATION, CLASS_NAME_COLUMN).value    = class_name
                ws.cell(WRITE_LOCATION, TEACHER_NAME_COLUMN).value  = teacher_name
                ws.cell(WRITE_LOCATION, STUDENT_NAME_COLUMN).value  = student_name
                WRITE_LOCATION += 1
            
            # 시험별 평균
            ws.cell(WRITE_LOCATION, CLASS_NAME_COLUMN).value    = class_name
            ws.cell(WRITE_LOCATION, TEACHER_NAME_COLUMN).value  = teacher_name
            ws.cell(WRITE_LOCATION, STUDENT_NAME_COLUMN).value  = "시험 평균"

            for col in range(1, AVERAGE_SCORE_COLUMN+1):
                ws.cell(WRITE_LOCATION, col).border = BORDER_TOP_THIN_9090_BOTTOM_MEDIUM_000

            WRITE_LOCATION += 1

        class_wb.close()
        del class_wb

        # 정렬
        for row in range(WRITE_RANGE, ws.max_row + 1):
            for col in range(1, AVERAGE_SCORE_COLUMN + 1):
                ws.cell(row, col).alignment = ALIGN_CENTER

        # 필터 범위 재지정
        ws.auto_filter.ref = f"A:{gcl(AVERAGE_SCORE_COLUMN)}"

    if prog:
        prog.step("함수 서식 범위 재조정 중...")

    return rescope_formulas(wb)

def add_student(student_name:str, target_class_name:str, wb:xl.Workbook=None):
    """
    학생 추가
    
    `move_student` 작업 시 `wb`로 작업중인 파일 정보 전달
    """
    file_validation()

    if wb is None:
        wb = open()

    ws = wb[DataFile.DEFAULT_SHEET_NAME]

    warnings = []

    for i in range(2):
        if i == 1:
            target_class_name += " (모의고사)"

        CLASS_NAME_COLUMN, TEACHER_NAME_COLUMN, STUDENT_NAME_COLUMN, AVERAGE_SCORE_COLUMN = find_dynamic_columns(ws)

        # 목표 반에 학생 추가
        for row in range(2, ws.max_row+1):
            if ws.cell(row, CLASS_NAME_COLUMN).value == target_class_name:
                class_index = row+2
                break
        else:
            continue

        already_exists = False
        while ws.cell(class_index, STUDENT_NAME_COLUMN).value != "시험 평균":
            if ws.cell(class_index, STUDENT_NAME_COLUMN).value > student_name:
                break
            elif ws.cell(class_index, STUDENT_NAME_COLUMN).font.strike:
                class_index += 1
            elif ws.cell(class_index, STUDENT_NAME_COLUMN).font.color is not None and ws.cell(class_index, STUDENT_NAME_COLUMN).font.color.rgb == "FFFF0000":
                class_index += 1
            elif ws.cell(class_index, STUDENT_NAME_COLUMN).value == student_name:
                warning = f"{student_name} 학생이 이미 존재합니다."
                if warning not in warnings:
                    warnings.append(warning)
                already_exists = True
                break
            else:
                class_index += 1

        if already_exists:
            continue

        ws.insert_rows(class_index)
        ws.cell(class_index, CLASS_NAME_COLUMN).value        = ws.cell(class_index-1, CLASS_NAME_COLUMN).value
        ws.cell(class_index, TEACHER_NAME_COLUMN).value      = ws.cell(class_index-1, TEACHER_NAME_COLUMN).value
        ws.cell(class_index, STUDENT_NAME_COLUMN).value      = student_name

        ws.cell(class_index, CLASS_NAME_COLUMN).alignment    = ALIGN_CENTER
        ws.cell(class_index, TEACHER_NAME_COLUMN).alignment  = ALIGN_CENTER
        ws.cell(class_index, STUDENT_NAME_COLUMN).alignment  = ALIGN_CENTER

        ws.cell(class_index, AVERAGE_SCORE_COLUMN).alignment = ALIGN_CENTER
        ws.cell(class_index, AVERAGE_SCORE_COLUMN).font      = FONT_BOLD

    rescope_formulas(wb)

    return warnings

def delete_student(class_name:str, student_name:str):
    """
    학생 퇴원 처리
    
    퇴원 처리된 학생은 모든 데이터에 취소선 적용
    """
    file_validation()

    wb = open()
    ws = wb[DataFile.DEFAULT_SHEET_NAME]

    CLASS_NAME_COLUMN, _, STUDENT_NAME_COLUMN, AVERAGE_SCORE_COLUMN = find_dynamic_columns(ws)

    for row in range(2, ws.max_row+1):
        if ws.cell(row, STUDENT_NAME_COLUMN).value == student_name and ws.cell(row, CLASS_NAME_COLUMN).value in (class_name, class_name + " (모의고사)"):
            for col in range(1, ws.max_column+1):
                if ws.cell(row, col).font.bold:
                    ws.cell(row, col).font = FONT_BOLD_STRIKE
                else:
                    ws.cell(row, col).font = FONT_STRIKE
            
            # 퇴원한 학생이 반 평균에 영향을 주지 않도록 수정
            ws.cell(row, AVERAGE_SCORE_COLUMN).value = ""

    save(wb)

def move_student(student_name:str, target_class_name:str, current_class_name:str):
    """
    학생 반 이동

    학생의 기존 반 데이터 글꼴 색을 빨간색으로 변경 후 목표 반에 학생 추가
    """
    file_validation()

    wb = open()
    ws = wb[DataFile.DEFAULT_SHEET_NAME]

    CLASS_NAME_COLUMN, _, STUDENT_NAME_COLUMN, _ = find_dynamic_columns(ws)

    if current_class_name == target_class_name:
        wb.close()
        raise InvalidOperationError("현재 반과 이동할 반이 같습니다.")

    # 기존 반 데이터 빨간색 처리
    source_found = False
    for row in range(2, ws.max_row+1):
        if ws.cell(row, STUDENT_NAME_COLUMN).value == student_name and ws.cell(row, CLASS_NAME_COLUMN).value in (current_class_name, current_class_name+" (모의고사)"):
            student_cell = ws.cell(row, STUDENT_NAME_COLUMN)
            is_red = (
                student_cell.font.color is not None
                and student_cell.font.color.rgb == "FFFF0000"
            )
            if student_cell.font.strike or is_red:
                continue
            source_found = True
            for col in range(1, ws.max_column+1):
                if ws.cell(row, col).font.bold:
                    ws.cell(row, col).font = FONT_BOLD_RED
                else:
                    ws.cell(row, col).font = FONT_RED
            # break

    if not source_found:
        wb.close()
        raise InvalidOperationError(
            f"데이터 파일의 {current_class_name} 반에서 {student_name} 학생을 찾을 수 없습니다."
        )

    return add_student(student_name, target_class_name, wb)

def rescope_formulas(wb:xl.Workbook=None):
    """
    데이터 파일 내 평균 산출 수식의 범위 재조정
    """
    file_validation()

    if wb is None:
        wb = open()

    ws = wb[DataFile.DEFAULT_SHEET_NAME]

    _, _, STUDENT_NAME_COLUMN, AVERAGE_SCORE_COLUMN = find_dynamic_columns(ws)

    # 평균 범위 재지정
    for row in range(2, ws.max_row+1):
        if ws.cell(row, STUDENT_NAME_COLUMN).value is None:
            break
        striked = False
        colored = False
        if ws.cell(row, STUDENT_NAME_COLUMN).font.strike:
            striked = True
        if ws.cell(row, STUDENT_NAME_COLUMN).font.color is not None:
            if ws.cell(row, STUDENT_NAME_COLUMN).font.color.rgb == "FFFF0000":
                colored = True

        if ws.cell(row, STUDENT_NAME_COLUMN).value == "날짜":
            DATE_ROW = row
            CLASS_START = row+2
        elif ws.cell(row, STUDENT_NAME_COLUMN).value == "시험 평균":
            CLASS_END = row-1
            ws[f"{gcl(AVERAGE_SCORE_COLUMN)}{row}"] = ArrayFormula(
                f"{gcl(AVERAGE_SCORE_COLUMN)}{row}",
                f"=ROUND(SUM(IFERROR({gcl(AVERAGE_SCORE_COLUMN)}{CLASS_START}:{gcl(AVERAGE_SCORE_COLUMN)}{CLASS_END},0))/COUNT({gcl(AVERAGE_SCORE_COLUMN)}{CLASS_START}:{gcl(AVERAGE_SCORE_COLUMN)}{CLASS_END}),0)",
            )
            if CLASS_START >= CLASS_END:
                continue
            for col in range(AVERAGE_SCORE_COLUMN+1, ws.max_column+1):
                if ws.cell(DATE_ROW, col).value is None:
                    break
                ws.cell(row, col).value = f"=ROUND(AVERAGE({gcl(col)}{CLASS_START}:{gcl(col)}{CLASS_END}), 0)"
                ws.cell(row, col).font  = FONT_BOLD
        elif ws.cell(row, STUDENT_NAME_COLUMN).value == "시험명":
            continue
        else:
            ws.cell(row, AVERAGE_SCORE_COLUMN).value = f"=ROUND(AVERAGE({gcl(AVERAGE_SCORE_COLUMN+1)}{row}:XFD{row}), 0)"

        if striked:
            ws.cell(row, AVERAGE_SCORE_COLUMN).font = FONT_BOLD_STRIKE
        elif colored:
            ws.cell(row, AVERAGE_SCORE_COLUMN).font = FONT_BOLD_RED
        else:
            ws.cell(row, AVERAGE_SCORE_COLUMN).font = FONT_BOLD

    save(wb)

def change_class_info(target_class_name:str, target_teacher_name:str):
    """
    특정 반의 담당 선생님 변경
    """
    wb = open()
    ws = wb[DataFile.DEFAULT_SHEET_NAME]

    CLASS_NAME_COLUMN, TEACHER_NAME_COLUMN, _, _ = find_dynamic_columns(ws)

    for row in range(2, ws.max_row+1):
        if ws.cell(row, CLASS_NAME_COLUMN).value in (target_class_name, target_class_name+" (모의고사)"):
            ws.cell(row, TEACHER_NAME_COLUMN).value = target_teacher_name

    save(wb)
