"""Common workbook structure validation."""

from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path

from openpyxl.worksheet.worksheet import Worksheet

from tdm.excel.data_file_errors import NoReservedColumnError
from tdm.excel.workbook_io import closing_workbook, load_workbook, require_worksheet


def find_required_columns(
    worksheet: Worksheet,
    headers: Iterable[str],
    *,
    header_row: int = 1,
) -> dict[str, int]:
    positions = {
        str(worksheet.cell(header_row, column).value): column
        for column in range(1, worksheet.max_column + 1)
        if worksheet.cell(header_row, column).value is not None
    }
    result: dict[str, int] = {}
    for header in headers:
        if header not in positions:
            raise NoReservedColumnError(
                f"{worksheet.title} 시트에 '{header}' 열이 없습니다."
            )
        result[header] = positions[header]
    return result


def validate_workbook_structure(
    path: str | Path,
    *,
    sheet_name: str,
    required_headers: Iterable[str] = (),
) -> None:
    with closing_workbook(load_workbook(path, read_only=True)) as workbook:
        worksheet = require_worksheet(workbook, sheet_name)
        if required_headers:
            find_required_columns(worksheet, required_headers)

