"""Read-only queries over the main data workbook."""

from datetime import datetime

from openpyxl.worksheet.worksheet import Worksheet

import tdm.excel.class_info
from tdm.domain.models import DataFile
from tdm.excel.data_file_storage import open
from tdm.excel.validation import find_required_columns

def get_data_sorted_dict(mocktest=False):
    """Return student and test indexes grouped by class."""
    workbook = open()
    class_workbook = tdm.excel.class_info.open()
    try:
        worksheet = workbook[DataFile.DEFAULT_SHEET_NAME]
        class_worksheet = tdm.excel.class_info.open_worksheet(class_workbook)
        class_column, _, student_column, average_column = find_dynamic_columns(
            worksheet
        )
        class_students: dict[str, dict[object, int]] = {}
        class_tests: dict[str, dict[str, int]] = {}

        class_names = tdm.excel.class_info.get_class_names(
            class_worksheet,
            mocktest=mocktest,
        )
        for class_name in class_names:
            student_indexes: dict[object, int] = {}
            test_indexes: dict[str, int] = {}
            for row in range(2, worksheet.max_row + 1):
                if worksheet.cell(row, class_column).value != class_name:
                    continue
                if worksheet.cell(row, student_column).value == "날짜":
                    for column in range(average_column + 1, worksheet.max_column + 1):
                        test_date = worksheet.cell(row, column).value
                        test_name = worksheet.cell(row + 1, column).value
                        if test_date is None and test_name is None:
                            break
                        if isinstance(test_date, datetime):
                            date_text = test_date.strftime("%y.%m.%d")
                        else:
                            date_text = (
                                str(test_date)
                                .split()[0][2:10]
                                .replace("-", ".")
                                .replace(",", ".")
                                .replace("/", ".")
                            )
                        test_indexes[f"[{date_text}] {test_name}"] = column
                    continue

                student_cell = worksheet.cell(row, student_column)
                if student_cell.value in ("시험명", "시험 평균"):
                    continue
                if student_cell.font.strike:
                    continue
                if (
                    student_cell.font.color is not None
                    and student_cell.font.color.rgb == "FFFF0000"
                ):
                    continue
                student_indexes[student_cell.value] = row

            class_students[class_name] = student_indexes
            class_tests[class_name] = dict(
                sorted(test_indexes.items(), reverse=True)
            )

        return dict(sorted(class_students.items())), class_tests
    finally:
        workbook.close()
        class_workbook.close()

def find_dynamic_columns(ws:Worksheet):
    """
    파일 열(column) 정보 동적 탐색

    '반' 열, '담당' 열, '이름' 열, '학생 평균' 열

    return `CLASS_NAME_COLUMN`, `TEACHER_NAME_COLUMN`, `STUDENT_NAME_COLUMN`, `AVERAGE_SCORE_COLUMN`
    """

    columns = find_required_columns(ws, ("반", "담당", "이름", "학생 평균"))
    return columns["반"], columns["담당"], columns["이름"], columns["학생 평균"]

def is_cell_empty(row:int, col:int) -> bool:
    """
    데이터 파일이 열려있지 않을 때 특정 셀의 값이 비어있는 지 확인

    데일리테스트 시트 한정 기능
    """
    wb = open(data_only=True, read_only=True)
    try:
        ws = wb[DataFile.DEFAULT_SHEET_NAME]
        value = ws.cell(row, col).value
        return value is None, value
    finally:
        wb.close()

def get_class_names(ws:Worksheet):
    class_names = []

    CLASS_NAME_COLUMN, _, _, _ = find_dynamic_columns(ws)

    for row in range(2, ws.max_row+1):
        if ws.cell(row, CLASS_NAME_COLUMN).value  not in class_names:
            class_names.append(ws.cell(row, CLASS_NAME_COLUMN).value)

    return class_names

def check_student_exists(student_name: str) -> bool:
    """데이터 파일의 어느 반에든 활성 상태인 학생이 있는지 확인."""
    wb = open(data_only=True, read_only=True)
    try:
        ws = wb[DataFile.DEFAULT_SHEET_NAME]
        _, _, STUDENT_NAME_COLUMN, _ = find_dynamic_columns(ws)

        for row in range(2, ws.max_row + 1):
            student_cell = ws.cell(row, STUDENT_NAME_COLUMN)
            if student_cell.value != student_name:
                continue
            if student_cell.font.strike:
                continue
            if student_cell.font.color is not None and student_cell.font.color.rgb == "FFFF0000":
                continue
            return True
        return False
    finally:
        wb.close()

# 파일 작업
