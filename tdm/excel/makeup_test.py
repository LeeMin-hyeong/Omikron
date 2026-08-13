import openpyxl as xl

from datetime import datetime
from openpyxl.utils.cell import get_column_letter as gcl

import tdm.excel.class_info
import tdm.excel.data_form
import tdm.excel.student_info

from tdm.domain.models import MakeupTestList, DataForm
from tdm.domain.errors import FileOpenException, TDMError
from tdm.excel.utils import calculate_makeup_test_schedule
from tdm.domain.progress import Progress
from tdm.excel.atomic import atomic_save_workbook
from tdm.excel.paths import WorkbookPaths
from tdm.excel.styles import ALIGN_CENTER, ALIGN_CENTER_WRAP, FILL_NEW_STUDENT, BORDER_ALL
from tdm.excel.workbook_io import load_workbook, require_worksheet


# 파일 기본 작업
def create_workbook() -> xl.Workbook:
    wb = xl.Workbook()
    ws = wb.worksheets[0]
    ws.title = MakeupTestList.DEFAULT_NAME
    ws[gcl(MakeupTestList.TEST_DATE_COLUMN)+"1"]          = "응시일"
    ws[gcl(MakeupTestList.CLASS_NAME_COLUMN)+"1"]         = "반"
    ws[gcl(MakeupTestList.TEACHER_NAME_COLUMN)+"1"]       = "담당T"
    ws[gcl(MakeupTestList.STUDENT_NAME_COLUMN)+"1"]       = "이름"
    ws[gcl(MakeupTestList.TEST_NAME_COLUMN)+"1"]          = "시험명"
    ws[gcl(MakeupTestList.MAKEUPTEST_DATE_COLUMN)+"1"]    = "재시 날짜"
    ws[gcl(MakeupTestList.MAKEUPTEST_SCORE_COLUMN)+"1"]   = "재시 점수"
    ws[gcl(MakeupTestList.ETC_COLUMN)+"1"]                = "비고"

    ws.column_dimensions[gcl(MakeupTestList.TEST_DATE_COLUMN)].width = 14
    ws.auto_filter.ref = "A:"+gcl(MakeupTestList.MAX)
    ws.freeze_panes    = "A2"

    for col in range(1, DataForm.MAX+1):
        ws.cell(1, col).alignment = ALIGN_CENTER_WRAP
        ws.cell(1, col).border    = BORDER_ALL

    return wb


def make_file():
    wb = create_workbook()
    try:
        atomic_save_workbook(wb, WorkbookPaths.current().makeup_test)
    finally:
        wb.close()

def open(data_only:bool=False) -> xl.Workbook:
    path = WorkbookPaths.current().makeup_test
    try:
        return load_workbook(
            path,
            display_name=MakeupTestList.DEFAULT_NAME,
            data_only=data_only,
        )
    except FileNotFoundError as exc:
        raise FileNotFoundError(
            f"{MakeupTestList.DEFAULT_NAME} 파일이 존재하지 않습니다.\n"
            "데이터 저장 시 재시험자가 발생하면 자동으로 생성됩니다."
        ) from exc

def open_worksheet(wb:xl.Workbook):
    return require_worksheet(wb, MakeupTestList.DEFAULT_NAME)

def save(wb:xl.Workbook):
    try:
        atomic_save_workbook(wb, WorkbookPaths.current().makeup_test)
    except TDMError:
        raise
    except Exception as exc:
        raise FileOpenException(
            f"{MakeupTestList.DEFAULT_NAME} 파일을 닫은 뒤 다시 시도해주세요"
        ) from exc
    finally:
        wb.close()

# 파일 유틸리티
def get_student_test_index_dict():
    """
    1st key: 학생 이름

    2nd key: 시험명

    value: 행 인덱스
    """
    wb = open(True)
    ws = open_worksheet(wb)

    student_test_index_dict:dict[str, dict[str, int]] = {}
    for row in range(2, ws.max_row+1):
        if ws.cell(row, MakeupTestList.MAKEUPTEST_SCORE_COLUMN).value is None:
            class_name       = ws.cell(row, MakeupTestList.CLASS_NAME_COLUMN).value
            student_name     = ws.cell(row, MakeupTestList.STUDENT_NAME_COLUMN).value
            makeup_test_name = ws.cell(row, MakeupTestList.TEST_NAME_COLUMN).value
            student_test_index_dict.setdefault(student_name, {})

            test_name = f"({class_name}) {makeup_test_name}"
            student_test_index_dict[student_name][test_name] = row

    wb.close()
    return student_test_index_dict

# 파일 작업
def save_makeup_test_list(filepath: str, makeup_test_date: dict, prog: Progress):
    form_wb = None
    student_wb = None
    wb = None

    try:
        form_wb = tdm.excel.data_form.open(filepath)
        form_ws = tdm.excel.data_form.open_worksheet(form_wb)

        # 재시험 정보 파일 없으면 생성
        if WorkbookPaths.current().makeup_test.is_file():
            wb = open()
        else:
            wb = create_workbook()
        ws = open_worksheet(wb)

        # 학생 정보
        student_wb = tdm.excel.student_info.open(True)
        student_ws = tdm.excel.student_info.open_worksheet(student_wb)

        # ✅ 오늘 날짜 캐시 (루프 밖)
        today = datetime.today().date()
        today_key = today.strftime("%y%m%d")

        # 재시험 데이터 작성 시작 위치 탐색
        for row in range(ws.max_row + 1, 1, -1):
            if ws.cell(row - 1, MakeupTestList.TEST_DATE_COLUMN).value is not None:
                MAKEUP_TEST_RANGE = MAKEUP_TEST_WRITE_ROW = row
                break
        else:
            # 시트가 비어있는 특이 케이스 방어
            MAKEUP_TEST_RANGE = MAKEUP_TEST_WRITE_ROW = 2

        # ✅ (핵심) 중복 검사 캐시: "오늘 날짜인 행"만 스캔해서 set 구축
        #     기존 로직은 '오늘자 영역에서 같은 학생+반이면 duplicated'였음
        today_existing = set()  # (student_name, class_name)

        # 뒤에서 앞으로 훑되, 날짜가 오늘보다 과거로 내려가면 break
        check = ws.max_row
        while check > 1:
            test_date = ws.cell(check, MakeupTestList.TEST_DATE_COLUMN).value

            if test_date is None or not isinstance(test_date, datetime):
                check -= 1
                continue

            dkey = test_date.strftime("%y%m%d")
            if dkey == today_key:
                sname = ws.cell(check, MakeupTestList.STUDENT_NAME_COLUMN).value
                cname = ws.cell(check, MakeupTestList.CLASS_NAME_COLUMN).value
                if sname is not None and cname is not None:
                    today_existing.add((sname, cname))
                check -= 1
                continue

            if dkey < today_key:
                break

            check -= 1

        for test_type in range(2):
            if test_type == 0:
                TEST_NAME_COLUMN = DataForm.DAILYTEST_NAME_COLUMN
                TEST_SCORE_COLUMN = DataForm.DAILYTEST_SCORE_COLUMN
            else:
                TEST_NAME_COLUMN = DataForm.MOCKTEST_NAME_COLUMN
                TEST_SCORE_COLUMN = DataForm.MOCKTEST_SCORE_COLUMN

            # 데일리데이터 기록 양식 루프
            class_name = test_name = teacher_name = None

            for i in range(2, form_ws.max_row + 1):
                # 반/시험명 갱신
                c = form_ws.cell(i, DataForm.CLASS_NAME_COLUMN).value
                tn = form_ws.cell(i, TEST_NAME_COLUMN).value
                if c is not None and tn is not None:
                    class_name = c
                    test_name = tn
                    teacher_name = form_ws.cell(i, DataForm.TEACHER_NAME_COLUMN).value

                test_score = form_ws.cell(i, TEST_SCORE_COLUMN).value
                if test_score is None or type(test_score) not in (int, float) or test_score >= 80:
                    continue

                makeup_test_check = form_ws.cell(i, DataForm.MAKEUP_TEST_CHECK_COLUMN).value
                if makeup_test_check in ("x", "X"):
                    continue

                student_name = form_ws.cell(i, DataForm.STUDENT_NAME_COLUMN).value
                if not student_name or not class_name:
                    continue

                # ✅ O(1) 중복 검사 (기존 while check 루프 제거)
                key = (student_name, class_name)
                if key in today_existing:
                    continue

                # 학생 재시험 정보 검색
                complete, makeup_test_weekday, _, new_student = tdm.excel.student_info.get_student_info(student_ws, student_name)
                if not complete:
                    prog.warning(f"{student_name}의 학생 정보가 존재하지 않습니다.")

                ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.TEST_DATE_COLUMN).value = today
                ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.CLASS_NAME_COLUMN).value = class_name
                ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.TEACHER_NAME_COLUMN).value = teacher_name
                ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.STUDENT_NAME_COLUMN).value = student_name
                ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.TEST_NAME_COLUMN).value = test_name

                if new_student:
                    ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.STUDENT_NAME_COLUMN).fill = FILL_NEW_STUDENT

                if makeup_test_weekday is not None:
                    ok, calculated_schedule, _ = calculate_makeup_test_schedule(makeup_test_weekday, makeup_test_date)
                    if not ok:
                        prog.warning(f"{student_name}의 재시험 일정이 올바른 양식이 아닙니다.")

                    ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.MAKEUPTEST_DATE_COLUMN).value = calculated_schedule
                    ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.MAKEUPTEST_DATE_COLUMN).number_format = "mm월 dd일(aaa)"

                # ✅ 오늘자 중복 캐시에 즉시 반영(같은 실행에서 중복 추가 방지)
                today_existing.add(key)

                MAKEUP_TEST_WRITE_ROW += 1

        # ✅ 정렬 및 테두리: "추가된 행 범위만" 적용
        for row in range(MAKEUP_TEST_RANGE, MAKEUP_TEST_WRITE_ROW):
            for col in range(1, MakeupTestList.MAX + 1):
                cell = ws.cell(row, col)
                cell.alignment = ALIGN_CENTER
                cell.border = BORDER_ALL

        return wb
    finally:
        # ✅ close (잠김/메모리 누수 방지)
        try:
            if form_wb is not None:
                form_wb.close()
        except Exception:
            pass
        try:
            if student_wb is not None:
                student_wb.close()
        except Exception:
            pass

def save_makeup_test_result(target_row:int, makeup_test_score:str) -> bool:
    wb = open()
    ws = open_worksheet(wb)

    ws.cell(target_row, MakeupTestList.MAKEUPTEST_SCORE_COLUMN).value = makeup_test_score

    save(wb)

    return True

def prepare_individual_makeup_test(student_name:str, class_name:str, test_name:str, test_score:int|float, makeup_test_date:dict, prog:Progress) -> xl.Workbook:
    if WorkbookPaths.current().makeup_test.is_file():
        wb = open()
    else:
        wb = create_workbook()
    ws = open_worksheet(wb)

    student_wb = tdm.excel.student_info.open(True)
    student_ws = tdm.excel.student_info.open_worksheet(student_wb)

    class_wb = tdm.excel.class_info.open(True)
    class_ws = tdm.excel.class_info.open_worksheet(class_wb)

    try:
        for row in range(ws.max_row+1, 1, -1):
            if ws.cell(row-1, MakeupTestList.TEST_DATE_COLUMN).value is not None:
                MAKEUP_TEST_WRITE_ROW = row
                break

        exist, teacher_name, _, _, _ = tdm.excel.class_info.get_class_info(class_name, class_ws)
        if not exist:
            prog.warning(f"{class_name}의 반 정보가 존재하지 않습니다.")

        exist, makeup_test_weekday, _, new_student = tdm.excel.student_info.get_student_info(student_ws, student_name)
        if not exist:
            prog.warning(f"{student_name}의 학생 정보가 존재하지 않습니다.")

        ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.TEST_DATE_COLUMN).value = datetime.today().date()
        ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.CLASS_NAME_COLUMN).value = class_name
        ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.TEACHER_NAME_COLUMN).value = teacher_name
        ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.STUDENT_NAME_COLUMN).value = student_name
        ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.TEST_NAME_COLUMN).value = test_name

        if new_student:
            ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.STUDENT_NAME_COLUMN).fill = FILL_NEW_STUDENT

        if makeup_test_weekday is not None:
            complete, calculated_schedule, _ = calculate_makeup_test_schedule(makeup_test_weekday, makeup_test_date)
            if not complete:
                prog.warning(f"{student_name}의 재시험 일정이 올바른 양식이 아닙니다.")

            ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.MAKEUPTEST_DATE_COLUMN).value = calculated_schedule
            ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.MAKEUPTEST_DATE_COLUMN).number_format = "mm월 dd일(aaa)"

        for col in range(1, MakeupTestList.MAX + 1):
            ws.cell(MAKEUP_TEST_WRITE_ROW, col).alignment = ALIGN_CENTER
            ws.cell(MAKEUP_TEST_WRITE_ROW, col).border = BORDER_ALL

        return wb
    except Exception:
        wb.close()
        raise
    finally:
        student_wb.close()
        class_wb.close()


def save_individual_makeup_test(student_name:str, class_name:str, test_name:str, test_score:int|float, makeup_test_date:dict, prog:Progress):
    wb = prepare_individual_makeup_test(
        student_name,
        class_name,
        test_name,
        test_score,
        makeup_test_date,
        prog,
    )
    try:
        save(wb)
    finally:
        wb.close()
