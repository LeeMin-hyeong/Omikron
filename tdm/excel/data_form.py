import openpyxl as xl

from openpyxl.styles import Protection
from openpyxl.utils.cell import get_column_letter as gcl
from openpyxl.worksheet.datavalidation import DataValidation

import tdm.aisosik.reader
import tdm.excel.class_info

from tdm.domain.errors import UserActionableError
from tdm.domain.models import DataForm
from tdm.excel.atomic import atomic_save_workbook
from tdm.excel.paths import WorkbookPaths
from tdm.excel.styles import BORDER_ALL, ALIGN_CENTER, ALIGN_CENTER_WRAP
from tdm.excel.workbook_io import load_workbook, require_worksheet

class DataValidationException(UserActionableError):
    code = "DATA_FORM_INVALID"

# 파일 기본 작업
def make_file() -> bool:
    wb = xl.Workbook()
    ws = wb.worksheets[0]
    ws.title = DataForm.DEFAULT_NAME
    ws[gcl(DataForm.CLASS_WEEKDAY_COLUMN)+"1"]     = "요일"
    ws[gcl(DataForm.TEST_TIME_COLUMN)+"1"]         = "시간"
    ws[gcl(DataForm.CLASS_NAME_COLUMN)+"1"]        = "반"
    ws[gcl(DataForm.STUDENT_NAME_COLUMN)+"1"]      = "이름"
    ws[gcl(DataForm.TEACHER_NAME_COLUMN)+"1"]      = "담당T"
    ws[gcl(DataForm.DAILYTEST_NAME_COLUMN)+"1"]    = "시험명"
    ws[gcl(DataForm.DAILYTEST_SCORE_COLUMN)+"1"]   = "점수"
    ws[gcl(DataForm.DAILYTEST_AVERAGE_COLUMN)+"1"] = "평균"
    ws[gcl(DataForm.MOCKTEST_NAME_COLUMN)+"1"]     = "모의고사 시험명"
    ws[gcl(DataForm.MOCKTEST_SCORE_COLUMN)+"1"]    = "모의고사 점수"
    ws[gcl(DataForm.MOCKTEST_AVERAGE_COLUMN)+"1"]  = "모의고사 평균"
    ws[gcl(DataForm.MAKEUP_TEST_CHECK_COLUMN)+"1"] = "재시험 응시 여부"
    ws["Y1"] = "X"
    ws["Z1"] = "x"
    ws.column_dimensions.group("Y", "Z", hidden=True)
    ws.auto_filter.ref = "A:"+gcl(DataForm.TEST_TIME_COLUMN)
    ws.freeze_panes    = "A2"
    
    for col in range(1, DataForm.MAX+1):
        ws.cell(1, col).alignment = ALIGN_CENTER_WRAP
        ws.cell(1, col).border    = BORDER_ALL

    class_wb = tdm.excel.class_info.open(True)
    class_ws = tdm.excel.class_info.open_worksheet(class_wb)

    for class_name, student_names in tdm.aisosik.reader.get_class_student_dict().items():
        if len(student_names) == 0:
            continue

        exist, teacher_name, class_weekday, test_time, _ = tdm.excel.class_info.get_class_info(class_name, ws=class_ws)
        if not exist:
            continue

        WRITE_LOCATION = start = ws.max_row + 1

        ws.cell(WRITE_LOCATION, DataForm.CLASS_NAME_COLUMN).value   = class_name
        ws.cell(WRITE_LOCATION, DataForm.TEACHER_NAME_COLUMN).value = teacher_name

        #학생 루프
        for student_name in student_names:
            ws.cell(WRITE_LOCATION, DataForm.CLASS_WEEKDAY_COLUMN).value = class_weekday
            ws.cell(WRITE_LOCATION, DataForm.TEST_TIME_COLUMN).value     = test_time
            ws.cell(WRITE_LOCATION, DataForm.STUDENT_NAME_COLUMN).value  = student_name
            dv = DataValidation(type="list", formula1="=Y1:Z1", showDropDown=True, allow_blank=True, showErrorMessage=True)
            dv.error = "이 셀의 값은 'x' 또는 'X'이어야 합니다."
            ws.add_data_validation(dv)
            dv.add(ws.cell(WRITE_LOCATION,DataForm.MAKEUP_TEST_CHECK_COLUMN))
            WRITE_LOCATION = ws.max_row + 1
        
        end = WRITE_LOCATION - 1

        # 시험 평균
        ws.cell(start, DataForm.DAILYTEST_AVERAGE_COLUMN).value = f"=ROUND(AVERAGE({gcl(DataForm.DAILYTEST_SCORE_COLUMN)}{start}:{gcl(DataForm.DAILYTEST_SCORE_COLUMN)}{end}), 0)"
        # 모의고사 평균
        ws.cell(start, DataForm.MOCKTEST_AVERAGE_COLUMN).value  = f"=ROUND(AVERAGE({gcl(DataForm.MOCKTEST_SCORE_COLUMN)}{start}:{gcl(DataForm.MOCKTEST_SCORE_COLUMN)}{end}), 0)"
        
        # 정렬 및 테두리
        for row in range(start, end + 1):
            for col in range(1, DataForm.MAX+1):
                ws.cell(row, col).alignment = ALIGN_CENTER
                ws.cell(row, col).border    = BORDER_ALL
        
        # 셀 병합
        if start < end:
            ws.merge_cells(f"{gcl(DataForm.CLASS_NAME_COLUMN)}{start}:{gcl(DataForm.CLASS_NAME_COLUMN)}{end}")
            ws.merge_cells(f"{gcl(DataForm.TEACHER_NAME_COLUMN)}{start}:{gcl(DataForm.TEACHER_NAME_COLUMN)}{end}")
            ws.merge_cells(f"{gcl(DataForm.DAILYTEST_NAME_COLUMN)}{start}:{gcl(DataForm.DAILYTEST_NAME_COLUMN)}{end}")
            ws.merge_cells(f"{gcl(DataForm.DAILYTEST_AVERAGE_COLUMN)}{start}:{gcl(DataForm.DAILYTEST_AVERAGE_COLUMN)}{end}")
            ws.merge_cells(f"{gcl(DataForm.MOCKTEST_NAME_COLUMN)}{start}:{gcl(DataForm.MOCKTEST_NAME_COLUMN)}{end}")
            ws.merge_cells(f"{gcl(DataForm.MOCKTEST_AVERAGE_COLUMN)}{start}:{gcl(DataForm.MOCKTEST_AVERAGE_COLUMN)}{end}")
        
    ws.protection.sheet         = True
    ws.protection.autoFilter    = False
    ws.protection.formatColumns = False
    for row in range(2, ws.max_row + 1):
        ws.cell(row, DataForm.CLASS_NAME_COLUMN).alignment         = ALIGN_CENTER_WRAP
        ws.cell(row, DataForm.DAILYTEST_NAME_COLUMN).alignment     = ALIGN_CENTER_WRAP
        ws.cell(row, DataForm.MOCKTEST_NAME_COLUMN).alignment      = ALIGN_CENTER_WRAP
        ws.cell(row, DataForm.DAILYTEST_NAME_COLUMN).protection    = Protection(locked=False)
        ws.cell(row, DataForm.DAILYTEST_SCORE_COLUMN).protection   = Protection(locked=False)
        ws.cell(row, DataForm.MOCKTEST_NAME_COLUMN).protection     = Protection(locked=False)
        ws.cell(row, DataForm.MOCKTEST_SCORE_COLUMN).protection    = Protection(locked=False)
        ws.cell(row, DataForm.MAKEUP_TEST_CHECK_COLUMN).protection = Protection(locked=False)

    class_wb.close()
    try:
        atomic_save_workbook(wb, WorkbookPaths.current().next_daily_form_path())
    finally:
        wb.close()

    return True

def open(filepath, data_only=True) -> xl.Workbook:
    return load_workbook(filepath, data_only=data_only)

def open_worksheet(wb:xl.Workbook):
    return require_worksheet(wb, DataForm.DEFAULT_NAME)

# 파일 유틸리티
def data_validation(filepath:str) -> bool:
    """
    데이터 입력 양식의 데이터가 올바르게 입력되었는지 확인
    """
    wb = open(filepath)
    try:
        ws = open_worksheet(wb)
        errors: list[str] = []
        form_checked = True
        dailytest_checked = False
        mocktest_checked = False
        class_name = None
        dailytest_name = None
        mocktest_name = None

        for row in range(1, ws.max_row + 1):
            if ws.cell(row, DataForm.CLASS_NAME_COLUMN).value is not None:
                class_name = ws.cell(row, DataForm.CLASS_NAME_COLUMN).value
                dailytest_checked = False
                mocktest_checked = False
                dailytest_name = ws.cell(row, DataForm.DAILYTEST_NAME_COLUMN).value
                mocktest_name = ws.cell(row, DataForm.MOCKTEST_NAME_COLUMN).value

            if dailytest_checked and mocktest_checked:
                continue

            if (
                not dailytest_checked
                and ws.cell(row, DataForm.DAILYTEST_SCORE_COLUMN).value is not None
                and dailytest_name is None
            ):
                errors.append(f"{class_name} 반의 시험명이 작성되지 않았습니다.")
                dailytest_checked = True
                form_checked = False
            if (
                not mocktest_checked
                and ws.cell(row, DataForm.MOCKTEST_SCORE_COLUMN).value is not None
                and mocktest_name is None
            ):
                errors.append(f"{class_name} 반의 모의고사명이 작성되지 않았습니다.")
                mocktest_checked = True
                form_checked = False

        if errors:
            raise DataValidationException("\n".join(errors))
        return form_checked
    finally:
        wb.close()
