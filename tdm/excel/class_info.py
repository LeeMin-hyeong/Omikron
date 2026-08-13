import openpyxl as xl

from openpyxl.utils.cell import get_column_letter as gcl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation

import tdm.aisosik.reader

from tdm.domain.models import ClassInfo
from tdm.domain.errors import FileOpenException, InvalidOperationError, TDMError
from tdm.domain.progress import Progress
from tdm.excel.atomic import (
    FileRevision,
    atomic_save_workbook,
    capture_file_revision,
    track_workbook_source,
    workbook_source_revision,
)
from tdm.excel.backup import create_backup
from tdm.excel.metadata import atomic_write_json, read_json
from tdm.excel.paths import WorkbookPaths
from tdm.excel.styles import BORDER_ALL, ALIGN_CENTER, ALIGN_CENTER_WRAP
from tdm.excel.workbook_io import load_workbook, require_worksheet

# 파일 기본 작업
def make_file():
    wb = xl.Workbook()
    ws = wb.worksheets[0]
    ws.title = ClassInfo.DEFAULT_NAME
    ws[gcl(ClassInfo.CLASS_NAME_COLUMN)+"1"]     = "반명"
    ws[gcl(ClassInfo.TEACHER_NAME_COLUMN)+"1"]   = "선생님명"
    ws[gcl(ClassInfo.CLASS_WEEKDAY_COLUMN)+"1"]  = "요일"
    ws[gcl(ClassInfo.TEST_TIME_COLUMN)+"1"]      = "시간"
    ws[gcl(ClassInfo.MOCKTEST_CHECK_COLUMN)+"1"] = "모의고사 응시여부"

    ws["Z1"] = "Y"
    ws.auto_filter.ref = "A:"+gcl(ClassInfo.MAX)
    ws.freeze_panes    = "A2"
    ws.column_dimensions.group("Z", hidden=True)

    # 반 루프
    for class_name in tdm.aisosik.reader.get_class_names():
        WRITE_LOCATION = ws.max_row + 1
        ws.cell(WRITE_LOCATION, 1).value = class_name

        dv = DataValidation(type="list", formula1="=Z1", allow_blank=True, errorStyle="stop", showErrorMessage=True)
        dv.error = "이 셀의 값은 'Y'이어야 합니다."
        ws.add_data_validation(dv)
        dv.add(ws.cell(WRITE_LOCATION, ClassInfo.MOCKTEST_CHECK_COLUMN))

    # 정렬 및 테두리
    for row in range(1, ws.max_row + 1):
        for col in range(1, ClassInfo.MAX + 1):
            ws.cell(row, col).alignment = ALIGN_CENTER
            ws.cell(row, col).border    = BORDER_ALL

    ws.cell(1, ClassInfo.MOCKTEST_CHECK_COLUMN).alignment = ALIGN_CENTER_WRAP

    save(wb)

def open(data_only:bool=True, read_only:bool=True) -> xl.Workbook:
    return load_workbook(
        WorkbookPaths.current().class_info,
        display_name=ClassInfo.DEFAULT_NAME,
        data_only=data_only,
        read_only=read_only,
    )

def open_temp(data_only:bool=True, read_only:bool=True) -> xl.Workbook:
    paths = WorkbookPaths.current()
    workbook = load_workbook(
        paths.class_info_temp,
        display_name=ClassInfo.TEMP_FILE_NAME,
        data_only=data_only,
        read_only=read_only,
    )
    if not read_only:
        try:
            revision = FileRevision.from_dict(
                read_json(paths.class_info_temp_revision).get("sourceRevision")
            )
        except (OSError, ValueError, KeyError, TypeError):
            revision = capture_file_revision(paths.class_info)
        track_workbook_source(workbook, paths.class_info, revision)
    return workbook

def open_worksheet(wb:xl.Workbook):
    return require_worksheet(wb, ClassInfo.DEFAULT_NAME)

def save(wb:xl.Workbook):
    try:
        atomic_save_workbook(wb, WorkbookPaths.current().class_info)
    except TDMError:
        raise
    except Exception as exc:
        raise FileOpenException(
            f"{ClassInfo.DEFAULT_NAME} 파일을 닫은 뒤 다시 시도해주세요"
        ) from exc
    finally:
        wb.close()

def save_to_temp(wb:xl.Workbook):
    paths = WorkbookPaths.current()
    revision = workbook_source_revision(wb, paths.class_info)
    if not isinstance(revision, (FileRevision, type(None))):
        revision = capture_file_revision(paths.class_info)
    try:
        atomic_save_workbook(
            wb,
            paths.class_info_temp,
            update_source=False,
        )
        atomic_write_json(
            paths.class_info_temp_revision,
            {
                "source": str(paths.class_info.absolute()),
                "sourceRevision": revision.to_dict() if revision else None,
            },
        )
    finally:
        wb.close()

def delete_temp():
    paths = WorkbookPaths.current()
    try:
        paths.class_info_temp.unlink(missing_ok=True)
        paths.class_info_temp_revision.unlink(missing_ok=True)
    except OSError as exc:
        raise FileOpenException(
            f"{ClassInfo.DEFAULT_NAME} 파일이 열려 있어 삭제에 실패했습니다."
        ) from exc

# 파일 유틸리티
def make_backup_file():
    paths = WorkbookPaths.current()
    return create_backup(
        paths.class_info,
        stem=ClassInfo.DEFAULT_NAME,
        backup_dir=paths.backup_dir,
    )

def get_class_info(class_name:str, ws:Worksheet = None):
    """
    반 정보 파일로부터 특정 반의 정보 추출

    return `반 정보 존재 여부`, `담당 선생님`, `수업 요일`, `테스트 응시 시간`
    """
    owned_workbook = None
    if ws is None:
        owned_workbook = open()
        ws = open_worksheet(owned_workbook)

    try:
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, ClassInfo.CLASS_NAME_COLUMN).value == class_name:
                teacher_name  = ws.cell(row, ClassInfo.TEACHER_NAME_COLUMN).value
                class_weekday = ws.cell(row, ClassInfo.CLASS_WEEKDAY_COLUMN).value
                test_time     = ws.cell(row, ClassInfo.TEST_TIME_COLUMN).value
                mock_test_check = ws.cell(row, ClassInfo.MOCKTEST_CHECK_COLUMN).value == "Y"
                return True, teacher_name, class_weekday, test_time, mock_test_check
        return False, None, None, None, False
    finally:
        if owned_workbook is not None:
            owned_workbook.close()

def get_class_names(ws:Worksheet = None, mocktest = False) -> list[str]:
    """
    반 정보 기준 반 이름 리스트 추출
    """
    owned_workbook = None
    if ws is None:
        owned_workbook = open()
        ws = open_worksheet(owned_workbook)

    try:
        class_names = []
        for row in range(2, ws.max_row + 1):
            class_name = ws.cell(row, ClassInfo.CLASS_NAME_COLUMN).value
            if class_name is not None:
                class_names.append(class_name)
            if mocktest and ws.cell(row, ClassInfo.MOCKTEST_CHECK_COLUMN).value == "Y":
                class_names.append(class_name + " (모의고사)")
        return sorted(class_names)
    finally:
        if owned_workbook is not None:
            owned_workbook.close()

def get_new_class_names():
    """
    임시 반 정보 파일에서 새 반 리스트를 리턴
    """
    temp_wb = open_temp()
    temp_ws = open_worksheet(temp_wb)

    res = get_class_names(temp_ws, True)

    temp_wb.close()
    del temp_wb

    return res

# 파일 작업
def make_temp_file_for_update(new_class_list:list[str]):
    """
    반 업데이트 작업에 필요한 임시 반 정보 파일 생성

    팝업창을 기준으로 업데이트 된 반을 추가하고 삭제된 반을 삭제함
    """
    make_backup_file()

    wb = open(read_only=False)
    ws = open_worksheet(wb)

    class_names = set(get_class_names(ws))

    unregistered_class_names = sorted(list(set(new_class_list).difference(class_names)))

    for row in range(2, ws.max_row+1):
        while ws.cell(row, ClassInfo.CLASS_NAME_COLUMN).value is not None and ws.cell(row, ClassInfo.CLASS_NAME_COLUMN).value not in new_class_list:
            ws.delete_rows(row)

    temp_path = str(WorkbookPaths.current().class_info_temp.absolute())

    if len(unregistered_class_names) == 0:
        save_to_temp(wb)
        return temp_path

    for row in range(ws.max_row+1, 1, -1):
        if ws.cell(row-1, ClassInfo.CLASS_NAME_COLUMN).value is not None:
            WRITE_RANGE = WRITE_ROW = row
            break

    for row, class_name in enumerate(unregistered_class_names, start=WRITE_ROW):
        ws.cell(row, ClassInfo.CLASS_NAME_COLUMN).value = class_name

    for row in range(WRITE_RANGE, ws.max_row + 1):
        if ws.cell(row, ClassInfo.CLASS_NAME_COLUMN).value is None:
            break
        for col in range(1, ClassInfo.MAX + 1):
            ws.cell(row, col).alignment = ALIGN_CENTER
            ws.cell(row, col).border    = BORDER_ALL

    save_to_temp(wb)

    return temp_path

def change_class_info(target_class_name:str, target_teacher_name:str):
    """
    특정 반의 담당 선생님 변경
    """
    make_backup_file()

    wb = open(read_only=False)
    ws = open_worksheet(wb)

    for row in range(2, ws.max_row + 1):
        if ws.cell(row, ClassInfo.CLASS_NAME_COLUMN).value == target_class_name:
            ws.cell(row, ClassInfo.TEACHER_NAME_COLUMN).value = target_teacher_name
            break
    else:
        raise InvalidOperationError(f"'{target_class_name}' 반이 존재하지 않습니다.")

    save(wb)

def update_class(prog: Progress | None = None):
    if prog:
        prog.step("반 정보 업데이트 중...")
    save(open_temp(read_only=False))
