import os.path
import openpyxl as xl
import zipfile

from openpyxl.utils.cell import get_column_letter as gcl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation

import tdm.chrome
import tdm.config

from tdm.defs import StudentInfo
from tdm.exception import NoMatchingSheetException, FileOpenException, ReopenFileException
from tdm.style import ALIGN_CENTER, ALIGN_CENTER_WRAP, BORDER_ALL
from tdm.sparse_worksheet import named_rows, delete_rows_sparse as _delete_student_rows

# 파일 기본 작업
def make_file() -> bool:
    wb = xl.Workbook()
    ws = wb.worksheets[0]
    ws.title = StudentInfo.DEFAULT_NAME

    ws[gcl(StudentInfo.STUDENT_NAME_COLUMN)+"1"]       = "이름"
    ws[gcl(StudentInfo.MAKEUPTEST_WEEKDAY_COLUMN)+"1"] = "재시험 응시 요일"
    ws[gcl(StudentInfo.MAKEUPTEST_TIME_COLUMN)+"1"]    = "재시험 응시 시간"
    ws[gcl(StudentInfo.NEW_STUDENT_CHECK_COLUMN)+"1"]  = "기수 신규생"

    ws["Z1"] = "N"
    ws.auto_filter.ref = "A:"+gcl(StudentInfo.MAX)
    ws.freeze_panes    = "A2"
    ws.column_dimensions.group("Z", hidden=True)

    # 첫 행 정렬 및 자동 줄 바꿈
    for col in range(1, StudentInfo.MAX+1):
        ws.cell(1, col).alignment = ALIGN_CENTER_WRAP
        ws.cell(1, col).border    = BORDER_ALL

    return update_student(wb)

def open(data_only:bool=False) -> xl.Workbook:
    try:
        return xl.load_workbook(f"{tdm.config.DATA_DIR}/{StudentInfo.DEFAULT_NAME}.xlsx", data_only=data_only)
    except PermissionError:
        raise ReopenFileException(f"{StudentInfo.DEFAULT_NAME} 파일에 접근할 수 없습니다.\n파일을 직접 연 후 닫으면 문제가 해결될 수 있습니다.")
    except zipfile.BadZipFile:
        raise ReopenFileException(f"{StudentInfo.DEFAULT_NAME} 파일을 직접 연 후 닫으면 문제가 해결될 수 있습니다.")

def open_worksheet(wb:xl.Workbook):
    try:
        return wb[StudentInfo.DEFAULT_NAME]
    except:
        raise NoMatchingSheetException(f"'{StudentInfo.DEFAULT_NAME}.xlsx'의 시트명을 '{StudentInfo.DEFAULT_NAME}'으로 변경해 주세요.")

def save(wb:xl.Workbook):
    try:
        wb.save(f"{tdm.config.DATA_DIR}/{StudentInfo.DEFAULT_NAME}.xlsx")
    except:
        raise FileOpenException()

def isopen() -> bool:
    return os.path.isfile(f"{tdm.config.DATA_DIR}/~${StudentInfo.DEFAULT_NAME}.xlsx")

# 파일 유틸리티
def _student_rows(ws: Worksheet) -> list[tuple[int, str]]:
    """빈 셀을 만들지 않고 이름이 저장된 행만 순서대로 찾는다."""
    return named_rows(ws, StudentInfo.STUDENT_NAME_COLUMN)


def get_student_info(ws:Worksheet, student_name:str):
    """
    학생 정보 파일로부터 학생 정보 추출

    return 파일 내 학생 존재 여부, 재시험 요일, 재시험 시간, 신규생 여부
    """
    for row, name in _student_rows(ws):
        if name == student_name:
            makeup_test_weekday = ws.cell(row, StudentInfo.MAKEUPTEST_WEEKDAY_COLUMN).value
            makeup_test_time    = ws.cell(row, StudentInfo.MAKEUPTEST_TIME_COLUMN).value
            new_studnet         = ws.cell(row, StudentInfo.NEW_STUDENT_CHECK_COLUMN).value
            break
    else:
        return False, None, None, False
    
    return True, makeup_test_weekday, makeup_test_time, new_studnet == 'N'

# 파일 작업
def add_student(target_student_name:str):
    """
    학생 정보 파일 내 신규생 추가
    """
    wb = open()
    ws = open_worksheet(wb)

    students = _student_rows(ws)
    row = students[-1][0] + 1 if students else 2
    ws.cell(row, StudentInfo.STUDENT_NAME_COLUMN).value = target_student_name
    ws.cell(row, StudentInfo.NEW_STUDENT_CHECK_COLUMN).value = "N"
    for col in range(1, StudentInfo.MAX+1):
        ws.cell(row, col).alignment = ALIGN_CENTER
        ws.cell(row, col).border = BORDER_ALL

    save(wb)

def delete_student(target_student_name:str):
    """
    학생 정보 파일에서 학생 정보 삭제
    """
    wb = open()
    ws = open_worksheet(wb)

    _delete_student_rows(ws, [row for row, name in _student_rows(ws) if name == target_student_name])

    save(wb)

def update_student(wb:xl.Workbook=None):
    latest_student_names = tdm.chrome.get_student_names()

    if wb is None:
        wb = open()

    ws = open_worksheet(wb)

    students = _student_rows(ws)
    student_names = {name for _, name in students}
    
    deleted_student_names      = set(student_names).difference(latest_student_names)
    unregistered_student_names = list(set(latest_student_names).difference(student_names))
    
    WRITE_ROW = students[-1][0] + 1 if students else 2
    
    for student_name in sorted(unregistered_student_names):
        ws.cell(WRITE_ROW, StudentInfo.STUDENT_NAME_COLUMN).value = student_name

        dv = DataValidation(type="list", formula1="=Z1", allow_blank=True, errorStyle="stop", showErrorMessage=True)
        dv.error = "이 셀의 값은 'N'이어야 합니다."
        ws.add_data_validation(dv)
        dv.add(ws.cell(WRITE_ROW, StudentInfo.NEW_STUDENT_CHECK_COLUMN))

        for col in range(1, StudentInfo.MAX+1):
            ws.cell(WRITE_ROW, col).alignment = ALIGN_CENTER
            ws.cell(WRITE_ROW, col).border    = BORDER_ALL

        WRITE_ROW += 1

    _delete_student_rows(ws, [row for row, name in students if name in deleted_student_names])

    save(wb)
