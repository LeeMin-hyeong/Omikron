# Omikron v1.2.0-alpha
import json
import os.path
import calendar
import threading
import tkinter as tk
import openpyxl as xl
import win32com.client # only works in Windows

from omikronconst import *
from tkinter import filedialog
from datetime import date as DATE, datetime, timedelta
from dateutil.relativedelta import relativedelta
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Color, PatternFill, Side, Font
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from win32process import CREATE_NO_WINDOW
from webdriver_manager.chrome import ChromeDriverManager

config = json.load(open('./config.json', encoding='UTF8'))
os.environ['WDM_PROGRESS_BAR'] = '0'
service = Service(ChromeDriverManager().install())
service.creation_flags = CREATE_NO_WINDOW

class GUI():
    def __init__(self, ui):
        self.ui = ui
        self.width = 320
        self.height = 435 # button +25
        self.x = int((self.ui.winfo_screenwidth()/4) - (self.width/2))
        self.y = int((self.ui.winfo_screenheight()/2) - (self.height/2))
        self.ui.geometry(f'{self.width}x{self.height}+{self.x}+{self.y}')
        self.ui.title('Omikron')
        self.ui.resizable(False, False)

        tk.Label(self.ui, text='Omikron 데이터 프로그램').pack()
        self.scroll = tk.Scrollbar(self.ui, orient='vertical')
        self.log = tk.Listbox(self.ui, yscrollcommand=self.scroll.set, width=51, height=5)
        self.scroll.config(command=self.log.yview)
        self.log.pack()
        
        tk.Label(self.ui, text='< 기수 변경 관련 >').pack()

        self.class_info_button = tk.Button(self.ui, text='반 정보 기록 양식 생성', width=40, command=lambda: self.class_info_thread())
        self.class_info_button.pack()
        if os.path.isfile('반 정보.xlsx'):
            self.class_info_button['state'] = tk.DISABLED

        self.student_info_button = tk.Button(self.ui, text='학생 정보 기록 양식 생성', width=40, command=lambda: self.student_info_thread())
        self.student_info_button.pack()
        if os.path.isfile('학생 정보.xlsx'):
            self.student_info_button['state'] = tk.DISABLED

        self.make_data_file_button = tk.Button(self.ui, text='데이터 파일 생성', width=40, command=lambda: self.make_data_file_thread())
        self.make_data_file_button.pack()
        if os.path.isfile('./data/' + config['dataFileName'] + '.xlsx'):
            self.make_data_file_button['state'] = tk.DISABLED

        self.update_class_button = tk.Button(self.ui, text='반 업데이트', width=40, command=lambda: self.update_class_thread())
        self.update_class_button.pack()

        tk.Label(self.ui, text='\n< 데이터 저장 및 문자 전송 >').pack()

        self.make_data_form_button = tk.Button(self.ui, text='데일리 테스트 기록 양식 생성', width=40, command=lambda: self.mkae_data_form_thread())
        self.make_data_form_button.pack()

        self.save_data_button = tk.Button(self.ui, text='데이터 엑셀 파일에 저장', width=40, command=lambda: self.save_data_thread())
        self.save_data_button.pack()
        
        self.send_message_button = tk.Button(self.ui, text='시험 결과 전송', width=40, command=lambda: self.send_message_thread())
        self.send_message_button.pack()

        tk.Label(self.ui, text='\n< 데이터 관리 >').pack()

        self.apply_color_button = tk.Button(self.ui, text='데이터 엑셀 파일 조건부 서식 재지정', width=40, command=lambda: apply_color(self))
        self.apply_color_button.pack()

        self.student_menagement_button = tk.Button(self.ui, text='신규 등록 / 퇴원 관리', width=40, command=None)
        self.student_menagement_button.pack()

    def append_log(self, msg:str):
        self.log.insert(tk.END, msg)
        self.log.update()
        self.log.see(tk.END)

    def class_info_thread(self):
        self.class_info_button['state'] = tk.DISABLED
        thread = threading.Thread(target=lambda: class_info(self))
        thread.daemon = True
        thread.start()

    def make_data_file_thread(self):
        self.make_data_file_button['state'] = tk.DISABLED
        thread = threading.Thread(target=lambda: make_data_file(self))
        thread.daemon = True
        thread.start()

    def save_data_thread(self):
        thread = threading.Thread(target=lambda: save_data(self))
        thread.daemon = True
        thread.start()

    def mkae_data_form_thread(self):
        thread = threading.Thread(target=lambda: make_data_form(self))
        thread.daemon = True
        thread.start()

    def send_message_thread(self):
        thread = threading.Thread(target=lambda: send_message(self))
        thread.daemon = True
        thread.start()

    def student_info_thread(self):
        self.student_info_button['state'] = tk.DISABLED
        thread = threading.Thread(target=lambda: student_info(self))
        thread.daemon = True
        thread.start()

    def update_class_thread(self):
        thread = threading.Thread(target=lambda: update_class(self))
        thread.daemon = True
        thread.start()

def make_data_file(gui:GUI):
    gui.make_data_file_button['state'] = tk.DISABLED
    if not os.path.isfile('./data/' + config['dataFileName'] + '.xlsx'):
        gui.append_log('데이터파일 생성 중...')

        ini_wb = xl.Workbook()
        ini_ws = ini_wb.active
        ini_ws.title = '데일리테스트'
        ini_ws[get_column_letter(DataFile.TEST_TIME_COLUMN)+'1'] = '시간'
        ini_ws[get_column_letter(DataFile.DATE_COLUMN)+'1'] = '요일'
        ini_ws[get_column_letter(DataFile.CLASS_NAME_COLUMN)+'1'] = '반'
        ini_ws[get_column_letter(DataFile.TEACHER_COLUMN)+'1'] = '담당'
        ini_ws[get_column_letter(DataFile.STUDENT_NAME_COLUMN)+'1'] = '이름'
        ini_ws[get_column_letter(DataFile.AVERAGE_SCORE_COLUMN)+'1'] = '학생 평균'
        ini_ws.freeze_panes = get_column_letter(DataFile.DATA_COLUMN) + '2'
        ini_ws.auto_filter.ref = 'A:' + get_column_letter(DataFile.MAX)

        # 반 정보 확인
        if not os.path.isfile('./반 정보.xlsx'):
            gui.append_log('[오류] 반 정보.xlsx 파일이 존재하지 않습니다.')
            return
        class_wb = xl.load_workbook("./반 정보.xlsx")
        try:
            class_ws = class_wb['반 정보']
        except:
            gui.append_log('[오류] \'반 정보.xlsx\'의 시트명을')
            gui.append_log('\'반 정보\'로 변경해 주세요.')
            gui.make_data_file_button['state'] = tk.NORMAL
            return

        options = webdriver.ChromeOptions()
        options.add_argument('headless')
        driver = webdriver.Chrome(service = service, options = options)

        # 아이소식 접속
        driver.get(config['url'])
        table_names = driver.find_elements(By.CLASS_NAME, 'style1')

        # 반 루프
        for i in range(3, len(table_names)):
            trs = driver.find_element(By.ID, 'table_' + str(i)).find_elements(By.CLASS_NAME, 'style12')
            write_location = ini_ws.max_row + 1

            class_name = table_names[i].text.rstrip()
            time = ''
            date = ''
            teacher = ''
            is_class_exist = False
            for j in range(2, class_ws.max_row + 1):
                if class_ws.cell(j, ClassInfo.CLASS_NAME_COLUMN).value == class_name:
                    teacher = class_ws.cell(j, ClassInfo.TEACHER_COLUMN).value
                    date = class_ws.cell(j, ClassInfo.DATE_COLUMN).value
                    time = class_ws.cell(j, ClassInfo.TEST_TIME_COLUMN).value
                    is_class_exist = True
            if not is_class_exist:
                continue
            
            # 시험명
            ini_ws.cell(write_location, DataFile.TEST_TIME_COLUMN).value = time
            ini_ws.cell(write_location, DataFile.DATE_COLUMN).value = date
            ini_ws.cell(write_location, DataFile.CLASS_NAME_COLUMN).value = class_name
            ini_ws.cell(write_location, DataFile.TEACHER_COLUMN).value = teacher
            ini_ws.cell(write_location, DataFile.STUDENT_NAME_COLUMN).value = '날짜'
            
            write_location = ini_ws.max_row + 1
            ini_ws.cell(write_location, DataFile.TEST_TIME_COLUMN).value = time
            ini_ws.cell(write_location, DataFile.DATE_COLUMN).value = date
            ini_ws.cell(write_location, DataFile.CLASS_NAME_COLUMN).value = class_name
            ini_ws.cell(write_location, DataFile.TEACHER_COLUMN).value = teacher
            ini_ws.cell(write_location, DataFile.STUDENT_NAME_COLUMN).value = '시험명'
            start = write_location + 1

            # 학생 루프
            for tr in trs:
                write_location = ini_ws.max_row + 1
                ini_ws.cell(write_location, DataFile.TEST_TIME_COLUMN).value = time
                ini_ws.cell(write_location, DataFile.DATE_COLUMN).value = date
                ini_ws.cell(write_location, DataFile.CLASS_NAME_COLUMN).value = class_name
                ini_ws.cell(write_location, DataFile.TEACHER_COLUMN).value = teacher
                ini_ws.cell(write_location, DataFile.STUDENT_NAME_COLUMN).value = tr.find_element(By.CLASS_NAME, 'style9').text
                ini_ws.cell(write_location, DataFile.AVERAGE_SCORE_COLUMN).value = '=ROUND(AVERAGE(G' + str(write_location) + ':XFD' + str(write_location) + '), 0)'
                ini_ws.cell(write_location, DataFile.AVERAGE_SCORE_COLUMN).font = Font(bold=True)
            
            # 시험별 평균
            write_location = ini_ws.max_row + 1
            end = write_location - 1
            ini_ws.cell(write_location, DataFile.TEST_TIME_COLUMN).value = time
            ini_ws.cell(write_location, DataFile.DATE_COLUMN).value = date
            ini_ws.cell(write_location, DataFile.CLASS_NAME_COLUMN).value = class_name
            ini_ws.cell(write_location, DataFile.TEACHER_COLUMN).value = teacher
            ini_ws.cell(write_location, DataFile.STUDENT_NAME_COLUMN).value = '시험 평균'
            ini_ws.cell(write_location, DataFile.AVERAGE_SCORE_COLUMN).value = '=ROUND(AVERAGE(F' + str(start) + ':F' + str(end) + '), 0)'
            ini_ws.cell(write_location, DataFile.AVERAGE_SCORE_COLUMN).font = Font(bold=True)

            for j in range(1, DataFile.DATA_COLUMN):
                ini_ws.cell(write_location, j).border = Border(bottom = Side(border_style='medium', color='000000'))

        # 정렬
        for i in range(1, ini_ws.max_row + 1):
            for j in range(1, ini_ws.max_column + 1):
                ini_ws.cell(i, j).alignment = Alignment(horizontal='center', vertical='center')
        
        # 모의고사 sheet 생성
        copy_ws = ini_wb.copy_worksheet(ini_wb['데일리테스트'])
        copy_ws.title = '모의고사'
        copy_ws.freeze_panes = get_column_letter(DataFile.DATA_COLUMN) + '2'
        copy_ws.auto_filter.ref = 'A:' + get_column_letter(DataFile.STUDENT_NAME_COLUMN)

        ini_wb.save('./data/' + config['dataFileName'] + '.xlsx')
        gui.append_log('데이터 파일을 생성했습니다.')
        gui.ui.wm_attributes("-topmost", 1)
        gui.ui.wm_attributes("-topmost", 0)

    else:
        gui.append_log('이미 파일이 존재합니다')

def make_data_form(gui:GUI):
    gui.append_log('데일리테스트 기록 양식 생성 중...')

    ini_wb = xl.Workbook()
    ini_ws = ini_wb.active
    ini_ws.title = '데일리테스트 기록 양식'
    ini_ws[get_column_letter(DataForm.DATE_COLUMN)+'1'] = '요일'
    ini_ws[get_column_letter(DataForm.TEST_TIME_COLUMN)+'1'] = '시간'
    ini_ws[get_column_letter(DataForm.CLASS_NAME_COLUMN)+'1'] = '반'
    ini_ws[get_column_letter(DataForm.STUDENT_NAME_COLUMN)+'1'] = '이름'
    ini_ws[get_column_letter(DataForm.TEACHER_COLUMN)+'1'] = '담당T'
    ini_ws[get_column_letter(DataForm.DAILYTEST_TEST_NAME_COLUMN)+'1'] = '시험명'
    ini_ws[get_column_letter(DataForm.DAILYTEST_SCORE_COLUMN)+'1'] = '점수'
    ini_ws[get_column_letter(DataForm.DAILYTEST_AVERAGE_COLUMN)+'1'] = '평균'
    ini_ws[get_column_letter(DataForm.MOCKTEST_TEST_NAME_COLUMN)+'1'] = '시험대비 모의고사명'
    ini_ws[get_column_letter(DataForm.MOCKTEST_SCORE_COLUMN)+'1'] = '모의고사 점수'
    ini_ws[get_column_letter(DataForm.MOCKTEST_AVERAGE_COLUMN)+'1'] = '모의고사 평균'
    ini_ws[get_column_letter(DataForm.MAKEUP_TEST_CHECK_COLUMN)+'1'] = '재시문자 X'
    ini_ws['Y1'] = 'X'
    ini_ws['Z1'] = 'x'
    ini_ws.column_dimensions.group('Y', 'Z', hidden=True)
    ini_ws.auto_filter.ref = 'A:'+get_column_letter(DataForm.TEST_TIME_COLUMN)
    if not os.path.isfile('./반 정보.xlsx'):
        gui.append_log('[오류] 반 정보.xlsx 파일이 존재하지 않습니다.')
        return
    
    class_wb = xl.load_workbook("반 정보.xlsx")
    try:
        class_ws = class_wb['반 정보']
    except:
        gui.append_log('[오류] \'반 정보.xlsx\'의 시트명을')
        gui.append_log('\'반 정보\'로 변경해 주세요.')
        return

    options = webdriver.ChromeOptions()
    options.add_argument('headless')
    driver = webdriver.Chrome(service = service, options = options)

    # 아이소식 접속
    driver.get(config['url'])
    table_names = driver.find_elements(By.CLASS_NAME, 'style1')
    #반 루프
    for i in range(3, len(table_names)):
        trs = driver.find_element(By.ID, 'table_' + str(i)).find_elements(By.CLASS_NAME, 'style12')
        write_location = start = ini_ws.max_row + 1

        class_name = table_names[i].text.rstrip()
        teacher = ''
        date = ''
        time = ''
        is_class_exist = False

        for j in range(2, class_ws.max_row + 1):
            if class_ws.cell(j, ClassInfo.CLASS_NAME_COLUMN).value == class_name:
                teacher = class_ws.cell(j, ClassInfo.TEACHER_COLUMN).value
                date = class_ws.cell(j, ClassInfo.DATE_COLUMN).value
                time = class_ws.cell(j, ClassInfo.TEST_TIME_COLUMN).value
        if not is_class_exist:
            continue
        ini_ws.cell(write_location, DataForm.CLASS_NAME_COLUMN).value = class_name
        ini_ws.cell(write_location, DataForm.TEACHER_COLUMN).value = teacher

        #학생 루프
        for tr in trs:
            ini_ws.cell(write_location, DataForm.DATE_COLUMN).value = date
            ini_ws.cell(write_location, DataForm.TEST_TIME_COLUMN).value = time
            ini_ws.cell(write_location, DataForm.STUDENT_NAME_COLUMN).value = tr.find_element(By.CLASS_NAME, 'style9').text
            dv = DataValidation(type='list', formula1='=Y1:Z1', showDropDown=True, allow_blank=True, showErrorMessage=True)
            ini_ws.add_data_validation(dv)
            dv.add(ini_ws.cell(write_location,DataForm.MAKEUP_TEST_CHECK_COLUMN))
            write_location = ini_ws.max_row + 1
        
        end = write_location - 1

        # 시험 평균
        ini_ws.cell(start, DataForm.DAILYTEST_AVERAGE_COLUMN).value = '=ROUND(AVERAGE('+get_column_letter(DataForm.DAILYTEST_SCORE_COLUMN)+str(start)+':'+get_column_letter(DataForm.DAILYTEST_SCORE_COLUMN)+str(end)+'), 0)'
        # 모의고사 평균
        ini_ws.cell(start, DataForm.MOCKTEST_AVERAGE_COLUMN).value = '=ROUND(AVERAGE('+get_column_letter(DataForm.MOCKTEST_SCORE_COLUMN)+str(start)+':'+get_column_letter(DataForm.MOCKTEST_SCORE_COLUMN)+str(end)+'), 0)'
        
        # 정렬 및 테두리
        for j in range(1, ini_ws.max_row + 1):
            for k in range(1, ini_ws.max_column + 1):
                ini_ws.cell(j, k).alignment = Alignment(horizontal='center', vertical='center')
                ini_ws.cell(j, k).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # 셀 병합
        if start < end:
            ini_ws.merge_cells(get_column_letter(DataForm.CLASS_NAME_COLUMN)+str(start)+':'+get_column_letter(DataForm.CLASS_NAME_COLUMN)+str(end))
            ini_ws.merge_cells(get_column_letter(DataForm.TEACHER_COLUMN)+str(start)+':'+get_column_letter(DataForm.TEACHER_COLUMN)+str(end))
            ini_ws.merge_cells(get_column_letter(DataForm.DAILYTEST_TEST_NAME_COLUMN)+str(start)+':'+get_column_letter(DataForm.DAILYTEST_TEST_NAME_COLUMN)+str(end))
            ini_ws.merge_cells(get_column_letter(DataForm.DAILYTEST_AVERAGE_COLUMN)+str(start)+':'+get_column_letter(DataForm.DAILYTEST_AVERAGE_COLUMN)+str(end))
            ini_ws.merge_cells(get_column_letter(DataForm.MOCKTEST_TEST_NAME_COLUMN)+str(start)+':'+get_column_letter(DataForm.MOCKTEST_TEST_NAME_COLUMN)+str(end))
            ini_ws.merge_cells(get_column_letter(DataForm.MOCKTEST_AVERAGE_COLUMN)+str(start)+':'+get_column_letter(DataForm.MOCKTEST_AVERAGE_COLUMN)+str(end))
        
    if os.path.isfile('./데일리테스트 기록 양식.xlsx'):
        i = 1
        while True:
            if not os.path.isfile('./데일리테스트 기록 양식(' + str(i) +').xlsx'):
                ini_wb.save('./데일리테스트 기록 양식(' + str(i) +').xlsx')
                break
            i += 1
    else:
        ini_wb.save('./데일리테스트 기록 양식.xlsx')
    gui.append_log('데일리테스트 기록 양식 생성을 완료했습니다.')
    gui.ui.wm_attributes("-topmost", 1)
    gui.ui.wm_attributes("-topmost", 0)

def save_data(gui:GUI):
    makeup_test_date = holiday_dialog(gui, gui.save_data_button)

    if gui.save_data_button['state'] == tk.NORMAL: return
    # 입력 양식 엑셀
    filepath = filedialog.askopenfilename(initialdir='./', title='데일리테스트 기록 양식 선택', filetypes=(('Excel files', '*.xlsx'),('all files', '*.*')))
    if filepath == '':
        gui.save_data_button['state'] = tk.NORMAL
        return
    
    form_wb = xl.load_workbook(filepath, data_only=True)
    form_ws = form_wb['데일리테스트 기록 양식']

    # 올바른 양식이 아닙니다.
    if not check_data_form(gui, form_ws):
        gui.append_log('데이터 저장이 중단되었습니다.')
        gui.save_data_button['state'] = tk.NORMAL
        return
    
    # 데이터 저장 엑셀
    if not os.path.isfile('./data/' + config['dataFileName'] + '.xlsx'):
        gui.append_log('[오류] ' + config['dataFileName'] + '.xlsx' + '파일이 존재하지 않습니다.')
        gui.save_data_button['state'] = tk.NORMAL
        return

    # 학생 정보 열기
    if not os.path.isfile('./학생 정보.xlsx'):
        gui.append_log('[오류] 학생 정보.xlsx 파일이 존재하지 않습니다.')
        gui.save_data_button['state'] = tk.NORMAL
        return
    student_wb = xl.load_workbook("./학생 정보.xlsx")
    try:
        student_ws = student_wb['학생 정보']
    except:
        gui.append_log('[오류] \'학생 정보.xlsx\'의 시트명을')
        gui.append_log('\'학생 정보\'로 변경해 주세요.')
        gui.save_data_button['state'] = tk.NORMAL
        return

    # 재시험 명단 열기
    if not os.path.isfile('./data/재시험 명단.xlsx'):
        gui.append_log('재시험 명단 파일 생성 중...')

        ini_wb = xl.Workbook()
        ini_ws = ini_wb.active
        ini_ws.title = '재시험 명단'
        ini_ws[get_column_letter(MakeupTestList.TEST_DATE_COLUMN)+'1'] = '응시일'
        ini_ws[get_column_letter(MakeupTestList.CLASS_NAME_COLUMN)+'1'] = '반'
        ini_ws[get_column_letter(MakeupTestList.TEACHER_COLUMN)+'1'] = '담당T'
        ini_ws[get_column_letter(MakeupTestList.STUDENT_NAME_COLUMN)+'1'] = '이름'
        ini_ws[get_column_letter(MakeupTestList.TEST_NAME_COLUMN)+'1'] = '시험명'
        ini_ws[get_column_letter(MakeupTestList.TEST_SCORE_COLUMN)+'1'] = '시험 점수'
        ini_ws[get_column_letter(MakeupTestList.MAKEUP_TEST_WEEK_DATE_COLUMN)+'1'] = '재시 요일'
        ini_ws[get_column_letter(MakeupTestList.MAKEUP_TEST_TIME_COLUMN)+'1'] = '재시 시간'
        ini_ws[get_column_letter(MakeupTestList.MAKEUP_TEST_DATE_COLUMN)+'1'] = '재시 날짜'
        ini_ws[get_column_letter(MakeupTestList.MAKEUP_TEST_SCORE_COLUMN)+'1'] = '재시 점수'
        ini_ws[get_column_letter(MakeupTestList.ETC_COLUMN)+'1'] = '비고'
        ini_ws.auto_filter.ref = 'A:'+get_column_letter(MakeupTestList.MAX)
        ini_wb.save('./data/재시험 명단.xlsx')
    makeup_list_wb = xl.load_workbook('./data/재시험 명단.xlsx')
    try:
        makeup_list_ws = makeup_list_wb['재시험 명단']
    except:
        gui.append_log('[오류] \'재시험 명단.xlsx\'의 시트명을')
        gui.append_log('\'재시험 명단\'으로 변경해 주세요.')
        gui.save_data_button['state'] = tk.NORMAL
        return
    
    excel = win32com.client.Dispatch('Excel.Application')
    excel.Visible = False
    wb = excel.Workbooks.Open(os.getcwd() + '\\data\\' + config['dataFileName'] + '.xlsx')
    wb.Save()
    wb.Close()

    gui.append_log('데이터 저장 중...')
    data_file_wb = xl.load_workbook('./data/' + config['dataFileName'] + '.xlsx')
    data_file_wb.save('./data/' + config['dataFileName'] + '(' + datetime.today().strftime('%Y%m%d') + ').xlsx')

    # 데일리 테스트 작성
    data_file_ws = data_file_wb['데일리테스트']

    # 동적 열 탐색
    for i in range(1, data_file_ws.max_column+1):
        if data_file_ws.cell(1, i).value == '반':
            CLASS_COLUMN = i
            break
    for i in range(1, data_file_ws.max_column+1):
        if data_file_ws.cell(1, i).value == '이름':
            NAME_COLUMN = i
            break
    for i in range(1, data_file_ws.max_column+1):
        if data_file_ws.cell(1, i).value == '학생 평균':
            SCORE_COLUMN = i
            break
    
    for i in range(2, form_ws.max_row+1): # 데일리데이터 기록 양식 루프
        # 파일 끝 검사
        if form_ws.cell(i, 4).value is None:
            break
        
        # 반 필터링
        if (form_ws.cell(i, DataForm.CLASS_NAME_COLUMN).value is not None) and (form_ws.cell(i, DataForm.DAILYTEST_TEST_NAME_COLUMN).value is not None):
            class_name = form_ws.cell(i, DataForm.CLASS_NAME_COLUMN).value
            test_name = form_ws.cell(i, DataForm.DAILYTEST_TEST_NAME_COLUMN).value
            teacher = form_ws.cell(i, DataForm.TEACHER_COLUMN).value

            #반 시작 찾기
            for j in range(2, data_file_ws.max_row+1):
                if data_file_ws.cell(j, CLASS_COLUMN).value == class_name: # data class_name == form class_name
                    start = j # 데이터파일에서 반이 시작하는 행 번호
                    break
            # 반 끝 찾기
            for j in range(start, data_file_ws.max_row+1):
                if data_file_ws.cell(j, NAME_COLUMN).value == '시험 평균': # data name is 시험 평균
                    end = j # 데이터파일에서 반이 끝나는 행 번호
                    break
            
            # 데일리테스트 작성 열 위치 찾기
            for j in range(SCORE_COLUMN+1, data_file_ws.max_column+2):
                if data_file_ws.cell(start, j).value is None:
                    write_column = j
                    break
                if data_file_ws.cell(start, j).value.strftime('%y.%m.%d') == DATE.today().strftime('%y.%m.%d'):
                    write_column = j
                    break
            # 입력 틀 작성
            average = '=ROUND(AVERAGE(' + get_column_letter(write_column) + str(start + 2) + ':' + get_column_letter(write_column) + str(end - 1) + '), 0)'
            data_file_ws.cell(start, write_column).value = DATE.today()
            data_file_ws.cell(start, write_column).number_format = 'yyyy.mm.dd(aaa)'
            data_file_ws.cell(start, write_column).alignment = Alignment(horizontal='center', vertical='center')

            data_file_ws.cell(start + 1, write_column).value = test_name
            data_file_ws.cell(start + 1, write_column).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            data_file_ws.cell(end, write_column).value = average
            data_file_ws.cell(end, write_column).alignment = Alignment(horizontal='center', vertical='center')
            data_file_ws.cell(end, write_column).border = Border(bottom=Side(border_style='medium', color='000000'))
            
        score = form_ws.cell(i, DataForm.DAILYTEST_SCORE_COLUMN).value
        if score is None:
            continue # 점수 없으면 미응시 처리
        
        for j in range(start + 2, end):
            if data_file_ws.cell(j, NAME_COLUMN).value == form_ws.cell(i, DataForm.STUDENT_NAME_COLUMN).value: # data name == form name
                data_file_ws.cell(j, write_column).value = score
                data_file_ws.cell(j, write_column).alignment = Alignment(horizontal='center', vertical='center')
                break

        # 재시험 작성
        if (type(score) == int) and (score < 80) and (form_ws.cell(i, DataForm.MAKEUP_TEST_CHECK_COLUMN).value != 'x') and (form_ws.cell(i, DataForm.MAKEUP_TEST_CHECK_COLUMN).value != 'X'):
            check = makeup_list_ws.max_row
            duplicated = False
            while makeup_list_ws.cell(check, MakeupTestList.TEST_DATE_COLUMN).value.strftime('%y.%m.%d') == DATE.today().strftime('%y.%m.%d'):
                if makeup_list_ws.cell(check, MakeupTestList.STUDENT_NAME_COLUMN).value == form_ws.cell(i, DataForm.STUDENT_NAME_COLUMN).value:
                    duplicated = True
                    break
                check -= 1
            if duplicated: continue

            for j in range(2, student_ws.max_row+1):
                if student_ws.cell(j, StudentInfo.STUDENT_NAME_COLUMN).value == form_ws.cell(i, DataForm.STUDENT_NAME_COLUMN).value:
                    dates = student_ws.cell(j, StudentInfo.MAKEUP_TEST_WEEK_DATE_COLUMN).value
                    time = student_ws.cell(j, StudentInfo.MAKEUP_TEST_TIME_COLUMN).value
                    new_studnet = student_ws.cell(j, StudentInfo.NEW_STUDENT_CHECK_COLUMN).value
                    break
            for j in range(2, makeup_list_ws.max_row + 2):
                if makeup_list_ws.cell(j, MakeupTestList.TEST_DATE_COLUMN).value is None:
                    writeRow = j
                    break
            makeup_list_ws.cell(writeRow, MakeupTestList.TEST_DATE_COLUMN).value = DATE.today()
            makeup_list_ws.cell(writeRow, MakeupTestList.CLASS_NAME_COLUMN).value = class_name
            makeup_list_ws.cell(writeRow, MakeupTestList.TEACHER_COLUMN).value = teacher
            makeup_list_ws.cell(writeRow, MakeupTestList.STUDENT_NAME_COLUMN).value = form_ws.cell(i, DataForm.STUDENT_NAME_COLUMN).value
            if (new_studnet is not None) and (new_studnet == 'N'):
                makeup_list_ws.cell(writeRow, MakeupTestList.STUDENT_NAME_COLUMN).fill = PatternFill(fill_type='solid', fgColor=Color('FFFF00'))
            makeup_list_ws.cell(writeRow, MakeupTestList.TEST_NAME_COLUMN).value = test_name
            makeup_list_ws.cell(writeRow, MakeupTestList.TEST_SCORE_COLUMN).value = score
            if dates is not None:
                makeup_list_ws.cell(writeRow, MakeupTestList.MAKEUP_TEST_WEEK_DATE_COLUMN).value = dates
                date_list = dates.split('/')
                result = makeup_test_date[date_list[0].replace(' ', '')]
                for d in date_list:
                    if result > makeup_test_date[d.replace(' ', '')]:
                        result = makeup_test_date[d.replace(' ', '')]
                if time is not None:
                    makeup_list_ws.cell(writeRow, MakeupTestList.MAKEUP_TEST_TIME_COLUMN).value = time
                makeup_list_ws.cell(writeRow, MakeupTestList.MAKEUP_TEST_DATE_COLUMN).value = result
                makeup_list_ws.cell(writeRow, MakeupTestList.MAKEUP_TEST_DATE_COLUMN).number_format = 'mm월 dd일(aaa)'
        
    # 모의고사 작성
    data_file_ws = data_file_wb['모의고사']

    # 동적 열 탐색
    for i in range(1, data_file_ws.max_column+1):
        if data_file_ws.cell(1, i).value == '반':
            CLASS_COLUMN = i
            break
    for i in range(1, data_file_ws.max_column+1):
        if data_file_ws.cell(1, i).value == '이름':
            NAME_COLUMN = i
            break
    for i in range(1, data_file_ws.max_column+1):
        if data_file_ws.cell(1, i).value == '학생 평균':
            SCORE_COLUMN = i
            break
    
    for i in range(2, form_ws.max_row+1): # 데일리데이터 기록 양식 루프
        # 파일 끝 검사
        if form_ws.cell(i, DataForm.STUDENT_NAME_COLUMN).value is None:
            break
        
        # 반 필터링
        if (form_ws.cell(i, DataForm.CLASS_NAME_COLUMN).value is not None) and (form_ws.cell(i, DataForm.MOCKTEST_TEST_NAME_COLUMN).value is not None): # form class_name is not None and form mock_test_name is not None
            class_name = form_ws.cell(i, DataForm.CLASS_NAME_COLUMN).value
            test_name = form_ws.cell(i, DataForm.MOCKTEST_TEST_NAME_COLUMN).value
            teacher = form_ws.cell(i, DataForm.TEACHER_COLUMN).value

            #반 시작 찾기
            for j in range(2, data_file_ws.max_row+1):
                if data_file_ws.cell(j, CLASS_COLUMN).value == class_name: # data class_name == form class_name
                    start = j # 데이터파일에서 반이 시작하는 행 번호
                    break
            # 반 끝 찾기
            for j in range(start, data_file_ws.max_row+1):
                if data_file_ws.cell(j, NAME_COLUMN).value == '시험 평균': # data name is 시험 평균
                    end = j # 데이터파일에서 반이 끝나는 행 번호
                    break
            
            # 데일리테스트 작성 열 위치 찾기
            for j in range(SCORE_COLUMN+1, data_file_ws.max_column+2):
                if data_file_ws.cell(start, j).value is None:
                    write_column = j
                    break
                if data_file_ws.cell(start, j).value.strftime('%y.%m.%d') == DATE.today().strftime('%y.%m.%d'):
                    write_column = j
                    break
            # 입력 틀 작성
            average = '=ROUND(AVERAGE(' + get_column_letter(write_column) + str(start + 2) + ':' + get_column_letter(write_column) + str(end - 1) + '), 0)'
            data_file_ws.cell(start, write_column).value = DATE.today()
            data_file_ws.cell(start, write_column).number_format = 'yyyy.mm.dd(aaa)'
            data_file_ws.cell(start, write_column).alignment = Alignment(horizontal='center', vertical='center')

            data_file_ws.cell(start + 1, write_column).value = test_name
            data_file_ws.cell(start + 1, write_column).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            data_file_ws.cell(end, write_column).value = average
            data_file_ws.cell(end, write_column).alignment = Alignment(horizontal='center', vertical='center')
            data_file_ws.cell(end, write_column).border = Border(bottom=Side(border_style='medium', color='000000'))
            
        score = form_ws.cell(i, DataForm.MOCKTEST_SCORE_COLUMN).value
        if score is None:
            continue # 점수 없으면 미응시 처리
        
        for j in range(start + 2, end):
            if data_file_ws.cell(j, NAME_COLUMN).value == form_ws.cell(i, DataForm.STUDENT_NAME_COLUMN).value: # data name == form name
                data_file_ws.cell(j, write_column).value = score
                data_file_ws.cell(j, write_column).alignment = Alignment(horizontal='center', vertical='center')
                break
        
        # 재시험 작성
        if (type(score) == int) and (score < 80) and (form_ws.cell(i, DataForm.MAKEUP_TEST_CHECK_COLUMN).value != 'x') and (form_ws.cell(i, DataForm.MAKEUP_TEST_CHECK_COLUMN).value != 'X'):
            check = makeup_list_ws.max_row
            duplicated = False
            while makeup_list_ws.cell(check, MakeupTestList.TEST_DATE_COLUMN).value.strftime('%y.%m.%d') == DATE.today().strftime('%y.%m.%d'):
                if makeup_list_ws.cell(check, MakeupTestList.STUDENT_NAME_COLUMN).value == form_ws.cell(i, DataForm.STUDENT_NAME_COLUMN).value:
                    duplicated = True
                    break
                check -= 1
            if duplicated: continue

            for j in range(2, student_ws.max_row+1):
                if student_ws.cell(j, StudentInfo.STUDENT_NAME_COLUMN).value == form_ws.cell(i, DataForm.STUDENT_NAME_COLUMN).value:
                    dates = student_ws.cell(j, StudentInfo.MAKEUP_TEST_WEEK_DATE_COLUMN).value
                    time = student_ws.cell(j, StudentInfo.MAKEUP_TEST_TIME_COLUMN).value
                    new_studnet = student_ws.cell(j, StudentInfo.NEW_STUDENT_CHECK_COLUMN).value
                    break
            for j in range(2, makeup_list_ws.max_row + 2):
                if makeup_list_ws.cell(j, MakeupTestList.TEST_DATE_COLUMN).value is None:
                    writeRow = j
                    break
            makeup_list_ws.cell(writeRow, MakeupTestList.TEST_DATE_COLUMN).value = DATE.today()
            makeup_list_ws.cell(writeRow, MakeupTestList.CLASS_NAME_COLUMN).value = class_name
            makeup_list_ws.cell(writeRow, MakeupTestList.TEACHER_COLUMN).value = teacher
            makeup_list_ws.cell(writeRow, MakeupTestList.STUDENT_NAME_COLUMN).value = form_ws.cell(i, DataForm.STUDENT_NAME_COLUMN).value
            if (new_studnet is not None) and (new_studnet == 'N'):
                makeup_list_ws.cell(writeRow, MakeupTestList.STUDENT_NAME_COLUMN).fill = PatternFill(fill_type='solid', fgColor=Color('FFFF00'))
            makeup_list_ws.cell(writeRow, MakeupTestList.TEST_NAME_COLUMN).value = test_name
            makeup_list_ws.cell(writeRow, MakeupTestList.TEST_SCORE_COLUMN).value = score
            if dates is not None:
                makeup_list_ws.cell(writeRow, MakeupTestList.MAKEUP_TEST_WEEK_DATE_COLUMN).value = dates
                date_list = dates.split('/')
                result = makeup_test_date[date_list[0].replace(' ', '')]
                for d in date_list:
                    if result > makeup_test_date[d.replace(' ', '')]:
                        result = makeup_test_date[d.replace(' ', '')]
                if time is not None:
                    makeup_list_ws.cell(writeRow, MakeupTestList.MAKEUP_TEST_TIME_COLUMN).value = time
                makeup_list_ws.cell(writeRow, MakeupTestList.MAKEUP_TEST_DATE_COLUMN).value = result
                makeup_list_ws.cell(writeRow, MakeupTestList.MAKEUP_TEST_DATE_COLUMN).number_format = 'mm월 dd일(aaa)'

    gui.append_log('재시험 명단 작성 중...')
    for j in range(1, makeup_list_ws.max_row + 1):
        if makeup_list_ws.cell(j, 1).value is None: break
        for k in range(1, makeup_list_ws.max_column + 1):
            makeup_list_ws.cell(j, k).alignment = Alignment(horizontal='center', vertical='center')
            makeup_list_ws.cell(j, k).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    gui.append_log('백업 파일 생성중...')
    form_wb.save('./data/backup/데일리테스트 기록 양식(' + datetime.today().strftime('%Y%m%d') + ').xlsx')
    
    try:
        data_file_wb.save('./data/' + config['dataFileName'] + '.xlsx')
    except:
        gui.append_log('데이터 파일 창을 끄고 다시 실행해 주세요.')
        gui.save_data_button['state'] = tk.NORMAL
        return

    data_file_ws = data_file_wb['데일리테스트']
    data_file_wb.save('./data/' + config['dataFileName'] + '.xlsx')
    try:
        makeup_list_wb.save('./data/재시험 명단.xlsx')
    except:
        gui.append_log('재시험 명단 파일 창을 끄고 다시 실행해 주세요.')
        gui.save_data_button['state'] = tk.NORMAL
        return
    
    apply_color(gui)

    gui.append_log('데이터 저장을 완료했습니다.')
    gui.ui.wm_attributes("-topmost", 1)
    gui.ui.wm_attributes("-topmost", 0)
    gui.save_data_button['state'] = tk.NORMAL

def class_info(gui:GUI):
    if not os.path.isfile('반 정보.xlsx'):
        gui.append_log('반 정보 입력 파일 생성 중...')

        ini_wb = xl.Workbook()
        ini_ws = ini_wb.active
        ini_ws.title = '반 정보'
        ini_ws[get_column_letter(ClassInfo.CLASS_NAME_COLUMN)+'1'] = '반명'
        ini_ws[get_column_letter(ClassInfo.TEACHER_COLUMN)+'1'] = '선생님명'
        ini_ws[get_column_letter(ClassInfo.DATE_COLUMN)+'1'] = '요일'
        ini_ws[get_column_letter(ClassInfo.TEST_TIME_COLUMN)+'1'] = '시간'

        options = webdriver.ChromeOptions()
        options.add_argument('headless')
        driver = webdriver.Chrome(service = service, options = options)

        # 아이소식 접속
        driver.get(config['url'])
        table_names = driver.find_elements(By.CLASS_NAME, 'style1')

        # 반 루프
        for tableName in table_names:
            write_location = ini_ws.max_row + 1
            ini_ws.cell(write_location, 1).value = tableName.text.rstrip()

        # 정렬 및 테두리
        for j in range(1, ini_ws.max_row + 1):
            for k in range(1, ini_ws.max_column + 1):
                ini_ws.cell(j, k).alignment = Alignment(horizontal='center', vertical='center')
                ini_ws.cell(j, k).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ini_wb.save('./반 정보.xlsx')
        gui.class_info_button['state'] = tk.DISABLED
        gui.append_log('반 정보 입력 파일 생성을 완료했습니다.')
        gui.append_log('반 정보를 입력해 주세요.')
        gui.ui.wm_attributes("-topmost", 1)
        gui.ui.wm_attributes("-topmost", 0)

    else:
        gui.append_log('이미 파일이 존재합니다.')
        gui.class_info_button['state'] = tk.DISABLED

def send_message(gui:GUI):
    
    makeup_test_date = holiday_dialog(gui, gui.send_message_button)
    
    if gui.send_message_button['state'] == tk.NORMAL: return

    filepath = filedialog.askopenfilename(initialdir='./', title='데일리테스트 기록 양식 선택', filetypes=(('Excel files', '*.xlsx'),('all files', '*.*')))
    if filepath == '':
        gui.send_message_button['state'] = tk.NORMAL
        return
    
    # 점수기록표 xlsx
    form_ws = xl.load_workbook(filepath, data_only=True)
    form_ws = form_ws['데일리테스트 기록 양식']
    
    # 올바른 양식이 아닙니다.
    if not check_data_form(gui, form_ws):
        gui.append_log('시험 점수 전송이 중단되었습니다.')
        gui.send_message_button['state'] = tk.NORMAL
        return

    try:
        gui.append_log('크롬을 실행시키는 중...')
        options = webdriver.ChromeOptions()
        options.add_experimental_option("detach", True)
        driver = webdriver.Chrome(service=service, options=options)

        if not os.path.isfile('./학생 정보.xlsx'):
            gui.append_log('[오류] 학생 정보.xlsx 파일이 존재하지 않습니다.')
            return
        student_wb = xl.load_workbook("./학생 정보.xlsx")
        try:
            student_ws = student_wb['학생 정보']
        except:
            gui.append_log('[오류] \'학생 정보.xlsx\'의 시트명을')
            gui.append_log('\'학생 정보\'로 변경해 주세요.')
            gui.send_message_button['state'] = tk.NORMAL
            return
        
        # 아이소식 접속
        driver.get(config['url'])
        driver.find_element(By.XPATH, '//*[@id="ctitle"]').send_keys(config['dailyTest'])
        
        driver.execute_script('window.open('');')
        driver.switch_to.window(driver.window_handles[1])
        driver.get(config['url'])
        driver.find_element(By.XPATH, '//*[@id="ctitle"]').send_keys(config['makeupTest'])

        driver.execute_script('window.open('');')
        driver.switch_to.window(driver.window_handles[2])
        driver.get(config['url'])
        driver.find_element(By.XPATH, '//*[@id="ctitle"]').send_keys(config['makeupTestDate'])

        gui.append_log('메시지 작성 중...')
        for i in range(2, form_ws.max_row+1):
            driver.switch_to.window(driver.window_handles[0])
            name = form_ws.cell(i, DataForm.STUDENT_NAME_COLUMN).value
            daily_test_score = form_ws.cell(i, DataForm.DAILYTEST_SCORE_COLUMN).value
            mock_test_score = form_ws.cell(i, DataForm.MOCKTEST_SCORE_COLUMN).value
            if form_ws.cell(i, 3).value is not None:
                class_name = form_ws.cell(i, DataForm.CLASS_NAME_COLUMN).value
                daily_test_name = form_ws.cell(i, DataForm.DAILYTEST_TEST_NAME_COLUMN).value
                mock_test_name = form_ws.cell(i, DataForm.MOCKTEST_TEST_NAME_COLUMN).value
                daily_test_average = form_ws.cell(i, DataForm.DAILYTEST_AVERAGE_COLUMN).value
                mock_test_average = form_ws.cell(i, DataForm.MOCKTEST_AVERAGE_COLUMN).value

            # 시험 미응시시 건너뛰기
            if daily_test_score is not None:
                test_name = daily_test_name
                score = daily_test_score
                average = daily_test_average
            elif mock_test_score is not None:
                test_name = mock_test_name
                score = mock_test_score
                average = mock_test_average
            else:
                continue

            if type(score) != int:
                continue
            
            table_names = driver.find_elements(By.CLASS_NAME, 'style1')
            for j in range(len(table_names)):
                if class_name in table_names[j].text:
                    index = j
                    break

            trs = driver.find_element(By.ID, 'table_' + str(index)).find_elements(By.CLASS_NAME, 'style12')
            for tr in trs:
                if tr.find_element(By.CLASS_NAME, 'style9').text == name:
                    tds = tr.find_elements(By.TAG_NAME, 'td')
                    tds[0].find_element(By.TAG_NAME, 'input').send_keys(test_name)
                    tds[1].find_element(By.TAG_NAME, 'input').send_keys(score)
                    tds[2].find_element(By.TAG_NAME, 'input').send_keys(average)
                    break
            
            if (type(score) == int) and (score < 80) and (form_ws.cell(i, 12).value != 'x') and (form_ws.cell(i, 12).value != 'X'):
                for j in range(2, student_ws.max_row+1):
                    if student_ws.cell(j, 1).value == name:
                        date = student_ws.cell(j, 4).value
                        time = student_ws.cell(j, 5).value
                        break
                if date is None:
                    driver.switch_to.window(driver.window_handles[1])
                    trs = driver.find_element(By.ID, 'table_' + str(index)).find_elements(By.CLASS_NAME, 'style12')
                    for tr in trs:
                        if tr.find_element(By.CLASS_NAME, 'style9').text == name:
                            tds = tr.find_elements(By.TAG_NAME, 'td')
                            tds[0].find_element(By.TAG_NAME, 'input').send_keys(test_name)
                else:
                    date_list = date.split('/')
                    result = makeup_test_date[date_list[0].replace(' ', '')]
                    timeIndex = 0
                    for i in range(len(date_list)):
                        if result > makeup_test_date[date_list[i].replace(' ', '')]:
                            result = makeup_test_date[date_list[i].replace(' ', '')]
                            timeIndex = i
                    driver.switch_to.window(driver.window_handles[2])
                    trs = driver.find_element(By.ID, 'table_' + str(index)).find_elements(By.CLASS_NAME, 'style12')
                    for tr in trs:
                        if tr.find_element(By.CLASS_NAME, 'style9').text == name:
                            tds = tr.find_elements(By.TAG_NAME, 'td')
                            tds[0].find_element(By.TAG_NAME, 'input').send_keys(test_name)
                            try:
                                if time is not None:
                                    if "/" in str(time):
                                        tds[1].find_element(By.TAG_NAME, 'input').send_keys(result.strftime('%m월 %d일') + ' ' + str(time).split('/')[timeIndex] + '시')
                                    else:
                                        tds[1].find_element(By.TAG_NAME, 'input').send_keys(result.strftime('%m월 %d일') + ' ' + str(time)+ '시')
                            except:
                                gui.append_log(name + "의 재시험 일정을 요일별 시간으로 설정하거나")
                                gui.append_log("한 시간으로 통일해 주세요.")
                                gui.append_log("중단되었습니다.")
                                gui.send_message_button['state'] = tk.NORMAL
                                driver.quit()
                                return

        gui.send_message_button['state'] = tk.NORMAL
        gui.append_log('메시지 입력을 완료했습니다.')
        gui.append_log('메시지 확인 후 전송해주세요.')
        gui.ui.wm_attributes("-topmost", 1)
        gui.ui.wm_attributes("-topmost", 0)
    except:
        gui.append_log('중단되었습니다.')
        gui.send_message_button['state'] = tk.NORMAL
        return

def student_info(gui:GUI):
    if not os.path.isfile('./학생 정보.xlsx'):
        gui.append_log('학생 정보 파일 생성 중...')

        ini_wb = xl.Workbook()
        ini_ws = ini_wb.active
        ini_ws.title = '학생 정보'
        ini_ws[get_column_letter(StudentInfo.STUDENT_NAME_COLUMN)+'1'] = '이름'
        ini_ws[get_column_letter(StudentInfo.CLASS_NAME_COLUMN)+'1'] = '반명'
        ini_ws[get_column_letter(StudentInfo.TEACHER_COLUMN)+'1'] = '담당'
        ini_ws[get_column_letter(StudentInfo.MAKEUP_TEST_WEEK_DATE_COLUMN)+'1'] = '요일'
        ini_ws[get_column_letter(StudentInfo.MAKEUP_TEST_TIME_COLUMN)+'1'] = '시간'
        ini_ws[get_column_letter(StudentInfo.NEW_STUDENT_CHECK_COLUMN)+'1'] = '기수 신규생'
        ini_ws['Z1'] = 'N'
        ini_ws.auto_filter.ref = 'A:'+get_column_letter(StudentInfo.MAX)
        ini_ws.column_dimensions.group('Z', hidden=True)

        # 반 정보 확인
        if not os.path.isfile('./반 정보.xlsx'):
            gui.append_log('[오류] 반 정보.xlsx 파일이 존재하지 않습니다.')
            return
        class_wb = xl.load_workbook("./반 정보.xlsx")
        try:
            class_ws = class_wb['반 정보']
        except:
            gui.append_log('[오류] \'반 정보.xlsx\'의 시트명을')
            gui.append_log('\'반 정보\'로 변경해 주세요.')
            gui.makeupTestInfoButton['state'] = tk.NORMAL
            return

        options = webdriver.ChromeOptions()
        options.add_argument('headless')
        driver = webdriver.Chrome(service = service, options = options)

        # 아이소식 접속
        driver.get(config['url'])
        table_names = driver.find_elements(By.CLASS_NAME, 'style1')

        # 반 루프
        for i in range(3, len(table_names)):
            trs = driver.find_element(By.ID, 'table_' + str(i)).find_elements(By.CLASS_NAME, 'style12')
            write_location = ini_ws.max_row + 1
            teacher = ''

            class_name = table_names[i].text.rstrip()
            is_class_exist = False
            for j in range(2, class_ws.max_row + 1):
                if class_ws.cell(j, 1).value == class_name:
                    teacher = class_ws.cell(j, 2).value
                    is_class_exist = True
            if not is_class_exist:
                continue

            # 학생 루프
            for tr in trs:
                write_location = ini_ws.max_row + 1
                ini_ws.cell(write_location, StudentInfo.STUDENT_NAME_COLUMN).value = tr.find_element(By.CLASS_NAME, 'style9').text
                ini_ws.cell(write_location, StudentInfo.CLASS_NAME_COLUMN).value = class_name
                ini_ws.cell(write_location, StudentInfo.TEACHER_COLUMN).value = teacher
                dv = DataValidation(type='list', formula1='=Z1',  allow_blank=True, errorStyle='stop', showErrorMessage=True)
                ini_ws.add_data_validation(dv)
                dv.add(ini_ws.cell(write_location, StudentInfo.NEW_STUDENT_CHECK_COLUMN))

        # 정렬 및 테두리
        for j in range(1, ini_ws.max_row + 1):
            for k in range(1, StudentInfo.MAX + 1):
                ini_ws.cell(j, k).alignment = Alignment(horizontal='center', vertical='center')
                ini_ws.cell(j, k).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        ini_wb.save('./학생 정보.xlsx')
        gui.append_log('학생 정보 파일을 생성했습니다.')
        gui.ui.wm_attributes("-topmost", 1)
        gui.ui.wm_attributes("-topmost", 0)

    else:
        gui.append_log('이미 파일이 존재합니다')
        gui.makeupTestInfoButton['state'] = tk.DISABLED

def apply_color(gui:GUI):
    gui.apply_color_button['state'] = tk.DISABLED
    try:
        if not os.path.isfile('./data/' + config['dataFileName'] + '.xlsx'):
            gui.append_log('[오류] '+ config['dataFileName'] +'.xlsx 파일이 존재하지 않습니다.')
            gui.apply_color_button['state'] = tk.NORMAL
            return
        
        if not os.path.isfile('./학생 정보.xlsx'):
            gui.append_log('[오류] 학생 정보.xlsx 파일이 존재하지 않습니다.')
            return
        student_wb = xl.load_workbook("./학생 정보.xlsx")
        try:
            student_ws = student_wb['학생 정보']
        except:
            gui.append_log('[오류] \'학생 정보.xlsx\'의 시트명을')
            gui.append_log('\'학생 정보\'로 변경해 주세요.')
            gui.apply_color_button['state'] = tk.NORMAL
            return
        
        excel = win32com.client.Dispatch('Excel.Application')
        excel.Visible = False
        wb = excel.Workbooks.Open(os.getcwd() + '\\data\\' + config['dataFileName'] + '.xlsx')
        wb.Save()
        wb.Close()

        gui.append_log('조건부 서식 적용중...')

        data_file_wb = xl.load_workbook('./data/' + config['dataFileName'] + '.xlsx')
        dataFileColorWb = xl.load_workbook('./data/' + config['dataFileName'] + '.xlsx', data_only=True)
    except:
        gui.append_log('데이터 파일 창을 끄고 다시 실행해 주세요.')
        gui.apply_color_button['state'] = tk.NORMAL
        return
    
    try:
        for sheetName in data_file_wb.sheetnames:
            data_file_ws = data_file_wb[sheetName]
            dataFileColorWs = dataFileColorWb[sheetName]

            for i in range(1, data_file_ws.max_column):
                if data_file_ws.cell(1, i).value == '이름':
                    NAME_COLUMN = i
                    break
            for i in range(1, data_file_ws.max_column):
                if data_file_ws.cell(1, i).value == '학생 평균':
                    SCORE_COLUMN = i
                    break
            
            for i in range(2, dataFileColorWs.max_row+1):
                if dataFileColorWs.cell(i, NAME_COLUMN).value is None:
                    break
                for j in range(1, dataFileColorWs.max_column+1):
                    data_file_ws.column_dimensions[get_column_letter(j)].width = 14
                    if data_file_ws.cell(i, NAME_COLUMN).value == '시험 평균' and data_file_ws.cell(i, j).value is not None:
                        data_file_ws.cell(i, j).border = Border(bottom=Side(border_style='medium', color='000000'))
                    if j > SCORE_COLUMN:    
                        if data_file_ws.cell(i, NAME_COLUMN).value == '날짜' and data_file_ws.cell(i, j).value is not None:
                            data_file_ws.cell(i, j).border = Border(top=Side(border_style='medium', color='000000'))
                        if type(dataFileColorWs.cell(i, j).value) == int:
                            if dataFileColorWs.cell(i, j).value < 60:
                                data_file_ws.cell(i, j).fill = PatternFill(fill_type='solid', fgColor=Color('EC7E31'))
                            elif dataFileColorWs.cell(i, j).value < 70:
                                data_file_ws.cell(i, j).fill = PatternFill(fill_type='solid', fgColor=Color('F5AF85'))
                            elif dataFileColorWs.cell(i, j).value < 80:
                                data_file_ws.cell(i, j).fill = PatternFill(fill_type='solid', fgColor=Color('FCE4D6'))
                            elif dataFileColorWs.cell(i, NAME_COLUMN).value == '시험 평균':
                                data_file_ws.cell(i, j).fill = PatternFill(fill_type='solid', fgColor=Color('DDEBF7'))
                            else:
                                data_file_ws.cell(i, j).fill = PatternFill(fill_type=None, fgColor=Color('00FFFFFF'))
                        else:
                            data_file_ws.cell(i, j).fill = PatternFill(fill_type=None, fgColor=Color('00FFFFFF'))
                        if dataFileColorWs.cell(i, NAME_COLUMN).value == '시험 평균':
                            data_file_ws.cell(i, j).font = Font(bold=True)

                # 학생별 평균 조건부 서식
                data_file_ws.cell(i, SCORE_COLUMN).font = Font(bold=True)
                if type(dataFileColorWs.cell(i, SCORE_COLUMN).value) == int:
                    if dataFileColorWs.cell(i, SCORE_COLUMN).value < 60:
                        data_file_ws.cell(i, SCORE_COLUMN).fill = PatternFill(fill_type='solid', fgColor=Color('EC7E31'))
                    elif dataFileColorWs.cell(i, SCORE_COLUMN).value < 70:
                        data_file_ws.cell(i, SCORE_COLUMN).fill = PatternFill(fill_type='solid', fgColor=Color('F5AF85'))
                    elif dataFileColorWs.cell(i, SCORE_COLUMN).value < 80:
                        data_file_ws.cell(i, SCORE_COLUMN).fill = PatternFill(fill_type='solid', fgColor=Color('FCE4D6'))
                    elif dataFileColorWs.cell(i, NAME_COLUMN).value == '시험 평균':
                        data_file_ws.cell(i, SCORE_COLUMN).fill = PatternFill(fill_type='solid', fgColor=Color('DDEBF7'))
                    else:
                        data_file_ws.cell(i, SCORE_COLUMN).fill = PatternFill(fill_type='solid', fgColor=Color('E2EFDA'))
                name = data_file_ws.cell(i, NAME_COLUMN)
                if (name.value is not None) or (name.value != '날짜') or (name.value != '시험명') or (name.value != '시험 평균'):
                    for j in range(2, student_ws.max_row+1):
                        if (name.value == student_ws.cell(j, 1).value) and (student_ws.cell(j, 6).value == 'N'):
                            name.fill = PatternFill(fill_type='solid', fgColor=Color('FFFF00'))
                            break
                        else:
                            name.fill = PatternFill(fill_type=None, fgColor=Color('00FFFFFF'))

    except:
        gui.append_log('이 데이터 양식에는 조건부 서식을 지정할 수 없습니다.')
        gui.apply_color_button['state'] = tk.NORMAL
        return
    
    try:
        data_file_wb.save('./data/' + config['dataFileName'] + '.xlsx')
        gui.append_log('조건부 서식 지정을 완료했습니다.')
        gui.ui.wm_attributes("-topmost", 1)
        gui.ui.wm_attributes("-topmost", 0)
        gui.apply_color_button['state'] = tk.NORMAL
        excel.Visible = True
        wb = excel.Workbooks.Open(os.getcwd() + '\\data\\' + config['dataFileName'] + '.xlsx')
        gui.ui.wm_attributes("-topmost", 1)
        gui.ui.wm_attributes("-topmost", 0)
    except:
        gui.append_log('데이터 파일 창을 끄고 다시 실행해 주세요.')
        gui.apply_color_button['state'] = tk.NORMAL
        return

def check_data_form(gui:GUI, form_ws:Worksheet) -> bool:
    gui.append_log('양식이 올바른지 확인 중...')
    # 올바른 양식이 아닙니다.
    if (form_ws.title != '데일리테스트 기록 양식') or \
        (form_ws[get_column_letter(DataForm.DATE_COLUMN)+'1'].value != '요일') or \
            (form_ws[get_column_letter(DataForm.TEST_TIME_COLUMN)+'1'].value != '시간') or \
                (form_ws[get_column_letter(DataForm.CLASS_NAME_COLUMN)+'1'].value != '반') or \
                    (form_ws[get_column_letter(DataForm.STUDENT_NAME_COLUMN)+'1'].value != '이름') or \
                        (form_ws[get_column_letter(DataForm.TEACHER_COLUMN)+'1'].value != '담당T') or \
                            (form_ws[get_column_letter(DataForm.DAILYTEST_TEST_NAME_COLUMN)+'1'].value != '시험명') or \
                                (form_ws[get_column_letter(DataForm.DAILYTEST_SCORE_COLUMN)+'1'].value != '점수') or \
                                    (form_ws[get_column_letter(DataForm.DAILYTEST_AVERAGE_COLUMN)+'1'].value != '평균') or \
                                        (form_ws[get_column_letter(DataForm.MOCKTEST_TEST_NAME_COLUMN)+'1'].value != '시험대비 모의고사명') or \
                                            (form_ws[get_column_letter(DataForm.MOCKTEST_SCORE_COLUMN)+'1'].value != '모의고사 점수') or \
                                                (form_ws[get_column_letter(DataForm.MOCKTEST_AVERAGE_COLUMN)+'1'].value != '모의고사 평균') or \
                                                    (form_ws[get_column_letter(DataForm.MAKEUP_TEST_CHECK_COLUMN)+'1'].value != '재시문자 X'):
        gui.append_log('올바른 기록 양식이 아닙니다.')
        return False
    
    form_checked = True
    dailytest_checked = False
    mocktest_checked = False
    for i in range(1, form_ws.max_row+1):
        if form_ws.cell(i, DataForm.CLASS_NAME_COLUMN).value is not None:
            class_name = form_ws.cell(i, DataForm.CLASS_NAME_COLUMN).value
            dailytest_checked = False
            mocktest_checked = False
            dailytest_name = form_ws.cell(i, DataForm.DAILYTEST_TEST_NAME_COLUMN).value
            mocktest_name = form_ws.cell(i, DataForm.MOCKTEST_TEST_NAME_COLUMN).value
        
        if dailytest_checked and mocktest_checked: continue
        
        if not dailytest_checked and form_ws.cell(i, DataForm.DAILYTEST_SCORE_COLUMN).value is not None and dailytest_name is None:
            gui.append_log(f'{class_name}의 시험명이 작성되지 않았습니다.')
            dailytest_checked = True
            form_checked = False
        if not mocktest_checked and form_ws.cell(i, DataForm.MOCKTEST_SCORE_COLUMN).value is not None and mocktest_name is None:
            gui.append_log(f'{class_name}의 모의고사명이 작성되지 않았습니다.')
            mocktest_checked = True
            form_checked = False
    
    if not form_checked: return False

    return True

def holiday_dialog(gui:GUI, button:tk.Button) -> dict:
    def quitEvent():
        window.destroy()
        button['state'] = tk.NORMAL
        gui.ui.wm_attributes("-disabled", False)
        gui.ui.wm_attributes("-topmost", 1)
        gui.ui.wm_attributes("-topmost", 0)
    button['state'] = tk.DISABLED
    gui.ui.wm_attributes("-disabled", True)
    window=tk.Tk()
    width = 200
    height = 300
    x = int((window.winfo_screenwidth()/4) - (width/2))
    y = int((window.winfo_screenheight()/2) - (height/2))
    window.geometry(f'{width}x{height}+{x}+{y}')
    window.title('휴일 선택')
    window.resizable(False, False)
    window.protocol("WM_DELETE_WINDOW", quitEvent)

    today = DATE.today()

    makeup_test_date={}
    if today == today + relativedelta(weekday=calendar.MONDAY):
        makeup_test_date['월'] = today + timedelta(days=7)
    else:
        makeup_test_date['월'] = today + relativedelta(weekday=calendar.MONDAY)

    if today == today + relativedelta(weekday=calendar.TUESDAY):
        makeup_test_date['화'] = today + timedelta(days=7)
    else:
        makeup_test_date['화'] = today + relativedelta(weekday=calendar.TUESDAY)

    if today == today + relativedelta(weekday=calendar.WEDNESDAY):
        makeup_test_date['수'] = today + timedelta(days=7)
    else:
        makeup_test_date['수'] = today + relativedelta(weekday=calendar.WEDNESDAY)

    if today == today + relativedelta(weekday=calendar.THURSDAY):
        makeup_test_date['목'] = today + timedelta(days=7)
    else:
        makeup_test_date['목'] = today + relativedelta(weekday=calendar.THURSDAY)

    if today == today + relativedelta(weekday=calendar.FRIDAY):
        makeup_test_date['금'] = today + timedelta(days=7)
    else:
        makeup_test_date['금'] = today + relativedelta(weekday=calendar.FRIDAY)

    if today == today + relativedelta(weekday=calendar.SATURDAY):
        makeup_test_date['토'] = today + timedelta(days=7)
    else:
        makeup_test_date['토'] = today + relativedelta(weekday=calendar.SATURDAY)

    if today == today + relativedelta(weekday=calendar.SUNDAY):
        makeup_test_date['일'] = today + timedelta(days=7)
    else:
        makeup_test_date['일'] = today + relativedelta(weekday=calendar.SUNDAY)

    mon = tk.IntVar()
    tue = tk.IntVar()
    wed = tk.IntVar()
    thu = tk.IntVar()
    fri = tk.IntVar()
    sat = tk.IntVar()
    sun = tk.IntVar()

    tk.Label(window, text='\n다음 중 휴일을 선택해주세요\n').pack()
    date_calculated = DATE.today() + timedelta(days=1)
    for i in range(0, 8):
        for j in range(0, 8):
            if date_calculated == makeup_test_date['월']:
                tk.Checkbutton(window, text=str(makeup_test_date['월'])+' (월)', variable=mon).pack()
                date_calculated += timedelta(days=1)
        for j in range(0, 8):
            if date_calculated == makeup_test_date['화']:
                tk.Checkbutton(window, text=str(makeup_test_date['화'])+' (화)', variable=tue).pack()
                date_calculated += timedelta(days=1)
        for j in range(0, 8):
            if date_calculated == makeup_test_date['수']:
                tk.Checkbutton(window, text=str(makeup_test_date['수'])+' (수)', variable=wed).pack()
                date_calculated += timedelta(days=1)
        for j in range(0, 8):
            if date_calculated == makeup_test_date['목']:
                tk.Checkbutton(window, text=str(makeup_test_date['목'])+' (목)', variable=thu).pack()
                date_calculated += timedelta(days=1)
        for j in range(0, 8):
            if date_calculated == makeup_test_date['금']:
                tk.Checkbutton(window, text=str(makeup_test_date['금'])+' (금)', variable=fri).pack()
                date_calculated += timedelta(days=1)
        for j in range(0, 8):
            if date_calculated == makeup_test_date['토']:
                tk.Checkbutton(window, text=str(makeup_test_date['토'])+' (토)', variable=sat).pack()
                date_calculated += timedelta(days=1)
        for j in range(0, 8):
            if date_calculated == makeup_test_date['일']:
                tk.Checkbutton(window, text=str(makeup_test_date['일'])+' (일)', variable=sun).pack()
                date_calculated += timedelta(days=1)
    tk.Label(window, text='\n').pack()
    tk.Button(window, text="확인", width=10 , command=window.destroy).pack()
    
    window.mainloop()
    if mon.get() == 1:
        makeup_test_date['월'] += timedelta(days=7)
    if tue.get() == 1:
        makeup_test_date['화'] += timedelta(days=7)
    if wed.get() == 1:
        makeup_test_date['수'] += timedelta(days=7)
    if thu.get() == 1:
        makeup_test_date['목'] += timedelta(days=7)
    if fri.get() == 1:
        makeup_test_date['금'] += timedelta(days=7)
    if sat.get() == 1:
        makeup_test_date['토'] += timedelta(days=7)
    if sun.get() == 1:
        makeup_test_date['일'] += timedelta(days=7)
    gui.ui.wm_attributes("-disabled", False)

    return makeup_test_date

def update_class(gui:GUI):
    gui.update_class_button['state'] = tk.DISABLED
    # 반 정보 확인
    if not os.path.isfile('./반 정보.xlsx'):
        gui.append_log('[오류] 반 정보.xlsx 파일이 존재하지 않습니다.')
        return
    class_wb = xl.load_workbook("./반 정보.xlsx")
    try:
        class_ws = class_wb['반 정보']
    except:
        gui.append_log('[오류] \'반 정보.xlsx\'의 시트명을')
        gui.append_log('\'반 정보\'로 변경해 주세요.')
        gui.update_class_button['state'] = tk.NORMAL
        return

    # 데이터 저장 엑셀
    if not os.path.isfile('./data/' + config['dataFileName'] + '.xlsx'):
        gui.append_log('[오류] ' + config['dataFileName'] + '.xlsx' + '파일이 존재하지 않습니다.')
        gui.updateDataButton['state'] = tk.NORMAL
        return

    data_file_wb = xl.load_workbook('./data/' + config['dataFileName'] + '.xlsx')
    data_file_ws = data_file_wb['데일리테스트']

    options = webdriver.ChromeOptions()
    options.add_argument('headless')
    driver = webdriver.Chrome(service = service, options = options)

    gui.append_log('아이소식으로부터 반 정보를 업데이트 하는 중...')
    # 아이소식 접속
    driver.get(config['url'])
    table_names = driver.find_elements(By.CLASS_NAME, 'style1')

    # 반 루프
    unregistered = {}
    for i in range(3, len(table_names)):
        is_class_exist = False
        for j in range(1, class_ws.max_row+1):
            if class_ws.cell(j, 1).value == table_names[i].text.rstrip():
                is_class_exist = True
                break
        if is_class_exist: continue
        unregistered[table_names[i].text.rstrip()] = i
        
    if len(unregistered) == 0:
        gui.append_log('업데이트된 사항이 없습니다.')
        gui.update_class_button['state'] = tk.NORMAL
        return

    for new_class, new_class_index in unregistered.items():
        gui.append_log(str(new_class))
    # trs = driver.find_element(By.ID, 'table_' + str(i)).find_elements(By.CLASS_NAME, 'style12')
    # write_location = start = ini_ws.max_row + 1
    # ini_ws.cell(write_location, 1).value = table_names[i].text.rstrip()

    gui.update_class_button['state'] = tk.NORMAL
    return

ui = tk.Tk()
gui = GUI(ui)
ui.mainloop()
