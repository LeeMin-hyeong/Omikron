import json
import os.path
import openpyxl as xl

from datetime import datetime

config = json.load(open('config.json'))

formWb = xl.load_workbook("test.xlsx", data_only=True)
formWs = formWb.active

for i in range(2, formWs.max_row+1):
    name = formWs.cell(i, 4).value
    dailyTestScore = formWs.cell(i, 7).value
    mockTestScore = formWs.cell(i, 10).value
    if formWs.cell(i, 3).value is not None:
        className = formWs.cell(i, 3).value
        teacher = formWs.cell(i, 5).value
        dailyTestName = formWs.cell(i, 6).value
        mockTestName = formWs.cell(i, 9).value
        dailyTestAverage = formWs.cell(i, 8).value
        mockTestAverage = formWs.cell(i, 11).value

    # 시험 미응시시 건너뛰기
    if dailyTestScore is not None:
        testName = dailyTestName
        score = dailyTestScore
        average = dailyTestAverage
    elif mockTestScore is not None:
        testName = mockTestName
        score = mockTestScore
        average = mockTestAverage
    else:
        continue
    
    # 해당 학생 엑셀 파일이 존재하지 않으면 생성
    if not os.path.isfile(config['dailyTestPersonalFilePath'] + str(formWs.cell(i, 4).value) + '.xlsx'):
        iniWb = xl.Workbook()
        iniWs = iniWb.active
        iniWs['A1'] = '응시일'
        iniWs['B1'] = '반'
        iniWs['C1'] = '담당T'
        iniWs['D1'] = '시험명'
        iniWs['E1'] = '점수'
        iniWs['F1'] = '반평균'
        iniWb.save(config['dailyTestPersonalFilePath'] + str(formWs.cell(i, 4).value) + '.xlsx')

    # 해당 학생 파일에 응시 결과 입력
    studentWb = xl.load_workbook(config['dailyTestPersonalFilePath'] + str(formWs.cell(i, 4).value) + '.xlsx', data_only=True)
    studentWs = studentWb.active
    writeLocation = studentWs.max_row + 1

    #중복방지
    if str(studentWs.cell(writeLocation-1, 4).value) == testName: continue

    studentWs.cell(writeLocation, 1).value = datetime.today().strftime('%Y.%m.%d')
    studentWs.cell(writeLocation, 2).value = className
    studentWs.cell(writeLocation, 3).value = teacher
    studentWs.cell(writeLocation, 4).value = testName
    studentWs.cell(writeLocation, 5).value = score
    studentWs.cell(writeLocation, 6).value = average
    studentWb.save(config['dailyTestPersonalFilePath'] + str(formWs.cell(i, 3).value) + '.xlsx')