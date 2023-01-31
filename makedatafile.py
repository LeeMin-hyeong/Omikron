import json
import os.path
import openpyxl as xl

from openpyxl.styles import Alignment, Border, Side

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service

from webdriver_manager.chrome import ChromeDriverManager

config = json.load(open('config.json', encoding='UTF8'))

if True:
# if not os.path.isfile('./data/'+config['dailyTestFileName']):
    print('Making Data File...')

    iniWb = xl.Workbook()
    iniWs = iniWb.active
    iniWs.title = 'DailyTest'
    iniWs['A1'] = '시간'
    iniWs['B1'] = '요일'
    iniWs['C1'] = '반'
    iniWs['D1'] = '담당'
    iniWs['E1'] = '이름'
    iniWs['F1'] = '학생 평균'
    iniWs.freeze_panes = 'G2'
    iniWs.auto_filter.ref = 'A:E'

    # 반 정보 확인
    classWb = xl.load_workbook("class.xlsx")
    classWs = classWb.active

    options = webdriver.ChromeOptions()
    options.add_argument('headless')
    driver = webdriver.Chrome(service = Service(ChromeDriverManager().install()), options = options)

    # 아이소식 접속
    driver.get(config['url'])
    tableNames = driver.find_elements(By.CLASS_NAME, 'style1')

    # 반 루프
    for i in range(3, len(tableNames)):
        trs = driver.find_element(By.ID, 'table_' + str(i)).find_elements(By.CLASS_NAME, 'style12')
        writeLocation = iniWs.max_row + 1

        className = tableNames[i].text.split('(')[0].rstrip()
        for j in range(2, classWs.max_row + 1):
            if classWs.cell(j, 1).value == className:
                teacher = classWs.cell(j, 2).value
                date = classWs.cell(j, 3).value
                time = classWs.cell(j, 4).value
        
        # 시험명
        iniWs.cell(writeLocation, 2).value = date
        iniWs.cell(writeLocation, 3).value = className
        iniWs.cell(writeLocation, 4).value = teacher
        iniWs.cell(writeLocation, 5).value = '시험명'

        # 학생 루프
        for tr in trs:
            writeLocation = iniWs.max_row + 1
            iniWs.cell(writeLocation, 2).value = date
            iniWs.cell(writeLocation, 3).value = className
            iniWs.cell(writeLocation, 4).value = teacher
            iniWs.cell(writeLocation, 5).value = tr.find_element(By.CLASS_NAME, 'style9').text
            iniWs.cell(writeLocation, 6).value = '=ROUND(AVERAGE(G' + str(writeLocation) + ':XFD' + str(writeLocation) + '), 0)'
        
        # 시험별 평균
        writeLocation = iniWs.max_row + 1
        iniWs.cell(writeLocation, 2).value = date
        iniWs.cell(writeLocation, 3).value = className
        iniWs.cell(writeLocation, 4).value = teacher
        iniWs.cell(writeLocation, 5).value = '시험 평균'
        iniWs.cell(writeLocation, 6).value = '=ROUND(AVERAGE(G' + str(writeLocation) + ':XFD' + str(writeLocation) + '), 0)'

        for j in range(1, 7):
            iniWs.cell(writeLocation, j).border = Border(bottom = Side(border_style='medium', color='000000'))

    # 정렬
    for j in range(1, iniWs.max_row + 1):
            for k in range(1, iniWs.max_column + 1):
                iniWs.cell(j, k).alignment = Alignment(horizontal='center', vertical='center')
    
    # 모의고사 sheet 생성
    copyWs = iniWb.copy_worksheet(iniWb['DailyTest'])
    copyWs.title = '모의고사'
    copyWs.freeze_panes = 'G2'
    copyWs.auto_filter.ref = 'A:E'

    iniWb.save('./data/'+config['dailyTestFileName'])
    print('Done')

else:
    print('이미 파일이 존재합니다')