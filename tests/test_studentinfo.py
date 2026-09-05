import unittest
from io import BytesIO
from unittest.mock import patch

import openpyxl
from openpyxl.styles import PatternFill

import tdm.studentinfo as studentinfo


class StudentInfoSparseRowsTests(unittest.TestCase):
    def make_workbook(self, last_row=1_048_576):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "학생 정보"
        ws.append(["이름", "재시험 응시 요일", "재시험 응시 시간", "기수 신규생"])
        ws.append(["Alpha", "월", 18, "N"])
        ws.cell(55, 1, "Hidden")
        ws.row_dimensions[55].hidden = True
        ws.cell(last_row, 10).fill = PatternFill(fill_type="solid", fgColor="FFFF00")
        self.addCleanup(wb.close)
        return wb

    def test_lookup_does_not_materialize_empty_rows(self):
        wb = self.make_workbook()
        ws = wb.active
        ws.cell(800_000, 1, "AfterGap")
        count = len(ws._cells)
        self.assertEqual(studentinfo.get_student_info(ws, "Alpha"), (True, "월", 18, True))
        self.assertTrue(studentinfo.get_student_info(ws, "Hidden")[0])
        self.assertTrue(studentinfo.get_student_info(ws, "AfterGap")[0])
        self.assertEqual(studentinfo.get_student_info(ws, "Missing"), (False, None, None, False))
        self.assertLessEqual(len(ws._cells), count + 6)

    def test_add_uses_last_student_row(self):
        wb = self.make_workbook()
        with patch.object(studentinfo, "open", return_value=wb), patch.object(studentinfo, "save"):
            studentinfo.add_student("New")
        self.assertEqual(wb.active["A56"].value, "New")
        self.assertEqual(wb.active.max_row, 1_048_576)
        self.assertLess(len(wb.active._cells), 30)

    def test_update_preserves_blank_and_formatted_rows(self):
        wb = self.make_workbook()
        with patch.object(studentinfo.tdm.chrome, "get_student_names", return_value=["Hidden", "New"]), patch.object(studentinfo, "save"):
            studentinfo.update_student(wb)
        ws = wb.active
        self.assertEqual(studentinfo._student_rows(ws), [(54, "Hidden"), (55, "New")])
        self.assertEqual(ws.cell(1_048_575, 10).fill.fgColor.rgb, "00FFFF00")
        self.assertLess(len(ws._cells), 30)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        restored = openpyxl.load_workbook(output)
        self.addCleanup(restored.close)
        self.assertEqual(studentinfo._student_rows(restored.active), [(54, "Hidden"), (55, "New")])

    def test_delete_removes_adjacent_duplicates_without_dense_cells(self):
        wb = self.make_workbook()
        wb.active["A3"] = "Alpha"
        with patch.object(studentinfo, "open", return_value=wb), patch.object(studentinfo, "save"):
            studentinfo.delete_student("Alpha")
        self.assertEqual(studentinfo._student_rows(wb.active), [(53, "Hidden")])
        self.assertLess(len(wb.active._cells), 30)

    def test_sparse_delete_matches_openpyxl_on_small_sheet(self):
        expected = self.make_workbook(last_row=60)
        actual = self.make_workbook(last_row=60)
        for wb in (expected, actual):
            wb.active["C55"] = "=C2+1"
        for row in (55, 2):
            expected.active.delete_rows(row)
        studentinfo._delete_student_rows(actual.active, [2, 55])

        def cells(ws):
            return {key: (cell.value, cell.style_id) for key, cell in ws._cells.items()
                    if cell.value is not None or cell.has_style}

        self.assertEqual(cells(actual.active), cells(expected.active))


if __name__ == "__main__":
    unittest.main()
