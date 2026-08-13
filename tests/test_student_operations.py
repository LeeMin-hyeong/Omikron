from __future__ import annotations

import openpyxl as xl

import tdm.excel.data_file as data_file
from tdm.domain.models import DataFile


def _student_rows(worksheet, student_name: str) -> dict[str, int]:
    return {
        worksheet.cell(row, 1).value: row
        for row in range(2, worksheet.max_row + 1)
        if worksheet.cell(row, 3).value == student_name
    }


def test_delete_only_affects_same_named_student_in_selected_class(
    make_data_workbook,
):
    paths = make_data_workbook({"A반": ["민수"], "B반": ["민수"]})

    data_file.delete_student("A반", "민수")

    workbook = xl.load_workbook(paths.data_file, data_only=False)
    worksheet = workbook[DataFile.DEFAULT_SHEET_NAME]
    rows = _student_rows(worksheet, "민수")
    assert worksheet.cell(rows["A반"], 3).font.strike is True
    assert not worksheet.cell(rows["B반"], 3).font.strike
    assert worksheet.cell(rows["A반"], 4).value is None
    assert worksheet.cell(rows["B반"], 4).value is not None
    workbook.close()

    assert data_file.check_student_exists("민수") is True


def test_move_marks_old_class_and_adds_active_student_to_target(
    make_data_workbook,
):
    paths = make_data_workbook({"A반": ["민수"], "B반": ["가영"]})

    warnings = data_file.move_student("민수", "B반", "A반")

    assert warnings == []
    workbook = xl.load_workbook(paths.data_file, data_only=False)
    worksheet = workbook[DataFile.DEFAULT_SHEET_NAME]
    rows = _student_rows(worksheet, "민수")
    assert worksheet.cell(rows["A반"], 3).font.color.rgb == "FFFF0000"
    assert worksheet.cell(rows["B반"], 3).font.color.type == "theme"
    assert "AVERAGE" in str(worksheet.cell(rows["B반"], 4).value)
    workbook.close()
