import unittest
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

import openpyxl
from openpyxl.styles import PatternFill

import tdm.classinfo as classinfo
from tdm.sparse_worksheet import named_rows


class TemporaryClassInfoTests(unittest.TestCase):
    def make_workbook(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "반 정보"
        ws.append(["반명", "선생님명", "요일", "시간", "모의고사 응시여부"])
        ws.append(["Remove", "Teacher1", "월", 18, "Y"])
        ws["A55"] = "Keep"
        ws["B55"] = "Teacher2"
        ws["E55"] = "Y"
        ws["A55"].fill = PatternFill(fill_type="solid", fgColor="00FF00")
        ws.auto_filter.ref = "A1:E1048576"
        ws.row_dimensions[55].hidden = True
        ws["J1048576"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
        self.addCleanup(wb.close)
        return wb

    def create_temp(self, wb, selected):
        source = BytesIO()
        wb.save(source)
        source.seek(0)
        source_wb = openpyxl.load_workbook(source, read_only=True, data_only=False)
        self.addCleanup(source_wb.close)
        output = BytesIO()
        with patch.object(classinfo, "make_backup_file") as backup, \
             patch.object(classinfo, "open", return_value=source_wb) as source_open, \
             patch.object(classinfo, "save_to_temp", side_effect=lambda book: book.save(output)), \
             patch.object(source_wb, "close", wraps=source_wb.close) as close:
            classinfo.make_temp_file_for_update(selected)
            backup.assert_called_once()
            source_open.assert_called_once_with(data_only=False, read_only=True)
            close.assert_called_once()
        self.assertLess(len(output.getvalue()), 50_000)
        output.seek(0)
        restored = openpyxl.load_workbook(output)
        self.addCleanup(restored.close)
        self.assertLess(len(restored.active._cells), 1000)
        self.assertEqual(restored.active.max_row, len(named_rows(restored.active, 1)) + 1)
        self.assertFalse(any(dim.hidden for dim in restored.active.row_dimensions.values()))
        return restored.active

    def test_append_after_last_remaining_class_with_inflated_dimension(self):
        ws = self.create_temp(self.make_workbook(), ["Keep", "Zulu", "Alpha"])
        self.assertEqual(named_rows(ws, 1), [(2, "Keep"), (3, "Alpha"), (4, "Zulu")])
        self.assertEqual(ws["B2"].value, "Teacher2")
        self.assertEqual(ws["E2"].value, "Y")
        self.assertEqual(ws["A2"].fill.fgColor.rgb, "0000FF00")
        self.assertEqual(ws["A3"].border, classinfo.BORDER_ALL)
        self.assertEqual(ws.auto_filter.ref, "A1:E4")
        self.assertEqual(ws["Z1"].value, "Y")
        self.assertTrue(ws.column_dimensions["Z"].hidden)
        self.assertEqual(str(ws.data_validations.dataValidation[0].sqref), "E2:E4")

    def test_no_new_classes(self):
        ws = self.create_temp(self.make_workbook(), ["Keep"])
        self.assertEqual(named_rows(ws, 1), [(2, "Keep")])

    def test_replace_all_classes_starts_at_row_two(self):
        ws = self.create_temp(self.make_workbook(), ["New"])
        self.assertEqual(named_rows(ws, 1), [(2, "New")])

    def test_remove_all_classes(self):
        ws = self.create_temp(self.make_workbook(), [])
        self.assertEqual(named_rows(ws, 1), [])
        self.assertEqual(ws["A1"].value, "반명")

    def test_more_than_one_hundred_classes_are_not_truncated(self):
        selected = [f"Class{i:03}" for i in range(120)]
        ws = self.create_temp(self.make_workbook(), selected)
        self.assertEqual([name for _, name in named_rows(ws, 1)], selected)

    def test_backup_copies_original_bytes_without_loading_workbook(self):
        with TemporaryDirectory() as directory:
            original = Path(directory) / "반 정보.xlsx"
            self.make_workbook().save(original)
            original_bytes = original.read_bytes()
            with patch.object(classinfo.tdm.config, "DATA_DIR", directory), \
                 patch.object(classinfo, "open", side_effect=AssertionError("Must not parse backup")):
                classinfo.make_backup_file()
            backups = list((Path(directory) / "data" / "backup").glob("*.xlsx"))
            self.assertEqual(len(backups), 1)
            self.assertEqual(backups[0].read_bytes(), original_bytes)
            self.assertEqual(original.read_bytes(), original_bytes)


if __name__ == "__main__":
    unittest.main()
