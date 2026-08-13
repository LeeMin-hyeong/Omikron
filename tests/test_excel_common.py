from __future__ import annotations

from datetime import datetime, timedelta
import os
import time

import openpyxl as xl
import pytest

import tdm.excel.class_info as class_info
from tdm.excel.atomic import atomic_save_workbook
from tdm.excel.backup import BackupRetentionPolicy, create_backup
from tdm.excel.data_file_errors import NoReservedColumnError
from tdm.excel.errors import ConcurrentWorkbookChangeError
from tdm.excel.storage_health import verify_directory_writable
from tdm.excel.transaction import cleanup_stale_staging_files
from tdm.excel.validation import find_required_columns


def test_required_column_validation_reports_missing_header():
    workbook = xl.Workbook()
    worksheet = workbook.active
    worksheet.append(["반", "이름"])

    with pytest.raises(NoReservedColumnError, match="학생 평균"):
        find_required_columns(worksheet, ("반", "이름", "학생 평균"))


def test_backup_retention_keeps_latest_count(workbook_paths):
    source = workbook_paths.data_file
    workbook = xl.Workbook()
    workbook.active["A1"] = "backup"
    atomic_save_workbook(workbook, source)

    policy = BackupRetentionPolicy(max_count=2, max_age_days=365)
    now = datetime.now()
    for days in (3, 2, 1):
        create_backup(
            source,
            stem="테스트 데이터",
            backup_dir=workbook_paths.backup_dir,
            policy=policy,
            now=now - timedelta(days=days),
        )

    backups = sorted(workbook_paths.backup_dir.glob("테스트 데이터(*).xlsx"))
    assert len(backups) == 2
    assert all("202" in path.name for path in backups)


def test_storage_probe_leaves_no_files(tmp_path):
    verify_directory_writable(tmp_path)
    assert list(tmp_path.iterdir()) == []


def test_stale_staging_cleanup_keeps_recent_file(tmp_path):
    stale = tmp_path / ".old.abc.tmp.xlsx"
    recent = tmp_path / ".new.abc.tmp.xlsx"
    stale.write_bytes(b"old")
    recent.write_bytes(b"new")
    old_time = time.time() - 48 * 60 * 60
    os.utime(stale, (old_time, old_time))

    removed = cleanup_stale_staging_files(tmp_path)

    assert removed == [stale]
    assert not stale.exists()
    assert recent.exists()


def test_class_update_draft_detects_original_changed_after_preparation(
    workbook_paths,
):
    original = xl.Workbook()
    original.active.title = "반 정보"
    original.active["A1"] = "반명"
    atomic_save_workbook(original, workbook_paths.class_info)

    draft = class_info.open(read_only=False)
    class_info.save_to_temp(draft)

    external = xl.Workbook()
    external.active.title = "반 정보"
    external.active["A1"] = "외부 변경"
    atomic_save_workbook(external, workbook_paths.class_info)

    prepared = class_info.open_temp(read_only=False)
    prepared.active["B1"] = "내 변경"
    with pytest.raises(ConcurrentWorkbookChangeError):
        class_info.save(prepared)
