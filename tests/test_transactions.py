from __future__ import annotations

from unittest.mock import patch

import openpyxl as xl
import pytest

from tdm.excel.atomic import atomic_save_workbook, commit_prepared_workbook
from tdm.excel.errors import ConcurrentWorkbookChangeError
from tdm.excel.transaction import (
    WorkbookSave,
    recover_pending_transactions,
    save_workbooks_transaction,
)
from tdm.excel.workbook_io import load_workbook


def _save_value(path, value: str) -> None:
    workbook = xl.Workbook()
    workbook.active["A1"] = value
    atomic_save_workbook(workbook, path)
    workbook.close()


def _read_value(path) -> str:
    workbook = xl.load_workbook(path)
    try:
        return workbook.active["A1"].value
    finally:
        workbook.close()


def test_two_workbook_transaction_commits_together(workbook_paths):
    first = workbook_paths.data_file
    second = workbook_paths.makeup_test
    _save_value(first, "old-a")
    _save_value(second, "old-b")

    first_workbook = load_workbook(first)
    second_workbook = load_workbook(second)
    first_workbook.active["A1"] = "new-a"
    second_workbook.active["A1"] = "new-b"

    save_workbooks_transaction(
        [WorkbookSave(first_workbook, first), WorkbookSave(second_workbook, second)],
        operation="transaction test",
        paths=workbook_paths,
    )

    assert _read_value(first) == "new-a"
    assert _read_value(second) == "new-b"
    assert not list(workbook_paths.transaction_dir.iterdir())


def test_commit_failure_rolls_back_already_replaced_file(workbook_paths):
    first = workbook_paths.data_file
    second = workbook_paths.makeup_test
    _save_value(first, "old-a")
    _save_value(second, "old-b")

    first_workbook = load_workbook(first)
    second_workbook = load_workbook(second)
    first_workbook.active["A1"] = "new-a"
    second_workbook.active["A1"] = "new-b"

    calls = 0

    def fail_second(prepared):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise OSError("second commit failed")
        return commit_prepared_workbook(prepared)

    with patch(
        "tdm.excel.transaction.commit_prepared_workbook",
        side_effect=fail_second,
    ):
        with pytest.raises(OSError, match="second commit failed"):
            save_workbooks_transaction(
                [
                    WorkbookSave(first_workbook, first),
                    WorkbookSave(second_workbook, second),
                ],
                operation="rollback test",
                paths=workbook_paths,
            )

    assert _read_value(first) == "old-a"
    assert _read_value(second) == "old-b"
    assert not list(workbook_paths.transaction_dir.iterdir())


def test_recovery_rolls_back_interrupted_partial_commit(workbook_paths):
    first = workbook_paths.data_file
    second = workbook_paths.makeup_test
    _save_value(first, "old-a")
    _save_value(second, "old-b")

    first_workbook = load_workbook(first)
    second_workbook = load_workbook(second)
    first_workbook.active["A1"] = "new-a"
    second_workbook.active["A1"] = "new-b"

    calls = 0

    def simulate_process_exit(prepared):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise KeyboardInterrupt()
        return commit_prepared_workbook(prepared)

    with patch(
        "tdm.excel.transaction.commit_prepared_workbook",
        side_effect=simulate_process_exit,
    ):
        with pytest.raises(KeyboardInterrupt):
            save_workbooks_transaction(
                [
                    WorkbookSave(first_workbook, first),
                    WorkbookSave(second_workbook, second),
                ],
                operation="crash recovery test",
                paths=workbook_paths,
            )

    assert _read_value(first) == "new-a"
    assert _read_value(second) == "old-b"

    results = recover_pending_transactions(workbook_paths)

    assert [result.action for result in results] == ["rolled_back"]
    assert _read_value(first) == "old-a"
    assert _read_value(second) == "old-b"
    assert not list(workbook_paths.transaction_dir.iterdir())


def test_transaction_rejects_external_change_without_overwriting_it(workbook_paths):
    first = workbook_paths.data_file
    second = workbook_paths.makeup_test
    _save_value(first, "old-a")
    _save_value(second, "old-b")

    first_workbook = load_workbook(first)
    second_workbook = load_workbook(second)
    first_workbook.active["A1"] = "mine-a"
    second_workbook.active["A1"] = "mine-b"

    _save_value(first, "other-pc")

    with pytest.raises(ConcurrentWorkbookChangeError):
        save_workbooks_transaction(
            [
                WorkbookSave(first_workbook, first),
                WorkbookSave(second_workbook, second),
            ],
            operation="conflict test",
            paths=workbook_paths,
        )

    assert _read_value(first) == "other-pc"
    assert _read_value(second) == "old-b"
    assert not list(workbook_paths.transaction_dir.iterdir())
