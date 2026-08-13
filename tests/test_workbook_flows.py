from __future__ import annotations

from datetime import datetime

import openpyxl as xl

import tdm.excel.class_info as class_info
import tdm.excel.data_file as data_file
import tdm.excel.data_file_operations as data_file_operations
import tdm.excel.makeup_test as makeup_test
from tdm.domain.models import ClassInfo, DataFile, DataForm, StudentInfo
from tdm.excel.atomic import atomic_save_workbook


class ProgressStub:
    def __init__(self) -> None:
        self.messages: list[str] = []

    def step(self, message: str) -> None:
        self.messages.append(message)

    def phase(self, current: int, total: int, message: str) -> None:
        self.messages.append(message)

    def warning(self, message: str) -> None:
        self.messages.append(message)


def _save_student_info(paths, students: list[str]) -> None:
    workbook = xl.Workbook()
    worksheet = workbook.active
    worksheet.title = StudentInfo.DEFAULT_NAME
    worksheet.append(["이름", "재시험 응시 요일", "재시험 응시 시간", "기수 신규생"])
    for student in students:
        worksheet.append([student, None, None, "Y"])
    atomic_save_workbook(workbook, paths.student_info)
    workbook.close()


def _save_class_info(paths, classes: list[str]) -> None:
    workbook = xl.Workbook()
    worksheet = workbook.active
    worksheet.title = ClassInfo.DEFAULT_NAME
    worksheet.append(["반명", "선생님명", "요일", "시간", "모의고사 응시여부"])
    for class_name in classes:
        worksheet.append([class_name, "담당자", None, None, None])
    atomic_save_workbook(workbook, paths.class_info)
    workbook.close()


def _save_daily_form(path, *, score: int = 70) -> None:
    workbook = xl.Workbook()
    worksheet = workbook.active
    worksheet.title = DataForm.DEFAULT_NAME
    worksheet.append(
        [
            "요일",
            "시간",
            "반",
            "이름",
            "담당T",
            "시험명",
            "점수",
            "평균",
            "모의고사 시험명",
            "모의고사 점수",
            "모의고사 평균",
            "재시험 응시 여부",
        ]
    )
    worksheet.append(
        [None, None, "A반", "민수", "담당자", "단원평가", score, score]
    )
    atomic_save_workbook(workbook, path)
    workbook.close()


def test_exam_result_flow_updates_score_and_preserves_valid_workbook(
    workbook_paths,
    make_data_workbook,
    monkeypatch,
):
    make_data_workbook({"A반": ["민수"]})
    _save_student_info(workbook_paths, ["민수"])
    form_path = workbook_paths.root / "입력 양식.xlsx"
    _save_daily_form(form_path, score=75)
    monkeypatch.setattr(data_file_operations, "recalculate_workbook", lambda _: None)

    prepared = data_file.save_test_data(str(form_path), ProgressStub())
    data_file.save(prepared)

    workbook = xl.load_workbook(workbook_paths.data_file, data_only=False)
    worksheet = workbook[DataFile.DEFAULT_SHEET_NAME]
    student_row = next(
        row
        for row in range(2, worksheet.max_row + 1)
        if worksheet.cell(row, 3).value == "민수"
    )
    assert worksheet.cell(student_row, 5).value == 75
    assert worksheet.cell(student_row - 1, 5).value == "단원평가"
    workbook.close()


def test_conditional_formatting_keeps_known_students_and_layout(
    workbook_paths,
    make_data_workbook,
    monkeypatch,
):
    make_data_workbook({"A반": ["민수"]})
    _save_student_info(workbook_paths, ["민수"])
    monkeypatch.setattr(data_file_operations, "recalculate_workbook", lambda _: None)

    warnings = data_file.conditional_formatting()

    assert warnings == []
    workbook = xl.load_workbook(workbook_paths.data_file, data_only=False)
    worksheet = workbook[DataFile.DEFAULT_SHEET_NAME]
    assert worksheet.column_dimensions["E"].width == 14
    assert worksheet.cell(2, 3).value == "날짜"
    assert worksheet.cell(2, 5).border.bottom.style == "medium"
    workbook.close()


def test_makeup_list_does_not_duplicate_same_student_and_class_today(
    workbook_paths,
):
    _save_student_info(workbook_paths, ["민수"])
    form_path = workbook_paths.root / "입력 양식.xlsx"
    _save_daily_form(form_path, score=70)
    progress = ProgressStub()

    first = makeup_test.save_makeup_test_list(str(form_path), {}, progress)
    makeup_test.save(first)
    second = makeup_test.save_makeup_test_list(str(form_path), {}, progress)
    makeup_test.save(second)

    workbook = xl.load_workbook(workbook_paths.makeup_test)
    worksheet = workbook.active
    matching_rows = [
        row
        for row in range(2, worksheet.max_row + 1)
        if worksheet.cell(row, 2).value == "A반"
        and worksheet.cell(row, 4).value == "민수"
        and worksheet.cell(row, 1).value.date() == datetime.today().date()
    ]
    assert matching_rows == [2]
    workbook.close()


def test_class_update_moves_removed_class_and_adds_new_class(
    workbook_paths,
    make_data_workbook,
    monkeypatch,
):
    make_data_workbook({"A반": ["민수"]})
    _save_class_info(workbook_paths, ["A반"])

    source = class_info.open(read_only=False)
    class_info.save_to_temp(source)
    draft = class_info.open_temp(read_only=False)
    worksheet = draft[ClassInfo.DEFAULT_NAME]
    worksheet.cell(2, ClassInfo.CLASS_NAME_COLUMN).value = "B반"
    class_info.save_to_temp(draft)

    monkeypatch.setattr(data_file_operations, "recalculate_workbook", lambda _: None)
    monkeypatch.setattr(
        data_file_operations.tdm.aisosik.reader,
        "get_class_student_dict",
        lambda: {"B반": ["민수"]},
    )

    data_file.update_class(ProgressStub())
    class_info.update_class(ProgressStub())

    data_workbook = xl.load_workbook(workbook_paths.data_file, data_only=False)
    data_classes = {
        data_workbook.active.cell(row, 1).value
        for row in range(2, data_workbook.active.max_row + 1)
    }
    assert "A반" not in data_classes
    assert "B반" in data_classes
    data_workbook.close()

    previous = xl.load_workbook(workbook_paths.previous_data, data_only=False)
    previous_classes = {
        previous.active.cell(row, 1).value
        for row in range(2, previous.active.max_row + 1)
    }
    assert "A반" in previous_classes
    previous.close()
