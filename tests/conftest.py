from __future__ import annotations

from collections.abc import Callable
from datetime import datetime

import openpyxl as xl
import pytest

import tdm.config
from tdm.domain.models import DataFile
from tdm.excel.atomic import atomic_save_workbook
from tdm.excel.paths import WorkbookPaths


@pytest.fixture
def workbook_paths(tmp_path, monkeypatch) -> WorkbookPaths:
    monkeypatch.setattr(tdm.config, "DATA_DIR", str(tmp_path))
    monkeypatch.setattr(tdm.config, "DATA_DIR_VALID", True)
    monkeypatch.setattr(tdm.config, "DATA_FILE_NAME", "테스트 데이터")
    paths = WorkbookPaths.current()
    paths.ensure_directories()
    return paths


@pytest.fixture
def make_data_workbook(
    workbook_paths: WorkbookPaths,
) -> Callable[[dict[str, list[str]]], WorkbookPaths]:
    def factory(classes: dict[str, list[str]]) -> WorkbookPaths:
        workbook = xl.Workbook()
        worksheet = workbook.active
        worksheet.title = DataFile.DEFAULT_SHEET_NAME
        worksheet.append(["반", "담당", "이름", "학생 평균", "시험 점수"])

        for class_name, students in classes.items():
            date_row = worksheet.max_row + 1
            worksheet.append(
                [class_name, "담당자", "날짜", None, datetime(2026, 8, 12)]
            )
            worksheet.append([class_name, "담당자", "시험명", None, "1차 시험"])
            first_student_row = worksheet.max_row + 1
            for student_name in students:
                row = worksheet.max_row + 1
                worksheet.append(
                    [
                        class_name,
                        "담당자",
                        student_name,
                        f"=ROUND(AVERAGE(E{row}:XFD{row}), 0)",
                        70,
                    ]
                )
            last_student_row = worksheet.max_row
            worksheet.append(
                [
                    class_name,
                    "담당자",
                    "시험 평균",
                    f"=ROUND(AVERAGE(D{first_student_row}:D{last_student_row}), 0)",
                    f"=ROUND(AVERAGE(E{first_student_row}:E{last_student_row}), 0)",
                ]
            )
            worksheet.cell(date_row, 5).number_format = "yyyy.mm.dd"

        atomic_save_workbook(workbook, workbook_paths.data_file)
        workbook.close()
        return workbook_paths

    return factory
