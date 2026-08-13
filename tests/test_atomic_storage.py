from __future__ import annotations

import os
from unittest.mock import patch

import openpyxl as xl
import pytest

from tdm.excel.atomic import atomic_save_workbook
from tdm.excel.errors import ConcurrentWorkbookChangeError
from tdm.excel.workbook_io import load_workbook


def _workbook(value: str) -> xl.Workbook:
    workbook = xl.Workbook()
    workbook.active["A1"] = value
    return workbook


def test_replace_failure_preserves_original_and_cleans_staging(tmp_path):
    target = tmp_path / "data.xlsx"
    original = _workbook("old")
    atomic_save_workbook(original, target)
    original_bytes = target.read_bytes()

    replacement = _workbook("new")
    with patch(
        "tdm.excel.atomic.os.replace",
        side_effect=OSError("replace failed"),
    ):
        with pytest.raises(OSError, match="replace failed"):
            atomic_save_workbook(replacement, target)

    assert target.read_bytes() == original_bytes
    assert not list(tmp_path.glob(".*.tmp.xlsx"))


def test_tracked_workbook_rejects_silent_overwrite(tmp_path):
    target = tmp_path / "data.xlsx"
    initial = _workbook("initial")
    atomic_save_workbook(initial, target)

    tracked = load_workbook(target)
    tracked.active["A1"] = "mine"

    external = _workbook("other pc")
    atomic_save_workbook(external, target)

    with pytest.raises(ConcurrentWorkbookChangeError):
        atomic_save_workbook(tracked, target)

    current = xl.load_workbook(target)
    assert current.active["A1"].value == "other pc"
    current.close()
    tracked.close()


def test_busy_replace_is_retried_before_success(tmp_path):
    target = tmp_path / "data.xlsx"
    workbook = _workbook("saved")
    real_replace = os.replace
    attempts = 0

    def flaky_replace(source, destination):
        nonlocal attempts
        attempts += 1
        if attempts < 3:
            raise PermissionError("temporarily busy")
        return real_replace(source, destination)

    with (
        patch("tdm.excel.atomic.os.replace", side_effect=flaky_replace),
        patch("tdm.excel.atomic.time.sleep"),
    ):
        atomic_save_workbook(workbook, target)

    assert attempts == 3
    saved = xl.load_workbook(target)
    assert saved.active["A1"].value == "saved"
    saved.close()
