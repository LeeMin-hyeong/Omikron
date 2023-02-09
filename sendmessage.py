import json
import os.path
import openpyxl as xl

from datetime import datetime
from tkinter import filedialog

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service

from webdriver_manager.chrome import ChromeDriverManager

config = json.load(open('config.json', encoding='UTF8'))
os.environ['WDM_PROGRESS_BAR'] = '0'
service = Service(ChromeDriverManager().install())

def sendMessage():
    filepath = filedialog.askopenfilename(initialdir='./', title='데일리테스트 기록 양식 선택', filetypes=(('Excel files', '*.xlsx'),('all files', '*.*')))
    
    options = webdriver.ChromeOptions()
    options.add_experimental_option("detach", True)
    driver = webdriver.Chrome(service=service, options=options)

    # 아이소식 접속
    driver.get(config['url'])
    driver.find_element(By.XPATH, '//*[@id="ctitle"]').send_keys(config['dailyTest'])
    
    driver.execute_script('window.open('');')
    driver.switch_to.window(driver.window_handles[1])
    driver.get(config['url'])
    driver.find_element(By.XPATH, '//*[@id="ctitle"]').send_keys(config['makeupTest'])

    driver.execute_script('window.open('');')
    driver.switch_to.window(driver.window_handles[2])
    driver.get(config['url'])
    driver.find_element(By.XPATH, '//*[@id="ctitle"]').send_keys(config['makeupTestDate'])

    # 점수기록표 xlsx
    formWb = xl.load_workbook(filepath, data_only=True)
    formWs = formWb.active

    if not os.path.isfile('./makeupTestInfo.xlsx'):
        print('[오류]makeupTestInfo.xlsx 파일이 존재하지 않습니다.')
        return
    makeupWb = xl.load_workbook("./makeupTestInfo.xlsx")
    makeupWs = makeupWb.active

    for i in range(2, formWs.max_row+1):
        driver.switch_to.window(driver.window_handles[0])
        name = formWs.cell(i, 4).value
        dailyTestScore = formWs.cell(i, 7).value
        mockTestScore = formWs.cell(i, 10).value
        if formWs.cell(i, 3).value is not None:
            className = formWs.cell(i, 3).value
            dailyTestName = formWs.cell(i, 6).value
            mockTestName = formWs.cell(i, 9).value
            dailyTestAverage = formWs.cell(i, 8).value
            mockTestAverage = formWs.cell(i, 11).value

        # 시험 미응시시 건너뛰기
        if dailyTestScore is not None:
            testName = dailyTestName
            score = dailyTestScore
            average = dailyTestAverage
        elif mockTestScore is not None:
            testName = mockTestName
            score = mockTestScore
            average = mockTestAverage
        else:
            continue

        tableNames = driver.find_elements(By.CLASS_NAME, 'style1')
        for j in range(len(tableNames)):
            if className in tableNames[j].text:
                index = j
                break

        trs = driver.find_element(By.ID, 'table_' + str(index)).find_elements(By.CLASS_NAME, 'style12')
        for tr in trs:
            if tr.find_element(By.CLASS_NAME, 'style9').text == name:
                tds = tr.find_elements(By.TAG_NAME, 'td')
                tds[0].find_element(By.TAG_NAME, 'input').send_keys(testName)
                tds[1].find_element(By.TAG_NAME, 'input').send_keys(score)
                tds[2].find_element(By.TAG_NAME, 'input').send_keys(average)
                break
        
        if score < 80:
            for j in range(2, makeupWs.max_row+1):
                if makeupWs.cell(j, 1).value == name:
                    date = makeupWs.cell(j, 4).value
                    time = makeupWs.cell(j, 5).value
                    break
            
            if date is None:
                driver.switch_to.window(driver.window_handles[1])
                trs = driver.find_element(By.ID, 'table_' + str(index)).find_elements(By.CLASS_NAME, 'style12')
                for tr in trs:
                    if tr.find_element(By.CLASS_NAME, 'style9').text == name:
                        tds = tr.find_elements(By.TAG_NAME, 'td')
                        tds[0].find_element(By.TAG_NAME, 'input').send_keys(testName)
            else:
                driver.switch_to.window(driver.window_handles[2])
                trs = driver.find_element(By.ID, 'table_' + str(index)).find_elements(By.CLASS_NAME, 'style12')
                for tr in trs:
                    if tr.find_element(By.CLASS_NAME, 'style9').text == name:
                        tds = tr.find_elements(By.TAG_NAME, 'td')
                        tds[0].find_element(By.TAG_NAME, 'input').send_keys(testName)
                        tds[1].find_element(By.TAG_NAME, 'input').send_keys('test')

    print('메시지 입력을 완료했습니다.')
    print('메시지 확인 후 전송해주세요.')

sendMessage()