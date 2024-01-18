import json
import queue
import os.path
import pythoncom # only works in Windows
import threading
import webbrowser
import tkinter as tk
import tkinter.messagebox
import openpyxl as xl
import win32com.client # only works in Windows

from copy import copy
from omikronconst import *
from tkinter import ttk, filedialog
from datetime import date as DATE, datetime, timedelta
from dateutil.relativedelta import relativedelta
from openpyxl.cell import Cell
from openpyxl.utils.cell import get_column_letter as gcl
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Protection, Side
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from win32process import CREATE_NO_WINDOW # only works in Windows
from webdriver_manager.chrome import ChromeDriverManager

config = json.load(open("./config.json", encoding="UTF8"))
os.environ["WDM_PROGRESS_BAR"] = "0"

service = Service(ChromeDriverManager().install())
service.creation_flags = CREATE_NO_WINDOW

if not os.path.exists("./data"):
    os.makedirs("./data")
if not os.path.exists("./data/backup"):
    os.makedirs("./data/backup")

class GUI():
    def __init__(self, ui:tk.Tk):
        self.q = queue.Queue()
        # self.q.put(VERSION)
        self.thread_end_flag = False
        self.ui = ui
        self.width = 320
        self.height = 585 # button +25
        self.x = int((self.ui.winfo_screenwidth()/4) - (self.width/2))
        self.y = int((self.ui.winfo_screenheight()/2) - (self.height/2))
        self.ui.geometry(f"{self.width}x{self.height}+{self.x}+{self.y}")
        self.ui.title(VERSION)
        self.ui.resizable(False, False)

        self.makeup_test_date = None

        tk.Label(self.ui, text="Omikron 데이터 프로그램").pack()
        
        def callback(url:str):
            webbrowser.open_new(url)
        link = tk.Label(self.ui, text="[ 사용법 및 도움말 ]", cursor="hand2")
        link.pack()
        link.bind("<Button-1>", lambda _: callback("https://omikron-db.notion.site/ad673cca64c146d28adb3deaf8c83a0d?pvs=4"))

        self.scroll = tk.Scrollbar(self.ui, orient="vertical")
        self.log = tk.Listbox(self.ui, yscrollcommand=self.scroll.set, width=51, height=5)
        self.scroll.config(command=self.log.yview)
        self.log.pack()
        
        tk.Label(self.ui, text="< 기수 변경 관련 >").pack()

        self.make_class_info_file_button = tk.Button(self.ui, cursor="hand2", text="반 정보 기록 양식 생성", width=40, command=lambda: self.make_class_info_file_thread())
        self.make_class_info_file_button.pack()

        self.make_student_info_file_button = tk.Button(self.ui, cursor="hand2", text="학생 정보 기록 양식 생성", width=40, command=lambda: self.make_student_info_file_thread())
        self.make_student_info_file_button.pack()

        self.make_data_file_button = tk.Button(self.ui, cursor="hand2", text="데이터 파일 생성", width=40, command=lambda: self.make_data_file_thread())
        self.make_data_file_button.pack()

        self.update_class_button = tk.Button(self.ui, cursor="hand2", text="반 업데이트", width=40, command=lambda: self.update_class_thread())
        self.update_class_button.pack()

        tk.Label(self.ui, text="\n< 데이터 저장 및 문자 전송 >").pack()

        self.make_data_form_button = tk.Button(self.ui, cursor="hand2", text="데일리 테스트 기록 양식 생성", width=40, command=lambda: self.make_data_form_thread())
        self.make_data_form_button.pack()

        self.save_data_button = tk.Button(self.ui, cursor="hand2", text="데이터 엑셀 파일에 저장", width=40, command=lambda: self.save_data_thread())
        self.save_data_button.pack()

        self.send_message_button = tk.Button(self.ui, cursor="hand2", text="시험 결과 전송", width=40, command=lambda: self.send_message_thread())
        self.send_message_button.pack()

        self.individual_record_button = tk.Button(self.ui, cursor="hand2", text="개별 시험 기록", width=40, command=lambda: self.individual_record_thread())
        self.individual_record_button.pack()

        self.makeup_test_record_button = tk.Button(self.ui, cursor="hand2", text="재시험 기록", width=40, command=lambda: self.makeup_test_record_thread())
        self.makeup_test_record_button.pack()

        tk.Label(self.ui, text="\n< 데이터 관리 >").pack()

        self.apply_color_button = tk.Button(self.ui, cursor="hand2", text="데이터 엑셀 파일 조건부 서식 재지정", width=40, command=lambda: apply_color(self))
        self.apply_color_button.pack()

        tk.Label(self.ui, text="< 학생 관리 >").pack()
        self.add_student_button = tk.Button(self.ui, cursor="hand2", text="신규생 추가", width=40, command=lambda: self.add_student_thread())
        self.add_student_button.pack()

        self.delete_student_button = tk.Button(self.ui, cursor="hand2", text="퇴원 처리", width=40, command=lambda: self.delete_student_thread())
        self.delete_student_button.pack()

        self.move_student_button = tk.Button(self.ui, cursor="hand2", text="학생 반 이동", width=40, command=lambda: self.move_student_thread())
        self.move_student_button.pack()

    # ui
    def thread_log(self):
        try:
            msg = self.q.get(block=False)
        except queue.Empty:
            self.ui.after(100, self.thread_log)
            return
        self.log.insert(tk.END, msg)
        self.log.update()
        self.log.see(tk.END)
        self.ui.after(100, self.thread_log)

    def check_files(self):
        check1 = check2 = check3 = False
        if os.path.isfile("반 정보.xlsx"):
            self.make_class_info_file_button["state"] = tk.DISABLED
            check1 = True
        else:
            self.make_class_info_file_button["state"]   = tk.NORMAL
            self.make_student_info_file_button["state"] = tk.DISABLED
            self.make_data_file_button["state"]         = tk.DISABLED
        if os.path.isfile("학생 정보.xlsx"):
            self.make_student_info_file_button["state"] = tk.DISABLED
            check2 = True
        else: 
            self.make_student_info_file_button["state"] = tk.NORMAL
        if os.path.isfile(f"./data/{config['dataFileName']}.xlsx"):
            self.make_data_file_button["state"] = tk.DISABLED
            check3 = True
        else:
            self.make_data_file_button["state"] = tk.NORMAL
        
        if check1 and check2 and check3:
            self.update_class_button["state"]       = tk.NORMAL
            self.make_data_form_button["state"]     = tk.NORMAL
            self.save_data_button["state"]          = tk.NORMAL
            self.send_message_button["state"]       = tk.NORMAL
            self.individual_record_button["state"]  = tk.NORMAL
            self.makeup_test_record_button["state"] = tk.NORMAL
            self.apply_color_button["state"]        = tk.NORMAL
            self.add_student_button["state"]        = tk.NORMAL
            self.delete_student_button["state"]     = tk.NORMAL
            self.move_student_button["state"]       = tk.NORMAL
        else:
            self.update_class_button["state"]       = tk.DISABLED
            self.make_data_form_button["state"]     = tk.DISABLED
            self.save_data_button["state"]          = tk.DISABLED
            self.send_message_button["state"]       = tk.DISABLED
            self.individual_record_button["state"]  = tk.DISABLED
            self.makeup_test_record_button["state"] = tk.DISABLED
            self.apply_color_button["state"]        = tk.DISABLED
            self.add_student_button["state"]        = tk.DISABLED
            self.delete_student_button["state"]     = tk.DISABLED
            self.move_student_button["state"]       = tk.DISABLED
        
        self.ui.after(100, self.check_files)

    def check_thread_end(self):
        if self.thread_end_flag:
            self.thread_end_flag = False
            self.ui.wm_attributes("-topmost", 1)
            self.ui.wm_attributes("-topmost", 0)
        self.ui.after(100, self.check_thread_end)

    # dialog
    def holiday_dialog(self) -> dict:
        def quitEvent():
            for i in range(7):
                if var_list[i].get():
                    makeup_test_date[weekday[i]] += timedelta(days=7)
            self.ui.wm_attributes("-disabled", False)
            popup.quit()
            popup.destroy()

        self.ui.wm_attributes("-disabled", True)
        popup = tk.Toplevel()
        width = 200
        height = 300
        x = int((popup.winfo_screenwidth()/4) - (width/2))
        y = int((popup.winfo_screenheight()/2) - (height/2))
        popup.geometry(f"{width}x{height}+{x}+{y}")
        popup.title("휴일 선택")
        popup.resizable(False, False)
        popup.protocol("WM_DELETE_WINDOW", quitEvent)

        today = DATE.today()
        weekday = ("월", "화", "수", "목", "금", "토", "일")
        makeup_test_date = {weekday[i] : today + relativedelta(weekday=i) for i in range(7)}
        for key, value in makeup_test_date.items():
            if value == today: makeup_test_date[key] += timedelta(days=7)

        mon = tk.BooleanVar()
        tue = tk.BooleanVar()
        wed = tk.BooleanVar()
        thu = tk.BooleanVar()
        fri = tk.BooleanVar()
        sat = tk.BooleanVar()
        sun = tk.BooleanVar()
        var_list = [mon, tue, wed, thu, fri, sat, sun]
        tk.Label(popup, text="\n다음 중 휴일을 선택해주세요\n").pack()
        sort = today.weekday()+1
        for i in range(7):
            tk.Checkbutton(popup, text=f"{str(makeup_test_date[weekday[(sort+i)%7]])} {weekday[(sort+i)%7]}", variable=var_list[(sort+i)%7]).pack()
        tk.Label(popup, text="\n").pack()
        tk.Button(popup, text="확인", width=10 , command=quitEvent).pack()
        
        popup.mainloop()    
        
        return makeup_test_date

    def delete_student_name_dialog(self) -> str:
        def quitEvent():
            self.ui.wm_attributes("-disabled", False)
            popup.quit()
            popup.destroy()

        self.ui.wm_attributes("-disabled", True)
        popup = tk.Toplevel()
        width = 250
        height = 140
        x = int((popup.winfo_screenwidth()/4) - (width/2))
        y = int((popup.winfo_screenheight()/2) - (height/2))
        popup.geometry(f"{width}x{height}+{x}+{y}")
        popup.title("퇴원 관리")
        popup.resizable(False, False)
        popup.protocol("WM_DELETE_WINDOW", quitEvent)

        # 반 정보 확인
        class_wb = xl.load_workbook("./반 정보.xlsx")
        try:
            class_ws = class_wb["반 정보"]
        except:
            self.q.put(r"'반 정보.xlsx'의 시트명을")
            self.q.put(r"'반 정보'로 변경해 주세요.")
            return
        
        data_file_wb = xl.load_workbook(f"./data/{config['dataFileName']}.xlsx")
        data_file_ws = data_file_wb.worksheets[0]
        for i in range(1, data_file_ws.max_column):
            if data_file_ws.cell(1, i).value == "이름":
                STUDENT_NAME_COLUMN = i
                break
        for i in range(1, data_file_ws.max_column):
            if data_file_ws.cell(1, i).value == "반":
                CLASS_NAME_COLUMN = i
                break

        class_dict = {}
        for i in range(2, class_ws.max_row + 1):
            class_name = class_ws.cell(i, ClassInfo.CLASS_NAME_COLUMN).value
            student_list = []
            for j in range(2, data_file_ws.max_row+1):
                if data_file_ws.cell(j, CLASS_NAME_COLUMN).value != class_name: continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).value in ("날짜", "시험명", "시험 평균"): continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.strike: continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.color is not None and data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.color.rgb == "FFFF0000": continue
                student_list.append(data_file_ws.cell(j, STUDENT_NAME_COLUMN).value)
            class_dict[class_name] = student_list
        class_dict = dict(sorted(class_dict.items()))
        
        tk.Label(popup).pack()
        def class_call_back(event):
            class_name = event.widget.get()
            student_combo.set("학생 선택")
            student_combo["values"] = class_dict[class_name]
        class_combo = ttk.Combobox(popup, values=list(class_dict.keys()), state="readonly")
        class_combo.set("반 선택")
        class_combo.bind("<<ComboboxSelected>>", class_call_back)
        class_combo.pack()

        tk.Label(popup).pack()
        selected_student = tk.StringVar()
        student_combo = ttk.Combobox(popup, values=None, state="readonly", textvariable=selected_student)
        student_combo.set("학생 선택")
        student_combo.pack()

        tk.Label(popup).pack()
        tk.Button(popup, text="퇴원", width=10 , command=quitEvent).pack()
        
        popup.mainloop()
        
        student_name = selected_student.get()
        if student_name == "학생 선택":
            return None
        else:
            return student_name

    def move_student_dialog(self):
        def quitEvent():
            self.ui.wm_attributes("-disabled", False)
            popup.quit()
            popup.destroy()

        self.ui.wm_attributes("-disabled", True)
        popup = tk.Toplevel()
        width = 250
        height = 160
        x = int((popup.winfo_screenwidth()/4) - (width/2))
        y = int((popup.winfo_screenheight()/2) - (height/2))
        popup.geometry(f"{width}x{height}+{x}+{y}")
        popup.title("학생 반 이동")
        popup.resizable(False, False)
        popup.protocol("WM_DELETE_WINDOW", quitEvent)

        # 반 정보 확인
        class_wb = xl.load_workbook("./반 정보.xlsx")
        try:
            class_ws = class_wb["반 정보"]
        except:
            self.q.put(r"'반 정보.xlsx'의 시트명을")
            self.q.put(r"'반 정보'로 변경해 주세요.")
            return
        
        data_file_wb = xl.load_workbook(f"./data/{config['dataFileName']}.xlsx")
        data_file_ws = data_file_wb.worksheets[0]
        for i in range(1, data_file_ws.max_column):
            if data_file_ws.cell(1, i).value == "이름":
                STUDENT_NAME_COLUMN = i
                break
        for i in range(1, data_file_ws.max_column):
            if data_file_ws.cell(1, i).value == "반":
                CLASS_NAME_COLUMN = i
                break

        class_dict = {}
        for i in range(2, class_ws.max_row + 1):
            class_name = class_ws.cell(i, ClassInfo.CLASS_NAME_COLUMN).value
            student_list = []
            for j in range(2, data_file_ws.max_row+1):
                if data_file_ws.cell(j, CLASS_NAME_COLUMN).value != class_name: continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).value in ("날짜", "시험명", "시험 평균"): continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.strike: continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.color is not None and data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.color.rgb == "FFFF0000": continue
                student_list.append(data_file_ws.cell(j, STUDENT_NAME_COLUMN).value)
            class_dict[class_name] = student_list
        class_dict = dict(sorted(class_dict.items()))

        tk.Label(popup).pack()
        def class_call_back(event):
            class_name = event.widget.get()
            student_combo.set("학생 선택")
            student_combo["values"] = class_dict[class_name]
        current_class_var = tk.StringVar()
        current_class_combo = ttk.Combobox(popup, values=list(class_dict.keys()), state="readonly", textvariable=current_class_var)
        current_class_combo.set("반 선택")
        current_class_combo.bind("<<ComboboxSelected>>", class_call_back)
        current_class_combo.pack()

        selected_student = tk.StringVar()
        student_combo = ttk.Combobox(popup, values=None, state="readonly", textvariable=selected_student)
        student_combo.set("학생 선택")
        student_combo.pack()

        tk.Label(popup).pack()
        target_class_var = tk.StringVar()
        current_class_combo = ttk.Combobox(popup, values=list(class_dict.keys()), state="readonly", textvariable=target_class_var)
        current_class_combo.set("이동할 반 선택")
        current_class_combo.pack()

        tk.Label(popup).pack()
        tk.Button(popup, text="반 이동", width=10 , command=quitEvent).pack()
        
        popup.mainloop()
        
        target_student_name = selected_student.get()
        target_class_name = target_class_var.get()
        current_class_name = current_class_var.get()
        if target_student_name == "학생 선택" or target_class_name == "이동할 반 선택":
            return None
        else:
            return target_student_name, target_class_name, current_class_name

    def add_student_dialog(self):
        def quitEvent():
            self.ui.wm_attributes("-disabled", False)
            popup.quit()
            popup.destroy()

        self.ui.wm_attributes("-disabled", True)
        popup = tk.Toplevel()
        width = 250
        height = 140
        x = int((popup.winfo_screenwidth()/4) - (width/2))
        y = int((popup.winfo_screenheight()/2) - (height/2))
        popup.geometry(f"{width}x{height}+{x}+{y}")
        popup.title("신규생 추가")
        popup.resizable(False, False)
        popup.protocol("WM_DELETE_WINDOW", quitEvent)

        # 반 정보 확인
        class_wb = xl.load_workbook("./반 정보.xlsx")
        try:
            class_ws = class_wb["반 정보"]
        except:
            self.q.put(r"'반 정보.xlsx'의 시트명을")
            self.q.put(r"'반 정보'로 변경해 주세요.")
            return
        
        class_names = sorted([class_ws.cell(i, ClassInfo.CLASS_NAME_COLUMN).value for i in range(2, class_ws.max_row + 1) if class_ws.cell(i, ClassInfo.CLASS_NAME_COLUMN).value is not None])

        tk.Label(popup).pack()
        target_class_var = tk.StringVar()
        class_combo = ttk.Combobox(popup, values=class_names, state="readonly", textvariable=target_class_var, width=25)
        class_combo.set("학생을 추가할 반 선택")
        class_combo.pack()

        tk.Label(popup).pack()
        target_student_var = tk.StringVar()
        tk.Entry(popup, textvariable=target_student_var, width=28).pack()

        tk.Label(popup).pack()
        tk.Button(popup, text="신규생 추가", width=10 , command=quitEvent).pack()
        
        popup.mainloop()
        
        target_class_name = target_class_var.get()
        target_student_name = target_student_var.get()
        if target_class_name == "학생을 추가할 반 선택" or target_student_name == "":
            return None
        else:
            return target_student_name, target_class_name

    def individual_record_dialog(self):
        def quitEvent():
            self.ui.wm_attributes("-disabled", False)
            popup.quit()
            popup.destroy()

        self.ui.wm_attributes("-disabled", True)
        popup = tk.Toplevel()
        width = 250
        height = 160
        x = int((popup.winfo_screenwidth()/4) - (width/2))
        y = int((popup.winfo_screenheight()/2) - (height/2))
        popup.geometry(f"{width}x{height}+{x}+{y}")
        popup.title("개별 점수 기록")
        popup.resizable(False, False)
        popup.protocol("WM_DELETE_WINDOW", quitEvent)

        # 반 정보 확인
        class_wb = xl.load_workbook("./반 정보.xlsx")
        try:
            class_ws = class_wb["반 정보"]
        except:
            self.q.put(r"'반 정보.xlsx'의 시트명을")
            self.q.put(r"'반 정보'로 변경해 주세요.")
            return
        
        data_file_wb = xl.load_workbook(f"./data/{config['dataFileName']}.xlsx")
        data_file_ws = data_file_wb.worksheets[0]
        for i in range(1, data_file_ws.max_column):
            if data_file_ws.cell(1, i).value == "이름":
                STUDENT_NAME_COLUMN = i
                break
        for i in range(1, data_file_ws.max_column):
            if data_file_ws.cell(1, i).value == "반":
                CLASS_NAME_COLUMN = i
                break
        for i in range(1, data_file_ws.max_column+1):
            if data_file_ws.cell(1, i).value == "학생 평균":
                AVERAGE_SCORE_COLUMN = i
                break

        class_dict1:dict[str, dict] = {}
        class_dict2:dict[str, dict] = {}
        for i in range(2, class_ws.max_row + 1):
            student_dict = {}
            class_name = class_ws.cell(i, ClassInfo.CLASS_NAME_COLUMN).value
            for j in range(2, data_file_ws.max_row+1):
                if data_file_ws.cell(j, CLASS_NAME_COLUMN).value != class_name: continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).value == "날짜":
                    test_name_dict = {str(data_file_ws.cell(j, k).value).split()[0][2:].replace("-", "/")+" "+str(data_file_ws.cell(j+1, k).value) : k for k in range(AVERAGE_SCORE_COLUMN+1, data_file_ws.max_row+1) if data_file_ws.cell(j, k).value is not None and data_file_ws.cell(j+1, k).value is not None}
                    continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).value in ("시험명", "시험 평균"): continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.strike: continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.color is not None and data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.color.rgb == "FFFF0000": continue
                student_dict[data_file_ws.cell(j, STUDENT_NAME_COLUMN).value] = j
            
            test_name_dict = dict(sorted(test_name_dict.items(), reverse=True))
            class_dict1[class_name] = student_dict
            class_dict2[class_name] = test_name_dict

        class_dict1 = dict(sorted(class_dict1.items()))
        tk.Label(popup).pack()
        def class_call_back(event):
            class_name = event.widget.get()
            student_combo.set("학생 선택")
            student_combo["values"] = list(class_dict1[class_name].keys())
            test_list_combo["values"] = list(class_dict2[class_name].keys())
        target_class_var = tk.StringVar()
        target_class_combo = ttk.Combobox(popup, values=list(class_dict1.keys()), state="readonly", textvariable=target_class_var, width=100)
        target_class_combo.set("반 선택")
        target_class_combo.bind("<<ComboboxSelected>>", class_call_back)
        target_class_combo.pack()

        target_studnet_var = tk.StringVar()
        student_combo = ttk.Combobox(popup, values=None, state="readonly", textvariable=target_studnet_var, width=100)
        student_combo.set("학생 선택")
        student_combo.pack()

        test_name_var = tk.StringVar()
        test_list_combo = ttk.Combobox(popup, values=list(class_dict2.keys()), state="readonly", textvariable=test_name_var, width=100)
        test_list_combo.set("시험 선택")
        test_list_combo.pack()

        score_var = tk.StringVar()
        tk.Entry(popup, textvariable=score_var, width=28).pack()

        tk.Label(popup).pack()
        tk.Button(popup, text="메세지 전송 및 저장", width=20 , command=quitEvent).pack()
        
        popup.mainloop()
        
        target_class_name   = target_class_var.get()
        target_student_name = target_studnet_var.get()
        test_name           = test_name_var.get()
        test_score          = score_var.get()
        try:
            if '.' in test_score:
                test_score = float(test_score)
            else:
                test_score = int(test_score)
        except:
            self.q.put("올바른 점수를 입력해 주세요.")
            return None
        
        if target_class_name == "반 선택" or target_student_name == "학생 선택" or test_name == "시험 선택":
            return None

        row = class_dict1[target_class_name][target_student_name]
        col = class_dict2[target_class_name][test_name]
        return target_student_name, target_class_name, test_name, row, col, test_score, data_file_ws.cell(row, col).value

    def makeup_test_record_dialog(self):
        def quitEvent():
            self.ui.wm_attributes("-disabled", False)
            popup.quit()
            popup.destroy()

        self.ui.wm_attributes("-disabled", True)
        popup = tk.Toplevel()
        width = 250
        height = 160
        x = int((popup.winfo_screenwidth()/4) - (width/2))
        y = int((popup.winfo_screenheight()/2) - (height/2))
        popup.geometry(f"{width}x{height}+{x}+{y}")
        popup.title("재시험 기록")
        popup.resizable(False, False)
        popup.protocol("WM_DELETE_WINDOW", quitEvent)

        # 반 정보 확인
        class_wb = xl.load_workbook("./반 정보.xlsx")
        try:
            class_ws = class_wb["반 정보"]
        except:
            self.q.put(r"'반 정보.xlsx'의 시트명을")
            self.q.put(r"'반 정보'로 변경해 주세요.")
            return
        
        makeup_list_wb = xl.load_workbook("./data/재시험 명단.xlsx")
        try:
            makeup_list_ws = makeup_list_wb["재시험 명단"]
        except:
            gui.q.put(r"'재시험 명단.xlsx'의 시트명을")
            gui.q.put(r"'재시험 명단'으로 변경해 주세요.")
            return
        
        data_file_wb = xl.load_workbook(f"./data/{config['dataFileName']}.xlsx")
        data_file_ws = data_file_wb.worksheets[0]
        for i in range(1, data_file_ws.max_column):
            if data_file_ws.cell(1, i).value == "이름":
                STUDENT_NAME_COLUMN = i
                break
        for i in range(1, data_file_ws.max_column):
            if data_file_ws.cell(1, i).value == "반":
                CLASS_NAME_COLUMN = i
                break

        class_dict:dict[str, list] = {}
        for i in range(2, class_ws.max_row + 1):
            class_name = class_ws.cell(i, ClassInfo.CLASS_NAME_COLUMN).value
            student_list = []
            for j in range(2, data_file_ws.max_row+1):
                if data_file_ws.cell(j, CLASS_NAME_COLUMN).value != class_name: continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).value in ("날짜", "시험명", "시험 평균"): continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.strike: continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.color is not None and data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.color.rgb == "FFFF0000": continue
                student_list.append(data_file_ws.cell(j, STUDENT_NAME_COLUMN).value)
            class_dict[class_name] = student_list
        class_dict = dict(sorted(class_dict.items()))

        student_dict:dict[str, dict[str, int]] = {}
        for i in range(2, makeup_list_ws.max_row+1):
            if makeup_list_ws.cell(i, MakeupTestList.MAKEUPTEST_SCORE_COLUMN).value is None:
                student_name = makeup_list_ws.cell(i, MakeupTestList.STUDENT_NAME_COLUMN).value
                makeup_test_name = makeup_list_ws.cell(i, MakeupTestList.TEST_NAME_COLUMN).value
                try:
                    student_dict[student_name]
                except:
                    student_dict[student_name] = {}
                
                student_dict[student_name][makeup_test_name] = i

        def class_call_back(event):
            class_name = event.widget.get()
            student_combo.set("학생 선택")
            student_combo["values"] = list(class_dict[class_name])
            makeup_test_list_combo.set("재시험 선택")
            makeup_test_list_combo["values"] = None
        target_class_var = tk.StringVar()
        target_class_combo = ttk.Combobox(popup, values=list(class_dict.keys()), state="readonly", textvariable=target_class_var, width=100)
        target_class_combo.set("반 선택")
        target_class_combo.bind("<<ComboboxSelected>>", class_call_back)
        target_class_combo.pack()

        def student_call_back(event):
            student_name = event.widget.get()
            makeup_test_list_combo.set("재시험 선택")
            try:
                makeup_test_list_combo["values"] = list(student_dict[student_name].keys())
            except:
                makeup_test_list_combo.set("재시험이 없습니다")
                makeup_test_list_combo["values"] = None
        target_studnet_var = tk.StringVar()
        student_combo = ttk.Combobox(popup, values=None, state="readonly", textvariable=target_studnet_var, width=100)
        student_combo.set("학생 선택")
        student_combo.bind("<<ComboboxSelected>>", student_call_back)
        student_combo.pack()

        makeup_test_name_var = tk.StringVar()
        makeup_test_list_combo = ttk.Combobox(popup, values=None, state="readonly", textvariable=makeup_test_name_var, width=100)
        makeup_test_list_combo.set("재시험 선택")
        makeup_test_list_combo.pack()

        makeup_test_score_var = tk.StringVar()
        tk.Entry(popup, textvariable=makeup_test_score_var, width=28).pack()

        tk.Label(popup).pack()
        tk.Button(popup, text="재시험 저장", width=10 , command=quitEvent).pack()
        
        popup.mainloop()

        target_class_name   = target_class_var.get()
        target_student_name = target_studnet_var.get()
        makeup_test_name    = makeup_test_name_var.get()
        makeup_test_score   = makeup_test_score_var.get()
        
        if target_class_name == "반 선택" or target_student_name == "학생 선택" or makeup_test_name == "재시험 선택" or makeup_test_name == "재시험이 없습니다" or makeup_test_score == "":
            return None
        
        row = student_dict[target_student_name][makeup_test_name]
        return row, makeup_test_score

    # threads
    def make_class_info_file_thread(self):
        thread = threading.Thread(target=lambda: make_class_info_file(self))
        thread.daemon = True
        thread.start()

    def make_student_info_file_thread(self):
        thread = threading.Thread(target=lambda: make_student_info_file(self))
        thread.daemon = True
        thread.start()

    def make_data_file_thread(self):
        thread = threading.Thread(target=lambda: make_data_file(self))
        thread.daemon = True
        thread.start()

    def update_class_thread(self):
        if os.path.isfile(f"./data/~${config['dataFileName']}.xlsx"):
            self.q.put(r"데이터 파일을 닫은 뒤 다시 시도해 주세요.")
            return
        if self.update_class_button["text"] == "반 업데이트":
            if os.path.isfile("./temp.xlsx"):
                os.remove("./temp.xlsx")
            self.update_class_button["state"] = tk.DISABLED
            global ret
            ret = check_update_class(self)
            excel = win32com.client.Dispatch("Excel.Application")
            try:
                excel.Visible = True
            except:
                self.q.put("모든 Excel 파일을 종료한 뒤 다시 시도해 주세요.")
            wb = excel.Workbooks.Open(f"{os.getcwd()}\\temp.xlsx")
            self.update_class_button["text"] = "반 정보 수정 후 반 업데이트 계속하기"
            if not tkinter.messagebox.askokcancel("반 정보 변경 확인", "반 정보 파일의 빈칸을 채운 뒤 Excel을 종료하고\n버튼을 눌러주세요.\n삭제할 반은 행을 삭제해 주세요.\n취소 선택 시 반 업데이트가 중단됩니다."):
                self.update_class_button["text"] = "반 업데이트"
                wb.Close()
                if os.path.isfile("./temp.xlsx"):
                    os.remove("./temp.xlsx")
                    self.q.put(r"반 업데이트를 중단합니다.")
        else:
            if os.path.isfile("./~$temp.xlsx"):
                self.q.put(r"임시 파일을 닫은 뒤 다시 시도해 주세요.")
                return
            if os.path.isfile(f"./data/~${config['dataFileName']}.xlsx"):
                self.q.put(r"데이터 파일을 닫은 뒤 다시 시도해 주세요.")
                return
            if not tkinter.messagebox.askokcancel("반 정보 변경 확인", "반 업데이트를 계속하시겠습니까?"):
                if os.path.isfile("./temp.xlsx"):
                    os.remove("./temp.xlsx")
                    self.q.put(r"반 업데이트를 중단합니다.")
            thread = threading.Thread(target=lambda: update_class(self, ret[0], ret[1]))
            thread.daemon = True
            thread.start()
            del ret
            self.update_class_button["text"] = "반 업데이트"
        self.update_class_button["state"] = tk.NORMAL

    def make_data_form_thread(self):
        self.make_data_form_button["state"] = tk.DISABLED
        thread = threading.Thread(target=lambda: make_data_form(self))
        thread.daemon = True
        thread.start()
        self.make_data_form_button['state'] = tk.NORMAL

    def save_data_thread(self):
        if os.path.isfile("./data/~$재시험 명단.xlsx"):
            self.q.put(r"재시험 명단 파일을 닫은 뒤 다시 시도해 주세요.")
            return
        if os.path.isfile(f"./data/~${config['dataFileName']}.xlsx"):
            self.q.put(r"데이터 파일을 닫은 뒤 다시 시도해 주세요.")
            return
        self.save_data_button["state"] = tk.DISABLED
        if self.makeup_test_date is None:
            self.makeup_test_date = self.holiday_dialog()
        filepath = filedialog.askopenfilename(initialdir="./", title="데일리테스트 기록 양식 선택", filetypes=(("Excel files", "*.xlsx"),("all files", "*.*")))
        if filepath == "": return
        thread = threading.Thread(target=lambda: save_data(self, filepath, self.makeup_test_date))
        thread.daemon = True
        thread.start()
        self.save_data_button["state"] = tk.NORMAL

    def send_message_thread(self):
        self.send_message_button["state"] = tk.DISABLED
        if self.makeup_test_date is None:
            self.makeup_test_date = self.holiday_dialog()
        filepath = filedialog.askopenfilename(initialdir="./", title="데일리테스트 기록 양식 선택", filetypes=(("Excel files", "*.xlsx"),("all files", "*.*")))
        if filepath == "": return
        thread = threading.Thread(target=lambda: send_message(self, filepath, self.makeup_test_date))
        thread.daemon = True
        thread.start()
        self.send_message_button["state"] = tk.NORMAL

    def individual_record_thread(self):
        if os.path.isfile("./data/~$재시험 명단.xlsx"):
            self.q.put(r"재시험 명단 파일을 닫은 뒤 다시 시도해 주세요.")
            return
        if os.path.isfile(f"./data/~${config['dataFileName']}.xlsx"):
            self.q.put(r"데이터 파일을 닫은 뒤 다시 시도해 주세요.")
            return
        self.individual_record_button["state"] = tk.DISABLED
        if self.makeup_test_date is None:
            self.makeup_test_date = self.holiday_dialog()
        ret = self.individual_record_dialog()
        if ret is None:
            self.individual_record_button["state"] = tk.NORMAL
            return
        student_name, class_name, test_name, row, col, test_score, cell_value = ret
        if cell_value is not None:
            if not tkinter.messagebox.askyesno("데이터 중복 확인", f"{student_name} 학생의 {test_name} 시험에 대한 점수({cell_value}점)가 이미 존재합니다.\n덮어쓰시겠습니까?"):
                self.q.put(r"개별 데이터 저장을 취소하였습니다.")
                self.individual_record_button["state"] = tk.NORMAL
                return
        thread = threading.Thread(target=lambda: individual_record(self, student_name, class_name, test_name, row, col, test_score, self.makeup_test_date))
        thread.daemon = True
        thread.start()
        self.individual_record_button["state"] = tk.NORMAL

    def makeup_test_record_thread(self):
        if os.path.isfile("./data/~$재시험 명단.xlsx"):
            self.q.put(r"재시험 명단 파일을 닫은 뒤 다시 시도해 주세요.")
            return
        self.makeup_test_record_button["state"] = tk.DISABLED
        ret = self.makeup_test_record_dialog()
        if ret is None:
            return
        row, makeup_test_score = ret
        makeup_list_wb = xl.load_workbook("./data/재시험 명단.xlsx")
        try:
            makeup_list_ws = makeup_list_wb["재시험 명단"]
        except:
            self.q.put(r"'재시험 명단.xlsx'의 시트명을")
            self.q.put(r"'재시험 명단'으로 변경해 주세요.")
            return
        makeup_list_ws.cell(row, MakeupTestList.MAKEUPTEST_SCORE_COLUMN).value = makeup_test_score
        # makeup_list_ws.sheet_view.topLeftCell = f"A{str(max(1, row-10))}"
        # makeup_list_ws.sheet_view.selection[0].sqref = f"{gcl(MakeupTestList.MAKEUPTEST_SCORE_COLUMN)}{str(row)}"
        makeup_list_wb.save("./data/재시험 명단.xlsx")
        self.q.put(f"{row} 행에 재시험 점수를 기록하였습니다.")
        # excel = win32com.client.Dispatch("Excel.Application")
        # excel.Visible = True
        # excel.Workbooks.Open(f"{os.getcwd()}\\data\\재시험 명단.xlsx")
        self.makeup_test_record_button["state"] = tk.NORMAL

    def add_student_thread(self):
        if os.path.isfile(f"./data/~${config['dataFileName']}.xlsx"):
            self.q.put(r"데이터 파일을 닫은 뒤 다시 시도해 주세요.")
            return
        if os.path.isfile("./data/~$학생 정보.xlsx"):
            self.q.put(r"학생 정보 파일을 닫은 뒤 다시 시도해 주세요.")
            return
        self.add_student_button["state"] = tk.DISABLED
        tmp = self.add_student_dialog()
        if tmp is not None:
            # 학생 추가 확인
            if not tkinter.messagebox.askyesno("학생 추가 확인", f"{tmp[0]} 학생을 {tmp[1]} 반에 추가하시겠습니까?"):
                self.add_student_button["state"] = tk.NORMAL
                return
            thread = threading.Thread(target=lambda: add_student(self, tmp[0], tmp[1]))
            thread.daemon = True
            thread.start()
        self.add_student_button["state"] = tk.NORMAL

    def delete_student_thread(self):
        if os.path.isfile(f"./data/~${config['dataFileName']}.xlsx"):
            self.q.put(r"데이터 파일을 닫은 뒤 다시 시도해 주세요.")
            return
        if os.path.isfile("./data/~$학생 정보.xlsx"):
            self.q.put(r"학생 정보 파일을 닫은 뒤 다시 시도해 주세요.")
            return
        self.delete_student_button["state"] = tk.DISABLED
        student_name = self.delete_student_name_dialog()
        if student_name is not None:
            # 퇴원 처리 확인
            if not tkinter.messagebox.askyesno("퇴원 확인", f"{student_name} 학생을 퇴원 처리하시겠습니까?"):
                self.delete_student_button["state"] = tk.NORMAL
                return
            thread = threading.Thread(target=lambda: delete_student(self, student_name))
            thread.daemon = True
            thread.start()
        self.delete_student_button["state"] = tk.NORMAL

    def move_student_thread(self):
        if os.path.isfile(f"./data/~${config['dataFileName']}.xlsx"):
            self.q.put(r"데이터 파일을 닫은 뒤 다시 시도해 주세요.")
            return
        if os.path.isfile("./data/~$학생 정보.xlsx"):
            self.q.put(r"학생 정보 파일을 닫은 뒤 다시 시도해 주세요.")
            return
        self.move_student_button["state"] = tk.DISABLED
        tmp = self.move_student_dialog()
        if tmp is not None:
            # 학생 반 이동 확인
            if tmp[1] == tmp[2]:
                self.q.put(r"학생의 현재 반과 이동할 반이 같아 취소되었습니다.")
                self.move_student_button["state"] = tk.NORMAL
                return
            if not tkinter.messagebox.askyesno("학생 반 이동 확인", f"{tmp[2]} 반의 {tmp[0]} 학생을\n{tmp[1]} 반으로 이동시키겠습니까?"):
                self.move_student_button["state"] = tk.NORMAL
                return
            thread = threading.Thread(target=lambda: move_student(self, tmp[0], tmp[1], tmp[2]))
            thread.daemon = True
            thread.start()
        self.move_student_button["state"] = tk.NORMAL

# tasks
def make_class_info_file(gui:GUI):
    gui.q.put("반 정보 입력 파일 생성 중...")

    ini_wb = xl.Workbook()
    ini_ws = ini_wb[ini_wb.sheetnames[0]]
    ini_ws.title = "반 정보"
    ini_ws[gcl(ClassInfo.CLASS_NAME_COLUMN)+"1"]    = "반명"
    ini_ws[gcl(ClassInfo.TEACHER_NAME_COLUMN)+"1"]  = "선생님명"
    ini_ws[gcl(ClassInfo.CLASS_WEEKDAY_COLUMN)+"1"] = "요일"
    ini_ws[gcl(ClassInfo.TEST_TIME_COLUMN)+"1"]     = "시간"

    options = webdriver.ChromeOptions()
    options.add_argument("headless")
    driver = webdriver.Chrome(service = service, options = options)

    # 아이소식 접속
    driver.get(config["url"])
    table_names = driver.find_elements(By.CLASS_NAME, "style1")

    # 반 루프
    for table_name in table_names:
        WRITE_LOCATION = ini_ws.max_row + 1
        ini_ws.cell(WRITE_LOCATION, 1).value = table_name.text.rstrip()

    # 정렬 및 테두리
    for j in range(1, ini_ws.max_row + 1):
        for k in range(1, ini_ws.max_column + 1):
            ini_ws.cell(j, k).alignment = Alignment(horizontal="center", vertical="center")
            ini_ws.cell(j, k).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    ini_wb.save("./반 정보.xlsx")
    gui.q.put("반 정보 입력 파일 생성을 완료했습니다.")
    gui.q.put("반 정보를 입력해 주세요.")
    gui.thread_end_flag = True

def make_student_info_file(gui:GUI):
    gui.q.put("학생 정보 파일 생성 중...")

    ini_wb = xl.Workbook()
    ini_ws = ini_wb[ini_wb.sheetnames[0]]
    ini_ws.title = "학생 정보"
    ini_ws[gcl(StudentInfo.STUDENT_NAME_COLUMN)+"1"]       = "이름"
    ini_ws[gcl(StudentInfo.CLASS_NAME_COLUMN)+"1"]         = "반명"
    ini_ws[gcl(StudentInfo.TEACHER_NAME_COLUMN)+"1"]       = "담당"
    ini_ws[gcl(StudentInfo.MAKEUPTEST_WEEKDAY_COLUMN)+"1"] = "재시험 응시 요일"
    ini_ws[gcl(StudentInfo.MAKEUPTEST_TIME_COLUMN)+"1"]    = "재시험 응시 시간"
    ini_ws[gcl(StudentInfo.NEW_STUDENT_CHECK_COLUMN)+"1"]  = "기수 신규생"
    ini_ws["Z1"] = "N"
    ini_ws.auto_filter.ref = "A:"+gcl(StudentInfo.MAX)
    ini_ws.column_dimensions.group("Z", hidden=True)
    ini_wb.save("./학생 정보.xlsx")

    student_wb = xl.load_workbook("./학생 정보.xlsx")
    try:
        student_ws = student_wb["학생 정보"]
    except:
        gui.q.put(r"'학생 정보.xlsx'의 시트명을")
        gui.q.put(r"'학생 정보'로 변경해 주세요.")
        return

    # 반 정보 확인
    class_wb = xl.load_workbook("./반 정보.xlsx")
    try:
        class_ws = class_wb["반 정보"]
    except:
        gui.q.put(r"'반 정보.xlsx'의 시트명을")
        gui.q.put(r"'반 정보'로 변경해 주세요.")
        return

    options = webdriver.ChromeOptions()
    options.add_argument("headless")
    driver = webdriver.Chrome(service = service, options = options)

    # 아이소식 접속
    driver.get(config["url"])
    table_names = driver.find_elements(By.CLASS_NAME, "style1")

    # 반 루프
    for i in range(3, len(table_names)):
        trs = driver.find_element(By.ID, f"table_{str(i)}").find_elements(By.CLASS_NAME, "style12")
        WRITE_LOCATION = student_ws.max_row + 1

        class_name = table_names[i].text.rstrip()
        for j in range(2, class_ws.max_row + 1):
            if class_ws.cell(j, ClassInfo.CLASS_NAME_COLUMN).value == class_name:
                teacher_name = class_ws.cell(j, ClassInfo.TEACHER_NAME_COLUMN).value
                break
        else: continue

        # 학생 루프
        for tr in trs:
            WRITE_LOCATION = ini_ws.max_row + 1
            ini_ws.cell(WRITE_LOCATION, StudentInfo.STUDENT_NAME_COLUMN).value = tr.find_element(By.CLASS_NAME, "style9").text
            ini_ws.cell(WRITE_LOCATION, StudentInfo.CLASS_NAME_COLUMN).value   = class_name
            ini_ws.cell(WRITE_LOCATION, StudentInfo.TEACHER_NAME_COLUMN).value = teacher_name
            dv = DataValidation(type="list", formula1="=Z1",  allow_blank=True, errorStyle="stop", showErrorMessage=True)
            student_ws.add_data_validation(dv)
            dv.add(ini_ws.cell(WRITE_LOCATION, StudentInfo.NEW_STUDENT_CHECK_COLUMN))

    # 정렬 및 테두리
    for j in range(1, ini_ws.max_row + 1):
        for k in range(1, StudentInfo.MAX + 1):
            ini_ws.cell(j, k).alignment = Alignment(horizontal="center", vertical="center")
            ini_ws.cell(j, k).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    ini_wb.save("./학생 정보.xlsx")
    gui.q.put("학생 정보 파일을 생성했습니다.")
    gui.thread_end_flag = True

def make_data_file(gui:GUI):
    gui.q.put("데이터파일 생성 중...")

    ini_wb = xl.Workbook()
    ini_ws = ini_wb[ini_wb.sheetnames[0]]
    ini_ws.title = "데일리테스트"
    ini_ws[gcl(DataFile.TEST_TIME_COLUMN)+"1"]     = "시간"
    ini_ws[gcl(DataFile.CLASS_WEEKDAY_COLUMN)+"1"] = "요일"
    ini_ws[gcl(DataFile.CLASS_NAME_COLUMN)+"1"]    = "반"
    ini_ws[gcl(DataFile.TEACHER_NAME_COLUMN)+"1"]  = "담당"
    ini_ws[gcl(DataFile.STUDENT_NAME_COLUMN)+"1"]  = "이름"
    ini_ws[gcl(DataFile.AVERAGE_SCORE_COLUMN)+"1"] = "학생 평균"
    ini_ws.freeze_panes = gcl(DataFile.DATA_COLUMN) + "2"
    ini_ws.auto_filter.ref = "A:" + gcl(DataFile.MAX)

    # 반 정보 확인
    class_wb = xl.load_workbook("./반 정보.xlsx")
    try:
        class_ws = class_wb["반 정보"]
    except:
        gui.q.put(r"'반 정보.xlsx'의 시트명을")
        gui.q.put(r"'반 정보'로 변경해 주세요.")
        return

    options = webdriver.ChromeOptions()
    options.add_argument("headless")
    driver = webdriver.Chrome(service = service, options = options)

    # 아이소식 접속
    driver.get(config["url"])
    table_names = driver.find_elements(By.CLASS_NAME, "style1")

    # 반 루프
    for i in range(3, len(table_names)):
        trs = driver.find_element(By.ID, f"table_{str(i)}").find_elements(By.CLASS_NAME, "style12")
        if len(trs) == 0:
            continue
        
        WRITE_LOCATION = ini_ws.max_row + 1

        class_name = table_names[i].text.rstrip()
        teacher_name  = ""
        class_weekday = ""
        test_time     = ""
        for j in range(2, class_ws.max_row + 1):
            if class_ws.cell(j, ClassInfo.CLASS_NAME_COLUMN).value == class_name:
                teacher_name  = class_ws.cell(j, ClassInfo.TEACHER_NAME_COLUMN).value
                class_weekday = class_ws.cell(j, ClassInfo.CLASS_WEEKDAY_COLUMN).value
                test_time     = class_ws.cell(j, ClassInfo.TEST_TIME_COLUMN).value
                break
        else:
            continue
        
        # 시험명
        ini_ws.cell(WRITE_LOCATION, DataFile.TEST_TIME_COLUMN).value     = test_time
        ini_ws.cell(WRITE_LOCATION, DataFile.CLASS_WEEKDAY_COLUMN).value = class_weekday
        ini_ws.cell(WRITE_LOCATION, DataFile.CLASS_NAME_COLUMN).value    = class_name
        ini_ws.cell(WRITE_LOCATION, DataFile.TEACHER_NAME_COLUMN).value  = teacher_name
        ini_ws.cell(WRITE_LOCATION, DataFile.STUDENT_NAME_COLUMN).value  = "날짜"
        
        WRITE_LOCATION = ini_ws.max_row + 1
        ini_ws.cell(WRITE_LOCATION, DataFile.TEST_TIME_COLUMN).value     = test_time
        ini_ws.cell(WRITE_LOCATION, DataFile.CLASS_WEEKDAY_COLUMN).value = class_weekday
        ini_ws.cell(WRITE_LOCATION, DataFile.CLASS_NAME_COLUMN).value    = class_name
        ini_ws.cell(WRITE_LOCATION, DataFile.TEACHER_NAME_COLUMN).value  = teacher_name
        ini_ws.cell(WRITE_LOCATION, DataFile.STUDENT_NAME_COLUMN).value  = "시험명"
        class_start = WRITE_LOCATION + 1

        # 학생 루프
        for tr in trs:
            WRITE_LOCATION = ini_ws.max_row + 1
            ini_ws.cell(WRITE_LOCATION, DataFile.TEST_TIME_COLUMN).value     = test_time
            ini_ws.cell(WRITE_LOCATION, DataFile.CLASS_WEEKDAY_COLUMN).value = class_weekday
            ini_ws.cell(WRITE_LOCATION, DataFile.CLASS_NAME_COLUMN).value    = class_name
            ini_ws.cell(WRITE_LOCATION, DataFile.TEACHER_NAME_COLUMN).value  = teacher_name
            ini_ws.cell(WRITE_LOCATION, DataFile.STUDENT_NAME_COLUMN).value  = tr.find_element(By.CLASS_NAME, "style9").text
            ini_ws.cell(WRITE_LOCATION, DataFile.AVERAGE_SCORE_COLUMN).value = f"=ROUND(AVERAGE(G{str(WRITE_LOCATION)}:XFD{str(WRITE_LOCATION)}), 0)"
            ini_ws.cell(WRITE_LOCATION, DataFile.AVERAGE_SCORE_COLUMN).font  = Font(bold=True)
        
        # 시험별 평균
        WRITE_LOCATION = ini_ws.max_row + 1
        class_end = WRITE_LOCATION - 1
        ini_ws.cell(WRITE_LOCATION, DataFile.TEST_TIME_COLUMN).value     = test_time
        ini_ws.cell(WRITE_LOCATION, DataFile.CLASS_WEEKDAY_COLUMN).value = class_weekday
        ini_ws.cell(WRITE_LOCATION, DataFile.CLASS_NAME_COLUMN).value    = class_name
        ini_ws.cell(WRITE_LOCATION, DataFile.TEACHER_NAME_COLUMN).value  = teacher_name
        ini_ws.cell(WRITE_LOCATION, DataFile.STUDENT_NAME_COLUMN).value  = "시험 평균"
        ini_ws[f"F{str(WRITE_LOCATION)}"] = ArrayFormula(f"F{str(WRITE_LOCATION)}", f"=ROUND(AVERAGE(IFERROR(F{str(class_start)}:F{str(class_end)}, \"\")), 0)")
        ini_ws.cell(WRITE_LOCATION, DataFile.AVERAGE_SCORE_COLUMN).font = Font(bold=True)

        for j in range(1, DataFile.DATA_COLUMN):
            ini_ws.cell(WRITE_LOCATION, j).border = Border(bottom = Side(border_style="medium", color="000000"))

    # 정렬
    for i in range(1, ini_ws.max_row + 1):
        for j in range(1, ini_ws.max_column + 1):
            ini_ws.cell(i, j).alignment = Alignment(horizontal="center", vertical="center")
    
    # 모의고사 sheet 생성
    copy_ws = ini_wb.copy_worksheet(ini_wb["데일리테스트"])
    copy_ws.title = "모의고사"
    copy_ws.freeze_panes = gcl(DataFile.DATA_COLUMN) + "2"
    copy_ws.auto_filter.ref = "A:" + gcl(DataFile.STUDENT_NAME_COLUMN)

    ini_wb.save(f"./data/{config['dataFileName']}.xlsx")
    gui.q.put("데이터 파일을 생성했습니다.")
    gui.thread_end_flag = True

def check_update_class(gui:GUI):
    # 반 정보 확인
    class_wb = xl.load_workbook("./반 정보.xlsx")
    try:
        class_ws = class_wb["반 정보"]
    except:
        gui.q.put(r"'반 정보.xlsx'의 시트명을")
        gui.q.put(r"'반 정보'로 변경해 주세요.")
        return

    options = webdriver.ChromeOptions()
    options.add_argument("headless")
    driver = webdriver.Chrome(service = service, options = options)

    gui.q.put("아이소식으로부터 반 정보를 업데이트 하는 중...")
    # 아이소식 접속
    driver.get(config["url"])
    table_names = driver.find_elements(By.CLASS_NAME, "style1")
    current_classes = [class_ws.cell(i, ClassInfo.CLASS_NAME_COLUMN).value for i in range(2, class_ws.max_row+1) if class_ws.cell(i, ClassInfo.CLASS_NAME_COLUMN).value is not None]

    unregistered_classes = {table_names[i].text.rstrip() : i for i in range(3, len(table_names)) if not table_names[i].text.rstrip() in current_classes}

    for row in range(class_ws.max_row+1, 1, -1):
        if class_ws.cell(row-1, ClassInfo.CLASS_NAME_COLUMN).value is not None:
            WRITE_LOCATION = row
            break
    for new_class_name in list(unregistered_classes.keys()):
        class_ws.cell(WRITE_LOCATION, ClassInfo.CLASS_NAME_COLUMN).value = new_class_name
        WRITE_LOCATION += 1
    
    # 정렬 및 테두리
    for row in range(1, class_ws.max_row + 1):
        for col in range(1, class_ws.max_column + 1):
            class_ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
            class_ws.cell(row, col).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    
    for row in range(class_ws.max_row, 1, -1):
        if class_ws.cell(row, ClassInfo.CLASS_NAME_COLUMN).value is None:
            class_ws.delete_rows(row)
        else: break
    
    class_wb.save("./temp.xlsx")
    return current_classes, unregistered_classes

def update_class(gui:GUI, current_classes:list, unregistered_classes:dict):
    # 백업 생성
    data_file_wb = xl.load_workbook(f"./data/{config['dataFileName']}.xlsx")
    data_file_wb.save(f"./data/backup/{config['dataFileName']}({datetime.today().strftime('%Y%m%d')}).xlsx")
    class_wb = xl.load_workbook("./반 정보.xlsx")
    class_wb.save(f"./data/backup/반 정보({datetime.today().strftime('%Y%m%d')}).xlsx")

    options = webdriver.ChromeOptions()
    options.add_argument("headless")
    driver = webdriver.Chrome(service = service, options = options)
    driver.get(config["url"])

    gui.q.put("수정된 반 정보를 바탕으로 업데이트 중...")
    class_wb_temp = xl.load_workbook("./temp.xlsx")

    class_temp_ws = class_wb_temp["반 정보"]
    update_class = [class_temp_ws.cell(i, ClassInfo.CLASS_NAME_COLUMN).value for i in range(2, class_temp_ws.max_row+1) if class_temp_ws.cell(i, ClassInfo.CLASS_NAME_COLUMN).value is not None]
    delete_class = [c for c in current_classes if not c in update_class]
    check_list = [c for c in list(unregistered_classes.keys()) if c in update_class]

    data_file_wb = xl.load_workbook(f"./data/{config['dataFileName']}.xlsx")
    data_file_ws = data_file_wb["데일리테스트"]

    if len(check_list) == 0 and len(delete_class) == 0:
        data_file_wb.save(f"./data/{config['dataFileName']}.xlsx")
        class_wb_temp.save("./반 정보.xlsx")
        os.remove("./temp.xlsx")
        gui.q.put("업데이트 된 항목이 없습니다.")
        return
    
    # 데이터 파일에서 이전 데이터 이동 및 삭제
    if len(delete_class) != 0:
        gui.q.put("이전 데이터 제거 중...")
        if not os.path.isfile("./data/지난 데이터.xlsx"):
            ini_wb = xl.Workbook()
            ini_ws = ini_wb[ini_wb.sheetnames[0]]
            ini_ws.title = "데일리테스트"
            ini_ws[gcl(DataFile.TEST_TIME_COLUMN)+"1"]     = "시간"
            ini_ws[gcl(DataFile.CLASS_WEEKDAY_COLUMN)+"1"] = "요일"
            ini_ws[gcl(DataFile.CLASS_NAME_COLUMN)+"1"]    = "반"
            ini_ws[gcl(DataFile.TEACHER_NAME_COLUMN)+"1"]  = "담당"
            ini_ws[gcl(DataFile.STUDENT_NAME_COLUMN)+"1"]  = "이름"
            ini_ws[gcl(DataFile.AVERAGE_SCORE_COLUMN)+"1"] = "학생 평균"
            ini_ws.freeze_panes = gcl(DataFile.DATA_COLUMN) + "2"
            ini_ws.auto_filter.ref = "A:" + gcl(DataFile.MAX)

            copy_ws = ini_wb.copy_worksheet(ini_wb["데일리테스트"])
            copy_ws.title = "모의고사"
            copy_ws.freeze_panes = gcl(DataFile.DATA_COLUMN) + "2"
            copy_ws.auto_filter.ref = "A:" + gcl(DataFile.STUDENT_NAME_COLUMN)

            ini_wb.save("./data/지난 데이터.xlsx")
        pythoncom.CoInitialize()
        excel = win32com.client.Dispatch("Excel.Application")
    
        try:
            wb = excel.Workbooks.Open(f"{os.getcwd()}\\data\\{config['dataFileName']}.xlsx")
            wb.Save()
            wb.Close()
        except:
            gui.q.put("데이터 파일 창을 끄고 다시 실행해 주세요.")
            return
        
        post_data_wb = xl.load_workbook("./data/지난 데이터.xlsx")
        data_file_temp_wb = xl.load_workbook(f"./data/{config['dataFileName']}.xlsx", data_only=True)
        for sheet_name in data_file_temp_wb.sheetnames:
            data_file_temp_ws = data_file_temp_wb[sheet_name]
            post_data_ws = post_data_wb[sheet_name]
            data_file_ws = data_file_wb[sheet_name]
            
            # 동적 열 탐색
            for i in range(1, data_file_temp_ws.max_column+1):
                temp = data_file_temp_ws.cell(1, i).value
                if temp == "시간":
                    TEST_TIME_COLUMN = i
                elif temp == "요일":
                    CLASS_WEEKDAY_COLUMN = i
                elif temp == "반":
                    CLASS_NAME_COLUMN = i
                elif temp == "담당":
                    TEACHER_NAME_COLUMN = i
                elif temp == "이름":
                    STUDENT_NAME_COLUMN = i
                elif temp == "학생 평균":
                    AVERAGE_SCORE_COLUMN = i

            for row in range(2, data_file_ws.max_row+1):
                data_file_ws.cell(row, AVERAGE_SCORE_COLUMN).value = ""
            
            # 지난 데이터 행 삭제
            for row in range(2, data_file_ws.max_row+1):
                while data_file_ws.cell(row, CLASS_NAME_COLUMN).value in delete_class:
                    data_file_ws.delete_rows(row)

            # 지난 데이터 행 복사
            for row in range(2, data_file_temp_ws.max_row+1):
                if data_file_temp_ws.cell(row, CLASS_NAME_COLUMN).value in delete_class:
                    POST_DATA_WRITE_ROW = post_data_ws.max_row+1
                    copy_cell(post_data_ws.cell(POST_DATA_WRITE_ROW, DataFile.TEST_TIME_COLUMN),     data_file_temp_ws.cell(row, TEST_TIME_COLUMN))
                    copy_cell(post_data_ws.cell(POST_DATA_WRITE_ROW, DataFile.CLASS_WEEKDAY_COLUMN), data_file_temp_ws.cell(row, CLASS_WEEKDAY_COLUMN))
                    copy_cell(post_data_ws.cell(POST_DATA_WRITE_ROW, DataFile.CLASS_NAME_COLUMN),    data_file_temp_ws.cell(row, CLASS_NAME_COLUMN))
                    copy_cell(post_data_ws.cell(POST_DATA_WRITE_ROW, DataFile.TEACHER_NAME_COLUMN),  data_file_temp_ws.cell(row, TEACHER_NAME_COLUMN))
                    copy_cell(post_data_ws.cell(POST_DATA_WRITE_ROW, DataFile.STUDENT_NAME_COLUMN),  data_file_temp_ws.cell(row, STUDENT_NAME_COLUMN))
                    copy_cell(post_data_ws.cell(POST_DATA_WRITE_ROW, DataFile.AVERAGE_SCORE_COLUMN), data_file_temp_ws.cell(row, AVERAGE_SCORE_COLUMN))
                    POST_DATA_WRITE_COLUMN = DataFile.MAX+1
                    for col in range(AVERAGE_SCORE_COLUMN+1, data_file_temp_ws.max_column+1):
                        copy_cell(post_data_ws.cell(POST_DATA_WRITE_ROW, POST_DATA_WRITE_COLUMN), data_file_temp_ws.cell(row, col))
                        POST_DATA_WRITE_COLUMN += 1

            # 필터 범위 재조정
            data_file_ws.auto_filter.ref = "A:" + gcl(AVERAGE_SCORE_COLUMN)
        
        post_data_wb.save("./data/지난 데이터.xlsx")
        data_file_wb.save(f"./data/{config['dataFileName']}.xlsx")

    # 데이터 파일에 새 반 추가
    if len(check_list) != 0:
        gui.q.put("신규 반 추가중...")
        data_file_wb = xl.load_workbook(f"./data/{config['dataFileName']}.xlsx")
        for sheet_name in data_file_wb.sheetnames:
            data_file_ws = data_file_wb[sheet_name]
            for i in range(1, data_file_ws.max_column+1):
                temp = data_file_ws.cell(1, i).value
                if temp == "시간":
                    TEST_TIME_COLUMN = i
                elif temp == "요일":
                    CLASS_WEEKDAY_COLUMN = i
                elif temp == "반":
                    CLASS_NAME_COLUMN = i
                elif temp == "담당":
                    TEACHER_NAME_COLUMN = i
                elif temp == "이름":
                    STUDENT_NAME_COLUMN = i
                elif temp == "학생 평균":
                    AVERAGE_SCORE_COLUMN = i
            
            for i in range(2, data_file_ws.max_row+2):
                if data_file_ws.cell(i, CLASS_NAME_COLUMN).value is None:
                    WRITE_LOCATION = i
                    break
            
            for new_class, new_class_index in unregistered_classes.items():
                if not new_class in update_class: continue
                
                trs = driver.find_element(By.ID, "table_" + str(new_class_index)).find_elements(By.CLASS_NAME, "style12")
                if len(trs) == 0:
                    continue

                class_name = new_class
                test_time = ""
                class_weekday = ""
                teacher_name = ""
                for j in range(2, class_temp_ws.max_row + 1):
                    if class_temp_ws.cell(j, ClassInfo.CLASS_NAME_COLUMN).value == class_name:
                        teacher_name = class_temp_ws.cell(j, ClassInfo.TEACHER_NAME_COLUMN).value
                        class_weekday = class_temp_ws.cell(j, ClassInfo.CLASS_WEEKDAY_COLUMN).value
                        test_time = class_temp_ws.cell(j, ClassInfo.TEST_TIME_COLUMN).value
                        break
                else: continue
                
                # 시험명
                data_file_ws.cell(WRITE_LOCATION, TEST_TIME_COLUMN).value     = test_time
                data_file_ws.cell(WRITE_LOCATION, CLASS_WEEKDAY_COLUMN).value = class_weekday
                data_file_ws.cell(WRITE_LOCATION, CLASS_NAME_COLUMN).value    = class_name
                data_file_ws.cell(WRITE_LOCATION, TEACHER_NAME_COLUMN).value  = teacher_name
                data_file_ws.cell(WRITE_LOCATION, STUDENT_NAME_COLUMN).value  = "날짜"
                WRITE_LOCATION += 1
                
                data_file_ws.cell(WRITE_LOCATION, TEST_TIME_COLUMN).value     = test_time
                data_file_ws.cell(WRITE_LOCATION, CLASS_WEEKDAY_COLUMN).value = class_weekday
                data_file_ws.cell(WRITE_LOCATION, CLASS_NAME_COLUMN).value    = class_name
                data_file_ws.cell(WRITE_LOCATION, TEACHER_NAME_COLUMN).value  = teacher_name
                data_file_ws.cell(WRITE_LOCATION, STUDENT_NAME_COLUMN).value  = "시험명"
                WRITE_LOCATION += 1

                # 학생 루프
                for tr in trs:
                    data_file_ws.cell(WRITE_LOCATION, TEST_TIME_COLUMN).value     = test_time
                    data_file_ws.cell(WRITE_LOCATION, CLASS_WEEKDAY_COLUMN).value = class_weekday
                    data_file_ws.cell(WRITE_LOCATION, CLASS_NAME_COLUMN).value    = class_name
                    data_file_ws.cell(WRITE_LOCATION, TEACHER_NAME_COLUMN).value  = teacher_name
                    data_file_ws.cell(WRITE_LOCATION, STUDENT_NAME_COLUMN).value  = tr.find_element(By.CLASS_NAME, "style9").text
                    WRITE_LOCATION += 1
                
                # 시험별 평균
                data_file_ws.cell(WRITE_LOCATION, TEST_TIME_COLUMN).value     = test_time
                data_file_ws.cell(WRITE_LOCATION, CLASS_WEEKDAY_COLUMN).value = class_weekday
                data_file_ws.cell(WRITE_LOCATION, CLASS_NAME_COLUMN).value    = class_name
                data_file_ws.cell(WRITE_LOCATION, TEACHER_NAME_COLUMN).value  = teacher_name
                data_file_ws.cell(WRITE_LOCATION, STUDENT_NAME_COLUMN).value  = "시험 평균"
                WRITE_LOCATION += 1

                for j in range(1, AVERAGE_SCORE_COLUMN+1):
                    data_file_ws.cell(WRITE_LOCATION-1, j).border = Border(bottom = Side(border_style="medium", color="000000"))

            # 정렬
            for i in range(1, data_file_ws.max_row + 1):
                for j in range(1, data_file_ws.max_column + 1):
                    data_file_ws.cell(i, j).alignment = Alignment(horizontal="center", vertical="center")

            # 필터 범위 재지정
            data_file_ws.auto_filter.ref = "A:" + gcl(AVERAGE_SCORE_COLUMN)
    
    # 변경 사항 저장
    data_file_wb.save(f"./data/{config['dataFileName']}.xlsx")
    
    rescoping_formula()

    class_wb_temp.save("./반 정보.xlsx")
    os.remove("./temp.xlsx")

    gui.q.put("반 업데이트를 완료하였습니다.")
    gui.thread_end_flag = True
    pythoncom.CoUninitialize()

def make_data_form(gui:GUI):
    gui.q.put("데일리테스트 기록 양식 생성 중...")

    ini_wb = xl.Workbook()
    ini_ws = ini_wb[ini_wb.sheetnames[0]]
    ini_ws.title = "데일리테스트 기록 양식"
    ini_ws[gcl(DataForm.CLASS_WEEKDAY_COLUMN)+"1"]     = "요일"
    ini_ws[gcl(DataForm.TEST_TIME_COLUMN)+"1"]         = "시간"
    ini_ws[gcl(DataForm.CLASS_NAME_COLUMN)+"1"]        = "반"
    ini_ws[gcl(DataForm.STUDENT_NAME_COLUMN)+"1"]      = "이름"
    ini_ws[gcl(DataForm.TEACHER_NAME_COLUMN)+"1"]      = "담당T"
    ini_ws[gcl(DataForm.DAILYTEST_NAME_COLUMN)+"1"]    = "시험명"
    ini_ws[gcl(DataForm.DAILYTEST_SCORE_COLUMN)+"1"]   = "점수"
    ini_ws[gcl(DataForm.DAILYTEST_AVERAGE_COLUMN)+"1"] = "평균"
    ini_ws[gcl(DataForm.MOCKTEST_NAME_COLUMN)+"1"]     = "모의고사 시험명"
    ini_ws[gcl(DataForm.MOCKTEST_SCORE_COLUMN)+"1"]    = "모의고사 점수"
    ini_ws[gcl(DataForm.MOCKTEST_AVERAGE_COLUMN)+"1"]  = "모의고사 평균"
    ini_ws[gcl(DataForm.MAKEUP_TEST_CHECK_COLUMN)+"1"] = "재시험 응시 여부"
    ini_ws["Y1"] = "X"
    ini_ws["Z1"] = "x"
    ini_ws.column_dimensions.group("Y", "Z", hidden=True)
    ini_ws.auto_filter.ref = "A:"+gcl(DataForm.TEST_TIME_COLUMN)
    
    for col in range(1, DataForm.MAX+1):
        ini_ws.cell(1, col).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        ini_ws.cell(1, col).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    
    class_wb = xl.load_workbook("./반 정보.xlsx")
    try:
        class_ws = class_wb["반 정보"]
    except:
        gui.q.put(r"'반 정보.xlsx'의 시트명을")
        gui.q.put(r"'반 정보'로 변경해 주세요.")
        return

    options = webdriver.ChromeOptions()
    options.add_argument("headless")
    driver = webdriver.Chrome(service = service, options = options)

    # 아이소식 접속
    driver.get(config["url"])
    table_names = driver.find_elements(By.CLASS_NAME, "style1")

    #반 루프
    for i in range(3, len(table_names)):
        trs = driver.find_element(By.ID, "table_" + str(i)).find_elements(By.CLASS_NAME, "style12")
        WRITE_LOCATION = start = ini_ws.max_row + 1

        class_name = table_names[i].text.rstrip()
        teacher_name = ""
        date = ""
        test_time = ""
        is_class_exist = False

        for row in range(2, class_ws.max_row + 1):
            if class_ws.cell(row, ClassInfo.CLASS_NAME_COLUMN).value == class_name:
                teacher_name = class_ws.cell(row, ClassInfo.TEACHER_NAME_COLUMN).value
                date         = class_ws.cell(row, ClassInfo.CLASS_WEEKDAY_COLUMN).value
                test_time    = class_ws.cell(row, ClassInfo.TEST_TIME_COLUMN).value
                is_class_exist = True
        if not is_class_exist or len(trs) == 0:
            continue
        ini_ws.cell(WRITE_LOCATION, DataForm.CLASS_NAME_COLUMN).value   = class_name
        ini_ws.cell(WRITE_LOCATION, DataForm.TEACHER_NAME_COLUMN).value = teacher_name

        #학생 루프
        for tr in trs:
            ini_ws.cell(WRITE_LOCATION, DataForm.CLASS_WEEKDAY_COLUMN).value = date
            ini_ws.cell(WRITE_LOCATION, DataForm.TEST_TIME_COLUMN).value     = test_time
            ini_ws.cell(WRITE_LOCATION, DataForm.STUDENT_NAME_COLUMN).value  = tr.find_element(By.CLASS_NAME, "style9").text
            dv = DataValidation(type="list", formula1="=Y1:Z1", showDropDown=True, allow_blank=True, showErrorMessage=True)
            dv.error = "이 셀의 값은 'x' 또는 'X'이어야 합니다."
            ini_ws.add_data_validation(dv)
            dv.add(ini_ws.cell(WRITE_LOCATION,DataForm.MAKEUP_TEST_CHECK_COLUMN))
            WRITE_LOCATION = ini_ws.max_row + 1
        
        end = WRITE_LOCATION - 1

        # 시험 평균
        ini_ws.cell(start, DataForm.DAILYTEST_AVERAGE_COLUMN).value = f"=ROUND(AVERAGE({gcl(DataForm.DAILYTEST_SCORE_COLUMN)+str(start)}:{gcl(DataForm.DAILYTEST_SCORE_COLUMN)+str(end)}), 0)"
        # 모의고사 평균
        ini_ws.cell(start, DataForm.MOCKTEST_AVERAGE_COLUMN).value = f"=ROUND(AVERAGE({gcl(DataForm.MOCKTEST_SCORE_COLUMN)+str(start)}:{gcl(DataForm.MOCKTEST_SCORE_COLUMN)+str(end)}), 0)"
        
        # 정렬 및 테두리
        for row in range(start, end + 1):
            for col in range(1, DataForm.MAX+1):
                ini_ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
                ini_ws.cell(row, col).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        
        # 셀 병합
        if start < end:
            ini_ws.merge_cells(f"{gcl(DataForm.CLASS_NAME_COLUMN)+str(start)}:{gcl(DataForm.CLASS_NAME_COLUMN)+str(end)}")
            ini_ws.merge_cells(f"{gcl(DataForm.TEACHER_NAME_COLUMN)+str(start)}:{gcl(DataForm.TEACHER_NAME_COLUMN)+str(end)}")
            ini_ws.merge_cells(f"{gcl(DataForm.DAILYTEST_NAME_COLUMN)+str(start)}:{gcl(DataForm.DAILYTEST_NAME_COLUMN)+str(end)}")
            ini_ws.merge_cells(f"{gcl(DataForm.DAILYTEST_AVERAGE_COLUMN)+str(start)}:{gcl(DataForm.DAILYTEST_AVERAGE_COLUMN)+str(end)}")
            ini_ws.merge_cells(f"{gcl(DataForm.MOCKTEST_NAME_COLUMN)+str(start)}:{gcl(DataForm.MOCKTEST_NAME_COLUMN)+str(end)}")
            ini_ws.merge_cells(f"{gcl(DataForm.MOCKTEST_AVERAGE_COLUMN)+str(start)}:{gcl(DataForm.MOCKTEST_AVERAGE_COLUMN)+str(end)}")
        
    ini_ws.protection.sheet         = True
    ini_ws.protection.autoFilter    = False
    ini_ws.protection.formatColumns = False
    for row in range(2, ini_ws.max_row + 1):
        ini_ws.cell(row, DataForm.CLASS_NAME_COLUMN).alignment         = Alignment(horizontal="center", vertical="center", wrapText=True)
        ini_ws.cell(row, DataForm.DAILYTEST_NAME_COLUMN).alignment     = Alignment(horizontal="center", vertical="center", wrapText=True)
        ini_ws.cell(row, DataForm.MOCKTEST_NAME_COLUMN).alignment      = Alignment(horizontal="center", vertical="center", wrapText=True)
        ini_ws.cell(row, DataForm.DAILYTEST_NAME_COLUMN).protection    = Protection(locked=False)
        ini_ws.cell(row, DataForm.DAILYTEST_SCORE_COLUMN).protection   = Protection(locked=False)
        ini_ws.cell(row, DataForm.MOCKTEST_NAME_COLUMN).protection     = Protection(locked=False)
        ini_ws.cell(row, DataForm.MOCKTEST_SCORE_COLUMN).protection    = Protection(locked=False)
        ini_ws.cell(row, DataForm.MAKEUP_TEST_CHECK_COLUMN).protection = Protection(locked=False)

    if os.path.isfile(f"./데일리테스트 기록 양식({datetime.today().strftime('%m.%d')}).xlsx"):
        i = 1
        while True:
            if not os.path.isfile(f"./데일리테스트 기록 양식({datetime.today().strftime('%m.%d')})({str(i)}).xlsx"):
                ini_wb.save(f"./데일리테스트 기록 양식({datetime.today().strftime('%m.%d')})({str(i)}).xlsx")
                break
            i += 1
    else:
        ini_wb.save(f"./데일리테스트 기록 양식({datetime.today().strftime('%m.%d')}).xlsx")
    gui.q.put("데일리테스트 기록 양식 생성을 완료했습니다.")
    gui.thread_end_flag = True

def save_data(gui:GUI, filepath:str, makeup_test_date:dict):
    form_wb = xl.load_workbook(filepath, data_only=True)
    form_ws = form_wb["데일리테스트 기록 양식"]

    # 올바른 양식이 아닙니다.
    if not data_validation(gui, form_ws):
        gui.q.put("데이터 저장이 중단되었습니다.")
        return
    
    # 학생 정보 열기
    student_wb = xl.load_workbook("./학생 정보.xlsx")
    try:
        student_ws = student_wb["학생 정보"]
    except:
        gui.q.put(r"'학생 정보.xlsx'의 시트명을")
        gui.q.put(r"'학생 정보'로 변경해 주세요.")
        return

    # 재시험 명단 열기
    if not os.path.isfile("./data/재시험 명단.xlsx"):
        gui.q.put("재시험 명단 파일 생성 중...")

        ini_wb = xl.Workbook()
        ini_ws = ini_wb[ini_wb.sheetnames[0]]
        ini_ws.title = "재시험 명단"
        ini_ws[gcl(MakeupTestList.TEST_DATE_COLUMN)+"1"]          = "응시일"
        ini_ws[gcl(MakeupTestList.CLASS_NAME_COLUMN)+"1"]         = "반"
        ini_ws[gcl(MakeupTestList.TEACHER_NAME_COLUMN)+"1"]       = "담당T"
        ini_ws[gcl(MakeupTestList.STUDENT_NAME_COLUMN)+"1"]       = "이름"
        ini_ws[gcl(MakeupTestList.TEST_NAME_COLUMN)+"1"]          = "시험명"
        ini_ws[gcl(MakeupTestList.TEST_SCORE_COLUMN)+"1"]         = "시험 점수"
        ini_ws[gcl(MakeupTestList.MAKEUPTEST_WEEKDAY_COLUMN)+"1"] = "재시 요일"
        ini_ws[gcl(MakeupTestList.MAKEUPTEST_TIME_COLUMN)+"1"]    = "재시 시간"
        ini_ws[gcl(MakeupTestList.MAKEUPTEST_DATE_COLUMN)+"1"]    = "재시 날짜"
        ini_ws[gcl(MakeupTestList.MAKEUPTEST_SCORE_COLUMN)+"1"]   = "재시 점수"
        ini_ws[gcl(MakeupTestList.ETC_COLUMN)+"1"] = "비고"
        ini_ws.auto_filter.ref = "A:"+gcl(MakeupTestList.MAX)
        ini_wb.save("./data/재시험 명단.xlsx")
    makeup_list_wb = xl.load_workbook("./data/재시험 명단.xlsx")
    try:
        makeup_list_ws = makeup_list_wb["재시험 명단"]
    except:
        gui.q.put(r"'재시험 명단.xlsx'의 시트명을")
        gui.q.put(r"'재시험 명단'으로 변경해 주세요.")
        return

    # 백업 생성
    data_file_wb = xl.load_workbook(f"./data/{config['dataFileName']}.xlsx")
    data_file_wb.save(f"./data/backup/{config['dataFileName']}({datetime.today().strftime('%Y%m%d')}).xlsx")
    
    gui.q.put("데이터 저장 및 재시험 명단 작성 중...")

    # 재시험 명단 작성 시작 위치 검색
    for row in range(makeup_list_ws.max_row+1, 1, -1):
        if makeup_list_ws.cell(row-1, MakeupTestList.TEST_DATE_COLUMN).value is not None:
            MAKEUP_TEST_RANGE = MAKEUP_TEST_WRITE_ROW = row
            break
    
    for sheet_name in data_file_wb.sheetnames:
        data_file_ws = data_file_wb[sheet_name]
        if sheet_name == "데일리테스트":
            TEST_NAME_COLUMN    = DataForm.DAILYTEST_NAME_COLUMN
            TEST_SCORE_COLUMN   = DataForm.DAILYTEST_SCORE_COLUMN
            TEST_AVERAGE_COLUMN = DataForm.DAILYTEST_AVERAGE_COLUMN
        elif sheet_name == "모의고사":
            TEST_NAME_COLUMN    = DataForm.MOCKTEST_NAME_COLUMN
            TEST_SCORE_COLUMN   = DataForm.MOCKTEST_SCORE_COLUMN
            TEST_AVERAGE_COLUMN = DataForm.MOCKTEST_AVERAGE_COLUMN
        else:
            # error
            return

        # 동적 열 탐색
        for i in range(1, data_file_ws.max_column+1):
            if data_file_ws.cell(1, i).value == "반":
                CLASS_NAME_COLUMN = i
                break
        else:
            # error
            return
        for i in range(1, data_file_ws.max_column+1):
            if data_file_ws.cell(1, i).value == "이름":
                STUDENT_NAME_COLUMN = i
                break
        else:
            # error
            return
        for i in range(1, data_file_ws.max_column+1):
            if data_file_ws.cell(1, i).value == "학생 평균":
                AVERAGE_SCORE_COLUMN = i
                break
        else:
            # error
            return
        
        for i in range(2, form_ws.max_row+1): # 데일리데이터 기록 양식 루프
            # 반 필터링
            if (form_ws.cell(i, DataForm.CLASS_NAME_COLUMN).value is not None) and (form_ws.cell(i, TEST_NAME_COLUMN).value is not None):
                class_name   = form_ws.cell(i, DataForm.CLASS_NAME_COLUMN).value
                test_name    = form_ws.cell(i, TEST_NAME_COLUMN).value
                teacher_name = form_ws.cell(i, DataForm.TEACHER_NAME_COLUMN).value
                test_average = form_ws.cell(i, TEST_AVERAGE_COLUMN).value
                
                #반 시작 찾기
                for row in range(2, data_file_ws.max_row+1):
                    if data_file_ws.cell(row, CLASS_NAME_COLUMN).value == class_name:
                        CLASS_START = row
                        break
                else:
                    # error
                    return
                # 반 끝 찾기
                for row in range(CLASS_START, data_file_ws.max_row+1):
                    if data_file_ws.cell(row, STUDENT_NAME_COLUMN).value == "시험 평균":
                        CLASS_END = row
                        break
                
                # 데일리테스트 작성 열 위치 찾기
                for col in range(AVERAGE_SCORE_COLUMN+1, data_file_ws.max_column+2): 
                    if data_file_ws.cell(CLASS_START, col).value is None:
                        WRITE_COLUMN = col
                        break
                    if str(data_file_ws.cell(CLASS_START, col).value) == DATE.today().strftime("%Y-%m-%d 00:00:00"):
                        WRITE_COLUMN = col
                        break
                
                # 입력 틀 작성
                AVERAGE_FORMULA = f"=ROUND(AVERAGE({gcl(WRITE_COLUMN)+str(CLASS_START + 2)}:{gcl(WRITE_COLUMN)+str(CLASS_END - 1)}), 0)"
                data_file_ws.column_dimensions[gcl(WRITE_COLUMN)].width    = 14
                data_file_ws.cell(CLASS_START, WRITE_COLUMN).value         = DATE.today()
                data_file_ws.cell(CLASS_START, WRITE_COLUMN).number_format = "yyyy.mm.dd(aaa)"
                data_file_ws.cell(CLASS_START, WRITE_COLUMN).alignment     = Alignment(horizontal="center", vertical="center")

                data_file_ws.cell(CLASS_START + 1, WRITE_COLUMN).value     = test_name
                data_file_ws.cell(CLASS_START + 1, WRITE_COLUMN).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

                data_file_ws.cell(CLASS_END, WRITE_COLUMN).value           = AVERAGE_FORMULA
                data_file_ws.cell(CLASS_END, WRITE_COLUMN).font            = Font(bold=True)
                data_file_ws.cell(CLASS_END, WRITE_COLUMN).alignment       = Alignment(horizontal="center", vertical="center")
                data_file_ws.cell(CLASS_END, WRITE_COLUMN).border          = Border(bottom=Side(border_style="medium", color="000000"))
                
                if type(test_average) == int:
                    if test_average < 60:
                        data_file_ws.cell(CLASS_END, WRITE_COLUMN).fill = PatternFill(fill_type="solid", fgColor=Color("EC7E31"))
                    elif test_average < 70:
                        data_file_ws.cell(CLASS_END, WRITE_COLUMN).fill = PatternFill(fill_type="solid", fgColor=Color("F5AF85"))
                    elif test_average < 80:
                        data_file_ws.cell(CLASS_END, WRITE_COLUMN).fill = PatternFill(fill_type="solid", fgColor=Color("FCE4D6"))
                    else:
                        data_file_ws.cell(CLASS_END, WRITE_COLUMN).fill = PatternFill(fill_type="solid", fgColor=Color("DDEBF7"))
            
            test_score = form_ws.cell(i, TEST_SCORE_COLUMN).value
            if test_score is None:
                continue # 점수 없으면 미응시 처리

            # 학생 찾기
            for row in range(CLASS_START + 2, CLASS_END):
                if data_file_ws.cell(row, STUDENT_NAME_COLUMN).value == form_ws.cell(i, DataForm.STUDENT_NAME_COLUMN).value:
                    data_file_ws.cell(row, WRITE_COLUMN).value = test_score
                    if type(test_score) == int or type(test_score) == float:
                        if test_score < 60:
                            data_file_ws.cell(row, WRITE_COLUMN).fill = PatternFill(fill_type="solid", fgColor=Color("EC7E31"))
                        elif test_score < 70:
                            data_file_ws.cell(row, WRITE_COLUMN).fill = PatternFill(fill_type="solid", fgColor=Color("F5AF85"))
                        elif test_score < 80:
                            data_file_ws.cell(row, WRITE_COLUMN).fill = PatternFill(fill_type="solid", fgColor=Color("FCE4D6"))
                    data_file_ws.cell(row, WRITE_COLUMN).alignment = Alignment(horizontal="center", vertical="center")
                    break
            else:
                gui.q.put(f"{class_name} 반에 {form_ws.cell(i, DataForm.STUDENT_NAME_COLUMN).value} 학생이 존재하지 않습니다.")
            
            # 재시험 작성
            if (type(test_score) == int or type(test_score) == float) and test_score < 80 and form_ws.cell(i, DataForm.MAKEUP_TEST_CHECK_COLUMN).value not in ("x", "X"):
                check = makeup_list_ws.max_row
                # 재시험 중복 작성 검사
                duplicated = False
                while check >= 1:
                    try:
                        if makeup_list_ws.cell(check, MakeupTestList.TEST_DATE_COLUMN).value is None:
                            check -= 1
                            continue
                        elif str(makeup_list_ws.cell(check, MakeupTestList.TEST_DATE_COLUMN).value) == DATE.today().strftime("%Y-%m-%d 00:00:00"):
                            if makeup_list_ws.cell(check, MakeupTestList.STUDENT_NAME_COLUMN).value == form_ws.cell(i, DataForm.STUDENT_NAME_COLUMN).value:
                                if makeup_list_ws.cell(check, MakeupTestList.CLASS_NAME_COLUMN).value == class_name:
                                    duplicated = True
                                    break
                        elif str(makeup_list_ws.cell(check, MakeupTestList.TEST_DATE_COLUMN).value) == (DATE.today()+timedelta(days=-1)).strftime("%Y-%m-%d 00:00:00"):
                            break
                    except:
                        pass
                    check -= 1
                    
                if duplicated: continue
                
                # 학생 재시험 정보 검색
                for row in range(2, student_ws.max_row+1):
                    if student_ws.cell(row, StudentInfo.STUDENT_NAME_COLUMN).value == form_ws.cell(i, DataForm.STUDENT_NAME_COLUMN).value:
                        makeup_test_weekday = student_ws.cell(row, StudentInfo.MAKEUPTEST_WEEKDAY_COLUMN).value
                        makeup_test_time    = student_ws.cell(row, StudentInfo.MAKEUPTEST_TIME_COLUMN).value
                        new_student         = student_ws.cell(row, StudentInfo.NEW_STUDENT_CHECK_COLUMN).value
                        break
                else:
                    makeup_test_weekday = None
                    makeup_test_time    = None
                    new_student         = None

                makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.TEST_DATE_COLUMN).value    = DATE.today()
                makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.CLASS_NAME_COLUMN).value   = class_name
                makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.TEACHER_NAME_COLUMN).value = teacher_name
                makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.STUDENT_NAME_COLUMN).value = form_ws.cell(i, DataForm.STUDENT_NAME_COLUMN).value
                makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.TEST_NAME_COLUMN).value    = test_name
                makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.TEST_SCORE_COLUMN).value   = test_score
                if (new_student is not None) and (new_student == "N"):
                    makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.STUDENT_NAME_COLUMN).fill = PatternFill(fill_type="solid", fgColor=Color("FFFF00"))
                if makeup_test_weekday is not None:
                    makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.MAKEUPTEST_WEEKDAY_COLUMN).value = makeup_test_weekday
                    date_list = makeup_test_weekday.split("/")
                    result    = makeup_test_date[date_list[0].replace(" ", "")]
                    for d in date_list:
                        if result > makeup_test_date[d.replace(" ", "")]:
                            result = makeup_test_date[d.replace(" ", "")]
                    if makeup_test_time is not None:
                        makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.MAKEUPTEST_TIME_COLUMN).value = makeup_test_time
                    makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.MAKEUPTEST_DATE_COLUMN).value         = result
                    makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.MAKEUPTEST_DATE_COLUMN).number_format = "mm월 dd일(aaa)"
                MAKEUP_TEST_WRITE_ROW += 1

    # 정렬 및 테두리
    for row in range(MAKEUP_TEST_RANGE, makeup_list_ws.max_row+1):
        if makeup_list_ws.cell(row, 1).value is None: break
        for col in range(1, makeup_list_ws.max_column + 1):
            makeup_list_ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
            makeup_list_ws.cell(row, col).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    data_file_ws = data_file_wb["데일리테스트"]
    data_file_wb.save(f"./data/{config['dataFileName']}.xlsx")
    makeup_list_wb.save("./data/재시험 명단.xlsx")

    # 조건부 서식 수식 로딩
    pythoncom.CoInitialize()
    excel = win32com.client.gencache.EnsureDispatch("Excel.Application")

    try:
        wb = excel.Workbooks.Open(f"{os.getcwd()}\\data\\{config['dataFileName']}.xlsx")
    except:
        gui.q.put("데이터 파일 창을 끄고 다시 실행해 주세요.")
        return
    wb.Save()
    wb.Close()

    data_file_wb       = xl.load_workbook(f"./data/{config['dataFileName']}.xlsx")
    data_file_color_wb = xl.load_workbook(f"./data/{config['dataFileName']}.xlsx", data_only=True)

    for sheet_name in data_file_wb.sheetnames:
        data_file_ws       = data_file_wb[sheet_name]
        data_file_color_ws = data_file_color_wb[sheet_name]

        for col in range(1, data_file_ws.max_column):
            if data_file_ws.cell(1, col).value == "이름":
                STUDENT_NAME_COLUMN = col
                break
        for col in range(1, data_file_ws.max_column):
            if data_file_ws.cell(1, col).value == "학생 평균":
                AVERAGE_SCORE_COLUMN = col
                break
        for row in range(2, data_file_color_ws.max_row+1):
            # 학생 별 평균 점수에 대한 조건부 서식
            student_average      = data_file_color_ws.cell(row, AVERAGE_SCORE_COLUMN).value
            student_average_cell = data_file_ws.cell(row, AVERAGE_SCORE_COLUMN)
            student_name_cell    = data_file_ws.cell(row, STUDENT_NAME_COLUMN)
            if type(student_average) == int:
                if student_average < 60:
                    student_average_cell.fill = PatternFill(fill_type="solid", fgColor=Color("EC7E31"))
                elif student_average < 70:
                    student_average_cell.fill = PatternFill(fill_type="solid", fgColor=Color("F5AF85"))
                elif student_average < 80:
                    student_average_cell.fill = PatternFill(fill_type="solid", fgColor=Color("FCE4D6"))
                elif student_name_cell.value == "시험 평균":
                    student_average_cell.fill = PatternFill(fill_type="solid", fgColor=Color("DDEBF7"))
                else:
                    student_average_cell.fill = PatternFill(fill_type="solid", fgColor=Color("E2EFDA"))
            # 신규생 하이라이트
            student_name_cell.fill = PatternFill(fill_type=None, fgColor=Color("00FFFFFF"))
            if (student_name_cell.value is not None) and student_name_cell.value not in ("날짜", "시험명", "시험 평균"):
                for row in range(2, student_ws.max_row+1):
                    if (student_name_cell.value == student_ws.cell(row, StudentInfo.STUDENT_NAME_COLUMN).value) and (student_ws.cell(row, StudentInfo.NEW_STUDENT_CHECK_COLUMN).value == "N"):
                        student_name_cell.fill = PatternFill(fill_type="solid", fgColor=Color("FFFF00"))
                        break
    
    data_file_ws = data_file_wb["데일리테스트"]
    data_file_wb.save(f"./data/{config['dataFileName']}.xlsx")

    gui.q.put("데이터 저장을 완료했습니다.")
    # excel.Visible = True
    # wb = excel.Workbooks.Open(f"{os.getcwd()}\\data\\{config['dataFileName']}.xlsx")
    pythoncom.CoUninitialize()
    gui.thread_end_flag = True

def send_message(gui:GUI, filepath:str, makeup_test_date:dict):
    form_wb = xl.load_workbook(filepath, data_only=True)
    form_ws = form_wb["데일리테스트 기록 양식"]

    options = webdriver.ChromeOptions()
    options.add_experimental_option("detach", True)
    driver = webdriver.Chrome(service=service, options=options)

    student_wb = xl.load_workbook("./학생 정보.xlsx")
    try:
        student_ws = student_wb["학생 정보"]
    except:
        gui.q.put("\"학생 정보.xlsx\"의 시트명을")
        gui.q.put("\"학생 정보\"로 변경해 주세요.")
        return
    
    # 아이소식 접속
    driver.get(config["url"])
    driver.execute_script(f"arguments[0].value = '{config['dailyTest']}'", driver.find_element(By.XPATH, '//*[@id="ctitle"]'))
    driver.find_element(By.XPATH, '//*[@id="ctitle"]').send_keys(' \b')

    driver.execute_script(f"window.open('{config['url']}')")
    driver.switch_to.window(driver.window_handles[1])
    driver.execute_script(f"arguments[0].value = '{config['makeupTest']}'", driver.find_element(By.XPATH, '//*[@id="ctitle"]'))
    driver.find_element(By.XPATH, '//*[@id="ctitle"]').send_keys(' \b')

    driver.execute_script(f"window.open('{config['url']}')")
    driver.switch_to.window(driver.window_handles[2])
    driver.execute_script(f"arguments[0].value = '{config['makeupTestDate']}'", driver.find_element(By.XPATH, '//*[@id="ctitle"]'))
    driver.find_element(By.XPATH, '//*[@id="ctitle"]').send_keys(' \b')

    driver.switch_to.window(driver.window_handles[0])
    tables = driver.find_elements(By.CLASS_NAME, "style1")
    table_names = [table.text for table in tables]
    class_search = 0

    gui.q.put("메시지 작성 중...")
    for i in range(2, form_ws.max_row+1):
        # 반 필터링
        if form_ws.cell(i, DataForm.CLASS_NAME_COLUMN).value is not None:
            class_name         = form_ws.cell(i, DataForm.CLASS_NAME_COLUMN).value
            daily_test_name    = form_ws.cell(i, DataForm.DAILYTEST_NAME_COLUMN).value
            mock_test_name     = form_ws.cell(i, DataForm.MOCKTEST_NAME_COLUMN).value
            daily_test_average = form_ws.cell(i, DataForm.DAILYTEST_AVERAGE_COLUMN).value
            mock_test_average  = form_ws.cell(i, DataForm.MOCKTEST_AVERAGE_COLUMN).value

            # 반 전체가 시험을 응시하지 않은 경우
            if daily_test_name is None and mock_test_name is None:
                keep_continue = True
                continue

            keep_continue = False
            student_search = 0
            
            # 테이블 인덱스 검색
            for idx in range(class_search, len(table_names)):
                if class_name == table_names[idx]:
                    class_index = idx
                    class_search = idx+1
                    break
                
        
        # 반 전체가 시험을 응시하지 않은 경우
        if keep_continue: continue

        student_name     = form_ws.cell(i, DataForm.STUDENT_NAME_COLUMN).value
        daily_test_score = form_ws.cell(i, DataForm.DAILYTEST_SCORE_COLUMN).value
        mock_test_score  = form_ws.cell(i, DataForm.MOCKTEST_SCORE_COLUMN).value

        # 시험 미응시 시 건너뛰기
        if daily_test_score is not None:
            test_name    = daily_test_name
            test_score   = daily_test_score
            test_average = daily_test_average
        elif mock_test_score is not None:
            test_name    = mock_test_name
            test_score   = mock_test_score
            test_average = mock_test_average
        else:
            continue

        if type(test_score) != int and type(test_score) != float:
            continue

        # 시험 결과 메시지 작성
        driver.switch_to.window(driver.window_handles[0])
        trs = driver.find_element(By.ID, f"table_{str(class_index)}").find_elements(By.CLASS_NAME, "style12")
        for idx in range(student_search, len(trs)):
            if trs[idx].find_element(By.CLASS_NAME, "style9").text == student_name:
                student_index = idx
                student_search = idx+1
                break

        tds = trs[student_index].find_elements(By.TAG_NAME, "td")
        driver.execute_script(f"arguments[0].value = '{test_name}'",  tds[0].find_element(By.TAG_NAME, "input"))
        driver.execute_script(f"arguments[0].value = '{test_score}'", tds[1].find_element(By.TAG_NAME, "input"))
        tds[2].find_element(By.TAG_NAME, "input").send_keys(test_average)

        # 재시험 메시지 작성
        if (test_score < 80) and form_ws.cell(i, DataForm.MAKEUP_TEST_CHECK_COLUMN).value not in ("x", "X"):
            for row in range(2, student_ws.max_row+1):
                if student_ws.cell(row, StudentInfo.STUDENT_NAME_COLUMN).value == student_name:
                    makeup_test_weekday = student_ws.cell(row, StudentInfo.MAKEUPTEST_WEEKDAY_COLUMN).value
                    makeup_test_time    = student_ws.cell(row, StudentInfo.MAKEUPTEST_TIME_COLUMN).value
                    break
            else:
                makeup_test_weekday = None
                makeup_test_time    = None
            
            if makeup_test_weekday is None:
                driver.switch_to.window(driver.window_handles[1])
                trs = driver.find_element(By.ID, "table_" + str(class_index)).find_elements(By.CLASS_NAME, "style12")
                tds = trs[student_index].find_elements(By.TAG_NAME, "td")
                driver.execute_script(f"arguments[0].value = '{test_name}'", tds[0].find_element(By.TAG_NAME, "input"))
                tds[1].find_element(By.TAG_NAME, "input").send_keys(' \b')
            else:
                weekday_list    = makeup_test_weekday.split("/")
                calculated_date = makeup_test_date[weekday_list[0].replace(" ", "")]
                time_index      = 0
                for tmp_idx in range(len(weekday_list)):
                    if calculated_date > makeup_test_date[weekday_list[tmp_idx].replace(" ", "")]:
                        calculated_date = makeup_test_date[weekday_list[tmp_idx].replace(" ", "")]
                        time_index = tmp_idx
                driver.switch_to.window(driver.window_handles[2])
                trs = driver.find_element(By.ID, "table_" + str(class_index)).find_elements(By.CLASS_NAME, "style12")
                tds = trs[student_index].find_elements(By.TAG_NAME, "td")
                driver.execute_script(f"arguments[0].value = '{test_name}'", tds[0].find_element(By.TAG_NAME, "input"))
                try:
                    if makeup_test_time is not None:
                        if "/" in str(makeup_test_time):
                            driver.execute_script(f"arguments[0].value = '{calculated_date.strftime('%m월 %d일'.encode('unicode-escape').decode()).encode().decode('unicode-escape')} {str(makeup_test_time).split('/')[time_index]}시'", tds[1].find_element(By.TAG_NAME, "input"))
                        else:
                            driver.execute_script(f"arguments[0].value = '{calculated_date.strftime('%m월 %d일'.encode('unicode-escape').decode()).encode().decode('unicode-escape')} {str(makeup_test_time)}시'", tds[1].find_element(By.TAG_NAME, "input"))
                    else:
                        driver.execute_script(f"arguments[0].value = '{calculated_date.strftime('%m월 %d일'.encode('unicode-escape').decode()).encode().decode('unicode-escape')}'", tds[1].find_element(By.TAG_NAME, "input"))
                except:
                    gui.q.put(f"{student_name}의 재시험 일정을 요일별 시간으로 설정하거나")
                    gui.q.put("하나의 시간으로 통일해 주세요.")
                    gui.q.put("중단되었습니다.")
                    driver.quit()
                    gui.thread_end_flag = True    
                    return
                tds[2].find_element(By.TAG_NAME, "input").send_keys(' \b')

    gui.q.put("메시지 입력을 완료했습니다.")
    gui.q.put("메시지 확인 후 전송해주세요.")
    gui.thread_end_flag = True

def apply_color(gui:GUI):
    student_wb = xl.load_workbook("./학생 정보.xlsx")
    try:
        student_ws = student_wb["학생 정보"]
    except:
        gui.q.put(r"'학생 정보.xlsx'의 시트명을")
        gui.q.put(r"'학생 정보'로 변경해 주세요.")
        return
    
    pythoncom.CoInitialize()
    excel = win32com.client.Dispatch("Excel.Application")

    try:
        wb = excel.Workbooks.Open(f"{os.getcwd()}\\data\\{config['dataFileName']}.xlsx")
    except:
        gui.q.put("데이터 파일 창을 끄고 다시 실행해 주세요.")
        return
    wb.Save()
    wb.Close()

    gui.q.put("조건부 서식 적용중...")

    data_file_wb       = xl.load_workbook(f"./data/{config['dataFileName']}.xlsx")
    data_file_color_wb = xl.load_workbook(f"./data/{config['dataFileName']}.xlsx", data_only=True)
    
    for sheet_name in data_file_wb.sheetnames:
        data_file_ws       = data_file_wb[sheet_name]
        data_file_color_ws = data_file_color_wb[sheet_name]

        for col in range(1, data_file_ws.max_column):
            if data_file_ws.cell(1, col).value == "이름":
                STUDENT_NAME_COLUMN = col
                break
        for col in range(1, data_file_ws.max_column):
            if data_file_ws.cell(1, col).value == "학생 평균":
                AVERAGE_SCORE_COLUMN = col
                break
        
        for row in range(2, data_file_color_ws.max_row+1):
            if data_file_color_ws.cell(row, STUDENT_NAME_COLUMN).value is None:
                break
            for col in range(1, data_file_color_ws.max_column+1):
                data_file_ws.column_dimensions[gcl(col)].width = 14
                if data_file_ws.cell(row, STUDENT_NAME_COLUMN).value == "시험 평균" and data_file_ws.cell(row, col).value is not None:
                    data_file_ws.cell(row, col).border = Border(bottom=Side(border_style="medium", color="000000"))
                if col > AVERAGE_SCORE_COLUMN:    
                    if data_file_ws.cell(row, STUDENT_NAME_COLUMN).value == "날짜" and data_file_ws.cell(row, col).value is not None:
                        data_file_ws.cell(row, col).border = Border(top=Side(border_style="medium", color="000000"))
                    if type(data_file_color_ws.cell(row, col).value) == int:
                        if data_file_color_ws.cell(row, col).value < 60:
                            data_file_ws.cell(row, col).fill = PatternFill(fill_type="solid", fgColor=Color("EC7E31"))
                        elif data_file_color_ws.cell(row, col).value < 70:
                            data_file_ws.cell(row, col).fill = PatternFill(fill_type="solid", fgColor=Color("F5AF85"))
                        elif data_file_color_ws.cell(row, col).value < 80:
                            data_file_ws.cell(row, col).fill = PatternFill(fill_type="solid", fgColor=Color("FCE4D6"))
                        elif data_file_color_ws.cell(row, STUDENT_NAME_COLUMN).value == "시험 평균":
                            data_file_ws.cell(row, col).fill = PatternFill(fill_type="solid", fgColor=Color("DDEBF7"))
                        else:
                            data_file_ws.cell(row, col).fill = PatternFill(fill_type=None, fgColor=Color("00FFFFFF"))
                    else:
                        data_file_ws.cell(row, col).fill = PatternFill(fill_type=None, fgColor=Color("00FFFFFF"))
                    if data_file_color_ws.cell(row, STUDENT_NAME_COLUMN).value == "시험 평균":
                        data_file_ws.cell(row, col).font = Font(bold=True)

            # 학생별 평균 조건부 서식
            data_file_ws.cell(row, AVERAGE_SCORE_COLUMN).font = Font(bold=True)
            if type(data_file_color_ws.cell(row, AVERAGE_SCORE_COLUMN).value) == int:
                if data_file_color_ws.cell(row, AVERAGE_SCORE_COLUMN).value < 60:
                    data_file_ws.cell(row, AVERAGE_SCORE_COLUMN).fill = PatternFill(fill_type="solid", fgColor=Color("EC7E31"))
                elif data_file_color_ws.cell(row, AVERAGE_SCORE_COLUMN).value < 70:
                    data_file_ws.cell(row, AVERAGE_SCORE_COLUMN).fill = PatternFill(fill_type="solid", fgColor=Color("F5AF85"))
                elif data_file_color_ws.cell(row, AVERAGE_SCORE_COLUMN).value < 80:
                    data_file_ws.cell(row, AVERAGE_SCORE_COLUMN).fill = PatternFill(fill_type="solid", fgColor=Color("FCE4D6"))
                elif data_file_color_ws.cell(row, STUDENT_NAME_COLUMN).value == "시험 평균":
                    data_file_ws.cell(row, AVERAGE_SCORE_COLUMN).fill = PatternFill(fill_type="solid", fgColor=Color("DDEBF7"))
                else:
                    data_file_ws.cell(row, AVERAGE_SCORE_COLUMN).fill = PatternFill(fill_type="solid", fgColor=Color("E2EFDA"))
            studnet_name_cell = data_file_ws.cell(row, STUDENT_NAME_COLUMN)
            studnet_name_cell.fill = PatternFill(fill_type=None, fgColor=Color("00FFFFFF"))
            if (studnet_name_cell.value is not None) or (studnet_name_cell.value != "날짜") or (studnet_name_cell.value != "시험명") or (studnet_name_cell.value != "시험 평균"):
                for row in range(2, student_ws.max_row+1):
                    if (studnet_name_cell.value == student_ws.cell(row, 1).value) and (student_ws.cell(row, StudentInfo.NEW_STUDENT_CHECK_COLUMN).value == "N"):
                        studnet_name_cell.fill = PatternFill(fill_type="solid", fgColor=Color("FFFF00"))
                        break
    
    data_file_wb.save(f"./data/{config['dataFileName']}.xlsx")
    gui.q.put("조건부 서식 지정을 완료했습니다.")
    # excel.Visible = True
    wb = excel.Workbooks.Open(f"{os.getcwd()}\\data\\{config['dataFileName']}.xlsx")
    pythoncom.CoUninitialize()

def delete_student(gui:GUI, student_name:str):
    data_file_wb = xl.load_workbook(f"./data/{config['dataFileName']}.xlsx")
    # 데이터 파일 취소선
    for sheet_name in data_file_wb.sheetnames:
        data_file_ws = data_file_wb[sheet_name]

        for col in range(1, data_file_ws.max_column+1):
            if data_file_ws.cell(1, col).value == "이름":
                STUDENT_NAME_COLUMN = col
                break
        for row in range(2, data_file_ws.max_row+1):
            if data_file_ws.cell(row, STUDENT_NAME_COLUMN).value == student_name:
                for col in range(1, data_file_ws.max_column+1):
                    data_file_ws.cell(row, col).font = Font(strike=True)
    
    student_wb = xl.load_workbook("./학생 정보.xlsx")
    try:
        student_ws = student_wb["학생 정보"]
    except:
        gui.q.put(r"'학생 정보.xlsx'의 시트명을")
        gui.q.put(r"'학생 정보'로 변경해 주세요.")
        return
    # 학생 정보 삭제
    for row in range(2, student_ws.max_row+1):
        if student_ws.cell(row, StudentInfo.STUDENT_NAME_COLUMN).value == student_name:
            student_ws.delete_rows(row)

    data_file_wb.save(f"./data/{config['dataFileName']}.xlsx")
    student_wb.save("./학생 정보.xlsx")
    gui.q.put(f"{student_name} 학생을 퇴원 처리하였습니다.")
    gui.thread_end_flag = True
    return

def add_student(gui:GUI, student_name:str, target_class_name:str):
    if not check_student_exists(gui, student_name, target_class_name):
        gui.q.put(r"아이소식 해당 반에 학생이 업데이트되지 않아")
        gui.q.put(r"신규생 추가를 중단합니다.")
        gui.thread_end_flag = True
        return
    
    # 학생 정보 파일에 학생 추가
    gui.q.put(r"학생 정보 파일에 학생 추가 중...")
    student_wb = xl.load_workbook("./학생 정보.xlsx")
    try:
        student_ws = student_wb["학생 정보"]
    except:
        gui.q.put(r"'학생 정보.xlsx'의 시트명을")
        gui.q.put(r"'학생 정보'로 변경해 주세요.")
        return
    for row in range(student_ws.max_row+1, 1, -1):
        if student_ws.cell(row-1, StudentInfo.STUDENT_NAME_COLUMN).value is not None:
            student_ws.cell(row, StudentInfo.STUDENT_NAME_COLUMN).value      = student_name
            student_ws.cell(row, StudentInfo.CLASS_NAME_COLUMN).value        = target_class_name
            student_ws.cell(row, StudentInfo.NEW_STUDENT_CHECK_COLUMN).value = "N"
            for col in range(1, StudentInfo.MAX+1):
                student_ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
                student_ws.cell(row, col).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
            break
    student_wb.save("./학생 정보.xlsx")

    # 데이터파일에 학생 추가
    gui.q.put(r"데이터 파일에 학생 추가 중...")
    data_file_wb = xl.load_workbook(f"./data/{config['dataFileName']}.xlsx")
    for sheet_name in data_file_wb.sheetnames:
        data_file_ws = data_file_wb[sheet_name]
        for i in range(1, data_file_ws.max_column+1):
            temp = data_file_ws.cell(1, i).value
            if temp == "시간":
                TEST_TIME_COLUMN = i
            elif temp == "요일":
                CLASS_WEEKDAY_COLUMN = i
            elif temp == "반":
                CLASS_NAME_COLUMN = i
            elif temp == "담당":
                TEACHER_NAME_COLUMN = i
            elif temp == "이름":
                STUDENT_NAME_COLUMN = i
            elif temp == "학생 평균":
                AVERAGE_SCORE_COLUMN = i
        
        for row in range(2, data_file_ws.max_row+1):
            if data_file_ws.cell(row, CLASS_NAME_COLUMN).value == target_class_name:
                class_index = row+2
                break
        else: continue # 목표 반이 없으면 건너뛰기

        while data_file_ws.cell(class_index, STUDENT_NAME_COLUMN).value != "시험 평균":
            if data_file_ws.cell(class_index, STUDENT_NAME_COLUMN).value > student_name:
                break
            elif data_file_ws.cell(class_index, STUDENT_NAME_COLUMN).value == student_name:
                gui.q.put(f"{student_name} 학생이 이미 존재합니다.")
                gui.q.put(r"신규생 추가를 중단합니다.")
                gui.thread_end_flag = True
                return
            else: class_index += 1
        data_file_ws.insert_rows(class_index)
        copy_cell(data_file_ws.cell(class_index, TEST_TIME_COLUMN), data_file_ws.cell(class_index-1, TEST_TIME_COLUMN))
        copy_cell(data_file_ws.cell(class_index, CLASS_WEEKDAY_COLUMN), data_file_ws.cell(class_index-1, CLASS_WEEKDAY_COLUMN))
        copy_cell(data_file_ws.cell(class_index, CLASS_NAME_COLUMN), data_file_ws.cell(class_index-1, CLASS_NAME_COLUMN))
        copy_cell(data_file_ws.cell(class_index, TEACHER_NAME_COLUMN), data_file_ws.cell(class_index-1, TEACHER_NAME_COLUMN))

        data_file_ws.cell(class_index, STUDENT_NAME_COLUMN).value = student_name
        data_file_ws.cell(class_index, STUDENT_NAME_COLUMN).alignment = Alignment(horizontal="center", vertical="center")
        data_file_ws.cell(class_index, AVERAGE_SCORE_COLUMN).alignment = Alignment(horizontal="center", vertical="center")
        data_file_ws.cell(class_index, AVERAGE_SCORE_COLUMN).font = Font(bold=True)

    data_file_wb.save(f"./data/{config['dataFileName']}.xlsx")

    rescoping_formula()

    gui.q.put(f"{student_name} 학생을 {target_class_name} 반에 추가하였습니다.")

    gui.thread_end_flag = True
    return

def move_student(gui:GUI, student_name:str, target_class_name:str, current_class_name:str):
    if not check_student_exists(gui, student_name, target_class_name):
        gui.q.put(r"아이소식 해당 반에 학생이 업데이트되지 않아")
        gui.q.put(r"학생 반 이동을 중단합니다.")
        gui.thread_end_flag = True
        return
    
    data_file_wb = xl.load_workbook(f"./data/{config['dataFileName']}.xlsx")
    for sheet_name in data_file_wb.sheetnames:
        data_file_ws = data_file_wb[sheet_name]
        for i in range(1, data_file_ws.max_column+1):
            temp = data_file_ws.cell(1, i).value
            if temp == "시간":
                TEST_TIME_COLUMN = i
            elif temp == "요일":
                CLASS_WEEKDAY_COLUMN = i
            elif temp == "반":
                CLASS_NAME_COLUMN = i
            elif temp == "담당":
                TEACHER_NAME_COLUMN = i
            elif temp == "이름":
                STUDENT_NAME_COLUMN = i
            elif temp == "학생 평균":
                AVERAGE_SCORE_COLUMN = i
        
        # 기존 반 데이터 빨간색 처리
        for row in range(2, data_file_ws.max_row+1):
            if data_file_ws.cell(row, STUDENT_NAME_COLUMN).value == student_name and data_file_ws.cell(row, CLASS_NAME_COLUMN).value == current_class_name:
                for col in range(1, data_file_ws.max_column+1):
                    data_file_ws.cell(row, col).font = Font(color="FF0000")
        
        # 목표 반에 학생 추가
        for row in range(2, data_file_ws.max_row+1):
            if data_file_ws.cell(row, CLASS_NAME_COLUMN).value == target_class_name:
                class_index = row+2
                break
        else: continue # 목표 반이 없으면 건너뛰기

        while data_file_ws.cell(class_index, STUDENT_NAME_COLUMN).value != "시험 평균":
            if data_file_ws.cell(class_index, STUDENT_NAME_COLUMN).value > student_name:
                break
            elif data_file_ws.cell(class_index, STUDENT_NAME_COLUMN).value == student_name:
                gui.q.put(f"{student_name} 학생이 이미 존재합니다.")
                gui.q.put(r"학생 반 이동을 중단합니다.")
                gui.thread_end_flag = True
                return
            else: class_index += 1
        data_file_ws.insert_rows(class_index)
        copy_cell(data_file_ws.cell(class_index, TEST_TIME_COLUMN), data_file_ws.cell(class_index-1, TEST_TIME_COLUMN))
        copy_cell(data_file_ws.cell(class_index, CLASS_WEEKDAY_COLUMN), data_file_ws.cell(class_index-1, CLASS_WEEKDAY_COLUMN))
        copy_cell(data_file_ws.cell(class_index, CLASS_NAME_COLUMN), data_file_ws.cell(class_index-1, CLASS_NAME_COLUMN))
        copy_cell(data_file_ws.cell(class_index, TEACHER_NAME_COLUMN), data_file_ws.cell(class_index-1, TEACHER_NAME_COLUMN))

        data_file_ws.cell(class_index, STUDENT_NAME_COLUMN).value = student_name
        data_file_ws.cell(class_index, STUDENT_NAME_COLUMN).alignment = Alignment(horizontal="center", vertical="center")
        data_file_ws.cell(class_index, AVERAGE_SCORE_COLUMN).alignment = Alignment(horizontal="center", vertical="center")
        data_file_ws.cell(class_index, AVERAGE_SCORE_COLUMN).font = Font(bold=True)

    data_file_wb.save(f"./data/{config['dataFileName']}.xlsx")

    rescoping_formula()

    gui.q.put(f"{student_name} 학생을 {current_class_name} 반에서")
    gui.q.put(f"{target_class_name} 반으로 이동하였습니다.")
    gui.thread_end_flag = True
    return

def data_validation(gui:GUI, form_ws:Worksheet) -> bool:
    gui.q.put("양식이 올바른지 확인 중...")
    if (form_ws.title != "데일리테스트 기록 양식"):
        gui.q.put("올바른 기록 양식이 아닙니다.")
        return False
    
    form_checked      = True
    dailytest_checked = False
    mocktest_checked  = False
    for i in range(1, form_ws.max_row+1):
        if form_ws.cell(i, DataForm.CLASS_NAME_COLUMN).value is not None:
            class_name = form_ws.cell(i, DataForm.CLASS_NAME_COLUMN).value
            dailytest_checked = False
            mocktest_checked  = False
            dailytest_name    = form_ws.cell(i, DataForm.DAILYTEST_NAME_COLUMN).value
            mocktest_name     = form_ws.cell(i, DataForm.MOCKTEST_NAME_COLUMN).value
        
        if dailytest_checked and mocktest_checked: continue
        
        if not dailytest_checked and form_ws.cell(i, DataForm.DAILYTEST_SCORE_COLUMN).value is not None and dailytest_name is None:
            gui.q.put(f"{class_name}의 시험명이 작성되지 않았습니다.")
            dailytest_checked = True
            form_checked      = False
        if not mocktest_checked and form_ws.cell(i, DataForm.MOCKTEST_SCORE_COLUMN).value is not None and mocktest_name is None:
            gui.q.put(f"{class_name}의 모의고사명이 작성되지 않았습니다.")
            mocktest_checked = True
            form_checked     = False

    return form_checked

def copy_cell(dst:Cell, src:Cell):
    dst.value         = src.value
    dst.font          = copy(src.font)
    dst.fill          = copy(src.fill)
    dst.border        = copy(src.border)
    dst.alignment     = copy(src.alignment)
    dst.number_format = copy(src.number_format)

def rescoping_formula():
    data_file_wb = xl.load_workbook(f"./data/{config['dataFileName']}.xlsx")
    for sheet_name in data_file_wb.sheetnames:
        data_file_ws = data_file_wb[sheet_name]
        for col in range(1, data_file_ws.max_column+1):
            temp = data_file_ws.cell(1, col).value
            if temp == "이름":
                STUDENT_NAME_COLUMN = col
            elif temp == "학생 평균":
                AVERAGE_SCORE_COLUMN = col

        # 평균 범위 재지정
        for row in range(2, data_file_ws.max_row+1):
            striked = False
            colored = False
            if data_file_ws.cell(row, STUDENT_NAME_COLUMN).font.strike:
                striked = True
            if data_file_ws.cell(row, STUDENT_NAME_COLUMN).font.color is not None:
                if data_file_ws.cell(row, STUDENT_NAME_COLUMN).font.color.rgb == "FFFF0000":
                    colored = True
            
            if data_file_ws.cell(row, STUDENT_NAME_COLUMN).value == "날짜":
                class_start = row+2
            elif data_file_ws.cell(row, STUDENT_NAME_COLUMN).value == "시험 평균":
                class_end = row-1
                data_file_ws[f"{gcl(AVERAGE_SCORE_COLUMN)}{str(row)}"] = ArrayFormula(f"{gcl(AVERAGE_SCORE_COLUMN)}{str(row)}", f"=ROUND(AVERAGE(IFERROR({gcl(AVERAGE_SCORE_COLUMN)}{str(class_start)}:{gcl(AVERAGE_SCORE_COLUMN)}{str(class_end)}, \"\")), 0)")
                data_file_ws.cell(row, AVERAGE_SCORE_COLUMN).font = Font(bold=True)
                if class_start >= class_end: continue
                for col in range(AVERAGE_SCORE_COLUMN+1, data_file_ws.max_column+1):
                    if data_file_ws.cell(class_start-2, col).value is None: break
                    data_file_ws.cell(row, col).value = f"=ROUND(AVERAGE({gcl(col)}{str(class_start)}:{gcl(col)}{str(class_end)}), 0)"
                    data_file_ws.cell(row, col).font = Font(bold=True)
            elif data_file_ws.cell(row, STUDENT_NAME_COLUMN).value == "시험명": continue
            else:
                data_file_ws.cell(row, AVERAGE_SCORE_COLUMN).value = f"=ROUND(AVERAGE({gcl(AVERAGE_SCORE_COLUMN+1)}{str(row)}:XFD{str(row)}), 0)"
                data_file_ws.cell(row, AVERAGE_SCORE_COLUMN).font = Font(bold=True)
            
            if striked:
                data_file_ws.cell(row, AVERAGE_SCORE_COLUMN).font = Font(strike=True)
            if colored:
                data_file_ws.cell(row, AVERAGE_SCORE_COLUMN).font = Font(color="00FF0000")

    data_file_wb.save(f"./data/{config['dataFileName']}.xlsx")

def check_student_exists(gui:GUI, target_student_name:str, target_class_name:str) -> bool:
    gui.q.put(r"아이소식으로부터 정보 받아오는 중...")
    options = webdriver.ChromeOptions()
    options.add_argument("headless")
    driver = webdriver.Chrome(service = service, options = options)

    # 아이소식 접속
    driver.get(config["url"])
    table_names = driver.find_elements(By.CLASS_NAME, "style1")

    # 반 루프
    for i in range(3, len(table_names)):
        if target_class_name == table_names[i].text.rstrip():
            trs = driver.find_element(By.ID, "table_" + str(i)).find_elements(By.CLASS_NAME, "style12")
            for tr in trs:
                if target_student_name == tr.find_element(By.CLASS_NAME, "style9").text:
                    return True
    return False

def individual_record(gui:GUI, student_name:str, class_name:int, test_name:int, row:int, col:int, test_score:int, makeup_test_date:dict):
    gui.q.put(f"{student_name} 학생 데이터 저장 중...")
    test_name = test_name[9:]
    
    # 학생 정보 열기
    student_wb = xl.load_workbook("./학생 정보.xlsx")
    try:
        student_ws = student_wb["학생 정보"]
    except:
        gui.q.put(r"'학생 정보.xlsx'의 시트명을")
        gui.q.put(r"'학생 정보'로 변경해 주세요.")
        return
    
    # 학생 정보 찾기
    for r in range(2, student_ws.max_row+1):
        if student_ws.cell(r, StudentInfo.STUDENT_NAME_COLUMN).value == student_name:
            makeup_test_weekday = student_ws.cell(r, StudentInfo.MAKEUPTEST_WEEKDAY_COLUMN).value
            makeup_test_time    = student_ws.cell(r, StudentInfo.MAKEUPTEST_TIME_COLUMN).value
            new_student         = student_ws.cell(r, StudentInfo.NEW_STUDENT_CHECK_COLUMN).value
            break
    else:
        makeup_test_weekday = None
        makeup_test_time    = None
        new_student         = None

    # 재시험 일정 계산
    if makeup_test_weekday is not None:
        weekday_list = makeup_test_weekday.split("/")
        calculated_date = makeup_test_date[weekday_list[0].replace(" ", "")]
        time_index = 0
        for tmp_idx in range(len(weekday_list)):
            if calculated_date > makeup_test_date[weekday_list[tmp_idx].replace(" ", "")]:
                calculated_date = makeup_test_date[weekday_list[tmp_idx].replace(" ", "")]
                time_index = tmp_idx
    
    # 데이터 저장
    data_file_wb = xl.load_workbook(f"./data/{config['dataFileName']}.xlsx")
    data_file_wb.save(f"./data/backup/{config['dataFileName']}({datetime.today().strftime('%Y%m%d')}).xlsx")
    data_file_ws = data_file_wb["데일리테스트"]

    data_file_ws.cell(row, col).value = test_score
    data_file_ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
    if test_score < 60:
        data_file_ws.cell(row, col).fill = PatternFill(fill_type="solid", fgColor=Color("EC7E31"))
    elif test_score < 70:
        data_file_ws.cell(row, col).fill = PatternFill(fill_type="solid", fgColor=Color("F5AF85"))
    elif test_score < 80:
        data_file_ws.cell(row, col).fill = PatternFill(fill_type="solid", fgColor=Color("FCE4D6"))

    data_file_wb.save(f"./data/{config['dataFileName']}.xlsx")

    pythoncom.CoInitialize()
    excel = win32com.client.Dispatch("Excel.Application")
    wb = excel.Workbooks.Open(f"{os.getcwd()}\\data\\{config['dataFileName']}.xlsx")
    wb.Save()
    wb.Close()

    data_file_wb       = xl.load_workbook(f"./data/{config['dataFileName']}.xlsx")
    data_file_color_wb = xl.load_workbook(f"./data/{config['dataFileName']}.xlsx", data_only=True)

    data_file_ws       = data_file_wb["데일리테스트"]
    data_file_color_ws = data_file_color_wb["데일리테스트"]

    for c in range(1, data_file_ws.max_column):
        if data_file_ws.cell(1, c).value == "이름":
            STUDENT_NAME_COLUMN = c
            break
    
    while data_file_color_ws.cell(row, STUDENT_NAME_COLUMN).value != "시험 평균":
        row += 1

    test_average = data_file_color_ws.cell(row, col).value
    if type(test_average) == int:
        if test_average < 60:
            data_file_ws.cell(row, col).fill = PatternFill(fill_type="solid", fgColor=Color("EC7E31"))
        elif test_average < 70:
            data_file_ws.cell(row, col).fill = PatternFill(fill_type="solid", fgColor=Color("F5AF85"))
        elif test_average < 80:
            data_file_ws.cell(row, col).fill = PatternFill(fill_type="solid", fgColor=Color("FCE4D6"))
        else:
            data_file_ws.cell(row, col).fill = PatternFill(fill_type="solid", fgColor=Color("DDEBF7"))
    
    data_file_wb.save(f"./data/{config['dataFileName']}.xlsx")

    gui.q.put("데이터 저장을 완료했습니다.")
    # excel.Visible = True
    # wb = excel.Workbooks.Open(f"{os.getcwd()}\\data\\{config['dataFileName']}.xlsx")
    pythoncom.CoUninitialize()
    
    # 재시험 명단 작성
    if test_score < 80:
        if not os.path.isfile("./data/재시험 명단.xlsx"):
            gui.q.put("재시험 명단 파일 생성 중...")

            ini_wb = xl.Workbook()
            ini_ws = ini_wb[ini_wb.sheetnames[0]]
            ini_ws.title = "재시험 명단"
            ini_ws[gcl(MakeupTestList.TEST_DATE_COLUMN)+"1"]          = "응시일"
            ini_ws[gcl(MakeupTestList.CLASS_NAME_COLUMN)+"1"]         = "반"
            ini_ws[gcl(MakeupTestList.TEACHER_NAME_COLUMN)+"1"]       = "담당T"
            ini_ws[gcl(MakeupTestList.STUDENT_NAME_COLUMN)+"1"]       = "이름"
            ini_ws[gcl(MakeupTestList.TEST_NAME_COLUMN)+"1"]          = "시험명"
            ini_ws[gcl(MakeupTestList.TEST_SCORE_COLUMN)+"1"]         = "시험 점수"
            ini_ws[gcl(MakeupTestList.MAKEUPTEST_WEEKDAY_COLUMN)+"1"] = "재시 요일"
            ini_ws[gcl(MakeupTestList.MAKEUPTEST_TIME_COLUMN)+"1"]    = "재시 시간"
            ini_ws[gcl(MakeupTestList.MAKEUPTEST_DATE_COLUMN)+"1"]    = "재시 날짜"
            ini_ws[gcl(MakeupTestList.MAKEUPTEST_SCORE_COLUMN)+"1"]   = "재시 점수"
            ini_ws[gcl(MakeupTestList.ETC_COLUMN)+"1"]                = "비고"
            ini_ws.auto_filter.ref = "A:"+gcl(MakeupTestList.MAX)
            ini_wb.save("./data/재시험 명단.xlsx")
        makeup_list_wb = xl.load_workbook("./data/재시험 명단.xlsx")
        try:
            makeup_list_ws = makeup_list_wb["재시험 명단"]
        except:
            gui.q.put(r"'재시험 명단.xlsx'의 시트명을")
            gui.q.put(r"'재시험 명단'으로 변경해 주세요.")
            return
        
        class_wb = xl.load_workbook("./반 정보.xlsx")
        try:
            class_ws = class_wb["반 정보"]
        except:
            gui.q.put(r"'반 정보.xlsx'의 시트명을")
            gui.q.put(r"'반 정보'로 변경해 주세요.")
            return
        
        for r in range(2, class_ws.max_row + 1):
            if class_ws.cell(r, ClassInfo.CLASS_NAME_COLUMN).value == class_name:
                teacher_name = class_ws.cell(r, ClassInfo.TEACHER_NAME_COLUMN).value
                break
        else: teacher_name = ""
        
        MAKEUP_TEST_WRITE_ROW = makeup_list_ws.max_row+1
        while makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.TEST_DATE_COLUMN).value is None:
            MAKEUP_TEST_WRITE_ROW -= 1
        MAKEUP_TEST_WRITE_ROW += 1

        makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.TEST_DATE_COLUMN).value = DATE.today()
        makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.CLASS_NAME_COLUMN).value = class_name
        makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.TEACHER_NAME_COLUMN).value = teacher_name
        makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.STUDENT_NAME_COLUMN).value = student_name
        if (new_student is not None) and (new_student == "N"):
            makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.STUDENT_NAME_COLUMN).fill = PatternFill(fill_type="solid", fgColor=Color("FFFF00"))
        makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.TEST_NAME_COLUMN).value = test_name
        makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.TEST_SCORE_COLUMN).value = test_score
        if makeup_test_weekday is not None:
            makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.MAKEUPTEST_WEEKDAY_COLUMN).value = makeup_test_weekday
            if makeup_test_time is not None:
                makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.MAKEUPTEST_TIME_COLUMN).value = makeup_test_time
            makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.MAKEUPTEST_DATE_COLUMN).value = calculated_date
            makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, MakeupTestList.MAKEUPTEST_DATE_COLUMN).number_format = "mm월 dd일(aaa)"
        for c in range(1, makeup_list_ws.max_column + 1):
            makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, c).alignment = Alignment(horizontal="center", vertical="center")
            makeup_list_ws.cell(MAKEUP_TEST_WRITE_ROW, c).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        makeup_list_wb.save("./data/재시험 명단.xlsx")

    # 개별 메시지 전송
    gui.q.put("메시지 작성 중...")
    options = webdriver.ChromeOptions()
    options.add_experimental_option("detach", True)
    driver = webdriver.Chrome(service=service, options=options)
    
    # 아이소식 접속
    driver.get(config["url"])
    driver.execute_script(f"arguments[0].value = '{config['dailyTest']}'", driver.find_element(By.XPATH, '//*[@id="ctitle"]'))
    driver.find_element(By.XPATH, '//*[@id="ctitle"]').send_keys(' \b')

    driver.switch_to.window(driver.window_handles[0])
    tables = driver.find_elements(By.CLASS_NAME, "style1")
    table_names = [table.text for table in tables]
    for idx in range(3, len(table_names)):
        if class_name == table_names[idx]:
            class_index = idx
            break
    
    # 시험 결과 메시지 작성
    trs = driver.find_element(By.ID, f"table_{str(class_index)}").find_elements(By.CLASS_NAME, "style12")
    for idx in range(0, len(trs)):
        if trs[idx].find_element(By.CLASS_NAME, "style9").text == student_name:
            student_index = idx
            break

    tds = trs[student_index].find_elements(By.TAG_NAME, "td")
    driver.execute_script(f"arguments[0].value = '{test_name}'",  tds[0].find_element(By.TAG_NAME, "input"))
    driver.execute_script(f"arguments[0].value = '{test_score}'", tds[1].find_element(By.TAG_NAME, "input"))
    tds[2].find_element(By.TAG_NAME, "input").send_keys(test_average)
    
    # 재시험 메시지 작성
    if test_score < 80:
        if makeup_test_weekday is None:
            driver.execute_script(f"window.open('{config['url']}')")
            driver.switch_to.window(driver.window_handles[1])
            driver.execute_script(f"arguments[0].value = '{config['makeupTest']}'", driver.find_element(By.XPATH, '//*[@id="ctitle"]'))
            driver.find_element(By.XPATH, '//*[@id="ctitle"]').send_keys(' \b')
            trs = driver.find_element(By.ID, "table_" + str(class_index)).find_elements(By.CLASS_NAME, "style12")
            tds = trs[student_index].find_elements(By.TAG_NAME, "td")
            driver.execute_script(f"arguments[0].value = '{test_name}'", tds[0].find_element(By.TAG_NAME, "input"))
            tds[1].find_element(By.TAG_NAME, "input").send_keys(' \b')
        else:
            driver.execute_script(f"window.open('{config['url']}')")
            driver.switch_to.window(driver.window_handles[1])
            driver.execute_script(f"arguments[0].value = '{config['makeupTestDate']}'", driver.find_element(By.XPATH, '//*[@id="ctitle"]'))
            driver.find_element(By.XPATH, '//*[@id="ctitle"]').send_keys(' \b')
            trs = driver.find_element(By.ID, "table_" + str(class_index)).find_elements(By.CLASS_NAME, "style12")
            tds = trs[student_index].find_elements(By.TAG_NAME, "td")
            driver.execute_script(f"arguments[0].value = '{test_name}'", tds[0].find_element(By.TAG_NAME, "input"))
            try:
                if makeup_test_time is not None:
                    if "/" in str(makeup_test_time):
                        driver.execute_script(f"arguments[0].value = '{calculated_date.strftime('%m월 %d일'.encode('unicode-escape').decode()).encode().decode('unicode-escape')} {str(makeup_test_time).split('/')[time_index]}시'", tds[1].find_element(By.TAG_NAME, "input"))
                    else:
                        driver.execute_script(f"arguments[0].value = '{calculated_date.strftime('%m월 %d일'.encode('unicode-escape').decode()).encode().decode('unicode-escape')} {str(makeup_test_time)}시'", tds[1].find_element(By.TAG_NAME, "input"))
                else:
                    driver.execute_script(f"arguments[0].value = '{calculated_date.strftime('%m월 %d일'.encode('unicode-escape').decode()).encode().decode('unicode-escape')}'", tds[1].find_element(By.TAG_NAME, "input"))
            except:
                gui.q.put(f"{student_name}의 재시험 일정을 요일별 시간으로 설정하거나")
                gui.q.put("하나의 시간으로 통일해 주세요.")
                gui.q.put("중단되었습니다.")
                driver.quit()
                gui.thread_end_flag = True    
                return
            tds[2].find_element(By.TAG_NAME, "input").send_keys(' \b')

    gui.q.put("메시지 입력을 완료했습니다.")
    gui.q.put("메시지 확인 후 전송해주세요.")
    gui.thread_end_flag = True

ui = tk.Tk()
gui = GUI(ui)
ui.after(100, gui.thread_log)
ui.after(100, gui.check_files)
ui.after(100, gui.check_thread_end)
ui.mainloop()
