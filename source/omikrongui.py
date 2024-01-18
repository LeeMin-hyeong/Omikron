import os
import queue
import webbrowser
import tkinter as tk
import openpyxl as xl

from tkinter import ttk, filedialog
from datetime import date as DATE, datetime, timedelta
from dateutil.relativedelta import relativedelta

import omikronthread
import omikronconfig
from omikronlog import OmikronLog
from omikrondefs import VERSION, LOG_INTERFACE_WIDTH

class GUI():
    def __init__(self, ui:tk.Tk):
        self.ui = ui
        
        # log queue
        self.log_q = OmikronLog.log_queue
        
        # 작업 종료 플래그
        self.thread_end_flag = omikronthread.thread_end_flag

        # 창 크기
        self.width = 320
        self.height = 585 # button +25
        # 창 위치
        self.x = int((self.ui.winfo_screenwidth()/4) - (self.width/2))
        self.y = int((self.ui.winfo_screenheight()/2) - (self.height/2))

        self.ui.geometry(f"{self.width}x{self.height}+{self.x}+{self.y}")
        self.ui.title(VERSION)
        self.ui.resizable(False, False)

        # 재시험 일정 초기화
        # 정해지지 않으면 최초 1회 사용자가 직접 설정
        self.makeup_test_date = None

        tk.Label(self.ui, text="Omikron 데이터 프로그램").pack()
        
        # Notion 사용 설명서 하이퍼링크
        def callback(url:str):
            webbrowser.open_new(url)
        link = tk.Label(self.ui, text="[ 사용법 및 도움말 ]", cursor="hand2")
        link.pack()
        link.bind("<Button-1>", lambda _: callback("https://omikron-db.notion.site/ad673cca64c146d28adb3deaf8c83a0d?pvs=4"))

        # 메세지 창
        self.scroll = tk.Scrollbar(self.ui, orient="vertical")
        self.log = tk.Listbox(self.ui, yscrollcommand=self.scroll.set, width=LOG_INTERFACE_WIDTH, height=5)
        self.scroll.config(command=self.log.yview)
        self.log.pack()
        
        # buttons
        tk.Label(self.ui, text="< 기수 변경 관련 >").pack()

        self.make_class_info_file_button = tk.Button(self.ui, cursor="hand2", text="반 정보 기록 양식 생성", width=40, command=omikronthread.make_class_info_file_thread)
        self.make_class_info_file_button.pack()

        self.make_student_info_file_button = tk.Button(self.ui, cursor="hand2", text="학생 정보 기록 양식 생성", width=40, command=lambda: self.make_student_info_file_thread())
        self.make_student_info_file_button.pack()

        self.make_data_file_button = tk.Button(self.ui, cursor="hand2", text="데이터 파일 생성", width=40, command=lambda: self.make_data_file_thread())
        self.make_data_file_button.pack()

        self.update_class_button = tk.Button(self.ui, cursor="hand2", text="반 업데이트", width=40, command=lambda: self.update_class_thread())
        self.update_class_button.pack()

        tk.Label(self.ui, text="\n< 데이터 저장 및 문자 전송 >").pack()

        self.make_data_form_button = tk.Button(self.ui, cursor="hand2", text="데일리 테스트 기록 양식 생성", width=40, command=lambda: self.make_data_form_thread())
        self.make_data_form_button.pack()

        self.save_data_button = tk.Button(self.ui, cursor="hand2", text="데이터 엑셀 파일에 저장", width=40, command=lambda: self.save_data_thread())
        self.save_data_button.pack()

        self.send_message_button = tk.Button(self.ui, cursor="hand2", text="시험 결과 전송", width=40, command=lambda: self.send_message_thread())
        self.send_message_button.pack()

        self.individual_record_button = tk.Button(self.ui, cursor="hand2", text="개별 시험 기록", width=40, command=lambda: self.individual_record_thread())
        self.individual_record_button.pack()

        self.makeup_test_record_button = tk.Button(self.ui, cursor="hand2", text="재시험 기록", width=40, command=lambda: self.makeup_test_record_thread())
        self.makeup_test_record_button.pack()

        tk.Label(self.ui, text="\n< 데이터 관리 >").pack()

        self.apply_color_button = tk.Button(self.ui, cursor="hand2", text="데이터 엑셀 파일 조건부 서식 재지정", width=40, command=lambda: apply_color(self))
        self.apply_color_button.pack()

        tk.Label(self.ui, text="< 학생 관리 >").pack()
        self.add_student_button = tk.Button(self.ui, cursor="hand2", text="신규생 추가", width=40, command=lambda: self.add_student_thread())
        self.add_student_button.pack()

        self.delete_student_button = tk.Button(self.ui, cursor="hand2", text="퇴원 처리", width=40, command=lambda: self.delete_student_thread())
        self.delete_student_button.pack()

        self.move_student_button = tk.Button(self.ui, cursor="hand2", text="학생 반 이동", width=40, command=lambda: self.move_student_thread())
        self.move_student_button.pack()

    # ui
    def print_log(self):
        try:
            msg = self.log_q.get(block=False)
        except queue.Empty:
            self.ui.after(100, self.print_log)
            return
        self.log.insert(tk.END, msg)
        self.log.update()
        self.log.see(tk.END)
        self.ui.after(100, self.print_log)

    def check_files(self):
        check1 = check2 = check3 = False
        if os.path.isfile("반 정보.xlsx"):
            self.make_class_info_file_button["state"] = tk.DISABLED
            check1 = True
        else:
            self.make_class_info_file_button["state"]   = tk.NORMAL
            self.make_student_info_file_button["state"] = tk.DISABLED
            self.make_data_file_button["state"]         = tk.DISABLED
        if os.path.isfile("학생 정보.xlsx"):
            self.make_student_info_file_button["state"] = tk.DISABLED
            check2 = True
        else: 
            self.make_student_info_file_button["state"] = tk.NORMAL
        if os.path.isfile(f"./data/{omikronconfig.config['dataFileName']}.xlsx"):
            self.make_data_file_button["state"] = tk.DISABLED
            check3 = True
        else:
            self.make_data_file_button["state"] = tk.NORMAL
        
        if check1 and check2 and check3:
            self.update_class_button["state"]       = tk.NORMAL
            self.make_data_form_button["state"]     = tk.NORMAL
            self.save_data_button["state"]          = tk.NORMAL
            self.send_message_button["state"]       = tk.NORMAL
            self.individual_record_button["state"]  = tk.NORMAL
            self.makeup_test_record_button["state"] = tk.NORMAL
            self.apply_color_button["state"]        = tk.NORMAL
            self.add_student_button["state"]        = tk.NORMAL
            self.delete_student_button["state"]     = tk.NORMAL
            self.move_student_button["state"]       = tk.NORMAL
        else:
            self.update_class_button["state"]       = tk.DISABLED
            self.make_data_form_button["state"]     = tk.DISABLED
            self.save_data_button["state"]          = tk.DISABLED
            self.send_message_button["state"]       = tk.DISABLED
            self.individual_record_button["state"]  = tk.DISABLED
            self.makeup_test_record_button["state"] = tk.DISABLED
            self.apply_color_button["state"]        = tk.DISABLED
            self.add_student_button["state"]        = tk.DISABLED
            self.delete_student_button["state"]     = tk.DISABLED
            self.move_student_button["state"]       = tk.DISABLED
        
        self.ui.after(100, self.check_files)

    def check_thread_end(self):
        if self.thread_end_flag:
            self.thread_end_flag = False
            self.ui.wm_attributes("-topmost", 1)
            self.ui.wm_attributes("-topmost", 0)
        self.ui.after(100, self.check_thread_end)

    # dialog
    def holiday_dialog(self) -> dict:
        def quit_event():
            for i in range(7):
                if var_list[i].get():
                    makeup_test_date[weekday[i]] += timedelta(days=7)
            self.ui.wm_attributes("-disabled", False)
            popup.quit()
            popup.destroy()

        self.ui.wm_attributes("-disabled", True)
        popup = tk.Toplevel()
        width = 200
        height = 300
        x = int((popup.winfo_screenwidth()/4) - (width/2))
        y = int((popup.winfo_screenheight()/2) - (height/2))
        popup.geometry(f"{width}x{height}+{x}+{y}")
        popup.title("휴일 선택")
        popup.resizable(False, False)
        popup.protocol("WM_DELETE_WINDOW", quit_event)

        today = DATE.today()
        weekday = ("월", "화", "수", "목", "금", "토", "일")
        makeup_test_date = {weekday[i] : today + relativedelta(weekday=i) for i in range(7)}
        for key, value in makeup_test_date.items():
            if value == today: makeup_test_date[key] += timedelta(days=7)

        mon = tk.BooleanVar()
        tue = tk.BooleanVar()
        wed = tk.BooleanVar()
        thu = tk.BooleanVar()
        fri = tk.BooleanVar()
        sat = tk.BooleanVar()
        sun = tk.BooleanVar()
        var_list = [mon, tue, wed, thu, fri, sat, sun]
        tk.Label(popup, text="\n다음 중 휴일을 선택해주세요\n").pack()
        sort = today.weekday()+1
        for i in range(7):
            tk.Checkbutton(popup, text=f"{str(makeup_test_date[weekday[(sort+i)%7]])} {weekday[(sort+i)%7]}", variable=var_list[(sort+i)%7]).pack()
        tk.Label(popup, text="\n").pack()
        tk.Button(popup, text="확인", width=10 , command=quit_event).pack()
        
        popup.mainloop()    
        
        return makeup_test_date

    def delete_student_name_dialog(self) -> str:
        def quit_event():
            self.ui.wm_attributes("-disabled", False)
            popup.quit()
            popup.destroy()

        self.ui.wm_attributes("-disabled", True)
        popup = tk.Toplevel()
        width = 250
        height = 140
        x = int((popup.winfo_screenwidth()/4) - (width/2))
        y = int((popup.winfo_screenheight()/2) - (height/2))
        popup.geometry(f"{width}x{height}+{x}+{y}")
        popup.title("퇴원 관리")
        popup.resizable(False, False)
        popup.protocol("WM_DELETE_WINDOW", quit_event)

        # 반 정보 확인
        class_wb = xl.load_workbook("./반 정보.xlsx")
        try:
            class_ws = class_wb["반 정보"]
        except:
            self.log_q.put(r"'반 정보.xlsx'의 시트명을")
            self.log_q.put(r"'반 정보'로 변경해 주세요.")
            return
        
        data_file_wb = xl.load_workbook(f"./data/{omikronconfig.config['dataFileName']}.xlsx")
        data_file_ws = data_file_wb.worksheets[0]
        for i in range(1, data_file_ws.max_column):
            if data_file_ws.cell(1, i).value == "이름":
                STUDENT_NAME_COLUMN = i
                break
        for i in range(1, data_file_ws.max_column):
            if data_file_ws.cell(1, i).value == "반":
                CLASS_NAME_COLUMN = i
                break

        class_dict = {}
        for i in range(2, class_ws.max_row + 1):
            class_name = class_ws.cell(i, ClassInfo.CLASS_NAME_COLUMN).value
            student_list = []
            for j in range(2, data_file_ws.max_row+1):
                if data_file_ws.cell(j, CLASS_NAME_COLUMN).value != class_name: continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).value in ("날짜", "시험명", "시험 평균"): continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.strike: continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.color is not None and data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.color.rgb == "FFFF0000": continue
                student_list.append(data_file_ws.cell(j, STUDENT_NAME_COLUMN).value)
            class_dict[class_name] = student_list
        class_dict = dict(sorted(class_dict.items()))
        
        tk.Label(popup).pack()
        def class_call_back(event):
            class_name = event.widget.get()
            student_combo.set("학생 선택")
            student_combo["values"] = class_dict[class_name]
        class_combo = ttk.Combobox(popup, values=list(class_dict.keys()), state="readonly")
        class_combo.set("반 선택")
        class_combo.bind("<<ComboboxSelected>>", class_call_back)
        class_combo.pack()

        tk.Label(popup).pack()
        selected_student = tk.StringVar()
        student_combo = ttk.Combobox(popup, values=None, state="readonly", textvariable=selected_student)
        student_combo.set("학생 선택")
        student_combo.pack()

        tk.Label(popup).pack()
        tk.Button(popup, text="퇴원", width=10 , command=quit_event).pack()
        
        popup.mainloop()
        
        student_name = selected_student.get()
        if student_name == "학생 선택":
            return None
        else:
            return student_name

    def move_student_dialog(self):
        def quit_event():
            self.ui.wm_attributes("-disabled", False)
            popup.quit()
            popup.destroy()

        self.ui.wm_attributes("-disabled", True)
        popup = tk.Toplevel()
        width = 250
        height = 160
        x = int((popup.winfo_screenwidth()/4) - (width/2))
        y = int((popup.winfo_screenheight()/2) - (height/2))
        popup.geometry(f"{width}x{height}+{x}+{y}")
        popup.title("학생 반 이동")
        popup.resizable(False, False)
        popup.protocol("WM_DELETE_WINDOW", quit_event)

        # 반 정보 확인
        class_wb = xl.load_workbook("./반 정보.xlsx")
        try:
            class_ws = class_wb["반 정보"]
        except:
            self.log_q.put(r"'반 정보.xlsx'의 시트명을")
            self.log_q.put(r"'반 정보'로 변경해 주세요.")
            return
        
        data_file_wb = xl.load_workbook(f"./data/{omikronconfig.config['dataFileName']}.xlsx")
        data_file_ws = data_file_wb.worksheets[0]
        for i in range(1, data_file_ws.max_column):
            if data_file_ws.cell(1, i).value == "이름":
                STUDENT_NAME_COLUMN = i
                break
        for i in range(1, data_file_ws.max_column):
            if data_file_ws.cell(1, i).value == "반":
                CLASS_NAME_COLUMN = i
                break

        class_dict = {}
        for i in range(2, class_ws.max_row + 1):
            class_name = class_ws.cell(i, ClassInfo.CLASS_NAME_COLUMN).value
            student_list = []
            for j in range(2, data_file_ws.max_row+1):
                if data_file_ws.cell(j, CLASS_NAME_COLUMN).value != class_name: continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).value in ("날짜", "시험명", "시험 평균"): continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.strike: continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.color is not None and data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.color.rgb == "FFFF0000": continue
                student_list.append(data_file_ws.cell(j, STUDENT_NAME_COLUMN).value)
            class_dict[class_name] = student_list
        class_dict = dict(sorted(class_dict.items()))

        tk.Label(popup).pack()
        def class_call_back(event):
            class_name = event.widget.get()
            student_combo.set("학생 선택")
            student_combo["values"] = class_dict[class_name]
        current_class_var = tk.StringVar()
        current_class_combo = ttk.Combobox(popup, values=list(class_dict.keys()), state="readonly", textvariable=current_class_var)
        current_class_combo.set("반 선택")
        current_class_combo.bind("<<ComboboxSelected>>", class_call_back)
        current_class_combo.pack()

        selected_student = tk.StringVar()
        student_combo = ttk.Combobox(popup, values=None, state="readonly", textvariable=selected_student)
        student_combo.set("학생 선택")
        student_combo.pack()

        tk.Label(popup).pack()
        target_class_var = tk.StringVar()
        current_class_combo = ttk.Combobox(popup, values=list(class_dict.keys()), state="readonly", textvariable=target_class_var)
        current_class_combo.set("이동할 반 선택")
        current_class_combo.pack()

        tk.Label(popup).pack()
        tk.Button(popup, text="반 이동", width=10 , command=quit_event).pack()
        
        popup.mainloop()
        
        target_student_name = selected_student.get()
        target_class_name = target_class_var.get()
        current_class_name = current_class_var.get()
        if target_student_name == "학생 선택" or target_class_name == "이동할 반 선택":
            return None
        else:
            return target_student_name, target_class_name, current_class_name

    def add_student_dialog(self):
        def quit_event():
            self.ui.wm_attributes("-disabled", False)
            popup.quit()
            popup.destroy()

        self.ui.wm_attributes("-disabled", True)
        popup = tk.Toplevel()
        width = 250
        height = 140
        x = int((popup.winfo_screenwidth()/4) - (width/2))
        y = int((popup.winfo_screenheight()/2) - (height/2))
        popup.geometry(f"{width}x{height}+{x}+{y}")
        popup.title("신규생 추가")
        popup.resizable(False, False)
        popup.protocol("WM_DELETE_WINDOW", quit_event)

        # 반 정보 확인
        class_wb = xl.load_workbook("./반 정보.xlsx")
        try:
            class_ws = class_wb["반 정보"]
        except:
            self.log_q.put(r"'반 정보.xlsx'의 시트명을")
            self.log_q.put(r"'반 정보'로 변경해 주세요.")
            return
        
        class_names = sorted([class_ws.cell(i, ClassInfo.CLASS_NAME_COLUMN).value for i in range(2, class_ws.max_row + 1) if class_ws.cell(i, ClassInfo.CLASS_NAME_COLUMN).value is not None])

        tk.Label(popup).pack()
        target_class_var = tk.StringVar()
        class_combo = ttk.Combobox(popup, values=class_names, state="readonly", textvariable=target_class_var, width=25)
        class_combo.set("학생을 추가할 반 선택")
        class_combo.pack()

        tk.Label(popup).pack()
        target_student_var = tk.StringVar()
        tk.Entry(popup, textvariable=target_student_var, width=28).pack()

        tk.Label(popup).pack()
        tk.Button(popup, text="신규생 추가", width=10 , command=quit_event).pack()
        
        popup.mainloop()
        
        target_class_name = target_class_var.get()
        target_student_name = target_student_var.get()
        if target_class_name == "학생을 추가할 반 선택" or target_student_name == "":
            return None
        else:
            return target_student_name, target_class_name

    def individual_record_dialog(self):
        def quit_event():
            self.ui.wm_attributes("-disabled", False)
            popup.quit()
            popup.destroy()

        self.ui.wm_attributes("-disabled", True)
        popup = tk.Toplevel()
        width = 250
        height = 160
        x = int((popup.winfo_screenwidth()/4) - (width/2))
        y = int((popup.winfo_screenheight()/2) - (height/2))
        popup.geometry(f"{width}x{height}+{x}+{y}")
        popup.title("개별 점수 기록")
        popup.resizable(False, False)
        popup.protocol("WM_DELETE_WINDOW", quit_event)

        # 반 정보 확인
        class_wb = xl.load_workbook("./반 정보.xlsx")
        try:
            class_ws = class_wb["반 정보"]
        except:
            self.log_q.put(r"'반 정보.xlsx'의 시트명을")
            self.log_q.put(r"'반 정보'로 변경해 주세요.")
            return
        
        data_file_wb = xl.load_workbook(f"./data/{omikronconfig.config['dataFileName']}.xlsx")
        data_file_ws = data_file_wb.worksheets[0]
        for i in range(1, data_file_ws.max_column):
            if data_file_ws.cell(1, i).value == "이름":
                STUDENT_NAME_COLUMN = i
                break
        for i in range(1, data_file_ws.max_column):
            if data_file_ws.cell(1, i).value == "반":
                CLASS_NAME_COLUMN = i
                break
        for i in range(1, data_file_ws.max_column+1):
            if data_file_ws.cell(1, i).value == "학생 평균":
                AVERAGE_SCORE_COLUMN = i
                break

        class_dict1:dict[str, dict] = {}
        class_dict2:dict[str, dict] = {}
        for i in range(2, class_ws.max_row + 1):
            student_dict = {}
            class_name = class_ws.cell(i, ClassInfo.CLASS_NAME_COLUMN).value
            for j in range(2, data_file_ws.max_row+1):
                if data_file_ws.cell(j, CLASS_NAME_COLUMN).value != class_name: continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).value == "날짜":
                    test_name_dict = {data_file_ws.cell(j, k).value.strftime("%y/%m/%d ")+str(data_file_ws.cell(j+1, k).value) : k for k in range(AVERAGE_SCORE_COLUMN+1, data_file_ws.max_row+1) if data_file_ws.cell(j, k).value is not None and data_file_ws.cell(j+1, k).value is not None}
                    continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).value in ("시험명", "시험 평균"): continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.strike: continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.color is not None and data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.color.rgb == "FFFF0000": continue
                student_dict[data_file_ws.cell(j, STUDENT_NAME_COLUMN).value] = j
            
            test_name_dict = dict(sorted(test_name_dict.items(), reverse=True))
            class_dict1[class_name] = student_dict
            class_dict2[class_name] = test_name_dict

        class_dict1 = dict(sorted(class_dict1.items()))
        tk.Label(popup).pack()
        def class_call_back(event):
            class_name = event.widget.get()
            student_combo.set("학생 선택")
            student_combo["values"] = list(class_dict1[class_name].keys())
            test_list_combo["values"] = list(class_dict2[class_name].keys())
        target_class_var = tk.StringVar()
        target_class_combo = ttk.Combobox(popup, values=list(class_dict1.keys()), state="readonly", textvariable=target_class_var, width=100)
        target_class_combo.set("반 선택")
        target_class_combo.bind("<<ComboboxSelected>>", class_call_back)
        target_class_combo.pack()

        target_studnet_var = tk.StringVar()
        student_combo = ttk.Combobox(popup, values=None, state="readonly", textvariable=target_studnet_var, width=100)
        student_combo.set("학생 선택")
        student_combo.pack()

        test_name_var = tk.StringVar()
        test_list_combo = ttk.Combobox(popup, values=list(class_dict2.keys()), state="readonly", textvariable=test_name_var, width=100)
        test_list_combo.set("시험 선택")
        test_list_combo.pack()

        score_var = tk.StringVar()
        tk.Entry(popup, textvariable=score_var, width=28).pack()

        tk.Label(popup).pack()
        tk.Button(popup, text="메세지 전송 및 저장", width=20 , command=quit_event).pack()
        
        popup.mainloop()
        
        target_class_name   = target_class_var.get()
        target_student_name = target_studnet_var.get()
        test_name           = test_name_var.get()
        test_score          = score_var.get()
        try:
            if '.' in test_score:
                test_score = float(test_score)
            else:
                test_score = int(test_score)
        except:
            self.log_q.put("올바른 점수를 입력해 주세요.")
            return None
        
        if target_class_name == "반 선택" or target_student_name == "학생 선택" or test_name == "시험 선택":
            return None

        row = class_dict1[target_class_name][target_student_name]
        col = class_dict2[target_class_name][test_name]
        return target_student_name, target_class_name, test_name, row, col, test_score, data_file_ws.cell(row, col).value

    def makeup_test_record_dialog(self):
        def quit_event():
            self.ui.wm_attributes("-disabled", False)
            popup.quit()
            popup.destroy()

        self.ui.wm_attributes("-disabled", True)
        popup = tk.Toplevel()
        width = 250
        height = 160
        x = int((popup.winfo_screenwidth()/4) - (width/2))
        y = int((popup.winfo_screenheight()/2) - (height/2))
        popup.geometry(f"{width}x{height}+{x}+{y}")
        popup.title("재시험 기록")
        popup.resizable(False, False)
        popup.protocol("WM_DELETE_WINDOW", quit_event)

        # 반 정보 확인
        class_wb = xl.load_workbook("./반 정보.xlsx")
        try:
            class_ws = class_wb["반 정보"]
        except:
            self.log_q.put(r"'반 정보.xlsx'의 시트명을")
            self.log_q.put(r"'반 정보'로 변경해 주세요.")
            return
        
        makeup_list_wb = xl.load_workbook("./data/재시험 명단.xlsx")
        try:
            makeup_list_ws = makeup_list_wb["재시험 명단"]
        except:
            gui.q.put(r"'재시험 명단.xlsx'의 시트명을")
            gui.q.put(r"'재시험 명단'으로 변경해 주세요.")
            return
        
        data_file_wb = xl.load_workbook(f"./data/{omikronconfig.config['dataFileName']}.xlsx")
        data_file_ws = data_file_wb.worksheets[0]
        for i in range(1, data_file_ws.max_column):
            if data_file_ws.cell(1, i).value == "이름":
                STUDENT_NAME_COLUMN = i
                break
        for i in range(1, data_file_ws.max_column):
            if data_file_ws.cell(1, i).value == "반":
                CLASS_NAME_COLUMN = i
                break

        class_dict:dict[str, list] = {}
        for i in range(2, class_ws.max_row + 1):
            class_name = class_ws.cell(i, ClassInfo.CLASS_NAME_COLUMN).value
            student_list = []
            for j in range(2, data_file_ws.max_row+1):
                if data_file_ws.cell(j, CLASS_NAME_COLUMN).value != class_name: continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).value in ("날짜", "시험명", "시험 평균"): continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.strike: continue
                if data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.color is not None and data_file_ws.cell(j, STUDENT_NAME_COLUMN).font.color.rgb == "FFFF0000": continue
                student_list.append(data_file_ws.cell(j, STUDENT_NAME_COLUMN).value)
            class_dict[class_name] = student_list
        class_dict = dict(sorted(class_dict.items()))

        student_dict:dict[str, dict[str, int]] = {}
        for i in range(2, makeup_list_ws.max_row+1):
            if makeup_list_ws.cell(i, MakeupTestList.MAKEUPTEST_SCORE_COLUMN).value is None:
                student_name = makeup_list_ws.cell(i, MakeupTestList.STUDENT_NAME_COLUMN).value
                makeup_test_name = makeup_list_ws.cell(i, MakeupTestList.TEST_NAME_COLUMN).value
                try:
                    student_dict[student_name]
                except:
                    student_dict[student_name] = {}
                
                student_dict[student_name][makeup_test_name] = i

        def class_call_back(event):
            class_name = event.widget.get()
            student_combo.set("학생 선택")
            student_combo["values"] = list(class_dict[class_name])
            makeup_test_list_combo.set("재시험 선택")
            makeup_test_list_combo["values"] = None
        target_class_var = tk.StringVar()
        target_class_combo = ttk.Combobox(popup, values=list(class_dict.keys()), state="readonly", textvariable=target_class_var, width=100)
        target_class_combo.set("반 선택")
        target_class_combo.bind("<<ComboboxSelected>>", class_call_back)
        target_class_combo.pack()

        def student_call_back(event):
            student_name = event.widget.get()
            makeup_test_list_combo.set("재시험 선택")
            try:
                makeup_test_list_combo["values"] = list(student_dict[student_name].keys())
            except:
                makeup_test_list_combo.set("재시험이 없습니다")
                makeup_test_list_combo["values"] = None
        target_studnet_var = tk.StringVar()
        student_combo = ttk.Combobox(popup, values=None, state="readonly", textvariable=target_studnet_var, width=100)
        student_combo.set("학생 선택")
        student_combo.bind("<<ComboboxSelected>>", student_call_back)
        student_combo.pack()

        makeup_test_name_var = tk.StringVar()
        makeup_test_list_combo = ttk.Combobox(popup, values=None, state="readonly", textvariable=makeup_test_name_var, width=100)
        makeup_test_list_combo.set("재시험 선택")
        makeup_test_list_combo.pack()

        makeup_test_score_var = tk.StringVar()
        tk.Entry(popup, textvariable=makeup_test_score_var, width=28).pack()

        tk.Label(popup).pack()
        tk.Button(popup, text="재시험 저장", width=10 , command=quit_event).pack()
        
        popup.mainloop()

        target_class_name   = target_class_var.get()
        target_student_name = target_studnet_var.get()
        makeup_test_name    = makeup_test_name_var.get()
        makeup_test_score   = makeup_test_score_var.get()
        
        if target_class_name == "반 선택" or target_student_name == "학생 선택" or makeup_test_name == "재시험 선택" or makeup_test_name == "재시험이 없습니다" or makeup_test_score == "":
            return None
        
        row = student_dict[target_student_name][makeup_test_name]
        return row, makeup_test_score
