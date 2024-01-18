import json
import queue
import os.path
import pythoncom # only works in Windows
import threading
import webbrowser
import tkinter as tk
import tkinter.messagebox
import openpyxl as xl
import win32com.client # only works in Windows

from copy import copy
from tkinter import ttk, filedialog
from datetime import date as DATE, datetime, timedelta
from dateutil.relativedelta import relativedelta
from openpyxl.cell import Cell
from openpyxl.utils.cell import get_column_letter as gcl
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Protection, Side
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from win32process import CREATE_NO_WINDOW # only works in Windows
from webdriver_manager.chrome import ChromeDriverManager

import omikronthread
from omikronlog import OmikronLog
from omikrondefs import DataForm, ClassInfo
from omikronconfig import config

service = Service(ChromeDriverManager().install())

# 아이소식과 프로그램
def check_update_class():
    # 반 정보 확인
    class_wb = xl.load_workbook("./반 정보.xlsx")
    try:
        class_ws = class_wb["반 정보"]
    except:
        OmikronLog.error(r"'반 정보.xlsx'의 시트명을 '반 정보'로 변경해 주세요.")
        return

    options = webdriver.ChromeOptions()
    options.add_argument("headless")
    driver = webdriver.Chrome(service = service, options = options)

    OmikronLog.log("아이소식으로부터 반 정보를 업데이트 하는 중...")
    # 아이소식 접속
    driver.get(config["url"])
    table_names = driver.find_elements(By.CLASS_NAME, "style1")
    current_classes = [class_ws.cell(i, ClassInfo.CLASS_NAME_COLUMN).value for i in range(2, class_ws.max_row+1) if class_ws.cell(i, ClassInfo.CLASS_NAME_COLUMN).value is not None]

    unregistered_classes = {table_names[i].text.rstrip() : i for i in range(3, len(table_names)) if not table_names[i].text.rstrip() in current_classes}

    for row in range(class_ws.max_row+1, 1, -1):
        if class_ws.cell(row-1, ClassInfo.CLASS_NAME_COLUMN).value is not None:
            WRITE_LOCATION = row
            break
    for new_class_name in list(unregistered_classes.keys()):
        class_ws.cell(WRITE_LOCATION, ClassInfo.CLASS_NAME_COLUMN).value = new_class_name
        WRITE_LOCATION += 1
    
    # 정렬 및 테두리
    for row in range(1, class_ws.max_row + 1):
        for col in range(1, class_ws.max_column + 1):
            class_ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
            class_ws.cell(row, col).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    
    for row in range(class_ws.max_row, 1, -1):
        if class_ws.cell(row, ClassInfo.CLASS_NAME_COLUMN).value is None:
            class_ws.delete_rows(row)
        else: break
    
    class_wb.save("./temp.xlsx")
    return current_classes, unregistered_classes

# 시험 결과 전송
def send_message(filepath:str, makeup_test_date:dict):
    form_wb = xl.load_workbook(filepath, data_only=True)
    form_ws = form_wb["데일리테스트 기록 양식"]

    options = webdriver.ChromeOptions()
    options.add_experimental_option("detach", True)
    driver = webdriver.Chrome(service=service, options=options)

    student_wb = xl.load_workbook("./학생 정보.xlsx")
    try:
        student_ws = student_wb["학생 정보"]
    except:
        OmikronLog.error("\"학생 정보.xlsx\"의 시트명을 \"학생 정보\"로 변경해 주세요.")
        return
    
    # 아이소식 접속
    driver.get(config["url"])
    driver.execute_script(f"arguments[0].value = '{config['dailyTest']}'", driver.find_element(By.XPATH, '//*[@id="ctitle"]'))
    driver.find_element(By.XPATH, '//*[@id="ctitle"]').send_keys(' \b')

    driver.execute_script(f"window.open('{config['url']}')")
    driver.switch_to.window(driver.window_handles[1])
    driver.execute_script(f"arguments[0].value = '{config['makeupTest']}'", driver.find_element(By.XPATH, '//*[@id="ctitle"]'))
    driver.find_element(By.XPATH, '//*[@id="ctitle"]').send_keys(' \b')

    driver.execute_script(f"window.open('{config['url']}')")
    driver.switch_to.window(driver.window_handles[2])
    driver.execute_script(f"arguments[0].value = '{config['makeupTestDate']}'", driver.find_element(By.XPATH, '//*[@id="ctitle"]'))
    driver.find_element(By.XPATH, '//*[@id="ctitle"]').send_keys(' \b')

    driver.switch_to.window(driver.window_handles[0])
    tables = driver.find_elements(By.CLASS_NAME, "style1")
    table_names = [table.text for table in tables]
    class_search = 0

    OmikronLog.log("메시지 작성 중...")
    for i in range(2, form_ws.max_row+1):
        # 반 필터링
        if form_ws.cell(i, DataForm.CLASS_NAME_COLUMN).value is not None:
            class_name         = form_ws.cell(i, DataForm.CLASS_NAME_COLUMN).value
            daily_test_name    = form_ws.cell(i, DataForm.DAILYTEST_NAME_COLUMN).value
            mock_test_name     = form_ws.cell(i, DataForm.MOCKTEST_NAME_COLUMN).value
            daily_test_average = form_ws.cell(i, DataForm.DAILYTEST_AVERAGE_COLUMN).value
            mock_test_average  = form_ws.cell(i, DataForm.MOCKTEST_AVERAGE_COLUMN).value

            # 반 전체가 시험을 응시하지 않은 경우
            if daily_test_name is None and mock_test_name is None:
                keep_continue = True
                continue

            keep_continue = False
            student_search = 0
            
            # 테이블 인덱스 검색
            for idx in range(class_search, len(table_names)):
                if class_name == table_names[idx]:
                    class_index = idx
                    class_search = idx+1
                    break
                
        
        # 반 전체가 시험을 응시하지 않은 경우
        if keep_continue: continue

        student_name     = form_ws.cell(i, DataForm.STUDENT_NAME_COLUMN).value
        daily_test_score = form_ws.cell(i, DataForm.DAILYTEST_SCORE_COLUMN).value
        mock_test_score  = form_ws.cell(i, DataForm.MOCKTEST_SCORE_COLUMN).value

        # 시험 미응시 시 건너뛰기
        if daily_test_score is not None:
            test_name    = daily_test_name
            test_score   = daily_test_score
            test_average = daily_test_average
        elif mock_test_score is not None:
            test_name    = mock_test_name
            test_score   = mock_test_score
            test_average = mock_test_average
        else:
            continue

        if type(test_score) != int and type(test_score) != float:
            continue

        # 시험 결과 메시지 작성
        driver.switch_to.window(driver.window_handles[0])
        trs = driver.find_element(By.ID, f"table_{str(class_index)}").find_elements(By.CLASS_NAME, "style12")
        for idx in range(student_search, len(trs)):
            if trs[idx].find_element(By.CLASS_NAME, "style9").text == student_name:
                student_index = idx
                student_search = idx+1
                break

        tds = trs[student_index].find_elements(By.TAG_NAME, "td")
        driver.execute_script(f"arguments[0].value = '{test_name}'",  tds[0].find_element(By.TAG_NAME, "input"))
        driver.execute_script(f"arguments[0].value = '{test_score}'", tds[1].find_element(By.TAG_NAME, "input"))
        tds[2].find_element(By.TAG_NAME, "input").send_keys(test_average)

        # 재시험 메시지 작성
        if (test_score < 80) and form_ws.cell(i, DataForm.MAKEUP_TEST_CHECK_COLUMN).value not in ("x", "X"):
            for row in range(2, student_ws.max_row+1):
                if student_ws.cell(row, StudentInfo.STUDENT_NAME_COLUMN).value == student_name:
                    makeup_test_weekday = student_ws.cell(row, StudentInfo.MAKEUPTEST_WEEKDAY_COLUMN).value
                    makeup_test_time    = student_ws.cell(row, StudentInfo.MAKEUPTEST_TIME_COLUMN).value
                    break
            else:
                makeup_test_weekday = None
                makeup_test_time    = None
            
            if makeup_test_weekday is None:
                driver.switch_to.window(driver.window_handles[1])
                trs = driver.find_element(By.ID, "table_" + str(class_index)).find_elements(By.CLASS_NAME, "style12")
                tds = trs[student_index].find_elements(By.TAG_NAME, "td")
                driver.execute_script(f"arguments[0].value = '{test_name}'", tds[0].find_element(By.TAG_NAME, "input"))
                tds[1].find_element(By.TAG_NAME, "input").send_keys(' \b')
            else:
                weekday_list    = makeup_test_weekday.split("/")
                calculated_date = makeup_test_date[weekday_list[0].replace(" ", "")]
                time_index      = 0
                for tmp_idx in range(len(weekday_list)):
                    if calculated_date > makeup_test_date[weekday_list[tmp_idx].replace(" ", "")]:
                        calculated_date = makeup_test_date[weekday_list[tmp_idx].replace(" ", "")]
                        time_index = tmp_idx
                driver.switch_to.window(driver.window_handles[2])
                trs = driver.find_element(By.ID, "table_" + str(class_index)).find_elements(By.CLASS_NAME, "style12")
                tds = trs[student_index].find_elements(By.TAG_NAME, "td")
                driver.execute_script(f"arguments[0].value = '{test_name}'", tds[0].find_element(By.TAG_NAME, "input"))
                try:
                    if makeup_test_time is not None:
                        if "/" in str(makeup_test_time):
                            driver.execute_script(f"arguments[0].value = '{calculated_date.strftime('%m월 %d일'.encode('unicode-escape').decode()).encode().decode('unicode-escape')} {str(makeup_test_time).split('/')[time_index]}시'", tds[1].find_element(By.TAG_NAME, "input"))
                        else:
                            driver.execute_script(f"arguments[0].value = '{calculated_date.strftime('%m월 %d일'.encode('unicode-escape').decode()).encode().decode('unicode-escape')} {str(makeup_test_time)}시'", tds[1].find_element(By.TAG_NAME, "input"))
                    else:
                        driver.execute_script(f"arguments[0].value = '{calculated_date.strftime('%m월 %d일'.encode('unicode-escape').decode()).encode().decode('unicode-escape')}'", tds[1].find_element(By.TAG_NAME, "input"))
                except:
                    OmikronLog.log(f"{student_name}의 재시험 일정을 요일별 시간으로 설정하거나")
                    OmikronLog.log("하나의 시간으로 통일해 주세요.")
                    OmikronLog.log("중단되었습니다.")
                    driver.quit()
                    gui.thread_end_flag = True    
                    return
                tds[2].find_element(By.TAG_NAME, "input").send_keys(' \b')

    OmikronLog.log("메시지 입력을 완료했습니다.")
    OmikronLog.log("메시지 확인 후 전송해주세요.")
    omikronthread.thread_end_flag = True

# 개별 응시 시험 전송
def individual_record_message(student_name:str, class_name:int, test_name:int, test_score:int, makeup_test_date:dict):
    options = webdriver.ChromeOptions()
    options.add_experimental_option("detach", True)
    driver = webdriver.Chrome(service=service, options=options)
    
    # 아이소식 접속
    driver.get(config["url"])
    driver.execute_script(f"arguments[0].value = '{config['dailyTest']}'", driver.find_element(By.XPATH, '//*[@id="ctitle"]'))
    driver.find_element(By.XPATH, '//*[@id="ctitle"]').send_keys(' \b')

    driver.switch_to.window(driver.window_handles[0])
    tables = driver.find_elements(By.CLASS_NAME, "style1")
    table_names = [table.text for table in tables]
    for idx in range(3, len(table_names)):
        if class_name == table_names[idx]:
            class_index = idx
            break
    
    # 시험 결과 메시지 작성
    trs = driver.find_element(By.ID, f"table_{str(class_index)}").find_elements(By.CLASS_NAME, "style12")
    for idx in range(0, len(trs)):
        if trs[idx].find_element(By.CLASS_NAME, "style9").text == student_name:
            student_index = idx
            break

    tds = trs[student_index].find_elements(By.TAG_NAME, "td")
    driver.execute_script(f"arguments[0].value = '{test_name}'",  tds[0].find_element(By.TAG_NAME, "input"))
    driver.execute_script(f"arguments[0].value = '{test_score}'", tds[1].find_element(By.TAG_NAME, "input"))
    tds[2].find_element(By.TAG_NAME, "input").send_keys(test_average)
    
    # 재시험 메시지 작성
    if test_score < 80:
        if makeup_test_weekday is None:
            driver.execute_script(f"window.open('{config['url']}')")
            driver.switch_to.window(driver.window_handles[1])
            driver.execute_script(f"arguments[0].value = '{config['makeupTest']}'", driver.find_element(By.XPATH, '//*[@id="ctitle"]'))
            driver.find_element(By.XPATH, '//*[@id="ctitle"]').send_keys(' \b')
            trs = driver.find_element(By.ID, "table_" + str(class_index)).find_elements(By.CLASS_NAME, "style12")
            tds = trs[student_index].find_elements(By.TAG_NAME, "td")
            driver.execute_script(f"arguments[0].value = '{test_name}'", tds[0].find_element(By.TAG_NAME, "input"))
            tds[1].find_element(By.TAG_NAME, "input").send_keys(' \b')
        else:
            driver.execute_script(f"window.open('{config['url']}')")
            driver.switch_to.window(driver.window_handles[1])
            driver.execute_script(f"arguments[0].value = '{config['makeupTestDate']}'", driver.find_element(By.XPATH, '//*[@id="ctitle"]'))
            driver.find_element(By.XPATH, '//*[@id="ctitle"]').send_keys(' \b')
            trs = driver.find_element(By.ID, "table_" + str(class_index)).find_elements(By.CLASS_NAME, "style12")
            tds = trs[student_index].find_elements(By.TAG_NAME, "td")
            driver.execute_script(f"arguments[0].value = '{test_name}'", tds[0].find_element(By.TAG_NAME, "input"))
            try:
                if makeup_test_time is not None:
                    if "/" in str(makeup_test_time):
                        driver.execute_script(f"arguments[0].value = '{calculated_date.strftime('%m월 %d일'.encode('unicode-escape').decode()).encode().decode('unicode-escape')} {str(makeup_test_time).split('/')[time_index]}시'", tds[1].find_element(By.TAG_NAME, "input"))
                    else:
                        driver.execute_script(f"arguments[0].value = '{calculated_date.strftime('%m월 %d일'.encode('unicode-escape').decode()).encode().decode('unicode-escape')} {str(makeup_test_time)}시'", tds[1].find_element(By.TAG_NAME, "input"))
                else:
                    driver.execute_script(f"arguments[0].value = '{calculated_date.strftime('%m월 %d일'.encode('unicode-escape').decode()).encode().decode('unicode-escape')}'", tds[1].find_element(By.TAG_NAME, "input"))
            except:
                OmikronLog.log(f"{student_name}의 재시험 일정을 요일별 시간으로 설정하거나")
                OmikronLog.log("하나의 시간으로 통일해 주세요.")
                OmikronLog.log("중단되었습니다.")
                driver.quit()
                gui.thread_end_flag = True    
                return
            tds[2].find_element(By.TAG_NAME, "input").send_keys(' \b')

# 아이소식에 등록된 정보에 기반하여
# 특정 반에 특정 학생이 존재하는지 확인
def check_student_exists(target_student_name:str, target_class_name:str) -> bool:
    options = webdriver.ChromeOptions()
    options.add_argument("headless")
    driver = webdriver.Chrome(service = service, options = options)

    # 아이소식 접속
    driver.get(config["url"])
    table_names = driver.find_elements(By.CLASS_NAME, "style1")

    # 반 루프
    for i in range(3, len(table_names)):
        if target_class_name == table_names[i].text.rstrip():
            trs = driver.find_element(By.ID, "table_" + str(i)).find_elements(By.CLASS_NAME, "style12")
            for tr in trs:
                if target_student_name == tr.find_element(By.CLASS_NAME, "style9").text:
                    return True
    return False

if __name__ == "__main__":
    print(check_student_exists("김경모", "계통수학B (23/2기)"))