import os.path
import threading
import tkinter as tk
import tkinter.messagebox
import openpyxl as xl
import win32com.client # only works in Windows

from copy import copy

from tkinter import ttk, filedialog
from datetime import date as DATE, datetime, timedelta
from dateutil.relativedelta import relativedelta
from openpyxl.cell import Cell
from openpyxl.utils.cell import get_column_letter as gcl
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Protection, Side
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from win32process import CREATE_NO_WINDOW # only works in Windows
from webdriver_manager.chrome import ChromeDriverManager

from omikronlog import OmikronLog

thread_end_flag = False

def make_class_info_file_thread():
    OmikronLog.error(r"'반 정보.xlsx'의 시트명을 '반 정보'로 변경해 주세요.")
    # thread = threading.Thread(target=make_class_info_file)
    # thread.daemon = True
    # thread.start()

def make_student_info_file_thread(self):
    thread = threading.Thread(target=lambda: make_student_info_file(self))
    thread.daemon = True
    thread.start()

def make_data_file_thread(self):
    thread = threading.Thread(target=lambda: make_data_file(self))
    thread.daemon = True
    thread.start()

def update_class_thread(self):
    if os.path.isfile(f"./data/~${config['dataFileName']}.xlsx"):
        self.q.put(r"데이터 파일을 닫은 뒤 다시 시도해 주세요.")
        return
    if self.update_class_button["text"] == "반 업데이트":
        if os.path.isfile("./temp.xlsx"):
            os.remove("./temp.xlsx")
        self.update_class_button["state"] = tk.DISABLED
        global ret
        ret = check_update_class(self)
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True
        wb = excel.Workbooks.Open(f"{os.getcwd()}\\temp.xlsx")
        self.update_class_button["text"] = "반 정보 수정 후 반 업데이트 계속하기"
        if not tkinter.messagebox.askokcancel("반 정보 변경 확인", "반 정보 파일의 빈칸을 채운 뒤 Excel을 종료하고\n버튼을 눌러주세요.\n삭제할 반은 행을 삭제해 주세요.\n취소 선택 시 반 업데이트가 중단됩니다."):
            self.update_class_button["text"] = "반 업데이트"
            wb.Close()
            if os.path.isfile("./temp.xlsx"):
                os.remove("./temp.xlsx")
                self.q.put(r"반 업데이트를 중단합니다.")
    else:
        if os.path.isfile("./~$temp.xlsx"):
            self.q.put(r"임시 파일을 닫은 뒤 다시 시도해 주세요.")
            return
        if os.path.isfile(f"./data/~${config['dataFileName']}.xlsx"):
            self.q.put(r"데이터 파일을 닫은 뒤 다시 시도해 주세요.")
            return
        if not tkinter.messagebox.askokcancel("반 정보 변경 확인", "반 업데이트를 계속하시겠습니까?"):
            if os.path.isfile("./temp.xlsx"):
                os.remove("./temp.xlsx")
                self.q.put(r"반 업데이트를 중단합니다.")
        thread = threading.Thread(target=lambda: update_class(self, ret[0], ret[1]))
        thread.daemon = True
        thread.start()
        del ret
        self.update_class_button["text"] = "반 업데이트"
    self.update_class_button["state"] = tk.NORMAL

def make_data_form_thread(self):
    self.make_data_form_button["state"] = tk.DISABLED
    thread = threading.Thread(target=lambda: make_data_form(self))
    thread.daemon = True
    thread.start()
    self.make_data_form_button['state'] = tk.NORMAL

def save_data_thread(self):
    if os.path.isfile("./data/~$재시험 명단.xlsx"):
        self.q.put(r"재시험 명단 파일을 닫은 뒤 다시 시도해 주세요.")
        return
    if os.path.isfile(f"./data/~${config['dataFileName']}.xlsx"):
        self.q.put(r"데이터 파일을 닫은 뒤 다시 시도해 주세요.")
        return
    self.save_data_button["state"] = tk.DISABLED
    if self.makeup_test_date is None:
        self.makeup_test_date = self.holiday_dialog()
    filepath = filedialog.askopenfilename(initialdir="./", title="데일리테스트 기록 양식 선택", filetypes=(("Excel files", "*.xlsx"),("all files", "*.*")))
    if filepath == "": return
    thread = threading.Thread(target=lambda: save_data(self, filepath, self.makeup_test_date))
    thread.daemon = True
    thread.start()
    self.save_data_button["state"] = tk.NORMAL

def send_message_thread(self):
    self.send_message_button["state"] = tk.DISABLED
    if self.makeup_test_date is None:
        self.makeup_test_date = self.holiday_dialog()
    filepath = filedialog.askopenfilename(initialdir="./", title="데일리테스트 기록 양식 선택", filetypes=(("Excel files", "*.xlsx"),("all files", "*.*")))
    if filepath == "": return
    thread = threading.Thread(target=lambda: send_message(self, filepath, self.makeup_test_date))
    thread.daemon = True
    thread.start()
    self.send_message_button["state"] = tk.NORMAL

def individual_record_thread(self):
    if os.path.isfile("./data/~$재시험 명단.xlsx"):
        self.q.put(r"재시험 명단 파일을 닫은 뒤 다시 시도해 주세요.")
        return
    if os.path.isfile(f"./data/~${config['dataFileName']}.xlsx"):
        self.q.put(r"데이터 파일을 닫은 뒤 다시 시도해 주세요.")
        return
    self.individual_record_button["state"] = tk.DISABLED
    if self.makeup_test_date is None:
        self.makeup_test_date = self.holiday_dialog()
    ret = self.individual_record_dialog()
    if ret is None:
        self.individual_record_button["state"] = tk.NORMAL
        return
    student_name, class_name, test_name, row, col, test_score, cell_value = ret
    if cell_value is not None:
        if not tkinter.messagebox.askyesno("데이터 중복 확인", f"{student_name} 학생의 {test_name} 시험에 대한 점수({cell_value}점)가 이미 존재합니다.\n덮어쓰시겠습니까?"):
            self.q.put(r"개별 데이터 저장을 취소하였습니다.")
            self.individual_record_button["state"] = tk.NORMAL
            return
    thread = threading.Thread(target=lambda: individual_record(self, student_name, class_name, test_name, row, col, test_score, self.makeup_test_date))
    thread.daemon = True
    thread.start()
    self.individual_record_button["state"] = tk.NORMAL

def makeup_test_record_thread(self):
    if os.path.isfile("./data/~$재시험 명단.xlsx"):
        self.q.put(r"재시험 명단 파일을 닫은 뒤 다시 시도해 주세요.")
        return
    self.makeup_test_record_button["state"] = tk.DISABLED
    ret = self.makeup_test_record_dialog()
    if ret is None:
        return
    row, makeup_test_score = ret
    makeup_list_wb = xl.load_workbook("./data/재시험 명단.xlsx")
    try:
        makeup_list_ws = makeup_list_wb["재시험 명단"]
    except:
        self.q.put(r"'재시험 명단.xlsx'의 시트명을")
        self.q.put(r"'재시험 명단'으로 변경해 주세요.")
        return
    makeup_list_ws.cell(row, MakeupTestList.MAKEUPTEST_SCORE_COLUMN).value = makeup_test_score
    # makeup_list_ws.sheet_view.topLeftCell = f"A{str(max(1, row-10))}"
    # makeup_list_ws.sheet_view.selection[0].sqref = f"{gcl(MakeupTestList.MAKEUPTEST_SCORE_COLUMN)}{str(row)}"
    makeup_list_wb.save("./data/재시험 명단.xlsx")
    self.q.put(f"{row} 행에 재시험 점수를 기록하였습니다.")
    # excel = win32com.client.Dispatch("Excel.Application")
    # excel.Visible = True
    # excel.Workbooks.Open(f"{os.getcwd()}\\data\\재시험 명단.xlsx")
    self.makeup_test_record_button["state"] = tk.NORMAL

def add_student_thread(self):
    if os.path.isfile(f"./data/~${config['dataFileName']}.xlsx"):
        self.q.put(r"데이터 파일을 닫은 뒤 다시 시도해 주세요.")
        return
    if os.path.isfile("./data/~$학생 정보.xlsx"):
        self.q.put(r"학생 정보 파일을 닫은 뒤 다시 시도해 주세요.")
        return
    self.add_student_button["state"] = tk.DISABLED
    tmp = self.add_student_dialog()
    if tmp is not None:
        # 학생 추가 확인
        if not tkinter.messagebox.askyesno("학생 추가 확인", f"{tmp[0]} 학생을 {tmp[1]} 반에 추가하시겠습니까?"):
            self.add_student_button["state"] = tk.NORMAL
            return
        thread = threading.Thread(target=lambda: add_student(self, tmp[0], tmp[1]))
        thread.daemon = True
        thread.start()
    self.add_student_button["state"] = tk.NORMAL

def delete_student_thread(self):
    if os.path.isfile(f"./data/~${config['dataFileName']}.xlsx"):
        self.q.put(r"데이터 파일을 닫은 뒤 다시 시도해 주세요.")
        return
    if os.path.isfile("./data/~$학생 정보.xlsx"):
        self.q.put(r"학생 정보 파일을 닫은 뒤 다시 시도해 주세요.")
        return
    self.delete_student_button["state"] = tk.DISABLED
    student_name = self.delete_student_name_dialog()
    if student_name is not None:
        # 퇴원 처리 확인
        if not tkinter.messagebox.askyesno("퇴원 확인", f"{student_name} 학생을 퇴원 처리하시겠습니까?"):
            self.delete_student_button["state"] = tk.NORMAL
            return
        thread = threading.Thread(target=lambda: delete_student(self, student_name))
        thread.daemon = True
        thread.start()
    self.delete_student_button["state"] = tk.NORMAL

def move_student_thread(self):
    if os.path.isfile(f"./data/~${config['dataFileName']}.xlsx"):
        self.q.put(r"데이터 파일을 닫은 뒤 다시 시도해 주세요.")
        return
    if os.path.isfile("./data/~$학생 정보.xlsx"):
        self.q.put(r"학생 정보 파일을 닫은 뒤 다시 시도해 주세요.")
        return
    self.move_student_button["state"] = tk.DISABLED
    tmp = self.move_student_dialog()
    if tmp is not None:
        # 학생 반 이동 확인
        if tmp[1] == tmp[2]:
            self.q.put(r"학생의 현재 반과 이동할 반이 같아 취소되었습니다.")
            self.move_student_button["state"] = tk.NORMAL
            return
        if not tkinter.messagebox.askyesno("학생 반 이동 확인", f"{tmp[2]} 반의 {tmp[0]} 학생을\n{tmp[1]} 반으로 이동시키겠습니까?"):
            self.move_student_button["state"] = tk.NORMAL
            return
        thread = threading.Thread(target=lambda: move_student(self, tmp[0], tmp[1], tmp[2]))
        thread.daemon = True
        thread.start()
    self.move_student_button["state"] = tk.NORMAL
