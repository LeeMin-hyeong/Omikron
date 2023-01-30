import json
import openpyxl as xl

config = json.load(open('config.json'))

print('Processing...')

#파일 위치 및 파일명 지정
dailyTestFile = config['dailyTestFilePath'] + config['dailyTestFileName']

dataFileColorWb = xl.load_workbook(filename=dailyTestFile, data_only=True)
dataFileColorWs = dataFileColorWb.active

for i in range(2, dataFileColorWs.max_row+1):
    #학생별 평균 조건부 서식
    print(dataFileColorWs.cell(i, 6).value)