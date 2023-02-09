import json
import os.path
import openpyxl as xl

from datetime import datetime
from tkinter import filedialog
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Alignment, Border, Color, PatternFill, Side

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service

from webdriver_manager.chrome import ChromeDriverManager

config = json.load(open('config.json', encoding='UTF8'))
os.environ['WDM_PROGRESS_BAR'] = '0'
service = Service(ChromeDriverManager().install())

def makeupTestInfo():
    if not os.path.isfile('./makeupTestInfo.xlsx'):
        print('재시험 정보 파일 생성 중...')

        iniWb = xl.Workbook()
        iniWs = iniWb.active
        iniWs.title = 'DailyTest'
        iniWs['A1'] = '이름'
        iniWs['B1'] = '반명'
        iniWs['C1'] = '담당'
        iniWs['D1'] = '요일'
        iniWs['E1'] = '시간'
        iniWs.auto_filter.ref = 'A:C'

        # 반 정보 확인
        if not os.path.isfile('./class.xlsx'):
            print('[오류]class.xlsx 파일이 존재하지 않습니다.')
            return
        classWb = xl.load_workbook("./class.xlsx")
        classWs = classWb.active

        options = webdriver.ChromeOptions()
        options.add_argument('headless')
        driver = webdriver.Chrome(service = service, options = options)

        # 아이소식 접속
        driver.get(config['url'])
        tableNames = driver.find_elements(By.CLASS_NAME, 'style1')

        # 반 루프
        for i in range(3, len(tableNames)):
            trs = driver.find_element(By.ID, 'table_' + str(i)).find_elements(By.CLASS_NAME, 'style12')
            writeLocation = iniWs.max_row + 1

            className = tableNames[i].text.split('(')[0].rstrip()
            for j in range(2, classWs.max_row + 1):
                if classWs.cell(j, 1).value == className:
                    teacher = classWs.cell(j, 2).value

            # 학생 루프
            for tr in trs:
                writeLocation = iniWs.max_row + 1
                iniWs.cell(writeLocation, 1).value = tr.find_element(By.CLASS_NAME, 'style9').text
                iniWs.cell(writeLocation, 2).value = className
                iniWs.cell(writeLocation, 3).value = teacher

        # 정렬 및 테두리
        for j in range(1, iniWs.max_row + 1):
                for k in range(1, iniWs.max_column + 1):
                    iniWs.cell(j, k).alignment = Alignment(horizontal='center', vertical='center')
                    iniWs.cell(j, k).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        iniWb.save('./makeupTestInfo.xlsx')
        print('재시험 정보 파일을 생성했습니다.')

    else:
        print('이미 파일이 존재합니다')

makeupTestInfo()