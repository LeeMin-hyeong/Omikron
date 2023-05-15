# Omikron v1.2.0-alpha
import json
import os.path
import threading
import tkinter as tk
import tkinter.messagebox
import openpyxl as xl
import win32com.client # only works in Windows

from omikronconst import *
from tkinter import ttk, filedialog
from datetime import date as DATE, datetime, timedelta
from dateutil.relativedelta import relativedelta
from openpyxl.utils.cell import get_column_letter as gcl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Color, PatternFill, Side, Font
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from win32process import CREATE_NO_WINDOW
from webdriver_manager.chrome import ChromeDriverManager

config = json.load(open("./config.json", encoding="UTF8"))
os.environ["WDM_PROGRESS_BAR"] = "0"
service = Service(ChromeDriverManager().install())
service.creation_flags = CREATE_NO_WINDOW

class OmikronExcel:
    def __init__(self):
        self.filepath = os.getcwd().replace('\\', '/')
        self.filename = ""
        self.file_open = False
    def open_workbook(self) -> xl.Workbook:
        self.workbook = xl.load_workbook(self.filepath)
        self.file_open = True
        return self.workbook
    def open_worksheet(self) -> Worksheet:
        if not self.file_open:
            self.open_workbook()
        self.worksheet = self.workbook[self.filename]
        return self.worksheet
    def is_exist(self) -> bool:
        return os.path.isfile(self.filepath)
    def require_file(self):
        if not self.is_exist():
            raise Exception(f"{self.filename} 파일이 존재하지 않습니다.")
        # try:
        return self.open_worksheet()
        # except:
        #     raise Exception(f"{self.filename}의 시트명을 {self.filename}으로 변경하세요.")
    def save_file(self, path:str=None):
        if path is None:
            path = self.filepath
        self.workbook.save(path)
        

class DataFile(OmikronExcel):
    TEST_TIME_COLUMN = 1
    DATE_COLUMN = 2
    CLASS_NAME_COLUMN = 3
    TEACHER_COLUMN = 4
    STUDENT_NAME_COLUMN = 5
    AVERAGE_SCORE_COLUMN = 6
    DATA_COLUMN = AVERAGE_SCORE_COLUMN + 1
    MAX = 6
    
    def __init__(self):
        super().__init__()
        self.filepath += f"/data/{config['dataFileName']}.xlsx"
        self.filename = config['dataFileName']
    def require_file(self): # datafile should be opened in Workbook. It has two worksheets.
        if not self.is_exist():
            raise Exception(f"{self.filename} 파일이 존재하지 않습니다.")
        return self.open_workbook()

class DataForm(OmikronExcel):
    DATE_COLUMN = 1
    TEST_TIME_COLUMN = 2
    CLASS_NAME_COLUMN = 3
    STUDENT_NAME_COLUMN = 4
    TEACHER_COLUMN = 5
    DAILYTEST_TEST_NAME_COLUMN = 6
    DAILYTEST_SCORE_COLUMN = 7
    DAILYTEST_AVERAGE_COLUMN = 8
    MOCKTEST_TEST_NAME_COLUMN = 9
    MOCKTEST_SCORE_COLUMN = 10
    MOCKTEST_AVERAGE_COLUMN = 11
    MAKEUP_TEST_CHECK_COLUMN = 12
    MAX = 12

    def open_workbook(self) -> xl.Workbook:
        self.filepath = filedialog.askopenfilename(initialdir="./", title="데일리테스트 기록 양식 선택", filetypes=(("Excel files", "*.xlsx"),("all files", "*.*")))
        self.filename = "데일리테스트 기록 양식"
        self.workbook = xl.load_workbook(self.filepath)
        self.file_open = True
        return self.workbook
    def form_validation(self):
        if not self.file_open: raise Exception("파일이 열리지 않았습니다.")
        if (self.worksheet[gcl(DataForm.DATE_COLUMN)+"1"].value != "요일") or \
                (self.worksheet[gcl(DataForm.TEST_TIME_COLUMN)+"1"].value != "시간") or \
                    (self.worksheet[gcl(DataForm.CLASS_NAME_COLUMN)+"1"].value != "반") or \
                        (self.worksheet[gcl(DataForm.STUDENT_NAME_COLUMN)+"1"].value != "이름") or \
                            (self.worksheet[gcl(DataForm.TEACHER_COLUMN)+"1"].value != "담당T") or \
                                (self.worksheet[gcl(DataForm.DAILYTEST_TEST_NAME_COLUMN)+"1"].value != "시험명") or \
                                    (self.worksheet[gcl(DataForm.DAILYTEST_SCORE_COLUMN)+"1"].value != "점수") or \
                                        (self.worksheet[gcl(DataForm.DAILYTEST_AVERAGE_COLUMN)+"1"].value != "평균") or \
                                            (self.worksheet[gcl(DataForm.MOCKTEST_TEST_NAME_COLUMN)+"1"].value != "시험대비 모의고사명") or \
                                                (self.worksheet[gcl(DataForm.MOCKTEST_SCORE_COLUMN)+"1"].value != "모의고사 점수") or \
                                                    (self.worksheet[gcl(DataForm.MOCKTEST_AVERAGE_COLUMN)+"1"].value != "모의고사 평균") or \
                                                        (self.worksheet[gcl(DataForm.MAKEUP_TEST_CHECK_COLUMN)+"1"].value != "재시문자 X"):
            return False

class MakeupTestList(OmikronExcel):
    TEST_DATE_COLUMN = 1
    CLASS_NAME_COLUMN = 2
    TEACHER_COLUMN = 3
    STUDENT_NAME_COLUMN = 4
    TEST_NAME_COLUMN = 5
    TEST_SCORE_COLUMN = 6
    MAKEUP_TEST_WEEK_DATE_COLUMN = 7
    MAKEUP_TEST_TIME_COLUMN = 8
    MAKEUP_TEST_DATE_COLUMN = 9
    MAKEUP_TEST_SCORE_COLUMN = 10
    ETC_COLUMN = 11
    MAX = 11

    def __init__(self):
        super().__init__()
        self.filepath += "/data/재시험 명단.xlsx"
        self.filename = "재시험 명단"
        print(self.filepath)

class ClassInfo(OmikronExcel):
    CLASS_NAME_COLUMN = 1
    TEACHER_COLUMN = 2
    DATE_COLUMN = 3
    TEST_TIME_COLUMN = 4
    MAX = 4

    def __init__(self):
        super().__init__()
        self.filepath += "/반 정보.xlsx"
        self.filename = "반 정보"

class StudentInfo(OmikronExcel):
    STUDENT_NAME_COLUMN = 1
    CLASS_NAME_COLUMN = 2
    TEACHER_COLUMN = 3
    MAKEUP_TEST_WEEK_DATE_COLUMN = 4
    MAKEUP_TEST_TIME_COLUMN = 5
    NEW_STUDENT_CHECK_COLUMN = 6
    MAX = 6

    def __init__(self):
        super().__init__()
        self.filepath += "/학생 정보.xlsx"
        self.filename = "학생 정보"
