import json
import openpyxl as xl
import win32com.client # only works in Windows

from datetime import datetime
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Border, Color, PatternFill, Side

config = json.load(open('config.json', encoding='UTF8'))

print('Saving Data...')

# 파일 위치 및 파일명 지정
dailyTestFile = config['dailyTestFilePath'] + config['dailyTestFileName']

# 입력 양식 엑셀
formWb = xl.load_workbook("dailyTestForm.xlsx", data_only=True)
formWs = formWb.active
# 데이터 저장 엑셀
dataFileWb = xl.load_workbook(dailyTestFile)
dataFileWs = dataFileWb.active

# 데이터 날짜 내림차순
writeColumn = 7
if str(dataFileWs.cell(1, writeColumn).value) != datetime.today().strftime('%Y.%m.%d'):
    dataFileWs.insert_cols(writeColumn)
    for i in range(2, dataFileWs.max_row+1):
        if dataFileWs.cell(i, 5).value == '시험 평균':
            dataFileWs.cell(i, 7).border = Border(bottom=Side(border_style='medium', color='000000'))

# 데이터 날짜 오름차순
# writeColumn = dataFileWs.max_column+1  

dataFileWs.cell(1, writeColumn).value = datetime.today().strftime('%Y.%m.%d')

for i in range(2, formWs.max_row+1):
    # 양식 파일 끝 검사
    if formWs.cell(i, 4).value is None:
        break
    
    # 반 필터링
    if formWs.cell(i, 3).value is not None: # form className is not None
        className = str(formWs.cell(i, 3).value)
        testName = str(formWs.cell(i, 6).value)
        if formWs.cell(i, 6).value is None:
            continue

        for j in range(2, dataFileWs.max_row+1):
            if str(dataFileWs.cell(j, 3).value) == className: # data className == form className
                dataFileWs.cell(j, writeColumn).value = testName
                start = j+1
                break
        
        for k in range(start, dataFileWs.max_row+1):
            if str(dataFileWs.cell(k, 5).value) == '시험 평균': # data name is 시험 평균
                dataFileWs.cell(k, writeColumn).value = '=ROUND(AVERAGE(' + get_column_letter(writeColumn) + str(start) + ':' + get_column_letter(writeColumn) + str(k-1) + '), 0)'
                end = k
                break

    #미응시자(점수 None) 필터링
    if formWs.cell(i, 7).value is None: # form score is None
        continue

    for j in range(start, end):
        if dataFileWs.cell(j, 5).value == formWs.cell(i, 4).value: # data name == form name
            dataFileWs.cell(j, writeColumn).value = formWs.cell(i, 7).value
            break

print('Clear Data From...')
formWb = xl.load_workbook("dailyTestForm.xlsx")
formWs = formWb.active
for i in range(2, formWs.max_row + 1):
    formWs.cell(i, 6).value = ''
    formWs.cell(i, 7).value = ''
    formWs.cell(i, 8).value = ''
    formWs.cell(i, 9).value = ''
formWb.save('./dailyTestForm.xlsx')

dataFileWb.save(dailyTestFile)

# 조건부서식
excel = win32com.client.Dispatch('Excel.Application')
excel.Visible = False
wb = excel.Workbooks.Open('C:/Users/lmhst/git/Omikron/data/dailyData(23_1).xlsx')
wb.Save()
wb.Close()
excel.Quit()

dataFileWb = xl.load_workbook(dailyTestFile)
dataFileWs = dataFileWb.active
dataFileColorWb = xl.load_workbook(filename=dailyTestFile, data_only=True)
dataFileColorWs = dataFileColorWb.active

for i in range(2, dataFileColorWs.max_row+1):
    # 입력 데이터 조건부 서식
    if type(dataFileColorWs.cell(i, writeColumn).value) == int:
        if dataFileColorWs.cell(i, writeColumn).value < 60:
            dataFileWs.cell(i, writeColumn).fill = PatternFill(fill_type='solid', fgColor=Color('EC7E31'))
        elif dataFileColorWs.cell(i, writeColumn).value < 70:
            dataFileWs.cell(i, writeColumn).fill = PatternFill(fill_type='solid', fgColor=Color('F5AF85'))
        elif dataFileColorWs.cell(i, writeColumn).value < 80:
            dataFileWs.cell(i, writeColumn).fill = PatternFill(fill_type='solid', fgColor=Color('FCE4D6'))
        elif dataFileColorWs.cell(i, 5).value == '시험 평균':
            dataFileWs.cell(i, 7).fill = PatternFill(fill_type='solid', fgColor=Color('DDEBF7'))

    # 학생별 평균 조건부 서식
    if type(dataFileColorWs.cell(i, 6).value) == int:
        if dataFileColorWs.cell(i, 6).value < 60:
            dataFileWs.cell(i, 6).fill = PatternFill(fill_type='solid', fgColor=Color('EC7E31'))
        elif dataFileColorWs.cell(i, 6).value < 70:
            dataFileWs.cell(i, 6).fill = PatternFill(fill_type='solid', fgColor=Color('F5AF85'))
        elif dataFileColorWs.cell(i, 6).value < 80:
            dataFileWs.cell(i, 6).fill = PatternFill(fill_type='solid', fgColor=Color('FCE4D6'))
        elif dataFileColorWs.cell(i, 5).value == '시험 평균':
            dataFileWs.cell(i, 6).fill = PatternFill(fill_type='solid', fgColor=Color('DDEBF7'))
        else:
            dataFileWs.cell(i, 6).fill = PatternFill(fill_type='solid', fgColor=Color('E2EFDA'))

dataFileWb.save(dailyTestFile)

print('Done')
wb = excel.Workbooks.Open('C:/Users/lmhst/git/Omikron/data/dailyData(23_1).xlsx')