import json
import os.path
import openpyxl as xl

from openpyxl.styles import Alignment, Border, Side

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service

from webdriver_manager.chrome import ChromeDriverManager

config = json.load(open('config.json', encoding='UTF8'))

# if True:
if not os.path.isfile('class.xlsx'):
    print('Making Class Info Form...')

    iniWb = xl.Workbook()
    iniWs = iniWb.active
    iniWs.title = 'DailyTestForm'
    iniWs['A1'] = '반명'
    iniWs['B1'] = '선생님명'
    iniWs['C1'] = '요일'
    iniWs['D1'] = '시간'

    options = webdriver.ChromeOptions()
    options.add_argument('headless')
    driver = webdriver.Chrome(service = Service(ChromeDriverManager().install()), options = options)

    # 아이소식 접속
    driver.get(config['url'])
    tableNames = driver.find_elements(By.CLASS_NAME, 'style1')

    # 반 루프
    for i in range(3, len(tableNames)):
        trs = driver.find_element(By.ID, 'table_' + str(i)).find_elements(By.CLASS_NAME, 'style12')
        writeLocation = start = iniWs.max_row + 1
        iniWs.cell(writeLocation, 1).value = tableNames[i].text.split('(')[0].rstrip()

    # 정렬 및 테두리
    for j in range(1, iniWs.max_row + 1):
            for k in range(1, iniWs.max_column + 1):
                iniWs.cell(j, k).alignment = Alignment(horizontal='center', vertical='center')
                iniWs.cell(j, k).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    iniWb.save('./class.xlsx')
    print('Done')

else:
    print('class.xlsx already exists')